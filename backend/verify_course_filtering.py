#!/usr/bin/env python3
import openpyxl

# Check sem5_CSE
print("\n" + "=" * 80)
print("SEM 5 CSE TIMETABLE:")
print("=" * 80)

wb = openpyxl.load_workbook('output_timetables/sem5_CSE_timetable.xlsx')
ws = wb['Course_Information']

core_courses = []
elective_courses = []

in_core = False
in_elective = False

for row in ws.iter_rows(min_row=1, max_row=100, values_only=True):
    if row[0] is None:
        continue
    
    row_str = str(row[0]).upper()
    
    if 'CORE COURSES' in row_str:
        in_core, in_elective = True, False
        continue
    if 'ELECTIVE COURSES' in row_str:
        in_core, in_elective = False, True
        continue
    if 'MINOR' in row_str:
        break
    
    if in_core and row[0] and row[0] not in ['Course Code']:
        if len(str(row[0])) > 2 and not str(row[0]).upper().startswith(('L-T-P', 'COURSE')):
            core_courses.append(str(row[0]))
    
    if in_elective and row[0] and row[0] not in ['Course Code']:
        if len(str(row[0])) > 2 and not str(row[0]).upper().startswith(('L-T-P', 'COURSE')):
            elective_courses.append(str(row[0]))

core_courses = sorted(set(core_courses))
elective_courses = sorted(set(elective_courses))

print(f"Core courses: {core_courses}")
print(f"Elective courses: {elective_courses}")

# Also check sem1_CSE for comparison
print("\n" + "=" * 80)
print("SEM 1 CSE TIMETABLE (for comparison):")
print("=" * 80)

wb1 = openpyxl.load_workbook('output_timetables/sem1_CSE_timetable.xlsx')
ws1 = wb1['Course_Information']

core_courses_1 = []
elective_courses_1 = []

in_core = False
in_elective = False

for row in ws1.iter_rows(min_row=1, max_row=100, values_only=True):
    if row[0] is None:
        continue
    
    row_str = str(row[0]).upper()
    
    if 'CORE COURSES' in row_str:
        in_core, in_elective = True, False
        continue
    if 'ELECTIVE COURSES' in row_str:
        in_core, in_elective = False, True
        continue
    if 'MINOR' in row_str:
        break
    
    if in_core and row[0] and row[0] not in ['Course Code']:
        if len(str(row[0])) > 2 and not str(row[0]).upper().startswith(('L-T-P', 'COURSE')):
            core_courses_1.append(str(row[0]))
    
    if in_elective and row[0] and row[0] not in ['Course Code']:
        if len(str(row[0])) > 2 and not str(row[0]).upper().startswith(('L-T-P', 'COURSE')):
            elective_courses_1.append(str(row[0]))

core_courses_1 = sorted(set(core_courses_1))
elective_courses_1 = sorted(set(elective_courses_1))

print(f"Core courses: {core_courses_1}")
print(f"Elective courses: {elective_courses_1}")

print("\n" + "=" * 80)
print("✅ VERIFICATION:")
print(f"Sem 1 and Sem 5 have DIFFERENT core courses: {core_courses != core_courses_1}")
print(f"Sem 1 and Sem 5 have DIFFERENT elective courses: {elective_courses != elective_courses_1}")
print("=" * 80)
