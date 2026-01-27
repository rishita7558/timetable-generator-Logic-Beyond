"""Detailed EC261 analysis"""
import openpyxl

wb = openpyxl.load_workbook(r'C:\timetable-generator-Logic-Beyond\timetable-generator-Logic-Beyond\backend\output_timetables\sem3_ECE_timetable.xlsx')
ws = wb['Regular_Timetable']

print("EC261 occurrence by (slot, day):")
print("="*80)

for row_idx in range(1, 20):
    for col_idx in range(1, 5):
        cell = ws.cell(row=row_idx, column=col_idx)
        if cell.value and isinstance(cell.value, str) and 'EC261' in cell.value:
            time_slot = ws.cell(row=row_idx, column=1).value
            day = ['', 'Mon', 'Tue', 'Wed', 'Thu'][col_idx]
            print(f"  Row {row_idx:2d} (slot={time_slot:15s}), Col {col_idx} ({day:3s}): {str(cell.value)[:40]}")

print("\n" + "="*80)
print("\nManual counting:")
print("="*80)

# Manual count
l_count = 0
t_count = 0
p_count = 0

for row_idx in range(1, 20):
    for col_idx in range(1, 5):
        cell = ws.cell(row=row_idx, column=col_idx)
        if cell.value and isinstance(cell.value, str) and 'EC261' in cell.value:
            cell_str = str(cell.value).lower()
            time_slot = str(ws.cell(row=row_idx, column=1).value).lower()
            day = ['', 'Mon', 'Tue', 'Wed', 'Thu'][col_idx]
            
            print(f"\n  [{row_idx}] {time_slot:15s} {day}: {cell.value}")
            
            if '(lab)' in cell_str:
                print(f"       -> LAB (1 hour)")
                p_count += 1
            elif '(tutorial)' in cell_str:
                print(f"       -> TUTORIAL (1 hour)")
                t_count += 1
            else:
                # Check if tutorial time
                if '14:30-15:30' in time_slot or '17:00-18:00' in time_slot or '18:00-18:30' in time_slot or '18:30-20:00' in time_slot:
                    print(f"       -> TUTORIAL (time-based)")
                    t_count += 1
                else:
                    print(f"       -> LECTURE (1.5 hours)")
                    l_count += 1

print(f"\nFinal counts: L={l_count}, T={t_count}, P={p_count}")
print(f"Required:     L=2, T=1, P=0")
