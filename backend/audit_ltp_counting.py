"""Comprehensive audit of lecture, lab, and tutorial counting"""
import openpyxl
import pandas as pd
from pathlib import Path

files = [
    'output_timetables/sem1_CSE_timetable.xlsx',
    'output_timetables/sem1_DSAI_timetable.xlsx',
    'output_timetables/sem1_ECE_timetable.xlsx',
    'output_timetables/sem3_CSE_timetable.xlsx',
    'output_timetables/sem3_DSAI_timetable.xlsx',
    'output_timetables/sem3_ECE_timetable.xlsx',
    'output_timetables/sem5_CSE_timetable.xlsx',
    'output_timetables/sem5_DSAI_timetable.xlsx',
    'output_timetables/sem5_ECE_timetable.xlsx',
    'output_timetables/sem7_CSE_timetable.xlsx',
    'output_timetables/sem7_DSAI_timetable.xlsx',
    'output_timetables/sem7_ECE_timetable.xlsx',
]

print("="*80)
print("COMPREHENSIVE AUDIT OF L/T/P COUNTING")
print("="*80)

issues_found = []

for file_path in files:
    file_name = Path(file_path).stem
    print(f"\n{'='*80}")
    print(f"CHECKING: {file_name}")
    print(f"{'='*80}")
    
    try:
        wb = openpyxl.load_workbook(file_path)
        
        # Find the timetable sheet (different names for different branches)
        timetable_sheet = None
        for sheet_name in ['Regular_Section_A', 'Regular_Timetable']:
            if sheet_name in wb.sheetnames:
                timetable_sheet = sheet_name
                break
        
        if not timetable_sheet:
            print(f"  ❌ No timetable sheet found. Available sheets: {wb.sheetnames}")
            continue
        
        ws = wb[timetable_sheet]
        
        # Find CORE COURSES legend
        core_row = None
        for row_idx in range(1, 100):
            cell = ws.cell(row=row_idx, column=1)
            if cell.value and 'CORE COURSES' in str(cell.value).upper():
                core_row = row_idx
                break
        
        if not core_row:
            print(f"  ❌ No CORE COURSES section found")
            continue
        
        header_row = core_row + 1
        headers = [ws.cell(row=header_row, column=i).value for i in range(1, 8)]
        
        # Verify headers
        expected_headers = ['Course Code', 'Course Name', 'L-T-P-S-C', 'Term Type', 'Lectures Hrs', 'Tutorials Hrs', 'Labs Hrs']
        if headers != expected_headers:
            print(f"  ⚠️  Header mismatch!")
            print(f"      Expected: {expected_headers}")
            print(f"      Got:      {headers}")
            issues_found.append(f"{file_name}: Header mismatch")
        
        # Check each course
        print(f"\n  Courses:")
        course_count = 0
        problematic = []
        
        for data_row in range(header_row + 1, header_row + 50):
            course_code = ws.cell(row=data_row, column=1).value
            if not course_code or not isinstance(course_code, str):
                break
            
            ltpsc = ws.cell(row=data_row, column=3).value
            lectures_hrs = ws.cell(row=data_row, column=5).value
            tutorials_hrs = ws.cell(row=data_row, column=6).value
            labs_hrs = ws.cell(row=data_row, column=7).value
            
            # Check format
            issues = []
            
            # Parse LTPSC
            try:
                parts = str(ltpsc).split('-')
                req_l = int(parts[0]) if len(parts) > 0 else 0
                req_t = int(parts[1]) if len(parts) > 1 else 0
                req_p = int(parts[2]) if len(parts) > 2 else 0
            except:
                req_l, req_t, req_p = 0, 0, 0
            
            # Parse scheduled/required format
            try:
                if lectures_hrs and '/' in str(lectures_hrs):
                    sch_l, req_l_check = map(int, str(lectures_hrs).split('/'))
                else:
                    sch_l = 0
                    
                if tutorials_hrs and '/' in str(tutorials_hrs):
                    sch_t, req_t_check = map(int, str(tutorials_hrs).split('/'))
                else:
                    sch_t = 0
                    
                if labs_hrs and '/' in str(labs_hrs):
                    sch_p, req_p_check = map(int, str(labs_hrs).split('/'))
                else:
                    sch_p = 0
            except Exception as e:
                issues.append(f"Format error: {e}")
                sch_l, sch_t, sch_p = 0, 0, 0
            
            # Check if scheduled <= required
            if sch_l > req_l:
                issues.append(f"L: {sch_l} > {req_l}")
            if sch_t > req_t:
                issues.append(f"T: {sch_t} > {req_t}")
            if sch_p > req_p:
                issues.append(f"P: {sch_p} > {req_p}")
            
            # Check for zeros when content exists
            if req_l > 0 and sch_l == 0:
                issues.append(f"L: requires {req_l} but scheduled 0")
            if req_t > 0 and sch_t == 0:
                issues.append(f"T: requires {req_t} but scheduled 0")
            if req_p > 0 and sch_p == 0:
                issues.append(f"P: requires {req_p} but scheduled 0")
            
            status = "✓" if not issues else "⚠️"
            print(f"    {status} {course_code}: L={lectures_hrs}, T={tutorials_hrs}, P={labs_hrs}", end="")
            
            if issues:
                print(f"  [{', '.join(issues)}]")
                problematic.append(f"    {course_code}: {issues}")
            else:
                print()
            
            course_count += 1
        
        print(f"\n  Total courses: {course_count}")
        if problematic:
            print(f"  ⚠️  Issues found:")
            for p in problematic:
                print(p)
                issues_found.append(f"{file_name}: {p}")
    
    except Exception as e:
        print(f"  ❌ Error processing file: {e}")
        issues_found.append(f"{file_name}: {e}")

print(f"\n{'='*80}")
print("SUMMARY")
print(f"{'='*80}")
if issues_found:
    print(f"\n❌ Issues found ({len(issues_found)}):")
    for issue in issues_found:
        print(f"  - {issue}")
else:
    print(f"\n✅ ALL CHECKS PASSED!")
    print(f"All lectures, labs, and tutorials are properly counted.")
