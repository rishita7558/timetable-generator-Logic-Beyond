"""Debug sem3_ECE timetable issue"""
import openpyxl

wb = openpyxl.load_workbook(r'C:\timetable-generator-Logic-Beyond\timetable-generator-Logic-Beyond\backend\output_timetables\sem3_ECE_timetable.xlsx')
ws = wb['Regular_Timetable']

print("Timetable structure:")
print("="*80)

# Show the timetable
for row_idx in range(1, 15):
    values = [ws.cell(row=row_idx, column=i).value for i in range(1, 6)]
    if any(values):
        print(f"Row {row_idx:2d}: {values}")

print("\n" + "="*80)
print("\nProblematic courses analysis:")
print("="*80)

test_courses = ['CS307', 'EC261', 'MA261', 'MA262']

for course_code in test_courses:
    print(f"\n{course_code}:")
    print("-" * 40)
    
    # Find all occurrences
    found = []
    for row_idx in range(1, 20):
        for col_idx in range(1, 6):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value and isinstance(cell.value, str) and course_code in cell.value:
                time_slot = ws.cell(row=row_idx, column=1).value
                col_name = ws.cell(row=1, column=col_idx).value if col_idx > 1 else time_slot
                day = ['', 'Mon', 'Tue', 'Wed', 'Thu', 'Fri'][col_idx]
                print(f"  [{row_idx:2d}] {str(time_slot):20s} {day:5s}: {str(cell.value)[:50]}")
                found.append({
                    'slot': time_slot,
                    'day': day,
                    'value': str(cell.value)
                })
    
    if not found:
        print("  NOT FOUND IN TIMETABLE")
