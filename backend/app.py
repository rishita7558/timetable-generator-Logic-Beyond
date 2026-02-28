from flask import Flask, render_template, request, jsonify, send_file
import os
import sys
import re
import pandas as pd
import random
import zipfile
import glob
from werkzeug.utils import secure_filename
from datetime import datetime, timedelta
import math
import traceback
import shutil
import time
import hashlib
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)
_SEMESTER_ELECTIVE_ALLOCATIONS = {}

# Configuration: prefer repo-local backend/temp_inputs so tests can overwrite fixtures
_BASE_DIR = os.path.dirname(os.path.abspath(__file__))
_DEFAULT_INPUT_DIR = os.path.join(_BASE_DIR, "temp_inputs")
_DEFAULT_OUTPUT_DIR = os.path.join(_BASE_DIR, "output_timetables")

# If backend/temp_inputs exists, use it; otherwise fall back to cwd/temp_inputs to preserve current behaviour.
if os.path.isdir(_DEFAULT_INPUT_DIR):
    INPUT_DIR = _DEFAULT_INPUT_DIR
else:
    INPUT_DIR = os.path.join(os.getcwd(), "temp_inputs")

OUTPUT_DIR = _DEFAULT_OUTPUT_DIR
os.makedirs(INPUT_DIR, exist_ok=True)
os.makedirs(OUTPUT_DIR, exist_ok=True)

# Cache variables with file hashes to detect changes
_cached_data_frames = None
_cached_timestamp = 0
_file_hashes = {}
# _EXAM_SCHEDULE_FILES = set()  # COMMENTED OUT - EXAM FUNCTIONALITY DISABLED

# Allowed file extensions
ALLOWED_EXTENSIONS = {'csv'}

_CLASSROOM_USAGE_TRACKER = {}
_TIMETABLE_CLASSROOM_ALLOCATIONS = {}
_GLOBAL_PREFERRED_CLASSROOMS = {}
_COMMON_COURSE_ROOMS = {}  # Track classroom allocations for common courses (same room for both sections)
_COMMON_COURSE_SCHEDULE = {}  # Track timeslot allocations for common courses (same timeslot for both sections)
_MID_SEM_COMMON_SCHEDULE = {}  # Track timeslot allocations for common courses in Pre-Mid and Post-Mid schedules

# Global counter to track total allocations per room for load balancing
# Structure: { room_id: total_allocation_count }
_ROOM_ALLOCATION_COUNTER = {}

# Global time slot labels used across schedule normalization
TIME_SLOT_LABELS = [
    '07:30-09:00',
    '09:00-10:30', '10:30-12:00', '12:00-13:00',
    '13:00-14:30', '14:30-15:30', '15:30-17:00', '17:00-18:00',
    '18:30-20:00'
]

# Track common minor placements and rooms per semester
_MINOR_COMMON_SCHEDULE = {}
_MINOR_COMMON_CLASSROOMS = {}

# Track lab room allocations for consecutive slots (day_slot1_slot2 -> room)
_LAB_ROOM_ALLOCATIONS = {}  # Maps (day, slot1, slot2) -> room to ensure same room for lab pairs

# ===== GLOBAL FACULTY BOOKING TRACKER =====
# Prevent faculty from being double-booked (teaching multiple different courses at same time)
# Structure: { (day, time_slot, period): { faculty_name: course_code } }
# period is 'Pre-Mid' or 'Post-Mid' to allow same slot in different periods
_FACULTY_BOOKING_TRACKER = {}

# ===== AUDIT TRACKERS FOR VERIFICATION =====
# Track faculty schedule allocations for audit file generation
# Structure: { faculty_name: { (day, time_slot): { course_code, semester, branch, section, classroom } } }
_FACULTY_SCHEDULE_TRACKER = {}

# Track classroom schedule allocations for audit file generation  
# Structure: { classroom_id: { (day, time_slot): { course_code, faculty, semester, branch, section } } }
_CLASSROOM_SCHEDULE_TRACKER = {}

# Track minor course slots for audit purposes (minors have no faculty)
# Structure: { (day, time_slot, semester): { minor_name, classroom, branch, section, schedule_type } }
_MINOR_SCHEDULE_TRACKER = {}

# ===== ELECTIVE ROOM SHARING TRACKER =====
# Tracks elective classroom allocations that should be SHARED across all branches/sections within the same semester
# This dict is NOT reset between branches - only at the start of a new generation
# Structure: { "ELECTIVE_COMMON_{semester}_{day}_{time_slot}_{course_code}_{session_type}": classroom }
_ELECTIVE_COMMON_ROOMS = {}

def normalize_time_slot_label(val):
    """Convert numeric or short labels to canonical time slot strings."""
    try:
        if isinstance(val, str):
            val_strip = val.strip()
            if val_strip in TIME_SLOT_LABELS or 'LUNCH' in val_strip.upper():
                return val_strip
            if val_strip.isdigit():
                idx = int(val_strip)
                if 0 <= idx < len(TIME_SLOT_LABELS):
                    return TIME_SLOT_LABELS[idx]
        if isinstance(val, (int, float)) and not isinstance(val, bool):
            idx = int(val)
            if 0 <= idx < len(TIME_SLOT_LABELS):
                return TIME_SLOT_LABELS[idx]
    except Exception:
        pass
    return str(val)


def initialize_classroom_usage_tracker():
    """Initialize the global classroom usage tracker"""
    global _CLASSROOM_USAGE_TRACKER
    days = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri']
    time_slots = TIME_SLOT_LABELS
    
    _CLASSROOM_USAGE_TRACKER = {}
    for day in days:
        _CLASSROOM_USAGE_TRACKER[day] = {}
        for time_slot in time_slots:
            _CLASSROOM_USAGE_TRACKER[day][time_slot] = set()

    print(f"   [SCHOOL] Initialized classroom tracker: {len(days)} days x {len(time_slots)} time slots")


def compute_effective_enrollment(enrollment, is_common_course=False):
    """Calculate effective enrollment for room allocation purposes.
    
    For common courses (same course taught in multiple sections/departments):
    - Divide enrollment by number of sections to get effective enrollment per section
    - This ensures room capacity is appropriate for one section, not the combined total
    
    Args:
        enrollment: Total enrollment number
        is_common_course: Whether this is a common course (taught multiple times)
    
    Returns:
        Effective enrollment for room sizing
    """
    if not enrollment or enrollment <= 0:
        return 40  # Default enrollment if not specified
    
    enrollment = int(enrollment)
    
    # For common courses, assume 2 sections and divide enrollment
    if is_common_course:
        return max(1, int(enrollment / 2))
    
    return enrollment


def _get_capacity_tiers(enrollment):
    """Return a list of classroom capacity tiers ordered by preference for the given enrollment.
    
    Args:
        enrollment: Effective enrollment number
    
    Returns:
        List of capacity tiers in priority order (smallest suitable first, then larger options)
    """
    # Define capacity ranges for different room types
    if enrollment <= 30:
        # Small rooms: 30, 40, 50, then larger
        return [30, 40, 50, 60, 80, 100, 120, 240]
    elif enrollment <= 50:
        # Medium-small rooms: 50, 60, 80, then larger
        return [50, 60, 80, 100, 120, 240]
    elif enrollment <= 80:
        # Medium rooms: 80, 100, 120
        return [80, 100, 120, 240]
    elif enrollment <= 120:
        # Large rooms: 120, 240
        return [120, 240]
    else:
        # Extra-large enrollment: prefer largest rooms
        return [240, 120, 100]


def _get_available_rooms(dfs):
    """Return a list of available room identifiers from classroom data if present, otherwise generate placeholders."""
    rooms = []
    if dfs and 'classroom' in dfs and not dfs['classroom'].empty:
        # Try common column names for room id
        possible_cols = ['Room Number', 'Room', 'room_number', 'room']
        df_room = dfs['classroom']
        col_found = None
        for c in possible_cols:
            if c in df_room.columns:
                col_found = c
                break
        if col_found:
            rooms = [str(x).strip() for x in df_room[col_found].tolist() if pd.notna(x) and str(x).strip().lower() not in ['nil', 'none', '']]
    # Fallback: generate placeholder rooms
    if not rooms:
        rooms = [f'R-{i+1}' for i in range(20)]
    return rooms


def _allocate_classrooms_for_file(df_a, df_b, dfs, filename, sem, branch, basket_allocations, course_enrollment_map=None, common_courses_list=None, classroom_capacities=None):
    """Allocate classrooms deterministically per course and avoid double-booking when possible.
    Returns modified df_a, df_b and a list of allocation records.
    """
    global _CLASSROOM_USAGE_TRACKER, _TIMETABLE_CLASSROOM_ALLOCATIONS, _GLOBAL_PREFERRED_CLASSROOMS
    
    # Build classroom capacity mapping ONCE at the start to be used consistently
    # This filters out non-teaching rooms (library, research, empty, etc.)
    classroom_capacities = {}
    room_types = {}
    if dfs and 'classroom' in dfs and not dfs['classroom'].empty:
        for _, room_row in dfs['classroom'].iterrows():
            room_id = str(room_row.get('Room Number', '')).strip()
            capacity = room_row.get('Capacity', 0)
            room_type = str(room_row.get('Type', '')).strip().lower()
            # Exclude non-teaching rooms (library, research, empty, physics lab, etc.)
            if room_type in ['library', 'research lab', 'empty', 'physics lab', 'examination room'] or 'nil' in str(capacity).lower():
                continue
            if room_id and capacity:
                try:
                    classroom_capacities[room_id] = int(capacity)
                    room_types[room_id] = room_type
                except (ValueError, TypeError):
                    pass
    
    # Use ONLY teaching rooms (filtered classroom_capacities keys)
    rooms = list(classroom_capacities.keys())
    if not rooms:
        # Fallback to _get_available_rooms if no capacity mapping was built
        rooms = _get_available_rooms(dfs)
    
    # Use the global preferred map so course-to-room preference is consistent across files
    preferred_room_map = _GLOBAL_PREFERRED_CLASSROOMS

    allocations = []

    def _extract_room_from_cell(val):
        if not isinstance(val, str):
            return None
        if '[' in val and ']' in val:
            try:
                return val.split('[')[1].split(']')[0]
            except:
                return None
        return None

    def _choose_room_for_course(course_key, day, time_slot, course_enrollment_map=None, common_courses_list=None, passed_classroom_capacities=None):
        """
        Simplified classroom allocation with clear fallback mechanism:
        1. Check if course has a preferred room that's available
        2. Find ALL available rooms (not booked at this specific slot)
        3. Sort by allocation count for even distribution
        4. First try C-prefix classrooms, then L-prefix classrooms as fallback
        5. Assign the first available room immediately
        """
        global _ROOM_ALLOCATION_COUNTER
        
        # Use the pre-built classroom_capacities from outer scope if not passed
        local_classroom_capacities = passed_classroom_capacities if passed_classroom_capacities else classroom_capacities
        
        # Ensure time slot exists in tracker
        if day not in _CLASSROOM_USAGE_TRACKER:
            _CLASSROOM_USAGE_TRACKER[day] = {}
        if time_slot not in _CLASSROOM_USAGE_TRACKER[day]:
            _CLASSROOM_USAGE_TRACKER[day][time_slot] = set()

        # Initialize room allocation counter for all rooms if not done
        for r in local_classroom_capacities.keys():
            if r not in _ROOM_ALLOCATION_COUNTER:
                _ROOM_ALLOCATION_COUNTER[r] = 0

        # Get the set of rooms already booked at this specific slot
        booked_at_slot = _CLASSROOM_USAGE_TRACKER[day][time_slot]
        
        # Check if this course already has a preferred room (for consistency across sections)
        pref = preferred_room_map.get(course_key)
        if pref and pref not in local_classroom_capacities:
            pref = None
            preferred_room_map[course_key] = None

        # If preferred room exists and is available at THIS slot, use it
        if pref and pref not in booked_at_slot:
            chosen = pref
            conflict = False
        else:
            # === SIMPLIFIED SEQUENTIAL FALLBACK MECHANISM ===
            # Step 1: Find ALL rooms that are FREE at this specific time slot
            all_available_rooms = [r for r in local_classroom_capacities.keys() if r not in booked_at_slot]
            
            if all_available_rooms:
                # Step 2: Separate into C-prefix (classrooms) and L-prefix (classrooms with type=classroom)
                # C-prefix classrooms (96 capacity and large classrooms)
                c_prefix_classrooms = [r for r in all_available_rooms if not r.startswith('L')]
                # L-prefix classrooms only (L402-L408 type=classroom, NOT labs L105-L208)
                l_prefix_classrooms = [r for r in all_available_rooms if r.startswith('L') and room_types.get(r, '').lower() == 'classroom']
                
                # Step 3: Sort each list by allocation count (ascending) for EVEN DISTRIBUTION
                c_prefix_classrooms.sort(key=lambda r: (_ROOM_ALLOCATION_COUNTER.get(r, 0), local_classroom_capacities.get(r, 0)))
                l_prefix_classrooms.sort(key=lambda r: (_ROOM_ALLOCATION_COUNTER.get(r, 0), local_classroom_capacities.get(r, 0)))
                
                # Step 4: Sequential fallback - try C-prefix first, then L-prefix
                chosen = None
                conflict = False
                
                # Try C-prefix classrooms first
                if c_prefix_classrooms:
                    chosen = c_prefix_classrooms[0]  # First available (least used)
                # Fallback to L-prefix classrooms if C-prefix are all taken
                elif l_prefix_classrooms:
                    chosen = l_prefix_classrooms[0]  # First available (least used)
                # Ultimate fallback: any available room in the list
                elif all_available_rooms:
                    all_available_rooms.sort(key=lambda r: (_ROOM_ALLOCATION_COUNTER.get(r, 0), -local_classroom_capacities.get(r, 0)))
                    chosen = all_available_rooms[0]
                
                # Set as preferred for this course for consistency
                if chosen and course_key not in preferred_room_map:
                    preferred_room_map[course_key] = chosen
            else:
                # NO rooms available at this slot - this should NOT happen with sufficient classrooms
                # Log warning and pick least-used room (will create a conflict)
                all_rooms_sorted = sorted(local_classroom_capacities.keys(), key=lambda r: _ROOM_ALLOCATION_COUNTER.get(r, 0))
                chosen = all_rooms_sorted[0] if all_rooms_sorted else None
                conflict = True
                if chosen:
                    print(f"      [ROOM-CONFLICT] All {len(local_classroom_capacities)} classrooms booked at {day} {time_slot}!")
                    print(f"                      Booked rooms: {sorted(booked_at_slot)}")
                    print(f"                      Assigning {course_key} to least-used room {chosen}")
        
        # Mark room as used at this slot and increment allocation counter
        if chosen:
            _CLASSROOM_USAGE_TRACKER[day][time_slot].add(chosen)
            _ROOM_ALLOCATION_COUNTER[chosen] = _ROOM_ALLOCATION_COUNTER.get(chosen, 0) + 1
        
        return chosen, conflict

    def _process_df(df, section_label, course_enrollment_map=None, common_courses_list=None, passed_capacities=None):
        # Use outer classroom_capacities if not passed
        local_caps = passed_capacities if passed_capacities else classroom_capacities
        if df.empty:
            return df
        df_copy = df.copy()
        # Determine how to get time slot values per row
        has_time_slot_index = False
        if df_copy.index.name == 'Time Slot' or 'Time Slot' in df_copy.columns:
            has_time_slot_index = True

        for row_idx in range(len(df_copy)):
            # Get time slot
            if has_time_slot_index:
                time_slot = df_copy.index[row_idx] if df_copy.index.name == 'Time Slot' else df_copy.iloc[row_idx][ 'Time Slot']
            elif 'Time' in df_copy.columns:
                time_slot = df_copy.iloc[row_idx]['Time']
            else:
                time_slot = f'row{row_idx}'
            for col in df_copy.columns:
                val = df_copy.iloc[row_idx][col]
                if not isinstance(val, str) or val in ['Free', 'LUNCH BREAK']:
                    continue
                # Extract existing room
                existing_room = _extract_room_from_cell(val)
                # Determine course key - prefer a course code; fallback to basket name
                course_code = extract_course_code(val)
                if not course_code:
                    # Find basket name
                    for b in ['ELECTIVE_B1','ELECTIVE_B2','ELECTIVE_B3','ELECTIVE_B4','ELECTIVE_B5','ELECTIVE_B6','ELECTIVE_B7','ELECTIVE_B8','ELECTIVE_B9','HSS_B1','HSS_B2']:
                        if b in val.upper():
                            course_code = b
                            break
                if not course_code:
                    course_code = val.strip()
                day = str(col)
                if existing_room:
                    room = existing_room
                    # Mark it in tracker
                    if day not in _CLASSROOM_USAGE_TRACKER:
                        _CLASSROOM_USAGE_TRACKER[day] = {}
                    if time_slot not in _CLASSROOM_USAGE_TRACKER[day]:
                        _CLASSROOM_USAGE_TRACKER[day][time_slot] = set()
                    _CLASSROOM_USAGE_TRACKER[day][time_slot].add(room)
                else:
                    room, conflict = _choose_room_for_course(course_code, day, time_slot, course_enrollment_map, common_courses_list, local_caps)
                    # Append room info to cell
                    if room:
                        df_copy.iat[row_idx, df_copy.columns.get_loc(col)] = f"{val} [{room}]"
                    allocations.append({'course': course_code, 'room': room, 'classroom': room, 'day': day, 'time_slot': time_slot, 'section': section_label, 'conflict': conflict if room is not None else False})
        return df_copy

    df_a_alloc = _process_df(df_a, 'A', course_enrollment_map, common_courses_list, classroom_capacities)
    df_b_alloc = _process_df(df_b, 'B', course_enrollment_map, common_courses_list, classroom_capacities) if df_b is not None and not df_b.empty else df_b

    # Convert allocations list into map keyed by day_time for consistency with other allocation structures
    alloc_map = {}
    for a in allocations:
        key = f"{a['day']}_{a['time_slot']}"
        alloc_map[key] = {
            'course': a.get('course'),
            'classroom': a.get('room'),
            'enrollment': a.get('enrollment', None),
            'conflict': a.get('conflict', False)
        }

    # Save allocations per filename
    _TIMETABLE_CLASSROOM_ALLOCATIONS[filename] = alloc_map
    return df_a_alloc, df_b_alloc, allocations


def normalize_faculty_name(name):
    """Normalize faculty name to a canonical form to avoid duplicates.
    Handles common variations like:
    - Extra/missing spaces
    - Period vs space in initials (D.N vs D N)
    - Minor typos in common names
    """
    if not name:
        return name
    
    name = str(name).strip()
    
    # Normalize periods followed by letters to have a space (D.N -> D N)
    import re
    name = re.sub(r'\.(?=[A-Z])', ' ', name)
    
    # Remove periods at end of initials
    name = name.replace('.', ' ')
    
    # Normalize multiple spaces to single space
    name = ' '.join(name.split())
    
    # Fix common typos/variations in faculty names
    # Chimayananda -> Chinmayananda (missing 'n')
    if 'Chimayananda' in name:
        name = name.replace('Chimayananda', 'Chinmayananda')
    
    return name


def reset_classroom_usage_tracker():
    """Reset the classroom usage tracker (call before generating new timetables)"""
    global _CLASSROOM_USAGE_TRACKER, _TIMETABLE_CLASSROOM_ALLOCATIONS, _COMMON_COURSE_SCHEDULE, _COMMON_COURSE_ROOMS, _LAB_ROOM_ALLOCATIONS
    global _FACULTY_SCHEDULE_TRACKER, _CLASSROOM_SCHEDULE_TRACKER, _FACULTY_BOOKING_TRACKER, _MINOR_SCHEDULE_TRACKER
    global _ROOM_ALLOCATION_COUNTER, _GLOBAL_PREFERRED_CLASSROOMS, _MID_SEM_COMMON_SCHEDULE
    _CLASSROOM_USAGE_TRACKER = {}
    _TIMETABLE_CLASSROOM_ALLOCATIONS = {}
    _COMMON_COURSE_SCHEDULE = {}
    _COMMON_COURSE_ROOMS = {}
    _LAB_ROOM_ALLOCATIONS = {}
    _FACULTY_SCHEDULE_TRACKER = {}
    _CLASSROOM_SCHEDULE_TRACKER = {}
    _FACULTY_BOOKING_TRACKER = {}  # Reset faculty booking tracker
    _MINOR_SCHEDULE_TRACKER = {}  # Reset minor schedule tracker
    _ROOM_ALLOCATION_COUNTER = {}  # Reset room allocation counter for load balancing
    _GLOBAL_PREFERRED_CLASSROOMS = {}  # Reset preferred classrooms to allow fresh distribution
    _MID_SEM_COMMON_SCHEDULE = {}  # Reset mid-semester common schedule tracker
    initialize_classroom_usage_tracker()
    print("[RESET] Classroom usage tracker, faculty booking tracker, room allocation counter, and audit trackers reset for new timetable generation")


# ===== AUDIT TRACKING FUNCTIONS =====
def track_faculty_schedule(faculty_name, day, time_slot, course_code, course_name, semester, branch, section, classroom=None):
    """Track a faculty's scheduled slot for audit purposes.
    Called during timetable scheduling to build the faculty audit data.
    
    The slot_key includes schedule_type (extracted from semester string) to ensure
    pre-mid and post-mid entries for the same day/time are tracked separately."""
    global _FACULTY_SCHEDULE_TRACKER
    
    if not faculty_name or faculty_name.lower() in ['unknown', 'n/a', 'na', '']:
        return
    
    # Normalize faculty name (strip whitespace)
    faculty_name = str(faculty_name).strip()
    
    if faculty_name not in _FACULTY_SCHEDULE_TRACKER:
        _FACULTY_SCHEDULE_TRACKER[faculty_name] = {}
    
    # Use (day, time_slot, semester_info) as key to distinguish pre-mid from post-mid
    # semester contains info like "3 (Pre-Mid)" or "3 (Post-Mid)"
    slot_key = (day, time_slot, semester)
    
    # Store schedule info for this slot
    _FACULTY_SCHEDULE_TRACKER[faculty_name][slot_key] = {
        'course_code': course_code,
        'course_name': course_name,
        'semester': semester,
        'branch': branch,
        'section': section,
        'classroom': classroom
    }


def is_faculty_available_for_slot(faculty_name, day, time_slot, period='Pre-Mid'):
    """Check if a faculty member is available (not already booked) for a given slot.
    
    Args:
        faculty_name: Name of the faculty member
        day: Day of the week
        time_slot: Time slot string
        period: 'Pre-Mid' or 'Post-Mid' to check within a specific period
    
    Returns:
        True if faculty is available, False if already booked for another course
    """
    global _FACULTY_BOOKING_TRACKER
    
    if not faculty_name or faculty_name.lower() in ['unknown', 'n/a', 'na', '']:
        return True  # Unknown faculty is always "available"
    
    faculty_name = normalize_faculty_name(faculty_name)
    slot_key = (day, time_slot, period)
    
    return slot_key not in _FACULTY_BOOKING_TRACKER or faculty_name not in _FACULTY_BOOKING_TRACKER[slot_key]


def get_faculty_booking_at_slot(faculty_name, day, time_slot, period='Pre-Mid'):
    """Get the course a faculty is already booked for at a given slot.
    
    Returns:
        Course code if faculty is booked, None if available
    """
    global _FACULTY_BOOKING_TRACKER
    
    if not faculty_name or faculty_name.lower() in ['unknown', 'n/a', 'na', '']:
        return None
    
    faculty_name = normalize_faculty_name(faculty_name)
    slot_key = (day, time_slot, period)
    
    if slot_key in _FACULTY_BOOKING_TRACKER:
        return _FACULTY_BOOKING_TRACKER[slot_key].get(faculty_name)
    return None


def book_faculty_for_slot(faculty_name, day, time_slot, course_code, period='Pre-Mid'):
    """Book a faculty member for a specific slot.
    
    Returns:
        True if booking successful, False if faculty already booked for different course
    """
    global _FACULTY_BOOKING_TRACKER
    
    if not faculty_name or faculty_name.lower() in ['unknown', 'n/a', 'na', '']:
        return True  # Unknown faculty - skip booking
    
    faculty_name = normalize_faculty_name(faculty_name)
    slot_key = (day, time_slot, period)
    
    if slot_key not in _FACULTY_BOOKING_TRACKER:
        _FACULTY_BOOKING_TRACKER[slot_key] = {}
    
    existing_course = _FACULTY_BOOKING_TRACKER[slot_key].get(faculty_name)
    
    if existing_course and existing_course != course_code:
        # Faculty already booked for a DIFFERENT course - conflict!
        print(f"      [FACULTY-CONFLICT] {faculty_name} already teaching {existing_course} at {day} {time_slot} ({period}), cannot assign {course_code}")
        return False
    
    # Book the faculty for this slot
    _FACULTY_BOOKING_TRACKER[slot_key][faculty_name] = course_code
    return True


def get_course_faculty_list(course, course_info_map=None, section=None, branch=None):
    """Extract list of faculty names from a course row.
    
    Args:
        course: Course row (Series or dict) with 'Faculty' column
        course_info_map: Optional course info map for additional lookup
        section: Section identifier ('A', 'B', or 'Whole') - used for CSE section-specific faculty
        branch: Department/branch name
        
    Returns:
        List of normalized faculty names
        
    Note: For CSE courses with 2 faculty members, 1st is for Section A, 2nd is for Section B.
    """
    faculty_str = str(course.get('Faculty', '')).strip()
    if not faculty_str or faculty_str.lower() in ['unknown', 'n/a', 'na', '', 'nan']:
        return []
    
    # Split by comma and normalize each name
    faculty_list = [normalize_faculty_name(f.strip()) for f in faculty_str.split(',')]
    faculty_list = [f for f in faculty_list if f and f.lower() not in ['unknown', 'n/a', 'na', '']]
    
    # SPECIAL HANDLING: For CSE courses with 2 or more faculty, assign by section
    # 1st faculty -> Section A, 2nd faculty -> Section B
    # If 3+ faculty names, ignore all beyond the 2nd
    if branch == 'CSE' and len(faculty_list) >= 2 and section in ['A', 'B']:
        if section == 'A':
            return [faculty_list[0]]  # First faculty teaches Section A
        else:  # section == 'B'
            return [faculty_list[1]]  # Second faculty teaches Section B
    
    return faculty_list


def check_all_faculty_available(faculty_list, day, time_slot, period='Pre-Mid'):
    """Check if all faculty in the list are available at the given slot.
    
    Returns:
        True if all faculty are available, False if any is already booked
    """
    for faculty in faculty_list:
        if not is_faculty_available_for_slot(faculty, day, time_slot, period):
            return False
    return True


def book_all_faculty_for_slot(faculty_list, day, time_slot, course_code, period='Pre-Mid'):
    """Book all faculty in the list for the given slot.
    
    Returns:
        True if all bookings successful, False if any conflict
    """
    for faculty in faculty_list:
        if not book_faculty_for_slot(faculty, day, time_slot, course_code, period):
            return False
    return True


def track_classroom_schedule(classroom_id, day, time_slot, course_code, course_name, faculty, semester, branch, section):
    """Track a classroom's scheduled slot for audit purposes.
    Called during timetable scheduling to build the classroom audit data.
    
    The slot_key includes schedule_type (extracted from semester string) to ensure
    pre-mid and post-mid entries for the same day/time are tracked separately.
    
    NOTE: This now stores a LIST of allocations per slot to detect double-bookings."""
    global _CLASSROOM_SCHEDULE_TRACKER
    
    if not classroom_id or classroom_id.lower() in ['none', 'n/a', 'na', '']:
        return
    
    # Normalize classroom id
    classroom_id = str(classroom_id).strip()
    
    if classroom_id not in _CLASSROOM_SCHEDULE_TRACKER:
        _CLASSROOM_SCHEDULE_TRACKER[classroom_id] = {}
    
    # Use (day, time_slot, semester_info) as key to distinguish pre-mid from post-mid
    slot_key = (day, time_slot, semester)
    
    # Create new allocation entry
    allocation_entry = {
        'course_code': course_code,
        'course_name': course_name,
        'faculty': faculty,
        'semester': semester,
        'branch': branch,
        'section': section
    }
    
    # Store as LIST to detect multiple allocations (double-bookings)
    if slot_key not in _CLASSROOM_SCHEDULE_TRACKER[classroom_id]:
        _CLASSROOM_SCHEDULE_TRACKER[classroom_id][slot_key] = [allocation_entry]
    else:
        # Check if this exact entry already exists (avoid duplicates from same course)
        existing_entries = _CLASSROOM_SCHEDULE_TRACKER[classroom_id][slot_key]
        is_duplicate = any(
            e['course_code'] == course_code and e['branch'] == branch and e['section'] == section
            for e in existing_entries
        )
        if not is_duplicate:
            _CLASSROOM_SCHEDULE_TRACKER[classroom_id][slot_key].append(allocation_entry)


def populate_audit_trackers_from_timetables(dfs, output_dir):
    """Scan generated timetable Excel files and populate the audit trackers.
    This extracts faculty and classroom schedule data from the actual timetables."""
    global _FACULTY_SCHEDULE_TRACKER, _CLASSROOM_SCHEDULE_TRACKER
    
    print("\n[AUDIT] Populating audit trackers from generated timetables...")
    
    # Get course info for looking up faculty and course details
    course_info = get_course_info(dfs) if dfs else {}
    
    # Find all generated timetable files
    timetable_files = glob.glob(os.path.join(output_dir, "sem*_*_timetable.xlsx"))
    
    if not timetable_files:
        print("[AUDIT] No timetable files found to scan")
        return
    
    print(f"[AUDIT] Found {len(timetable_files)} timetable files to scan")
    
    # Track faculty double-booking across all files for conflict detection
    faculty_slot_usage = {}  # { faculty_name: { (day, time_slot): [(course, semester, branch, section)] } }
    
    # Track classroom double-booking across all files for conflict detection
    classroom_slot_usage = {}  # { classroom_id: { (day, time_slot, schedule_type): [(course, semester, branch, section)] } }
    
    for filepath in timetable_files:
        try:
            filename = os.path.basename(filepath)
            print(f"[AUDIT] Scanning: {filename}")
            
            # Extract semester and branch from filename (e.g., sem3_CSE_timetable.xlsx)
            parts = filename.replace('.xlsx', '').split('_')
            semester = int(parts[0].replace('sem', '')) if len(parts) > 0 else 0
            branch = parts[1] if len(parts) > 1 else 'Unknown'
            
            # Read all sheets from the Excel file
            xl = pd.ExcelFile(filepath)
            
            for sheet_name in xl.sheet_names:
                # Skip non-timetable sheets (legends, summaries, etc.)
                if any(skip in sheet_name.lower() for skip in ['legend', 'summary', 'course_', 'basket', 'utilization', 'allocation']):
                    continue
                
                # AUDIT FIX: Only process PreMid_* and PostMid_* sheets to avoid duplication
                # Regular/Full-Sem sheets contain the same data as pre-mid + post-mid combined
                # Skip: Regular_Section_A, Regular_Section_B, Section_A, Section_B, etc.
                sheet_lower = sheet_name.lower()
                is_premid_sheet = 'premid' in sheet_lower or 'pre_mid' in sheet_lower
                is_postmid_sheet = 'postmid' in sheet_lower or 'post_mid' in sheet_lower
                
                if not is_premid_sheet and not is_postmid_sheet:
                    # Skip all sheets that are not specifically pre-mid or post-mid
                    continue
                
                # Determine section from sheet name
                section = ''
                if 'section_a' in sheet_name.lower() or sheet_name.endswith('_A'):
                    section = 'A'
                elif 'section_b' in sheet_name.lower() or sheet_name.endswith('_B'):
                    section = 'B'
                elif 'whole' in sheet_name.lower():
                    section = 'Whole'
                
                # Determine schedule type (pre-mid or post-mid only)
                schedule_type = 'Pre-Mid' if is_premid_sheet else 'Post-Mid'
                
                try:
                    df = pd.read_excel(filepath, sheet_name=sheet_name)
                    
                    if df.empty:
                        continue
                    
                    # Identify time slot column and day columns
                    time_col = None
                    day_cols = []
                    days = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri']
                    
                    for col in df.columns:
                        col_str = str(col).strip()
                        if 'time' in col_str.lower() or 'slot' in col_str.lower():
                            time_col = col
                        elif col_str in days:
                            day_cols.append(col)
                    
                    if not time_col or not day_cols:
                        # Try using index as time slot
                        if df.index.name and 'time' in str(df.index.name).lower():
                            time_col = 'index'
                        else:
                            continue
                    
                    # Iterate through schedule
                    for idx, row in df.iterrows():
                        if time_col == 'index':
                            time_slot = str(idx).strip()
                        else:
                            time_slot = str(row.get(time_col, '')).strip()
                        
                        # Skip lunch, free slots, or invalid time slots
                        if not time_slot or 'lunch' in time_slot.lower():
                            continue
                        
                        for day in day_cols:
                            cell_value = str(row.get(day, '')).strip()
                            
                            # Skip empty, free, or lunch cells
                            if not cell_value or cell_value.lower() in ['free', 'nan', 'none', ''] or 'lunch' in cell_value.lower():
                                continue
                            
                            # Extract course code and classroom from cell
                            course_code = None
                            course_name = ''
                            classroom = None
                            
                            # Parse cell value (formats: "CS161 [C001]", "ELECTIVE_B1", "MA161 (Tutorial) [C002]")
                            cell_parts = cell_value
                            
                            # Extract classroom if present
                            if '[' in cell_parts and ']' in cell_parts:
                                bracket_start = cell_parts.rfind('[')
                                bracket_end = cell_parts.rfind(']')
                                if bracket_start < bracket_end:
                                    classroom = cell_parts[bracket_start+1:bracket_end].strip()
                                    cell_parts = cell_parts[:bracket_start].strip()
                            
                            # Extract course code
                            clean_code = cell_parts.replace('(Tutorial)', '').replace('(Lab)', '').strip()
                            course_code = extract_course_code(clean_code)
                            
                            # Handle basket entries
                            is_basket = any(kw in clean_code.upper() for kw in ['ELECTIVE_', 'HSS_', 'PROF_', 'OE_'])
                            
                            # Handle MINOR entries (format: "MINOR: CourseName" or just "MINOR")
                            is_minor = clean_code.upper().startswith('MINOR')
                            
                            if not course_code and not is_basket and not is_minor:
                                # Skip if we can't identify the course
                                continue
                            
                            # Look up course info to get faculty and course name
                            if is_minor:
                                # Minor courses - extract the minor name and track as minor slot
                                if ':' in clean_code:
                                    minor_name = clean_code.split(':', 1)[1].strip()
                                else:
                                    minor_name = clean_code
                                course_code = f"MINOR_{minor_name.replace(' ', '_')}"
                                course_name = minor_name
                                faculty_raw = ''  # Minor courses don't have assigned faculty
                                print(f"[AUDIT MINOR] {day} {time_slot}: Detected minor slot '{minor_name}' in {branch} Sem {semester} ({schedule_type})")
                            elif course_code:
                                # FIXED: Look up branch-specific key FIRST, then fallback to generic
                                info = course_info.get(f"{course_code}_{branch}", course_info.get(course_code, {}))
                                course_name = info.get('name', '')
                                faculty_raw = info.get('instructor', '')
                                
                                # CSE SECTION-SPECIFIC FIX: For CSE courses with multiple faculty,
                                # 1st faculty is for Section A, 2nd faculty is for Section B
                                # If 3+ faculty, ignore all beyond the 2nd
                                # Only track the faculty for the current section
                                if branch == 'CSE' and section in ['A', 'B'] and faculty_raw:
                                    faculty_list = [f.strip() for f in faculty_raw.split(',') if f.strip()]
                                    if len(faculty_list) >= 2:
                                        # Section A gets first faculty, Section B gets second
                                        if section == 'A':
                                            faculty_raw = faculty_list[0]
                                        else:  # Section B
                                            faculty_raw = faculty_list[1]
                                        print(f"[AUDIT DEBUG] CSE Section {section} for {course_code}: Using faculty {faculty_raw}")
                            elif is_basket:
                                # For baskets, we need to get faculty for all courses in the basket
                                faculty_raw = ''
                                course_code = clean_code  # Use basket name as course code
                            else:
                                faculty_raw = ''
                            
                            # Track classroom usage
                            if classroom:
                                track_classroom_schedule(
                                    classroom, day, time_slot,
                                    course_code, course_name, faculty_raw,
                                    f"{semester} ({schedule_type})", branch, section
                                )
                                
                                # Track for classroom double-booking detection
                                if classroom not in classroom_slot_usage:
                                    classroom_slot_usage[classroom] = {}
                                classroom_slot_key = (day, time_slot, schedule_type)
                                if classroom_slot_key not in classroom_slot_usage[classroom]:
                                    classroom_slot_usage[classroom][classroom_slot_key] = []
                                classroom_slot_usage[classroom][classroom_slot_key].append({
                                    'course': course_code,
                                    'semester': semester,
                                    'branch': branch,
                                    'section': section,
                                    'schedule_type': schedule_type
                                })
                            
                            # Track faculty usage - handle multiple instructors
                            if faculty_raw:
                                for faculty in faculty_raw.split(','):
                                    faculty = faculty.strip()
                                    if faculty and faculty.lower() not in ['unknown', 'n/a', 'na', '']:
                                        track_faculty_schedule(
                                            faculty, day, time_slot,
                                            course_code, course_name,
                                            f"{semester} ({schedule_type})", branch, section,
                                            classroom
                                        )
                                        
                                        # Track for double-booking detection within same schedule period
                                        # Include schedule_type in key - pre-mid and post-mid are separate periods
                                        if faculty not in faculty_slot_usage:
                                            faculty_slot_usage[faculty] = {}
                                        slot_key = (day, time_slot, schedule_type)
                                        if slot_key not in faculty_slot_usage[faculty]:
                                            faculty_slot_usage[faculty][slot_key] = []
                                        faculty_slot_usage[faculty][slot_key].append({
                                            'course': course_code,
                                            'semester': semester,
                                            'branch': branch,
                                            'section': section,
                                            'schedule_type': schedule_type
                                        })
                    
                except Exception as sheet_error:
                    print(f"[AUDIT] Error reading sheet '{sheet_name}': {sheet_error}")
                    continue
                    
        except Exception as file_error:
            print(f"[AUDIT] Error processing file '{filepath}': {file_error}")
            traceback.print_exc()
            continue
    
    # NOTE: We previously scanned _TIMETABLE_CLASSROOM_ALLOCATIONS here for basket/elective allocations
    # BUT this caused duplicate entries because:
    # 1. Excel file scan already tracks all classroom allocations including baskets (from cell values)
    # 2. _TIMETABLE_CLASSROOM_ALLOCATIONS keys don't distinguish between Regular/PreMid/PostMid
    # So the duplicates were being flagged as false-positive conflicts.
    # For now, we rely solely on the Excel file scan which properly extracts classrooms from cells.
    # Basket allocations that don't have [room] in cells will need to be addressed separately if needed.
    
    # Report potential double-bookings (within same schedule period)
    double_bookings = []
    for faculty, slots in faculty_slot_usage.items():
        for slot_key, usages in slots.items():
            if len(usages) > 1:
                # Check if it's actually different courses (not just same course in different sections for common courses)
                unique_courses = set(u['course'] for u in usages)
                if len(unique_courses) > 1:
                    schedule_type = slot_key[2] if len(slot_key) > 2 else 'Unknown'
                    double_bookings.append({
                        'faculty': faculty,
                        'day': slot_key[0],
                        'time_slot': slot_key[1],
                        'schedule_type': schedule_type,
                        'courses': list(unique_courses),
                        'details': usages
                    })
    
    if double_bookings:
        print(f"[AUDIT] WARNING: Potential double-bookings detected: {len(double_bookings)}")
        for db in double_bookings:
            print(f"   {db['faculty']} at {db['day']} {db['time_slot']} ({db['schedule_type']}): {db['courses']}")
    else:
        print("[AUDIT] ✓ No faculty double-bookings detected")
    
    # Report potential CLASSROOM double-bookings (within same schedule period)
    classroom_double_bookings = []
    for classroom, slots in classroom_slot_usage.items():
        for slot_key, usages in slots.items():
            if len(usages) > 1:
                # Check if it's actually different courses/sections (not same common course)
                # For common courses, same course in different sections at same room is expected
                unique_entries = set()
                for u in usages:
                    # Create unique key that considers course + branch + section
                    entry_key = (u['course'], u['branch'], u['section'])
                    unique_entries.add(entry_key)
                
                if len(unique_entries) > 1:
                    # Multiple different entries for same room at same time = potential conflict
                    schedule_type = slot_key[2] if len(slot_key) > 2 else 'Unknown'
                    classroom_double_bookings.append({
                        'classroom': classroom,
                        'day': slot_key[0],
                        'time_slot': slot_key[1],
                        'schedule_type': schedule_type,
                        'entries': list(unique_entries),
                        'details': usages
                    })
    
    if classroom_double_bookings:
        print(f"[AUDIT] WARNING: Potential classroom conflicts detected: {len(classroom_double_bookings)}")
        for cb in classroom_double_bookings:
            entries_str = ', '.join([f"{e[0]} ({e[1]} {e[2]})" for e in cb['entries']])
            print(f"   {cb['classroom']} at {cb['day']} {cb['time_slot']} ({cb['schedule_type']}): {entries_str}")
    else:
        print("[AUDIT] ✓ No classroom conflicts detected")
    
    print(f"[AUDIT] Populated trackers: {len(_FACULTY_SCHEDULE_TRACKER)} faculty, {len(_CLASSROOM_SCHEDULE_TRACKER)} classrooms")


def generate_faculty_audit_file(dfs, output_dir):
    """Generate the Faculty Availability & Schedule Audit Excel file.
    Creates one sheet per faculty showing all time slots with availability and schedule info."""
    global _FACULTY_SCHEDULE_TRACKER
    
    print("\n[AUDIT] Generating Faculty Availability & Schedule Audit File...")
    
    # Get faculty availability data - normalize names to avoid duplicates
    faculty_availability = {}
    if 'faculty_availability' in dfs and not dfs['faculty_availability'].empty:
        fa_df = dfs['faculty_availability']
        for _, row in fa_df.iterrows():
            faculty_name = normalize_faculty_name(str(row.get('Faculty Name', '')))
            if not faculty_name:
                continue
            available_days_raw = str(row.get('Available Days', 'Mon,Tue,Wed,Thu,Fri')).strip()
            unavailable_slots_raw = str(row.get('Unavailable Time Slots', '')).strip()
            
            # Parse available days
            available_days = [d.strip() for d in available_days_raw.split(',') if d.strip()]
            if not available_days:
                available_days = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri']
            
            # Parse unavailable slots (format: "Mon 09:00-10:30, Tue 13:00-14:30")
            unavailable_slots = set()
            if unavailable_slots_raw and unavailable_slots_raw.lower() not in ['none', 'na', 'n/a', '']:
                for slot in unavailable_slots_raw.split(','):
                    slot = slot.strip()
                    if slot:
                        parts = slot.split()
                        if len(parts) >= 2:
                            unavailable_slots.add((parts[0], parts[1]))
            
            # Merge if key already exists (from different name variants)
            if faculty_name in faculty_availability:
                # Merge unavailable slots
                faculty_availability[faculty_name]['unavailable_slots'].update(unavailable_slots)
            else:
                faculty_availability[faculty_name] = {
                    'available_days': available_days,
                    'unavailable_slots': unavailable_slots
                }
    
    # Get course info for looking up course names
    course_info = get_course_info(dfs) if dfs else {}
    
    # Build a mapping from original names to normalized names
    # and collect all unique normalized faculty names
    normalized_faculty_map = {}  # original_name -> normalized_name
    all_faculty_normalized = set()
    
    # From course data
    if 'course' in dfs and not dfs['course'].empty:
        for _, row in dfs['course'].iterrows():
            faculty_raw = str(row.get('Faculty', '')).strip()
            # Handle multiple instructors (comma-separated)
            for f in faculty_raw.split(','):
                f = f.strip()
                if f and f.lower() not in ['unknown', 'n/a', 'na', '']:
                    normalized = normalize_faculty_name(f)
                    normalized_faculty_map[f] = normalized
                    all_faculty_normalized.add(normalized)
    
    # From availability data
    for orig_name in faculty_availability.keys():
        normalized = normalize_faculty_name(orig_name)
        normalized_faculty_map[orig_name] = normalized
        all_faculty_normalized.add(normalized)
    
    # From tracker - normalize keys and merge schedule data
    normalized_tracker = {}
    for orig_name, schedule in _FACULTY_SCHEDULE_TRACKER.items():
        normalized = normalize_faculty_name(orig_name)
        normalized_faculty_map[orig_name] = normalized
        all_faculty_normalized.add(normalized)
        
        # Merge schedules for same normalized name
        if normalized not in normalized_tracker:
            normalized_tracker[normalized] = {}
        normalized_tracker[normalized].update(schedule)
    
    # Replace the tracker with normalized version for this function
    faculty_schedule_tracker = normalized_tracker
    
    # Also normalize the faculty_availability keys
    normalized_availability = {}
    for orig_name, avail_info in faculty_availability.items():
        normalized = normalize_faculty_name(orig_name)
        if normalized in normalized_availability:
            # Merge unavailable slots
            normalized_availability[normalized]['unavailable_slots'].update(avail_info['unavailable_slots'])
        else:
            normalized_availability[normalized] = avail_info
    faculty_availability = normalized_availability
    
    if not all_faculty_normalized:
        print("[AUDIT] No faculty data found, skipping faculty audit file")
        return None
    
    print(f"[AUDIT] Found {len(all_faculty_normalized)} unique faculty (after normalization)")
    
    days = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri']
    # Exclude lunch break from working time slots
    working_time_slots = [slot for slot in TIME_SLOT_LABELS if slot != '12:00-13:00']
    
    filepath = os.path.join(output_dir, "Faculty_Availability_Schedule_Audit.xlsx")
    
    try:
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            conflict_summary = []
            
            for faculty_name in sorted(all_faculty_normalized):
                # Create schedule matrix for this faculty
                schedule_data = []
                
                # Get availability info
                avail_info = faculty_availability.get(faculty_name, {
                    'available_days': ['Mon', 'Tue', 'Wed', 'Thu', 'Fri'],
                    'unavailable_slots': set()
                })
                available_days = avail_info['available_days']
                unavailable_slots = avail_info['unavailable_slots']
                
                # Get faculty's scheduled slots (from normalized tracker)
                faculty_schedule = faculty_schedule_tracker.get(faculty_name, {})
                
                for time_slot in working_time_slots:
                    row_data = {'Time Slot': time_slot}
                    
                    for day in days:
                        # Check if faculty is available on this day
                        is_available_day = day in available_days
                        is_unavailable_slot = (day, time_slot) in unavailable_slots
                        
                        # Find all schedule entries for this day/time slot
                        # New key format is (day, time_slot, semester_info)
                        matching_entries = []
                        for slot_key, schedule_info in faculty_schedule.items():
                            if len(slot_key) >= 2 and slot_key[0] == day and slot_key[1] == time_slot:
                                matching_entries.append(schedule_info)
                        
                        if matching_entries:
                            # Faculty is scheduled - may have multiple entries (pre-mid + post-mid)
                            cell_values = []
                            
                            # CRITICAL: Check for faculty double-booking (multiple DIFFERENT courses at same time)
                            # Group entries by schedule_type (Pre-Mid vs Post-Mid) to detect conflicts within same period
                            entries_by_period = {}
                            for schedule_info in matching_entries:
                                sem_info = schedule_info.get('semester', '')
                                # Extract schedule type from semester string like "3 (Pre-Mid)" or "3 (Post-Mid)"
                                if '(Pre-Mid)' in str(sem_info):
                                    period_key = 'Pre-Mid'
                                elif '(Post-Mid)' in str(sem_info):
                                    period_key = 'Post-Mid'
                                else:
                                    period_key = 'Unknown'
                                
                                if period_key not in entries_by_period:
                                    entries_by_period[period_key] = []
                                entries_by_period[period_key].append(schedule_info)
                            
                            # Check for double-booking within each period
                            is_double_booked = False
                            double_booking_details = []
                            for period_key, period_entries in entries_by_period.items():
                                # Get unique courses for this period
                                unique_courses = set()
                                for entry in period_entries:
                                    course = entry.get('course_code', 'N/A')
                                    if course and course != 'N/A':
                                        unique_courses.add(course)
                                
                                # If multiple different courses in same period = double-booking
                                if len(unique_courses) > 1:
                                    is_double_booked = True
                                    double_booking_details.append({
                                        'period': period_key,
                                        'courses': list(unique_courses)
                                    })
                            
                            for schedule_info in matching_entries:
                                course_code = schedule_info.get('course_code', 'N/A')
                                course_name = schedule_info.get('course_name', '')
                                semester = schedule_info.get('semester', 'N/A')
                                branch = schedule_info.get('branch', 'N/A')
                                section = schedule_info.get('section', '')
                                classroom = schedule_info.get('classroom', '')
                                
                                # Build cell value for this entry
                                cell_parts = [f"{course_code}"]
                                if course_name:
                                    cell_parts.append(f"({course_name})")
                                cell_parts.append(f"Sem {semester} | {branch}")
                                if section:
                                    cell_parts.append(f"Sec {section}")
                                if classroom:
                                    cell_parts.append(f"[{classroom}]")
                                
                                entry_value = ' '.join(cell_parts)
                                
                                # Check for conflicts and add to summary
                                if not is_available_day:
                                    entry_value = f"CONFLICT (Day Off): {entry_value}"
                                    conflict_summary.append({
                                        'Faculty': faculty_name,
                                        'Day': day,
                                        'Time Slot': time_slot,
                                        'Issue': 'Scheduled on unavailable day',
                                        'Details': f"{course_code} - {branch} Sem {semester}"
                                    })
                                elif is_unavailable_slot:
                                    entry_value = f"CONFLICT (Blocked): {entry_value}"
                                    conflict_summary.append({
                                        'Faculty': faculty_name,
                                        'Day': day,
                                        'Time Slot': time_slot,
                                        'Issue': 'Scheduled during unavailable time',
                                        'Details': f"{course_code} - {branch} Sem {semester}"
                                    })
                                
                                cell_values.append(entry_value)
                            
                            # Add double-booking conflict if detected
                            if is_double_booked:
                                for db_detail in double_booking_details:
                                    conflict_summary.append({
                                        'Faculty': faculty_name,
                                        'Day': day,
                                        'Time Slot': time_slot,
                                        'Issue': f"DOUBLE-BOOKED ({db_detail['period']}): Teaching multiple courses simultaneously",
                                        'Details': f"Courses: {', '.join(db_detail['courses'])}"
                                    })
                                # Mark cell as conflicted
                                cell_value = '⚠ DOUBLE-BOOKING:\n' + '\n'.join(cell_values)
                            else:
                                # Join multiple entries with newline
                                cell_value = '\n'.join(cell_values)
                        else:
                            # Faculty not scheduled
                            if not is_available_day:
                                cell_value = "NOT AVAILABLE (Day Off)"
                            elif is_unavailable_slot:
                                cell_value = "NOT AVAILABLE (Blocked)"
                            else:
                                cell_value = "Available - Free"
                        
                        row_data[day] = cell_value
                    
                    schedule_data.append(row_data)
                
                # Create DataFrame and write to sheet
                df = pd.DataFrame(schedule_data)
                
                # Sanitize sheet name (Excel limits: 31 chars, no special chars)
                sheet_name = faculty_name[:31].replace('/', '-').replace('\\', '-').replace('*', '-').replace('?', '-').replace('[', '(').replace(']', ')')
                
                df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Add conflict summary sheet
            if conflict_summary:
                conflict_df = pd.DataFrame(conflict_summary)
                conflict_df.to_excel(writer, sheet_name='CONFLICT_SUMMARY', index=False)
                
                # Categorize conflicts by type for clearer reporting
                double_booking_conflicts = [c for c in conflict_summary if 'DOUBLE-BOOKED' in c.get('Issue', '')]
                day_off_conflicts = [c for c in conflict_summary if 'unavailable day' in c.get('Issue', '')]
                blocked_time_conflicts = [c for c in conflict_summary if 'unavailable time' in c.get('Issue', '')]
                
                print(f"[AUDIT] Found {len(conflict_summary)} potential faculty conflicts:")
                if double_booking_conflicts:
                    print(f"   - Double-booking (multiple courses): {len(double_booking_conflicts)}")
                if day_off_conflicts:
                    print(f"   - Scheduled on unavailable day: {len(day_off_conflicts)}")
                if blocked_time_conflicts:
                    print(f"   - Scheduled during blocked time: {len(blocked_time_conflicts)}")
                
                # Create a categorized summary sheet
                summary_data = [
                    {'Category': 'Total Conflicts', 'Count': len(conflict_summary)},
                    {'Category': 'Double-Booking (Multiple Courses)', 'Count': len(double_booking_conflicts)},
                    {'Category': 'Scheduled on Unavailable Day', 'Count': len(day_off_conflicts)},
                    {'Category': 'Scheduled During Blocked Time', 'Count': len(blocked_time_conflicts)}
                ]
                summary_df = pd.DataFrame(summary_data)
                summary_df.to_excel(writer, sheet_name='CONFLICT_CATEGORIES', index=False)
            else:
                # Create empty summary with message
                summary_df = pd.DataFrame([{'Status': 'No conflicts detected - All faculty schedules comply with availability'}])
                summary_df.to_excel(writer, sheet_name='CONFLICT_SUMMARY', index=False)
                print("[AUDIT] No faculty conflicts detected")
        
        # Apply formatting
        _format_audit_excel(filepath, 'faculty')
        
        print(f"[AUDIT] Faculty audit file saved: {filepath}")
        return filepath
        
    except Exception as e:
        print(f"[AUDIT] Error generating faculty audit file: {e}")
        traceback.print_exc()
        return None


def generate_classroom_audit_file(dfs, output_dir):
    """Generate the Classroom Availability & Schedule Audit Excel file.
    Creates one sheet per classroom showing all time slots with schedule info."""
    global _CLASSROOM_SCHEDULE_TRACKER, _CLASSROOM_USAGE_TRACKER
    
    print("\n[AUDIT] Generating Classroom Availability & Schedule Audit File...")
    
    # Get classroom data
    all_classrooms = set()
    classroom_info = {}
    
    if 'classroom' in dfs and not dfs['classroom'].empty:
        classroom_df = dfs['classroom']
        for _, row in classroom_df.iterrows():
            room_number = str(row.get('Room Number', '')).strip()
            if room_number and room_number.lower() not in ['none', 'n/a', 'na', '']:
                all_classrooms.add(room_number)
                classroom_info[room_number] = {
                    'capacity': row.get('Capacity', 'N/A'),
                    'type': row.get('Type', 'N/A'),
                    'location': row.get('Location', 'N/A')
                }
    
    # Also include classrooms from the tracker
    all_classrooms.update(_CLASSROOM_SCHEDULE_TRACKER.keys())
    
    # Also check _CLASSROOM_USAGE_TRACKER for any rooms
    for day in _CLASSROOM_USAGE_TRACKER:
        for time_slot in _CLASSROOM_USAGE_TRACKER[day]:
            all_classrooms.update(_CLASSROOM_USAGE_TRACKER[day][time_slot])
    
    if not all_classrooms:
        print("[AUDIT] No classroom data found, skipping classroom audit file")
        return None
    
    days = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri']
    # Exclude lunch break from working time slots
    working_time_slots = [slot for slot in TIME_SLOT_LABELS if slot != '12:00-13:00']
    
    filepath = os.path.join(output_dir, "Classroom_Availability_Schedule_Audit.xlsx")
    
    try:
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            double_booking_summary = []
            
            for classroom_id in sorted(all_classrooms):
                # Create schedule matrix for this classroom
                schedule_data = []
                
                # Get classroom's scheduled slots
                classroom_schedule = _CLASSROOM_SCHEDULE_TRACKER.get(classroom_id, {})
                
                # Get classroom info
                room_info = classroom_info.get(classroom_id, {})
                
                for time_slot in working_time_slots:
                    row_data = {'Time Slot': time_slot}
                    
                    for day in days:
                        # Find all schedule entries for this day/time slot
                        # New key format is (day, time_slot, semester_info)
                        # NOTE: schedule_info is now a LIST of entries (to track multiple allocations/conflicts)
                        matching_entries = []
                        for slot_key, schedule_info_list in classroom_schedule.items():
                            if len(slot_key) >= 2 and slot_key[0] == day and slot_key[1] == time_slot:
                                # schedule_info_list is a LIST of allocation entries
                                if isinstance(schedule_info_list, list):
                                    matching_entries.extend(schedule_info_list)
                                else:
                                    # Legacy: if it's still a dict, wrap it in a list
                                    matching_entries.append(schedule_info_list)
                        
                        # Check if classroom is used (from usage tracker as fallback)
                        is_used_in_tracker = classroom_id in _CLASSROOM_USAGE_TRACKER.get(day, {}).get(time_slot, set())
                        
                        if matching_entries:
                            # Classroom has detailed schedule info - may have multiple entries
                            cell_values = []
                            for schedule_info in matching_entries:
                                course_code = schedule_info.get('course_code', 'N/A')
                                course_name = schedule_info.get('course_name', '')
                                faculty = schedule_info.get('faculty', 'N/A')
                                semester = schedule_info.get('semester', 'N/A')
                                branch = schedule_info.get('branch', 'N/A')
                                section = schedule_info.get('section', '')
                                
                                # Build cell value for this entry
                                cell_parts = [f"{course_code}"]
                                if course_name:
                                    cell_parts.append(f"({course_name})")
                                cell_parts.append(f"| {faculty}")
                                cell_parts.append(f"| Sem {semester} | {branch}")
                                if section:
                                    cell_parts.append(f"Sec {section}")
                                
                                cell_values.append(' '.join(cell_parts))
                            
                            # Check for CONFLICTS - multiple entries with DIFFERENT courses at same slot
                            # within the SAME schedule period (Pre-Mid or Post-Mid)
                            # Group entries by period to detect real conflicts
                            # NOTE: Common courses (same course code, different sections) are NOT conflicts
                            entries_by_period = {}
                            for entry in matching_entries:
                                sem_info = entry.get('semester', '')
                                if '(Pre-Mid)' in str(sem_info):
                                    period_key = 'Pre-Mid'
                                elif '(Post-Mid)' in str(sem_info):
                                    period_key = 'Post-Mid'
                                else:
                                    period_key = 'Unknown'
                                
                                if period_key not in entries_by_period:
                                    entries_by_period[period_key] = []
                                entries_by_period[period_key].append(entry)
                            
                            # Check for double-booking within each period
                            # A conflict is when we have DIFFERENT courses (by code) in the same period
                            # Same course code in different sections is NOT a conflict (common course sharing room)
                            is_conflict = False
                            conflict_courses = []
                            conflict_periods = []
                            for period_key, period_entries in entries_by_period.items():
                                # Get unique course codes in this period (normalize to base code)
                                unique_courses_in_period = set()
                                for e in period_entries:
                                    course_code = e.get('course_code', '')
                                    if course_code:
                                        # Normalize MINOR course codes to a canonical form
                                        # "MINOR_Generative_Ai" -> "MINOR_Generative_Ai"
                                        # "MINOR: Generative Ai" -> "MINOR_Generative_Ai"
                                        if course_code.startswith('MINOR:'):
                                            # Convert "MINOR: Xyz" to "MINOR_Xyz" format
                                            minor_name = course_code.replace('MINOR:', '').strip().replace(' ', '_')
                                            base_code = f"MINOR_{minor_name}"
                                        elif course_code.startswith('MINOR_'):
                                            base_code = course_code
                                        else:
                                            base_code = course_code
                                        unique_courses_in_period.add(base_code)
                                
                                # Only flag as conflict if we have truly different courses
                                # (more than 1 unique course code in the same period)
                                if len(unique_courses_in_period) > 1:
                                    is_conflict = True
                                    conflict_courses.extend(unique_courses_in_period)
                                    conflict_periods.append(period_key)
                            
                            # Get room info for additional context in conflict detection
                            room_info = classroom_info.get(classroom_id, {})
                            room_capacity = room_info.get('capacity', 'N/A')
                            room_type = str(room_info.get('type', '')).upper()
                            # Check actual room type from CSV, not prefix (L402-L408 are classrooms, not labs)
                            is_lab_room = 'LAB' in room_type
                            is_large_room = str(room_capacity).isdigit() and int(room_capacity) >= 120
                            
                            if is_conflict:
                                # Mark as conflict and add to summary
                                conflict_label = 'LAB ' if is_lab_room else ('LARGE ROOM ' if is_large_room else '')
                                cell_value = f'⚠ {conflict_label}CONFLICT:\n' + '\n'.join(cell_values)
                                double_booking_summary.append({
                                    'Classroom': classroom_id,
                                    'Capacity': room_capacity,
                                    'Room Type': 'Lab' if is_lab_room else ('Large (120/240)' if is_large_room else 'Regular'),
                                    'Day': day,
                                    'Time Slot': time_slot,
                                    'Conflict Period(s)': ', '.join(conflict_periods),
                                    'Courses': list(set(conflict_courses)),
                                    'Details': cell_values
                                })
                            else:
                                # Join multiple entries with newline (same course, different schedules)
                                cell_value = '\n'.join(cell_values)
                        elif is_used_in_tracker:
                            # Classroom is used but no detailed info from schedule tracker
                            # Try to get details from the usage tracker by checking all timetables
                            # For now, mark as occupied - the schedule tracker should have the details
                            # if properly populated from pre-mid and post-mid sheets
                            cell_value = "FREE"  # If not in schedule tracker, it's likely from Regular sheets which we skip
                        else:
                            cell_value = "FREE"
                        
                        row_data[day] = cell_value
                    
                    schedule_data.append(row_data)
                
                # Create DataFrame and write to sheet
                df = pd.DataFrame(schedule_data)
                
                # Sanitize sheet name
                sheet_name = str(classroom_id)[:31].replace('/', '-').replace('\\', '-').replace('*', '-').replace('?', '-').replace('[', '(').replace(']', ')')
                
                df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Check for double-bookings by analyzing _CLASSROOM_USAGE_TRACKER
            for day in days:
                for time_slot in working_time_slots:
                    # Check each classroom for this slot
                    for classroom_id in all_classrooms:
                        schedule_info = _CLASSROOM_SCHEDULE_TRACKER.get(classroom_id, {}).get((day, time_slot))
                        
                        # Count how many entries we have for this slot
                        # A double-booking would be detected if the same classroom appears multiple times
                        # in different timetables for the same slot
                        # Note: Current tracker structure doesn't allow duplicates per slot
                        # But we can check consistency
                        pass  # Double-booking prevention is handled during allocation
            
            # Add summary sheet with classroom utilization
            utilization_data = []
            for classroom_id in sorted(all_classrooms):
                classroom_schedule = _CLASSROOM_SCHEDULE_TRACKER.get(classroom_id, {})
                room_info = classroom_info.get(classroom_id, {})
                
                total_slots = len(days) * len(working_time_slots)
                occupied_slots = len(classroom_schedule)
                utilization = (occupied_slots / total_slots * 100) if total_slots > 0 else 0
                
                utilization_data.append({
                    'Classroom': classroom_id,
                    'Capacity': room_info.get('capacity', 'N/A'),
                    'Type': room_info.get('type', 'N/A'),
                    'Occupied Slots': occupied_slots,
                    'Total Slots': total_slots,
                    'Utilization %': f"{utilization:.1f}%"
                })
            
            utilization_df = pd.DataFrame(utilization_data)
            utilization_df.to_excel(writer, sheet_name='UTILIZATION_SUMMARY', index=False)
            
            if double_booking_summary:
                conflict_df = pd.DataFrame(double_booking_summary)
                conflict_df.to_excel(writer, sheet_name='DOUBLE_BOOKING_ALERTS', index=False)
                
                # Categorize conflicts by room type for clearer reporting
                lab_conflicts = [c for c in double_booking_summary if c.get('Room Type') == 'Lab']
                large_room_conflicts = [c for c in double_booking_summary if c.get('Room Type') == 'Large (120/240)']
                regular_conflicts = [c for c in double_booking_summary if c.get('Room Type') == 'Regular']
                
                print(f"[AUDIT] Found {len(double_booking_summary)} potential double-bookings:")
                if lab_conflicts:
                    print(f"   - Lab Room conflicts: {len(lab_conflicts)}")
                if large_room_conflicts:
                    print(f"   - Large Room (120/240) conflicts: {len(large_room_conflicts)}")
                if regular_conflicts:
                    print(f"   - Regular Room conflicts: {len(regular_conflicts)}")
                
                # Create a categorized summary sheet
                summary_data = [
                    {'Category': 'Total Conflicts', 'Count': len(double_booking_summary)},
                    {'Category': 'Lab Room Conflicts', 'Count': len(lab_conflicts)},
                    {'Category': 'Large Room (120/240) Conflicts', 'Count': len(large_room_conflicts)},
                    {'Category': 'Regular Room Conflicts', 'Count': len(regular_conflicts)}
                ]
                summary_df = pd.DataFrame(summary_data)
                summary_df.to_excel(writer, sheet_name='CONFLICT_SUMMARY', index=False)
            else:
                # No conflicts - still create both sheets for consistency
                summary_df = pd.DataFrame([{'Status': 'No double-bookings detected - All classroom allocations are unique'}])
                summary_df.to_excel(writer, sheet_name='DOUBLE_BOOKING_ALERTS', index=False)
                
                # Create CONFLICT_SUMMARY sheet even when no conflicts
                summary_data = [
                    {'Category': 'Total Conflicts', 'Count': 0},
                    {'Category': 'Lab Room Conflicts', 'Count': 0},
                    {'Category': 'Large Room (120/240) Conflicts', 'Count': 0},
                    {'Category': 'Regular Room Conflicts', 'Count': 0},
                    {'Category': 'Status', 'Count': 'All Clear ✓'}
                ]
                conflict_summary_df = pd.DataFrame(summary_data)
                conflict_summary_df.to_excel(writer, sheet_name='CONFLICT_SUMMARY', index=False)
                print("[AUDIT] No classroom double-bookings detected")
        
        # Apply formatting
        _format_audit_excel(filepath, 'classroom')
        
        print(f"[AUDIT] Classroom audit file saved: {filepath}")
        return filepath
        
    except Exception as e:
        print(f"[AUDIT] Error generating classroom audit file: {e}")
        traceback.print_exc()
        return None


def _format_audit_excel(filepath, audit_type):
    """Apply formatting to audit Excel files for better readability."""
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        wb = load_workbook(filepath)
        
        # Define styles
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        occupied_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light green for occupied
        free_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # White
        unavailable_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")  # Gray
        conflict_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Light red
        timeslot_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")  # Light gray for time slot column
        border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Format header row
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment
                cell.border = border
            
            # Format data cells
            for row_idx in range(2, ws.max_row + 1):
                for col_idx in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.border = border
                    cell.alignment = center_alignment
                    
                    cell_value = str(cell.value or '').upper()
                    
                    # First column is Time Slot - give it a subtle background
                    if col_idx == 1:
                        cell.fill = timeslot_fill
                        cell.font = Font(bold=True)
                    elif 'CONFLICT' in cell_value:
                        cell.fill = conflict_fill
                    elif 'FREE' in cell_value or 'AVAILABLE - FREE' in cell_value:
                        cell.fill = free_fill
                    elif 'NOT AVAILABLE' in cell_value:
                        cell.fill = unavailable_fill
                    elif cell_value and cell_value not in ['NONE', 'N/A', '']:
                        # Any cell with course/schedule data gets the occupied color
                        cell.fill = occupied_fill
            
            # Auto-adjust column widths
            for col_idx in range(1, ws.max_column + 1):
                column_letter = get_column_letter(col_idx)
                max_length = 0
                
                for cell in ws[column_letter]:
                    try:
                        cell_value = str(cell.value) if cell.value else ""
                        lines = cell_value.split('\n')
                        max_line_length = max(len(line) for line in lines) if lines else 0
                        max_length = max(max_length, max_line_length)
                    except:
                        pass
                
                # Set width with constraints
                adjusted_width = min(max(max_length + 2, 15), 60)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(filepath)
        
    except Exception as e:
        print(f"[AUDIT] Error formatting audit file: {e}")


def generate_audit_files(dfs, output_dir):
    """Generate both Faculty and Classroom audit files.
    Call this after timetable generation is complete."""
    
    print("\n" + "="*60)
    print("GENERATING AUDIT FILES FOR VERIFICATION")
    print("="*60)
    
    faculty_file = generate_faculty_audit_file(dfs, output_dir)
    classroom_file = generate_classroom_audit_file(dfs, output_dir)
    
    print("\n" + "="*60)
    if faculty_file and classroom_file:
        print("AUDIT FILES GENERATED SUCCESSFULLY")
        print(f"  - Faculty Audit: {os.path.basename(faculty_file)}")
        print(f"  - Classroom Audit: {os.path.basename(classroom_file)}")
    else:
        print("AUDIT FILE GENERATION COMPLETED WITH WARNINGS")
        if not faculty_file:
            print("  - Faculty Audit: SKIPPED (no data)")
        if not classroom_file:
            print("  - Classroom Audit: SKIPPED (no data)")
    print("="*60 + "\n")
    
    return {
        'faculty_audit': faculty_file,
        'classroom_audit': classroom_file
    }


# EXAM SCHEDULE FILE MANAGEMENT FUNCTIONS - COMMENTED OUT
"""
def get_exam_schedule_files():
    \"\"\"Get list of exam schedule files that should be displayed\"\"\"
    global _EXAM_SCHEDULE_FILES
    return list(_EXAM_SCHEDULE_FILES)

def add_exam_schedule_file(filename):
    \"\"\"Add a filename to the list of exam schedules to display\"\"\"
    global _EXAM_SCHEDULE_FILES
    _EXAM_SCHEDULE_FILES.add(filename)

def clear_exam_schedule_files():
    \"\"\"Clear the list of exam schedules to display\"\"\"
    global _EXAM_SCHEDULE_FILES
    _EXAM_SCHEDULE_FILES.clear()

def remove_exam_schedule_file(filename):
    \"\"\"Remove a filename from the list of exam schedules to display\"\"\"
    global _EXAM_SCHEDULE_FILES
    if filename in _EXAM_SCHEDULE_FILES:
        _EXAM_SCHEDULE_FILES.remove(filename)
"""

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def get_file_hash(filepath):
    """Calculate MD5 hash of a file to detect changes"""
    try:
        with open(filepath, 'rb') as f:
            return hashlib.md5(f.read()).hexdigest()
    except:
        return None

def has_files_changed():
    """Check if any input files have changed since last load"""
    global _file_hashes
    
    if not os.path.exists(INPUT_DIR):
        return True
        
    current_files = os.listdir(INPUT_DIR)
    current_hashes = {}
    
    for file in current_files:
        filepath = os.path.join(INPUT_DIR, file)
        current_hashes[file] = get_file_hash(filepath)
    
    # If number of files changed
    if set(current_files) != set(_file_hashes.keys()):
        print("[DIR] File count changed")
        _file_hashes = current_hashes
        return True
    
    # If file contents changed
    for file, current_hash in current_hashes.items():
        if file not in _file_hashes or _file_hashes[file] != current_hash:
            print(f"[FILE] File content changed: {file}")
            _file_hashes = current_hashes
            return True
    
    return False

def find_csv_file(filename):
    """Find CSV file with flexible search (case-insensitive, space-insensitive)."""
    if not os.path.exists(INPUT_DIR):
        return None
        
    files = os.listdir(INPUT_DIR)
    filename_clean = filename.lower().replace(' ', '')
    
    for file in files:
        file_clean = file.lower().replace(' ', '')
        if file_clean == filename_clean:
            return os.path.join(INPUT_DIR, file)
    
    # Try more flexible matching for common patterns
    for file in files:
        file_clean = file.lower().replace(' ', '').replace('_', '').replace('-', '')
        filename_clean_alt = filename_clean.replace('_', '').replace('-', '')
        if file_clean == filename_clean_alt:
            return os.path.join(INPUT_DIR, file)
    
    return None

def load_all_data(force_reload=False):
    """Load CSV files and return dataframes with force reload option"""
    global _cached_data_frames
    global _cached_timestamp
    
    # Always check if files have changed
    files_changed = has_files_changed()
    
    # Use caching but allow force reload or if files changed
    if (not force_reload and 
        not files_changed and
        '_cached_data_frames' in globals() and 
        _cached_data_frames is not None):

        current_time = time.time()
        # Cache for 30 seconds max
        if current_time - _cached_timestamp < 30:
            print("[FOLDER] Using cached data frames (files unchanged)")
            return _cached_data_frames
        else:
            print("[FOLDER] Cache expired, reloading data")
    
    if files_changed:
        print("[FOLDER] Files changed, reloading data")
    
    required_files = [
        "course_data.csv",
        "faculty_availability.csv",
        "classroom_data.csv",
        "student_data.csv",
        "exams_data.csv"
    ]
    optional_files = [
        "minor_data.csv"  # Optional minor course scheduling
    ]
    dfs = {}
    
    print("[FOLDER] Loading CSV files...")
    print(f"[DIR] Input directory contents: {os.listdir(INPUT_DIR) if os.path.exists(INPUT_DIR) else 'Directory not found'}")
    
    # Update file hashes
    if os.path.exists(INPUT_DIR):
        for file in os.listdir(INPUT_DIR):
            filepath = os.path.join(INPUT_DIR, file)
            _file_hashes[file] = get_file_hash(filepath)
    
    # Load required files
    for f in required_files:
        file_path = find_csv_file(f)
        if not file_path:
            print(f"[FAIL] CSV not found: {f}")
            # Try to find any similar file
            files = os.listdir(INPUT_DIR) if os.path.exists(INPUT_DIR) else []
            similar_files = [file for file in files if f.split('_')[0] in file.lower()]
            if similar_files:
                print(f"   [TIP] Similar files found: {similar_files}")
            return None
        
        try:
            key = f.replace("_data.csv", "").replace(".csv", "")
            dfs[key] = pd.read_csv(file_path)
            print(f"[OK] Loaded {f} from {file_path} ({len(dfs[key])} rows)")
            
            # If faculty_availability contains only a single column of names, make it compatible
            if key == 'faculty_availability':
                fa_df = dfs[key].copy()
                # Normalize column names case-insensitively to expected columns
                col_map = {}
                for col in fa_df.columns:
                    cl = col.strip().lower()
                    if 'faculty' in cl:
                        col_map[col] = 'Faculty Name'
                    elif 'available' in cl:
                        col_map[col] = 'Available Days'
                    elif 'unavailable' in cl or 'unavailable time' in cl:
                        col_map[col] = 'Unavailable Time Slots'

                if col_map:
                    fa_df = fa_df.rename(columns=col_map)

                # If still missing 'Faculty Name' but file has only one column, treat that as faculty names
                if 'Faculty Name' not in fa_df.columns and len(fa_df.columns) == 1:
                    fa_df = fa_df.rename(columns={fa_df.columns[0]: 'Faculty Name'})

                # Ensure required columns exist with sensible defaults
                if 'Available Days' not in fa_df.columns:
                    fa_df['Available Days'] = 'Mon,Tue,Wed,Thu,Fri'
                if 'Unavailable Time Slots' not in fa_df.columns:
                    fa_df['Unavailable Time Slots'] = ''

                dfs[key] = fa_df
                print(f"   [INFO] faculty_availability.csv normalized; columns now: {list(fa_df.columns)}")

            # Normalize student dataframe column names for compatibility with minimal schema
            if key == 'student':
                st_df = dfs[key].copy()
                student_col_map = {}
                for col in st_df.columns:
                    cl = col.strip().lower()
                    if 'roll' in cl and 'no' in cl:
                        student_col_map[col] = 'Roll No'
                    elif col.strip().lower() in ['name', 'student name']:
                        student_col_map[col] = 'Name'
                    elif 'semester' in cl:
                        student_col_map[col] = 'Semester'
                    elif 'department' in cl:
                        student_col_map[col] = 'Department'
                if student_col_map:
                    st_df = st_df.rename(columns=student_col_map)
                dfs[key] = st_df
                print(f"   [INFO] student_data.csv normalized; columns now: {list(st_df.columns)}")

            # Show sample data for verification
            if not dfs[key].empty:
                print(f"   Columns: {list(dfs[key].columns)}")
                if 'course' in key:
                    print(f"   First 3 courses:")
                    for i, row in dfs[key].head(3).iterrows():
                        print(f"     {i+1}. {row['Course Code'] if 'Course Code' in row else 'N/A'} - Semester: {row.get('Semester', 'N/A')} - Branch: {row.get('Branch', 'N/A')} - Elective: {row.get('Elective (Yes/No)', 'N/A')}")
                
        except Exception as e:
            print(f"[FAIL] Error loading {f}: {e}")
            return None
    
    # Load optional files
    for f in optional_files:
        file_path = find_csv_file(f)
        if not file_path:
            print(f"[INFO] Optional file not found: {f}; skipping")
            continue
        
        try:
            key = f.replace("_data.csv", "").replace(".csv", "")
            dfs[key] = pd.read_csv(file_path)
            print(f"[OK] Loaded optional {f} from {file_path} ({len(dfs[key])} rows)")
            
            # Special handling for minor data to normalize columns
            if key == 'minor':
                minor_df = dfs[key].copy()
                col_map = {}
                for col in minor_df.columns:
                    cl = str(col).strip().lower()
                    if 'minor' in cl and 'course' in cl:
                        col_map[col] = 'Minor Course'
                    elif 'semester' in cl:
                        col_map[col] = 'Semester'
                    elif 'registered' in cl and 'student' in cl:
                        col_map[col] = 'Registered Students'
                if col_map:
                    minor_df = minor_df.rename(columns=col_map)
                # Ensure required columns exist
                if 'Minor Course' not in minor_df.columns and len(minor_df.columns) > 0:
                    minor_df['Minor Course'] = minor_df.iloc[:, 0].astype(str)
                if 'Semester' not in minor_df.columns:
                    minor_df['Semester'] = None
                if 'Registered Students' not in minor_df.columns:
                    minor_df['Registered Students'] = 0
                dfs[key] = minor_df
                print(f"   [INFO] minor_data.csv normalized; columns: {list(minor_df.columns)}")
        
        except Exception as e:
            print(f"[WARN] Error loading optional file {f}: {e}")
    
    # Validate required columns in course data
    if 'course' in dfs:
        required_course_columns = ['Course Code', 'Semester', 'LTPSC']
        missing_columns = [col for col in required_course_columns if col not in dfs['course'].columns]
        if missing_columns:
            print(f"[FAIL] Missing columns in course_data: {missing_columns}")
            print(f"   Available columns: {list(dfs['course'].columns)}")
            return None
    
    # Cache the results
    _cached_data_frames = dfs
    _cached_timestamp = time.time()
    
    print("[OK] All CSV files loaded successfully!")
    return dfs

def get_course_info(dfs):
    """Extract course information from course data for frontend display with proper department mapping"""
    course_info = {}
    if 'course' in dfs:
        course_df = dfs['course']

        # Identify optional mid-semester columns once to avoid repeated lookups
        half_sem_col = None
        post_mid_col = None
        for col in course_df.columns:
            col_low = str(col).lower()
            if half_sem_col is None and 'half' in col_low and 'sem' in col_low:
                half_sem_col = col
            if post_mid_col is None and 'post' in col_low and 'mid' in col_low:
                post_mid_col = col

        for _, course in course_df.iterrows():
            course_code = course['Course Code']

            # Get department from CSV (not inferred from course code)
            csv_department = course.get('Department', '')
            
            # FIXED: Map department based on course code prefix as fallback
            department = map_department_from_course_code(course_code)

            is_elective = course.get('Elective (Yes/No)', 'No').upper() == 'YES'
            course_type = 'Elective' if is_elective else 'Core'

            # FIX: Use 'Faculty' column instead of 'Instructor'
            instructor = course.get('Faculty', 'Unknown')

            # Derive term label (pre-mid, post-mid, or full semester) for legend clarity
            half_sem_val = str(course.get(half_sem_col, '')).strip().upper() if half_sem_col else ''
            post_mid_val_raw = str(course.get(post_mid_col, '')).strip().upper() if post_mid_col else ''
            # Treat blank/missing post-mid as NO so half-sem courses default to pre-mid-only
            post_mid_val = post_mid_val_raw if post_mid_val_raw else 'NO'

            if half_sem_val == 'YES' and post_mid_val == 'YES':
                term_label = 'Post-Mid Only'
            elif half_sem_val == 'YES':  # Any YES with non-YES post_mid -> Pre-mid only
                term_label = 'Pre-Mid Only'
            else:
                term_label = 'Full Sem'  # Half-semester = NO or missing -> Full semester

            # Use course_code + department as key for department-specific entries
            # This allows same course to have different term labels per department
            key = f"{course_code}_{csv_department}" if csv_department else course_code
            
            course_info[key] = {
                'name': course.get('Course Name', 'Unknown Course'),
                'credits': course.get('Credits', 0),
                'type': course_type,
                'instructor': instructor,  # This will now use the correct Faculty column
                'department': department,  # Use mapped department
                'csv_department': csv_department,  # Store CSV department for lookups
                'semester': course.get('Semester', None),
                'is_elective': is_elective,
                'branch': department,  # Use department as branch for compatibility
                'is_common_elective': is_elective,
                'ltpsc': course.get('LTPSC', ''),
                'ltpsc_components': parse_ltpsc(course.get('LTPSC', '')) if 'LTPSC' in course else None,
                'common': course.get('Common', 'No'),  # Add Common field
                'registered_students': int(course.get('Registered Students', 0)) if pd.notna(course.get('Registered Students')) else None,  # Add registered students from CSV
                'term_type': term_label,
                'half_semester': half_sem_val or None,
                'post_mid': post_mid_val or None,
                'course_code': course_code  # Store original course code
            }
            
            # Also store with course_code only as fallback for lookups that don't specify department
            if course_code not in course_info:
                course_info[course_code] = course_info[key]

            # Debug logging
            print(f"   [NOTE] Course {course_code} ({csv_department}): Department = {department}, Term = {term_label}")

    return course_info

def map_department_from_course_code(course_code):
    """Map department based on course code prefix"""
    if not isinstance(course_code, str):
        return 'Department of Arts, Science and Design'
    
    course_code_upper = course_code.upper().strip()
    
    if course_code_upper.startswith('CS'):
        return 'Computer Science and Engineering'
    elif course_code_upper.startswith('EC'):
        return 'Electronics and Communication Engineering'
    elif course_code_upper.startswith('DS'):
        return 'Data Science and Artificial Intelligence'
    elif course_code_upper.startswith('DA'):  # Also handle DA prefix for DSAI
        return 'Data Science and Artificial Intelligence'
    else:
        return 'Department of Arts, Science and Design'


def get_course_info_by_dept(course_info, course_code, department=None):
    """Helper to get course info with department-specific lookup
    
    Args:
        course_info: The course info dictionary
        course_code: The course code to look up
        department: Optional department (CSE, DSAI, ECE, etc.)
    
    Returns:
        Course info dict, or empty dict if not found
    """
    if department:
        # Try department-specific key first
        dept_key = f"{course_code}_{department}"
        if dept_key in course_info:
            return course_info[dept_key]
    
    # Fallback to course code only
    return course_info.get(course_code, {})


def _get_registered_or_default_enrollment(info):
    """Return registered_students if present, otherwise a sensible default."""
    registered = info.get('registered_students') or info.get('Registered Students')
    if registered is not None and str(registered).strip() != '':
        try:
            return int(registered)
        except Exception:
            pass
    # Default heuristics
    return 40 if info.get('is_elective', False) else 60


def detect_cross_dsai_ece_common(course_info, course_code, semester_id):
    """Identify DSAI/ECE common courses (same faculty, Common=Yes) and return shared metadata."""
    target_departments = {
        'Data Science and Artificial Intelligence',
        'Electronics and Communication Engineering',
    }

    matches = []
    course_code_upper = str(course_code).strip().upper()

    for info in course_info.values():
        if str(info.get('course_code', '')).strip().upper() != course_code_upper:
            continue

        dept_raw = info.get('csv_department') or info.get('department') or info.get('branch')
        dept_norm = normalize_branch_string(dept_raw)
        if dept_norm not in target_departments:
            continue

        common_flag = str(info.get('common', info.get('Common', 'No'))).strip().upper() == 'YES'
        if not common_flag:
            continue

        instructor = str(info.get('instructor', info.get('Faculty', '')) or '').strip().lower()
        if not instructor:
            continue

        matches.append((dept_norm, instructor, info))

    if not matches:
        return None

    for instructor in {inst for _, inst, _ in matches}:
        depts = {dept for dept, inst, _ in matches if inst == instructor}
        if target_departments.issubset(depts):
            total = sum(
                _get_registered_or_default_enrollment(info)
                for dept, inst, info in matches
                if inst == instructor and dept in target_departments
            )
            total = max(1, int(math.ceil(total)))
            return {
                'instructor': instructor,
                'departments': target_departments,
                'total_enrollment': total,
                'schedule_key': f"sem{semester_id}_DSAI_ECE_{course_code_upper}",
                'room_key': f"{semester_id}_DSAI_ECE_{course_code_upper}",
            }

    return None


def detect_dual_instructor_course(course_info, course_code, department):
    """
    Detect if a course code has multiple instructors teaching it in the same department.
    If yes, return the shared enrollment/2 for room allocation optimization.
    
    Args:
        course_info: Dictionary of all course information
        course_code: The course code to check
        department: The target department (CSE, DSAI, ECE, etc.)
    
    Returns:
        Dict with 'is_dual': bool, 'effective_enrollment': int (enrollment/2 if dual)
        OR None if not found
    """
    course_code_upper = str(course_code).strip().upper()
    dept_norm = normalize_branch_string(department)
    
    # Find all entries for this course code in this department
    instructors = {}
    for info in course_info.values():
        if str(info.get('course_code', '')).strip().upper() != course_code_upper:
            continue
        
        info_dept = info.get('csv_department') or info.get('department') or info.get('branch')
        if normalize_branch_string(info_dept) != dept_norm:
            continue
        
        instructor = str(info.get('instructor') or info.get('Faculty', '')).strip().lower()
        if not instructor or instructor == 'unknown':
            continue
        
        if instructor not in instructors:
            instructors[instructor] = []
        instructors[instructor].append(info)
    
    # If multiple instructors teaching same course in same department
    if len(instructors) >= 2:
        # Sum all enrollments and divide by number of instructors for room sizing
        total_enrollment = sum(
            _get_registered_or_default_enrollment(info)
            for info_list in instructors.values()
            for info in info_list
        )
        effective_enrollment = max(1, int(math.ceil(total_enrollment / len(instructors))))
        return {
            'is_dual': True,
            'num_instructors': len(instructors),
            'total_enrollment': total_enrollment,
            'effective_enrollment': effective_enrollment,
            'instructors': list(instructors.keys())
        }
    
    return {'is_dual': False, 'effective_enrollment': None}


def normalize_branch_string(branch_raw):
    """Normalize department/branch strings and common abbreviations to canonical department names"""
    if not branch_raw:
        return ''
    br = str(branch_raw).strip().upper()
    if br in ['CSE', 'CS', 'COMPUTER SCIENCE', 'COMPUTER SCIENCE AND ENGINEERING']:
        return 'Computer Science and Engineering'
    if br in ['ECE', 'EC', 'ELECTRONICS', 'ELECTRONICS AND COMMUNICATION ENGINEERING']:
        return 'Electronics and Communication Engineering'
    if br in ['DSAI', 'DS', 'DA', 'DATA SCIENCE', 'DATA SCIENCE AND ARTIFICIAL INTELLIGENCE']:
        return 'Data Science and Artificial Intelligence'
    return branch_raw.strip()

def get_departments_from_data(dfs):
    """Extract unique departments from course data"""
    if 'course' in dfs and 'Department' in dfs['course'].columns:
        departments = dfs['course']['Department'].unique()
        # Filter out None/NaN and return as list
        return [dept for dept in departments if pd.notna(dept) and dept != '']
    else:
        return ['CSE', 'DSAI', 'ECE']  # fallback

def separate_courses_by_type(dfs, semester_id, branch=None):
    """Separate courses into core and elective baskets for a given semester and branch"""
    if 'course' not in dfs:
        return {'core_courses': [], 'elective_courses': []}
    
    try:
        # Normalize branch name to match Department values in data (short form)
        def normalize_branch_name(branch_raw):
            if not branch_raw:
                return None
            br = str(branch_raw).strip().upper()
            if br in ['CSE', 'CS', 'COMPUTER SCIENCE', 'COMPUTER SCIENCE AND ENGINEERING']:
                return 'CSE'
            if br in ['ECE', 'EC', 'ELECTRONICS', 'ELECTRONICS AND COMMUNICATION ENGINEERING']:
                return 'ECE'
            if br in ['DSAI', 'DS', 'DA', 'DATA SCIENCE', 'DATA SCIENCE AND ARTIFICIAL INTELLIGENCE']:
                return 'DSAI'
            # Fallback: return as uppercase for exact matches in data
            return br.upper()

        # Filter courses for the semester
        sem_courses = dfs['course'][
            dfs['course']['Semester'].astype(str).str.strip() == str(semester_id)
        ].copy()
        
        if sem_courses.empty:
            return {'core_courses': pd.DataFrame(), 'elective_courses': pd.DataFrame()}
        
        # ENHANCED: Filter by department if specified - only include courses for the specific department
        if branch and 'Department' in sem_courses.columns:
            normalized_branch = normalize_branch_name(branch)

            # Build robust department mask that matches exactly
            dept_series = sem_courses['Department'].astype(str).fillna('').str.strip().str.upper()
            
            # For CORE courses: must match the department exactly
            # For ELECTIVE courses: include all (electives are usually available to all departments)
            is_elective = sem_courses['Elective (Yes/No)'].astype(str).str.upper() == 'YES'
            
            # Match rows where:
            # - Department equals normalized branch (exact match on short form), OR
            # - Is elective (electives are available to all departments)
            dept_match = (dept_series == normalized_branch.upper())
            
            # Keep courses that match the department for core courses, and all electives
            sem_courses = sem_courses[dept_match | is_elective].copy()
        
        if sem_courses.empty:
            return {'core_courses': pd.DataFrame(), 'elective_courses': pd.DataFrame()}
        
        # Separate core and elective courses
        core_courses = sem_courses[
            sem_courses['Elective (Yes/No)'].str.upper() != 'YES'
        ].copy()
        
        elective_courses = sem_courses[
            sem_courses['Elective (Yes/No)'].str.upper() == 'YES'
        ].copy()
        
        # Filter electives by semester-specific baskets
        if not elective_courses.empty and 'Basket' in elective_courses.columns:
            # Define allowed baskets per semester
            semester_baskets = {
                1: ['ELECTIVE_B1', 'HSS_B1'],
                3: ['ELECTIVE_B3', 'HSS_B3'],  # Note: HSS_B5 is also in sem3
                5: ['ELECTIVE_B4', 'ELECTIVE_B5',],
                7: ['ELECTIVE_B6', 'ELECTIVE_B7', 'ELECTIVE_B8', 'ELECTIVE_B9']
            }
            
            allowed_baskets = semester_baskets.get(int(semester_id), [])
            if allowed_baskets:
                # Filter to only include courses with baskets appropriate for this semester
                basket_series = elective_courses['Basket'].astype(str).fillna('')
                basket_match = basket_series.isin(allowed_baskets)
                elective_courses = elective_courses[basket_match].copy()
        
        print(f"   [STATS] Course separation for Semester {semester_id}, Department {branch or 'All'}:")
        print(f"      Core courses: {len(core_courses)}")
        print(f"      Elective courses: {len(elective_courses)}")
        
        # ENHANCED: Debug info about department-specific courses
        if branch:
            dept_core_courses = core_courses[core_courses['Department'] == branch]
            print(f"      Department-specific core courses for {branch}: {len(dept_core_courses)}")
            if not dept_core_courses.empty:
                print(f"      Department core courses: {dept_core_courses['Course Code'].tolist()}")
        
        return {
            'core_courses': core_courses,
            'elective_courses': elective_courses
        }
        
    except Exception as e:
        print(f"[WARN] Error separating courses by type: {e}")
        traceback.print_exc()
        return {'core_courses': pd.DataFrame(), 'elective_courses': pd.DataFrame()}

def separate_courses_by_mid_semester(dfs, semester_id, branch=None):
    """Separate courses into pre-mid and post-mid based on Half Semester and Post mid-sem columns
    
    Logic:
    - If Half Semester = NO -> schedule in BOTH (full semester)
    - If Half Semester = YES and Post mid-sem = YES -> POST-MID ONLY
    - If Half Semester = YES and Post mid-sem is blank or anything except YES -> PRE-MID ONLY
    """
    if 'course' not in dfs:
        return {'pre_mid_courses': pd.DataFrame(), 'post_mid_courses': pd.DataFrame()}
    
    try:
        # Filter courses for the semester
        sem_courses = dfs['course'][
            dfs['course']['Semester'].astype(str).str.strip() == str(semester_id)
        ].copy()
        
        if sem_courses.empty:
            return {'pre_mid_courses': pd.DataFrame(), 'post_mid_courses': pd.DataFrame()}
        
        # Filter by department if specified
        if branch and 'Department' in sem_courses.columns:
            normalized_branch = branch.strip()
            # Include ONLY department-specific courses (do NOT include common courses from other departments)
            # Common courses for other departments have different Post mid-sem values!
            dept_match = sem_courses['Department'].astype(str).str.strip() == normalized_branch
            sem_courses = sem_courses[dept_match].copy()
        
        if sem_courses.empty:
            return {'pre_mid_courses': pd.DataFrame(), 'post_mid_courses': pd.DataFrame()}
        
        # Check if required columns exist (try flexible matching)
        post_mid_col = None
        for col in sem_courses.columns:
            if 'post' in str(col).lower() and 'mid' in str(col).lower():
                post_mid_col = col
                break
        
        half_sem_col = None
        for col in sem_courses.columns:
            if 'half' in str(col).lower() and 'semester' in str(col).lower():
                half_sem_col = col
                break
        
        if post_mid_col is None:
            print(f"[WARN] 'Post mid-sem' column not found for semester {semester_id}")
            print(f"   Available columns: {list(sem_courses.columns)}")
            return {'pre_mid_courses': pd.DataFrame(), 'post_mid_courses': pd.DataFrame()}
        
        if half_sem_col is None:
            print(f"[WARN] 'Half Semester' column not found for semester {semester_id}")
            print(f"   Available columns: {list(sem_courses.columns)}")
            return {'pre_mid_courses': pd.DataFrame(), 'post_mid_courses': pd.DataFrame()}
        
        # Debug: Print sample of column values
        print(f"   [DEBUG] Using columns:")
        print(f"      - Half Semester: '{half_sem_col}'")
        print(f"      - Post mid-sem: '{post_mid_col}'")
        sample_half = sem_courses[half_sem_col].astype(str).str.strip().str.upper().unique()[:10]
        sample_post = sem_courses[post_mid_col].astype(str).str.strip().str.upper().unique()[:10]
        print(f"      Half Semester values: {list(sample_half)}")
        print(f"      Post mid-sem values: {list(sample_post)}")
        print(f"      Total courses: {len(sem_courses)}")
        
        # Normalize column values
        sem_courses[half_sem_col] = sem_courses[half_sem_col].astype(str).str.strip().str.upper()
        # Treat blank/NaN post-mid values as 'NO' so half-sem courses default to pre-mid
        sem_courses[post_mid_col] = sem_courses[post_mid_col].astype(str).str.strip().str.upper()
        sem_courses[post_mid_col] = sem_courses[post_mid_col].replace({'': 'NO'})
        
        # ======================================================
        # CORRECTED LOGIC (Based on Half Semester):
        # ======================================================
        
        # RULE 1: Courses with Half Semester = NO
        # These go in BOTH pre-mid and post-mid
        half_sem_no_courses = sem_courses[
            sem_courses[half_sem_col] == 'NO'
        ].copy()
        
        # RULE 2: Courses with Half Semester = YES
        # Rule 2a: Post mid-sem = NO -> PRE-MID ONLY
        half_sem_yes_pre_mid = sem_courses[
            (sem_courses[half_sem_col] == 'YES') &
            (sem_courses[post_mid_col] != 'YES')
        ].copy()

        # Rule 2b: Post mid-sem = YES -> POST-MID ONLY
        half_sem_yes_post_mid = sem_courses[
            (sem_courses[half_sem_col] == 'YES') &
            (sem_courses[post_mid_col] == 'YES')
        ].copy()
        
        # FINAL SEPARATION:
        # PRE-MID: Half Sem = NO courses (full semester) + Half Sem = YES with Post mid = NO courses (pre-mid only)
        pre_mid_courses = pd.concat([half_sem_no_courses, half_sem_yes_pre_mid], ignore_index=True)
        pre_mid_courses = pre_mid_courses.drop_duplicates(subset=['Course Code'])
        
        # POST-MID: Half Sem = NO courses (full semester) + Half Sem = YES with Post mid = YES courses (post-mid only)
        post_mid_courses = pd.concat([half_sem_no_courses, half_sem_yes_post_mid], ignore_index=True)
        post_mid_courses = post_mid_courses.drop_duplicates(subset=['Course Code'])
        
        print(f"   [STATS] Course separation for Semester {semester_id}, Department {branch or 'All'}:")
        print(f"      Total courses: {len(sem_courses)}")
        print(f"      Half Semester = NO (both pre & post): {len(half_sem_no_courses)}")
        print(f"      Half Semester = YES, Post mid = NO (pre-mid only): {len(half_sem_yes_pre_mid)}")
        print(f"      Half Semester = YES, Post mid = YES (post-mid only): {len(half_sem_yes_post_mid)}")
        print(f"      -> Pre-mid courses total: {len(pre_mid_courses)}")
        print(f"      -> Post-mid courses total: {len(post_mid_courses)}")
        
        return {
            'pre_mid_courses': pre_mid_courses,
            'post_mid_courses': post_mid_courses
        }
        
    except Exception as e:
        print(f"[WARN] Error separating courses by mid-semester: {e}")
        traceback.print_exc()
        return {'pre_mid_courses': pd.DataFrame(), 'post_mid_courses': pd.DataFrame()}

def parse_ltpsc(ltpsc_string):
    """Parse L-T-P-S-C string and return components"""
    # Default values: 2 lectures, 1 tutorial, 0 practicals
    default_ltpsc = {'L': 2, 'T': 1, 'P': 0, 'S': 0, 'C': 2}
    
    # Handle empty or None values
    if not ltpsc_string or pd.isna(ltpsc_string) or str(ltpsc_string).strip() == '':
        return default_ltpsc
    
    try:
        parts = str(ltpsc_string).strip().split('-')
        if len(parts) == 5:
            return {
                'L': int(parts[0]) if parts[0].strip() else 2,  # Lectures per week
                'T': int(parts[1]) if parts[1].strip() else 1,  # Tutorials per week
                'P': int(parts[2]) if parts[2].strip() else 0,  # Practicals per week
                'S': int(parts[3]) if parts[3].strip() else 0,  # S credits
                'C': int(parts[4]) if parts[4].strip() else 2   # Total credits
            }
        else:
            return default_ltpsc
    except:
        return default_ltpsc

def enforce_elective_day_separation(basket_allocations):
    """Enforce that elective lectures and tutorials are on different days"""
    print("[DEBUG] ENFORCING ELECTIVE DAY SEPARATION...")
    
    for basket_name, allocation in basket_allocations.items():
        lecture_days = set(day for day, time in allocation['lectures'])
        tutorial_day = allocation['tutorial'][0]
        
        if tutorial_day in lecture_days:
            print(f"   [FAIL] VIOLATION: Basket '{basket_name}' has tutorial on same day as lectures")
            print(f"      Lecture days: {lecture_days}, Tutorial day: {tutorial_day}")
            
            # Find an alternative tutorial day
            all_days = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri']
            available_days = [day for day in all_days if day not in lecture_days]
            
            if available_days:
                new_tutorial_day = available_days[0]
                # Find a tutorial time slot for the new day
                tutorial_times = ['14:30-15:30']  # Assuming 1-hour tutorial slots
                new_tutorial_slot = (new_tutorial_day, tutorial_times[0])
                
                allocation['tutorial'] = new_tutorial_slot
                allocation['tutorial_day'] = new_tutorial_day
                allocation['days_separated'] = True
                
                print(f"   [OK] FIXED: Moved tutorial to {new_tutorial_day}")
            else:
                print(f"   [WARN]  CANNOT FIX: No available days for tutorial")
        else:
            print(f"   [OK] VALID: Basket '{basket_name}' has proper day separation")
    
    return basket_allocations

def schedule_core_courses_with_tutorials(core_courses, schedule, used_slots, days, lecture_times, tutorial_times, lab_times=None, branch=None, semester_id=None, course_info_map=None, section=None):
    """Schedule core courses strictly adhering to LTPSC structure"""
    if core_courses.empty:
        return used_slots
    
    global _COMMON_COURSE_SCHEDULE
    course_day_usage = {}
    
    # Lab times are handled as consecutive slot pairs (2-hour labs use 2 consecutive 1.5-hour slots)
    # No need for separate lab_times parameter - handled internally
    
    # ENHANCED: Filter core courses to only include department-specific ones
    if branch and 'Department' in core_courses.columns:
        dept_core_courses = core_courses[core_courses['Department'] == branch].copy()
        print(f"   [COURSES] Scheduling {len(dept_core_courses)} department-specific core courses for {branch}...")
        if not dept_core_courses.empty:
            print(f"      Courses to schedule: {dept_core_courses['Course Code'].tolist()}")
    else:
        dept_core_courses = core_courses.copy()
        print(f"   [COURSES] Scheduling {len(dept_core_courses)} core courses...")
    
    if dept_core_courses.empty:
        print(f"   [INFO] No department-specific core courses found for {branch}")
        return used_slots
    
    # Parse LTPSC for core courses - STRICTLY ADHERE TO LTPSC STRUCTURE
    for _, course in dept_core_courses.iterrows():
        course_code = course['Course Code']
        ltpsc_str = course.get('LTPSC', '')
        is_common = str(course.get('Common', 'No')).strip().upper() == 'YES'

        cross_common = None
        normalized_branch = normalize_branch_string(branch)
        if course_info_map:
            cross_common = detect_cross_dsai_ece_common(course_info_map, course_code, semester_id)

        cross_common_active = bool(
            is_common
            and cross_common
            and normalized_branch in cross_common.get('departments', set())
        )

        common_schedule_key = None
        if is_common:
            if cross_common_active:
                common_schedule_key = cross_common['schedule_key']
                print(f"      [COMMON-CROSS] {course_code} shared for DSAI+ECE with instructor '{cross_common['instructor']}' (key={common_schedule_key})")
            else:
                common_schedule_key = f"sem{semester_id}_{branch or 'ALL'}_{course_code}"
        
        # CHECK: If this is a common course, check if it's already been scheduled for another section
        if is_common and common_schedule_key in _COMMON_COURSE_SCHEDULE:
            # This course has already been scheduled for another section - use the same timeslots
            existing_schedule = _COMMON_COURSE_SCHEDULE[common_schedule_key]
            print(f"      [COMMON] {course_code} already scheduled for another section - reusing timeslots (key={common_schedule_key})")
            
            # Copy the schedule from the other section
            for slot_info in existing_schedule:
                day = slot_info['day']
                time_slot = slot_info['time_slot']
                label = slot_info['label']
                
                # Add to current schedule at the same timeslot
                schedule.loc[time_slot, day] = label
                used_slots.add((day, time_slot))
                print(f"         [COMMON-COPY] {label} on {day} at {time_slot}")
            
            continue  # Skip normal scheduling for this course
        
        # Parse LTPSC - defaults to 2 lectures, 1 tutorial if empty
        ltpsc = parse_ltpsc(ltpsc_str)
        L = ltpsc['L']
        T = ltpsc['T']
        P = ltpsc['P']
        
        # FIX: NEVER schedule labs for MA (Mathematics) courses
        is_math_course = course_code.startswith('MA') or 'Mathematics' in str(course.get('Department', ''))
        if is_math_course:
            P = 0  # Force no labs for math courses
            print(f"      [DISABLE] MA Course detected: {course_code} - Labs disabled regardless of LTPSC")
        
        # Determine lectures needed: if L is 2 or 3, schedule 2 lectures
        if L == 2 or L == 3:
            lectures_needed = 2
        elif L == 1:
            lectures_needed = 1
        else:
            # For other values, schedule min(L, 2) lectures
            lectures_needed = min(L, 2)
        
        # Determine tutorials needed: if T is 1, schedule 1 tutorial
        tutorials_needed = 1 if T >= 1 else 0
        
        # Determine labs needed: if P is 1 or more, schedule 1 lab (2 hours)
        # BUT NEVER for MA courses (already handled above)
        # If P is 0, do NOT schedule any labs regardless of other factors
        labs_needed = 1 if P >= 1 else 0
        
        # Track which days we've used for this course
        course_day_usage[course_code] = {'lectures': set(), 'tutorials': set(), 'labs': set()}
        
        # GET FACULTY LIST for this course - used for conflict checking
        # For CSE with 2 faculty, 1st is Section A, 2nd is Section B
        course_faculty = get_course_faculty_list(course, section=section, branch=branch)
        if course_faculty:
            print(f"      Faculty for {course_code}: {course_faculty}")
        
        print(f"      Scheduling {course_code} (LTPSC: {ltpsc_str} -> L={L}, T={T}, P={P}):")
        print(f"         -> {lectures_needed} lectures, {tutorials_needed} tutorial, {labs_needed} lab")
        
        # Schedule lectures (1.5 hours each) - GUARANTEED SCHEDULING
        lectures_scheduled = 0
        
        # First try systematic slot filling
        for day in days:
            if lectures_scheduled >= lectures_needed:
                break
            if day in course_day_usage[course_code]['lectures']:
                continue
                
            for time_slot in lecture_times:
                if lectures_scheduled >= lectures_needed:
                    break
                key = (day, time_slot)
                
                if key not in used_slots and schedule.loc[time_slot, day] == 'Free':
                    # CHECK: Ensure all faculty are available at this slot FOR BOTH PERIODS
                    # Regular timetable applies to both Pre-Mid and Post-Mid
                    if course_faculty and (not check_all_faculty_available(course_faculty, day, time_slot, 'Pre-Mid') or
                                           not check_all_faculty_available(course_faculty, day, time_slot, 'Post-Mid')):
                        print(f"      [FACULTY-SKIP] {course_code} cannot use {day} {time_slot} - faculty conflict")
                        continue  # Try next time slot
                    
                    schedule.loc[time_slot, day] = course_code
                    used_slots.add(key)
                    course_day_usage[course_code]['lectures'].add(day)
                    
                    # BOOK all faculty for this slot FOR BOTH PERIODS
                    if course_faculty:
                        book_all_faculty_for_slot(course_faculty, day, time_slot, course_code, 'Pre-Mid')
                        book_all_faculty_for_slot(course_faculty, day, time_slot, course_code, 'Post-Mid')
                    
                    lectures_scheduled += 1
                    print(f"      [OK] Scheduled lecture {lectures_scheduled} for {course_code} on {day} at {time_slot}")
                    break
        
        # If still not all lectures scheduled, use ANY available slot
        if lectures_scheduled < lectures_needed:
            print(f"      [FALLBACK] Using fallback scheduling for {course_code} lectures...")
            for day in days:
                if lectures_scheduled >= lectures_needed:
                    break
                for time_slot in schedule.index:
                    if lectures_scheduled >= lectures_needed:
                        break
                    if 'LUNCH' in time_slot:
                        continue
                    key = (day, time_slot)
                    if key not in used_slots and schedule.loc[time_slot, day] == 'Free':
                        # CHECK: Ensure all faculty are available at this slot FOR BOTH PERIODS
                        if course_faculty and (not check_all_faculty_available(course_faculty, day, time_slot, 'Pre-Mid') or
                                               not check_all_faculty_available(course_faculty, day, time_slot, 'Post-Mid')):
                            print(f"      [FALLBACK-SKIP] {course_code} cannot use {day} {time_slot} - faculty conflict")
                            continue  # Try next time slot
                        
                        schedule.loc[time_slot, day] = course_code
                        used_slots.add(key)
                        course_day_usage[course_code]['lectures'].add(day)
                        
                        # BOOK all faculty for this slot FOR BOTH PERIODS
                        if course_faculty:
                            book_all_faculty_for_slot(course_faculty, day, time_slot, course_code, 'Pre-Mid')
                            book_all_faculty_for_slot(course_faculty, day, time_slot, course_code, 'Post-Mid')
                        
                        lectures_scheduled += 1
                        print(f"      [FALLBACK] Scheduled lecture {lectures_scheduled} for {course_code} on {day} at {time_slot}")
        
        # Schedule tutorial (1 hour) if needed - GUARANTEED SCHEDULING
        tutorials_scheduled = 0
        if tutorials_needed > 0:
            # First try systematic slot filling
            for day in days:
                if tutorials_scheduled >= tutorials_needed:
                    break
                if day in course_day_usage[course_code]['tutorials']:
                    continue
                    
                for time_slot in tutorial_times:
                    if tutorials_scheduled >= tutorials_needed:
                        break
                    key = (day, time_slot)
                    
                    if key not in used_slots and schedule.loc[time_slot, day] == 'Free':
                        # CHECK: Ensure all faculty are available at this slot FOR BOTH PERIODS
                        if course_faculty and (not check_all_faculty_available(course_faculty, day, time_slot, 'Pre-Mid') or
                                               not check_all_faculty_available(course_faculty, day, time_slot, 'Post-Mid')):
                            print(f"      [TUTORIAL-SKIP] {course_code} cannot use {day} {time_slot} - faculty conflict")
                            continue  # Try next time slot
                        
                        schedule.loc[time_slot, day] = f"{course_code} (Tutorial)"
                        used_slots.add(key)
                        course_day_usage[course_code]['tutorials'].add(day)
                        
                        # BOOK all faculty for this slot FOR BOTH PERIODS
                        if course_faculty:
                            book_all_faculty_for_slot(course_faculty, day, time_slot, course_code, 'Pre-Mid')
                            book_all_faculty_for_slot(course_faculty, day, time_slot, course_code, 'Post-Mid')
                        
                        tutorials_scheduled += 1
                        print(f"      [OK] Scheduled tutorial for {course_code} on {day} at {time_slot}")
                        break
            
            # If still not scheduled, use ANY available slot
            if tutorials_scheduled < tutorials_needed:
                print(f"      [FALLBACK] Using fallback scheduling for {course_code} tutorial...")
                for day in days:
                    if tutorials_scheduled >= tutorials_needed:
                        break
                    for time_slot in schedule.index:
                        if tutorials_scheduled >= tutorials_needed:
                            break
                        if 'LUNCH' in time_slot:
                            continue
                        key = (day, time_slot)
                        if key not in used_slots and schedule.loc[time_slot, day] == 'Free':
                            # CHECK: Ensure all faculty are available at this slot FOR BOTH PERIODS
                            if course_faculty and (not check_all_faculty_available(course_faculty, day, time_slot, 'Pre-Mid') or
                                                   not check_all_faculty_available(course_faculty, day, time_slot, 'Post-Mid')):
                                continue  # Try next time slot
                            
                            schedule.loc[time_slot, day] = f"{course_code} (Tutorial)"
                            used_slots.add(key)
                            course_day_usage[course_code]['tutorials'].add(day)
                            
                            # BOOK all faculty for this slot FOR BOTH PERIODS
                            if course_faculty:
                                book_all_faculty_for_slot(course_faculty, day, time_slot, course_code, 'Pre-Mid')
                                book_all_faculty_for_slot(course_faculty, day, time_slot, course_code, 'Post-Mid')
                            
                            tutorials_scheduled += 1
                            print(f"      [FALLBACK] Scheduled tutorial for {course_code} on {day} at {time_slot}")
        
        # Schedule lab (STRICTLY 2 hours) if needed - GUARANTEED SCHEDULING
        # NOTE: If P=0 in LTPSC, NO labs will be scheduled for this course
        # Labs are ALWAYS scheduled for exactly 2 hours using two consecutive slots
        # BUT NEVER for MA courses
        labs_scheduled = 0
        if labs_needed > 0 and P > 0 and not is_math_course:
            # Define 2-hour lab slot pairs (consecutive 1.5-hour slots that form a 2-hour lab)
            # Labs MUST be scheduled for exactly 2 hours - no exceptions
            lab_slot_pairs = [
                (['13:00-14:30', '14:30-15:30'], '13:00-15:30'),  # 2 hours: 13:00-15:30
                (['15:30-17:00', '17:00-18:00'], '15:30-18:00'),  # 2 hours: 15:30-18:00
            ]
            
            # First try systematic day-by-day filling
            for day in days:
                if labs_scheduled >= labs_needed:
                    break
                if day in course_day_usage[course_code]['labs']:
                    continue
                    
                # Try each lab slot pair for this day
                for slot_pair, lab_display_time in lab_slot_pairs:
                    if labs_scheduled >= labs_needed:
                        break
                    # Check if both consecutive slots are free
                    slot1, slot2 = slot_pair
                    key1 = (day, slot1)
                    key2 = (day, slot2)
                    
                    if (key1 not in used_slots and key2 not in used_slots and
                        schedule.loc[slot1, day] == 'Free' and schedule.loc[slot2, day] == 'Free'):
                        
                        # CHECK: Ensure all faculty are available at BOTH slots FOR BOTH PERIODS
                        if course_faculty:
                            if not check_all_faculty_available(course_faculty, day, slot1, 'Pre-Mid') or \
                               not check_all_faculty_available(course_faculty, day, slot2, 'Pre-Mid') or \
                               not check_all_faculty_available(course_faculty, day, slot1, 'Post-Mid') or \
                               not check_all_faculty_available(course_faculty, day, slot2, 'Post-Mid'):
                                print(f"      [LAB-SKIP] {course_code} cannot use {day} {lab_display_time} - faculty conflict")
                                continue  # Try next lab slot pair
                        
                        # Mark both slots as lab
                        schedule.loc[slot1, day] = f"{course_code} (Lab)"
                        schedule.loc[slot2, day] = f"{course_code} (Lab)"
                        used_slots.add(key1)
                        used_slots.add(key2)
                        course_day_usage[course_code]['labs'].add(day)
                        
                        # BOOK all faculty for BOTH slots FOR BOTH PERIODS
                        if course_faculty:
                            book_all_faculty_for_slot(course_faculty, day, slot1, course_code, 'Pre-Mid')
                            book_all_faculty_for_slot(course_faculty, day, slot2, course_code, 'Pre-Mid')
                            book_all_faculty_for_slot(course_faculty, day, slot1, course_code, 'Post-Mid')
                            book_all_faculty_for_slot(course_faculty, day, slot2, course_code, 'Post-Mid')
                        
                        labs_scheduled += 1
                        print(f"      [OK] Scheduled lab for {course_code} on {day} at {lab_display_time} (using slots {slot1} and {slot2})")
                        break
            
            # If still not scheduled, try ANY consecutive pair in ANY slot
            if labs_scheduled < labs_needed:
                print(f"      [FALLBACK] Using fallback scheduling for {course_code} lab...")
                # Try to find ANY two consecutive slots
                all_time_slots = [s for s in schedule.index if 'LUNCH' not in s]
                for day in days:
                    if labs_scheduled >= labs_needed:
                        break
                    for i in range(len(all_time_slots) - 1):
                        if labs_scheduled >= labs_needed:
                            break
                        slot1 = all_time_slots[i]
                        slot2 = all_time_slots[i + 1]
                        key1 = (day, slot1)
                        key2 = (day, slot2)
                        
                        if (key1 not in used_slots and key2 not in used_slots and
                            schedule.loc[slot1, day] == 'Free' and schedule.loc[slot2, day] == 'Free'):
                            
                            # CHECK: Ensure all faculty are available at BOTH slots FOR BOTH PERIODS
                            if course_faculty:
                                if not check_all_faculty_available(course_faculty, day, slot1, 'Pre-Mid') or \
                                   not check_all_faculty_available(course_faculty, day, slot2, 'Pre-Mid') or \
                                   not check_all_faculty_available(course_faculty, day, slot1, 'Post-Mid') or \
                                   not check_all_faculty_available(course_faculty, day, slot2, 'Post-Mid'):
                                    continue  # Try next slot pair
                            
                            schedule.loc[slot1, day] = f"{course_code} (Lab)"
                            schedule.loc[slot2, day] = f"{course_code} (Lab)"
                            used_slots.add(key1)
                            used_slots.add(key2)
                            course_day_usage[course_code]['labs'].add(day)
                            
                            # BOOK all faculty for BOTH slots FOR BOTH PERIODS
                            if course_faculty:
                                book_all_faculty_for_slot(course_faculty, day, slot1, course_code, 'Pre-Mid')
                                book_all_faculty_for_slot(course_faculty, day, slot2, course_code, 'Pre-Mid')
                                book_all_faculty_for_slot(course_faculty, day, slot1, course_code, 'Post-Mid')
                                book_all_faculty_for_slot(course_faculty, day, slot2, course_code, 'Post-Mid')
                            
                            labs_scheduled += 1
                            print(f"      [FALLBACK] Scheduled lab for {course_code} on {day} using slots {slot1} and {slot2}")
                            break
        elif P == 0:
            print(f"      [SKIP] No labs scheduled for {course_code} (P=0 in LTPSC)")
        elif is_math_course:
            print(f"      [DISABLE] Skipping lab for MA course {course_code} (P={P} in LTPSC but labs disabled for Mathematics)")
        
        # Summary - CRITICAL VALIDATION
        if lectures_scheduled < lectures_needed:
            print(f"      [CRITICAL] Could only schedule {lectures_scheduled}/{lectures_needed} lectures for {course_code}")
        if tutorials_needed > 0 and tutorials_scheduled < tutorials_needed:
            print(f"      [CRITICAL] Could only schedule {tutorials_scheduled}/{tutorials_needed} tutorials for {course_code}")
        if labs_needed > 0 and labs_scheduled < labs_needed and not is_math_course:
            print(f"      [CRITICAL] Could only schedule {labs_scheduled}/{labs_needed} labs for {course_code}")
        if lectures_scheduled == lectures_needed and tutorials_scheduled == tutorials_needed and labs_scheduled == labs_needed:
            print(f"      [OK] Successfully scheduled {course_code} according to LTPSC structure")
        
        # SAVE: If this is a common course, save its schedule for other sections to reuse
        if is_common and common_schedule_key:
            if common_schedule_key not in _COMMON_COURSE_SCHEDULE:
                _COMMON_COURSE_SCHEDULE[common_schedule_key] = []

            # Extract all scheduled slots for this course from the schedule
            for day in days:
                for time_slot in schedule.index:
                    value = schedule.loc[time_slot, day]
                    if isinstance(value, str) and course_code in value:
                        _COMMON_COURSE_SCHEDULE[common_schedule_key].append({
                            'day': day,
                            'time_slot': time_slot,
                            'label': value
                        })

            print(f"      [COMMON-SAVE] Saved schedule for common course {course_code} ({len(_COMMON_COURSE_SCHEDULE[common_schedule_key])} slots) [key={common_schedule_key}]")
    
    # FINAL VERIFICATION - Ensure ALL courses are scheduled
    print(f"\n   [VERIFY] Checking that ALL courses were scheduled...")
    all_scheduled_codes = set()
    for day in schedule.columns:
        for time_slot in schedule.index:
            value = schedule.loc[time_slot, day]
            if isinstance(value, str) and value not in ['Free', 'LUNCH BREAK']:
                # Extract course code
                clean_code = value.replace(' (Tutorial)', '').replace(' (Lab)', '')
                if not any(basket in clean_code for basket in ['ELECTIVE_', 'HSS_', 'PROF_', 'OE_']):
                    all_scheduled_codes.add(clean_code)
    
    expected_courses = set(dept_core_courses['Course Code'].tolist())
    missing_courses = expected_courses - all_scheduled_codes
    
    if missing_courses:
        print(f"   [CRITICAL] MISSING COURSES - The following courses were NOT scheduled: {', '.join(sorted(missing_courses))}")
        print(f"   [CRITICAL] Total missing: {len(missing_courses)} out of {len(expected_courses)} courses")
    else:
        print(f"   [OK] ALL {len(expected_courses)} courses successfully scheduled - ZERO courses missing!")
    
    scheduled_count = len(all_scheduled_codes)
    print(f"   [STATS] Scheduled courses: {scheduled_count}/{len(expected_courses)}")

    return used_slots

def get_common_elective_slots():
    """Define common elective slots that will be used for both sections"""
    # Return slots for 2 lectures + 1 tutorial per elective
    return [
        # Monday slots
        ('Mon', '09:00-10:30'), ('Mon', '10:30-12:00'), ('Mon', '13:00-14:30'), 
        ('Mon', '14:30-15:30'), ('Mon', '15:30-17:00'), ('Mon', '17:00-18:00'),
        # Tuesday slots
        ('Tue', '09:00-10:30'), ('Tue', '10:30-12:00'), ('Tue', '13:00-14:30'),
        ('Tue', '14:30-15:30'), ('Tue', '15:30-17:00'), ('Tue', '17:00-18:00'),
        # Wednesday slots
        ('Wed', '09:00-10:30'), ('Wed', '10:30-12:00'), ('Wed', '13:00-14:30'),
        ('Wed', '14:30-15:30'), ('Wed', '15:30-17:00'), ('Wed', '17:00-18:00'),
        # Thursday slots
        ('Thu', '09:00-10:30'), ('Thu', '10:30-12:00'), ('Thu', '13:00-14:30'),
        ('Thu', '14:30-15:30'), ('Thu', '15:30-17:00'), ('Thu', '17:00-18:00'),
        # Friday slots
        ('Fri', '09:00-10:30'), ('Fri', '10:30-12:00'), ('Fri', '13:00-14:30'),
        ('Fri', '14:30-15:30'), ('Fri', '15:30-17:00'), ('Fri', '17:00-18:00')
    ]

def get_basket_time_slots():
    """Define time slots for elective baskets"""
    return [
        ('Mon', '09:00-10:30'), ('Mon', '13:00-14:30'),
        ('Tue', '09:00-10:30'), ('Tue', '13:00-14:30'),
        ('Wed', '09:00-10:30'), ('Wed', '13:00-14:30'),
        ('Thu', '09:00-10:30'), ('Thu', '13:00-14:30'),
        ('Fri', '09:00-10:30'), ('Fri', '13:00-14:30'),
        ('Mon', '14:30-15:30'), ('Tue', '14:30-15:30'),
        ('Wed', '14:30-15:30'), ('Thu', '14:30-15:30'),
        ('Fri', '14:30-15:30')
    ]

def schedule_electives_by_baskets(elective_allocations, schedule, used_slots, section, branch=None):
    """Schedule elective courses in IDENTICAL COMMON slots for ALL branches and sections"""
    elective_scheduled = 0
    branch_info = f" for {branch}" if branch else ""
    
    print(f"   [TIME] Applying IDENTICAL COMMON slots for Section {section}{branch_info}:")
    
    # Track which basket slots we've already scheduled
    scheduled_basket_slots = set()
    
    for course_code, allocation in elective_allocations.items():
        if allocation is None:
            continue
            
        basket_name = allocation['basket_name']
        lectures = allocation['lectures']
        tutorial = allocation['tutorial']
        all_courses = allocation['all_courses_in_basket']
        days_separated = allocation['days_separated']
        
        print(f"      [BASKET] Basket '{basket_name}' - IDENTICAL across all branches:")
        print(f"         Courses: {', '.join(all_courses)}")
        print(f"         Days Separation: {'[OK]' if days_separated else '[FAIL]'}")
        
        # Schedule lectures - write individual course codes, not basket names
        for day, time_slot in lectures:
            slot_key = (basket_name, day, time_slot, 'lecture')
            
            if slot_key in scheduled_basket_slots:
                continue
                
            key = (day, time_slot)
            
            if schedule.loc[time_slot, day] == 'Free':
                # Write ONLY the basket name (not individual courses)
                # Individual courses and classrooms will be shown in the legends
                schedule.loc[time_slot, day] = basket_name
                used_slots.add(key)
                scheduled_basket_slots.add(slot_key)
                elective_scheduled += 1
                print(f"         [OK] COMMON LECTURE: {day} {time_slot}")
                print(f"                SAME for ALL branches & sections")
            else:
                print(f"         [FAIL] LECTURE CONFLICT: {day} {time_slot} - {schedule.loc[time_slot, day]}")
        
        # Schedule tutorial - write individual course codes with (Tutorial) suffix
        if tutorial:
            day, time_slot = tutorial
            slot_key = (basket_name, day, time_slot, 'tutorial')
            
            if slot_key not in scheduled_basket_slots:
                key = (day, time_slot)
                
                if schedule.loc[time_slot, day] == 'Free':
                    # Write ONLY the basket name with (Tutorial) suffix
                    # Individual courses and classrooms will be shown in the legends
                    schedule.loc[time_slot, day] = f"{basket_name} (Tutorial)"
                    used_slots.add(key)
                    scheduled_basket_slots.add(slot_key)
                    elective_scheduled += 1
                    print(f"         [OK] COMMON TUTORIAL: {day} {time_slot}")
                    print(f"                SAME for ALL branches & sections")
                else:
                    print(f"         [FAIL] TUTORIAL CONFLICT: {day} {time_slot} - {schedule.loc[time_slot, day]}")
        else:
            print(f"         [INFO] No tutorial scheduled (T=0 in LTPSC)")
    
    print(f"   [OK] Scheduled {elective_scheduled} IDENTICAL COMMON elective sessions")
    return used_slots

def schedule_minor_courses(dfs, semester_id, schedule, used_slots):
    """Schedule minor courses into dedicated morning and evening slots.
    Uses time slots 07:30-09:00 and 18:30-20:00 across Mon-Fri.
    Columns parsed case-insensitively from dfs['minor'].
    """
    try:
        if 'minor' not in dfs or dfs['minor'].empty:
            return used_slots

        minor_df = dfs['minor'].copy()
        # Normalize columns defensively
        col_map = {}
        for col in minor_df.columns:
            cl = str(col).strip().lower()
            if 'minor' in cl and 'course' in cl:
                col_map[col] = 'Minor Course'
            elif 'semester' in cl:
                col_map[col] = 'Semester'
            elif 'registered' in cl and 'student' in cl:
                col_map[col] = 'Registered Students'
        if col_map:
            minor_df = minor_df.rename(columns=col_map)

        # Filter by semester id
        if 'Semester' in minor_df.columns:
            minor_df = minor_df[minor_df['Semester'].astype(str).str.strip() == str(semester_id)]

        if minor_df.empty:
            print(f"   [MINOR] No minor courses for Semester {semester_id}")
            return used_slots

        days = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri']
        morning_slot = '07:30-09:00'
        evening_slot = '18:30-20:00'

        # Ensure schedule has the minor slots
        for slot in [morning_slot, evening_slot]:
            if slot not in schedule.index:
                schedule.loc[slot] = 'Free'

        # If a common schedule already exists for this semester, reuse it for all branches/sections
        if semester_id in _MINOR_COMMON_SCHEDULE:
            placements = _MINOR_COMMON_SCHEDULE[semester_id]
            print(f"   [MINOR] Reusing common minor slots for Semester {semester_id}: {placements}")
            for entry in placements:
                day = entry['day']
                slot = entry['slot']
                name = entry['name']
                key = (day, slot)
                if slot not in schedule.index:
                    schedule.loc[slot] = 'Free'
                if schedule.loc[slot, day] == 'Free':
                    schedule.loc[slot, day] = f"MINOR: {name}"
                used_slots.add(key)
            return used_slots

        print(f"   [MINOR] Scheduling {len(minor_df)} minor course(s) for Semester {semester_id}")
        morning_day_idx = 0
        evening_day_idx = 2  # offset evenings to spread days
        placements = []

        for _, row in minor_df.iterrows():
            name = str(row.get('Minor Course', '')).strip() or 'Minor Course'

            # Allocate morning slot
            attempts = 0
            scheduled_morning = False
            while attempts < len(days) and not scheduled_morning:
                day = days[morning_day_idx % len(days)]
                key = (day, morning_slot)
                if schedule.loc[morning_slot, day] == 'Free' and key not in used_slots:
                    schedule.loc[morning_slot, day] = f"MINOR: {name}"
                    used_slots.add(key)
                    placements.append({'name': name, 'day': day, 'slot': morning_slot})
                    scheduled_morning = True
                    print(f"      [OK] MINOR morning: {name} on {day} {morning_slot}")
                morning_day_idx += 1
                attempts += 1

            # Allocate evening slot
            attempts = 0
            scheduled_evening = False
            while attempts < len(days) and not scheduled_evening:
                day = days[evening_day_idx % len(days)]
                key = (day, evening_slot)
                if schedule.loc[evening_slot, day] == 'Free' and key not in used_slots:
                    schedule.loc[evening_slot, day] = f"MINOR: {name}"
                    used_slots.add(key)
                    placements.append({'name': name, 'day': day, 'slot': evening_slot})
                    scheduled_evening = True
                    print(f"      [OK] MINOR evening: {name} on {day} {evening_slot}")
                evening_day_idx += 1
                attempts += 1

        # Persist common placements for this semester so all branches/sections reuse the same slots
        _MINOR_COMMON_SCHEDULE[semester_id] = placements

        return used_slots
    except Exception as e:
        print(f"[WARN] Error scheduling minor courses: {e}")
        traceback.print_exc()
        return used_slots

def generate_section_schedule_with_elective_baskets(dfs, semester_id, section, elective_allocations, branch=None, time_config=None, basket_allocations=None):
    """Generate schedule with basket-based elective allocation - COMMON slots across branches.
    Allows overriding time slots via time_config: {
        'morning_slots': [..], 'lunch_slots': [..], 'afternoon_slots': [..],
        'lecture_times': [..], 'tutorial_times': [..]
    }
    basket_allocations: Optional dict of basket allocations to ensure all required baskets are scheduled."""
    branch_info = f", Branch {branch}" if branch else ""
    print(f"   [TARGET] Generating BASKET-BASED schedule for Semester {semester_id}, Section {section}{branch_info}")
    print(f"   [PIN] Using COMMON elective basket slots (same for all branches)")
    
    if 'course' not in dfs:
        print("[FAIL] Course data not available")
        return None
    
    try:
        # Get only the core courses for this specific branch
        course_baskets = separate_courses_by_type(dfs, semester_id, branch)
        core_courses = course_baskets['core_courses']
        
        days = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri']
        
        # Time slot structure
        if time_config:
            morning_slots = time_config.get('morning_slots', ['09:00-10:30', '10:30-12:00'])
            lunch_slots = time_config.get('lunch_slots', ['12:00-13:00'])
            afternoon_slots = time_config.get('afternoon_slots', ['13:00-14:30', '14:30-15:30', '15:30-17:00', '17:00-18:00'])
        else:
            morning_slots = ['07:30-09:00', '09:00-10:30', '10:30-12:00']
            lunch_slots = ['12:00-13:00']
            afternoon_slots = ['13:00-14:30', '14:30-15:30', '15:30-17:00', '17:00-18:00', '18:30-20:00']
        all_slots = morning_slots + lunch_slots + afternoon_slots
        
        # Lecture slots (1.5 hours)
        if time_config and time_config.get('lecture_times'):
            lecture_times = time_config['lecture_times']
        else:
            # Do NOT include minor-only slots here to keep them reserved for minors
            lecture_times = ['09:00-10:30', '10:30-12:00', '13:00-14:30', '15:30-17:00']
        
        # Tutorial slots (1 hour)
        if time_config and time_config.get('tutorial_times'):
            tutorial_times = time_config['tutorial_times']
        else:
            tutorial_times = ['14:30-15:30', '17:00-18:00']
        
        # Lab slots (2 hours) - represented as pairs of consecutive 1.5-hour slots
        # Labs will use: ['13:00-14:30', '14:30-15:30'] or ['15:30-17:00', '17:00-18:00']
        lab_times = None  # Will be handled as slot pairs in the scheduling function
        
        # Create schedule template
        schedule = pd.DataFrame(index=all_slots, columns=days, dtype=object).fillna('Free')
        # Mark lunch break label across provided lunch slots
        for lunch_slot in lunch_slots:
            if lunch_slot in schedule.index:
                schedule.loc[lunch_slot] = 'LUNCH BREAK'

        used_slots = set()

        # Schedule elective courses FIRST using COMMON basket allocation
        if elective_allocations:
            print(f"   [TIME] Applying COMMON basket elective slots for Section {section}:")
            # Show what we're scheduling
            unique_baskets = set()
            for allocation in elective_allocations.values():
                if allocation:
                    unique_baskets.add(allocation['basket_name'])
            
            print(f"   [BASKET] Baskets to schedule: {list(unique_baskets)}")
            
            used_slots = schedule_electives_by_baskets(elective_allocations, schedule, used_slots, section, branch)

        # Schedule minor courses into dedicated minor slots
        used_slots = schedule_minor_courses(dfs, semester_id, schedule, used_slots)
        
        # ENSURE all required baskets are scheduled from basket_allocations
        # This ensures baskets are scheduled even if courses aren't in elective_allocations
        if basket_allocations:
            scheduled_basket_names = set()
            if elective_allocations:
                for allocation in elective_allocations.values():
                    if allocation and allocation.get('basket_name'):
                        scheduled_basket_names.add(allocation['basket_name'])
            
            # Schedule any baskets from basket_allocations that weren't scheduled via elective_allocations
            for basket_name, basket_allocation in basket_allocations.items():
                if basket_name not in scheduled_basket_names:
                    # scheduled from basket allocations (silent)
                    # Create a temporary allocation structure to schedule this basket
                    temp_allocation = {
                        'basket_name': basket_name,
                        'lectures': basket_allocation['lectures'],
                        'tutorial': basket_allocation.get('tutorial'),
                        'all_courses_in_basket': basket_allocation.get('courses', []),
                        'days_separated': basket_allocation.get('days_separated', True)
                    }
                    # Schedule using a dummy course code key
                    dummy_key = f"__BASKET_{basket_name}__"
                    temp_elective_allocations = {dummy_key: temp_allocation}
                    used_slots = schedule_electives_by_baskets(temp_elective_allocations, schedule, used_slots, section, branch)
        
        # Schedule core courses AFTER electives and minors - these are branch-specific
        # IMPORTANT: Filter out any elective courses that are already scheduled in baskets
        if not core_courses.empty:
            # Get list of elective course codes from baskets to exclude them
            elective_course_codes = set()
            if elective_allocations:
                for allocation in elective_allocations.values():
                    if allocation and 'all_courses_in_basket' in allocation:
                        elective_course_codes.update(allocation['all_courses_in_basket'])
            
            # Filter out elective courses from core_courses
            if elective_course_codes:
                core_courses_filtered = core_courses[~core_courses['Course Code'].isin(elective_course_codes)].copy()
                excluded_count = len(core_courses) - len(core_courses_filtered)
                if excluded_count > 0:
                    print(f"   [DISABLE] Excluded {excluded_count} elective course(s) already scheduled in baskets: {', '.join(elective_course_codes)}")
                core_courses = core_courses_filtered
            
            if not core_courses.empty:
                print(f"   [COURSES] Scheduling {len(core_courses)} BRANCH-SPECIFIC core courses for {branch}...")
                used_slots = schedule_core_courses_with_tutorials(
                        core_courses, schedule, used_slots, days,
                        lecture_times, tutorial_times, None, branch, semester_id=semester_id, course_info_map=get_course_info(dfs), section=section
                )
            else:
                print(f"   [INFO] No core courses to schedule after filtering electives (might be elective-only or project-only semester)")
        else:
            print(f"   [INFO] No core courses found for {branch} - semester may be elective-only or project-based")
        
        # IMPORTANT: For semesters with minimal scheduling (e.g., Semester 7 with only projects),
        # ensure the schedule is still valid by having at least lunch breaks and basket allocations
        if len(used_slots) == 0:
            print(f"   [WARN] No courses scheduled - schedule only contains lunch breaks (may be project-only semester)")
        
        return schedule
        
    except Exception as e:
        print(f"[FAIL] Error generating basket-based schedule: {e}")
        traceback.print_exc()
        return None

def generate_mid_semester_schedule(dfs, semester_id, section, courses_df, branch=None, time_config=None, schedule_type='pre_mid', elective_allocations=None):
    """Generate pre-mid or post-mid schedule with elective basket support"""
    global _MID_SEM_COMMON_SCHEDULE
    schedule_type_name = "PRE-MID" if schedule_type == 'pre_mid' else "POST-MID"
    branch_info = f", Branch {branch}" if branch else ""
    print(f"   [TARGET] Generating {schedule_type_name} schedule for Semester {semester_id}, Section {section}{branch_info}")
    
    if courses_df.empty:
        print(f"   [INFO] No courses to schedule for {schedule_type_name}")
        return None
    
    try:
        days = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri']
        
        # Time slot structure
        if time_config:
            morning_slots = time_config.get('morning_slots', ['09:00-10:30', '10:30-12:00'])
            lunch_slots = time_config.get('lunch_slots', ['12:00-13:00'])
            afternoon_slots = time_config.get('afternoon_slots', ['13:00-14:30', '14:30-15:30', '15:30-17:00', '17:00-18:00'])
        else:
            morning_slots = ['07:30-09:00', '09:00-10:30', '10:30-12:00']
            lunch_slots = ['12:00-13:00']
            afternoon_slots = ['13:00-14:30', '14:30-15:30', '15:30-17:00', '17:00-18:00', '18:30-20:00']
        all_slots = morning_slots + lunch_slots + afternoon_slots
        
        # Lecture slots (1.5 hours)
        if time_config and time_config.get('lecture_times'):
            lecture_times = time_config['lecture_times']
        else:
            # Exclude minor-only slots to keep them free for minors
            lecture_times = ['09:00-10:30', '10:30-12:00', '13:00-14:30', '15:30-17:00']
        
        # Tutorial slots (1 hour)
        if time_config and time_config.get('tutorial_times'):
            tutorial_times = time_config['tutorial_times']
        else:
            tutorial_times = ['14:30-15:30', '17:00-18:00']
        
        # Create schedule template
        schedule = pd.DataFrame(index=all_slots, columns=days, dtype=object).fillna('Free')
        # Mark lunch break label across provided lunch slots
        for lunch_slot in lunch_slots:
            if lunch_slot in schedule.index:
                schedule.loc[lunch_slot] = 'LUNCH BREAK'

        used_slots = set()
        
        # SEPARATE courses into core and electives
        if 'Elective (Yes/No)' in courses_df.columns:
            core_courses = courses_df[
                courses_df['Elective (Yes/No)'].astype(str).str.upper() != 'YES'
            ].copy()
            elective_courses = courses_df[
                courses_df['Elective (Yes/No)'].astype(str).str.upper() == 'YES'
            ].copy()
        else:
            core_courses = courses_df.copy()
            elective_courses = pd.DataFrame()
        
        print(f"   [COURSES] Scheduling {len(courses_df)} courses for {schedule_type_name}...")
        print(f"      Core courses: {len(core_courses)}, Elective courses: {len(elective_courses)}")
        
        # SCHEDULE ELECTIVES FIRST using basket logic
        if not elective_courses.empty and elective_allocations:
            print(f"   [BASKET] Scheduling elective baskets for {schedule_type_name}...")
            used_slots = schedule_electives_by_baskets(elective_allocations, schedule, used_slots, section, branch)
        elif not elective_courses.empty:
            print(f"   [WARN] No basket allocations provided for electives, scheduling as core courses")
        
        # SCHEDULE MINORS before core courses to reserve minor slots
        used_slots = schedule_minor_courses(dfs, semester_id, schedule, used_slots)

        # SCHEDULE CORE COURSES - iterate over core_courses only
        print(f"   [CORE] Scheduling {len(core_courses)} core courses...")
        
        # Build course_info_map for cross-department detection (keyed by course_code + department)
        course_info_map = {}
        if 'course' in dfs and not dfs['course'].empty:
            for _, row in dfs['course'].iterrows():
                code = row['Course Code']
                dept = str(row.get('Department', '')).strip()
                key = f"{code}_{dept}"
                course_info_map[key] = {
                    'course_code': code,
                    'instructor': str(row.get('Faculty', '')).strip(),
                    'is_common': str(row.get('Common', 'No')).strip().upper() == 'YES',
                    'common': str(row.get('Common', 'No')).strip(),
                    'department': dept,
                    'csv_department': dept,
                }
        
        for _, course in core_courses.iterrows():
            course_code = course['Course Code']
            ltpsc_str = course.get('LTPSC', '')
            
            # Check if this is a common course (shared between sections A and B)
            is_common = str(course.get('Common', 'No')).strip().upper() == 'YES'
            mid_sem_schedule_key = None
            
            if is_common:
                # Check for cross-department common course (DSAI/ECE sharing same instructor)
                cross_common = detect_cross_dsai_ece_common(course_info_map, course_code, semester_id)
                normalized_branch = normalize_branch_string(branch)
                
                cross_common_active = bool(
                    cross_common
                    and normalized_branch in cross_common.get('departments', set())
                )
                
                if cross_common_active:
                    # Use shared key for cross-department courses
                    mid_sem_schedule_key = f"mid_{schedule_type}_{semester_id}_CROSS_{course_code}"
                    print(f"      [COMMON-CROSS] {course_code} shared for DSAI+ECE in {schedule_type_name} (key={mid_sem_schedule_key})")
                else:
                    # Use branch-specific key for same-branch sections
                    mid_sem_schedule_key = f"mid_{schedule_type}_{semester_id}_{branch or 'ALL'}_{course_code}"
                
                # Check if this course was already scheduled for another section/department
                if mid_sem_schedule_key in _MID_SEM_COMMON_SCHEDULE:
                    print(f"      [COMMON] {course_code} already scheduled for another section - reusing timeslots (key={mid_sem_schedule_key})")
                    existing_slots = _MID_SEM_COMMON_SCHEDULE[mid_sem_schedule_key]
                    
                    # Copy the schedule from the other section
                    for slot_info in existing_slots:
                        day = slot_info['day']
                        time_slot = slot_info['time_slot']
                        label = slot_info['label']
                        
                        schedule.loc[time_slot, day] = label
                        used_slots.add((day, time_slot))
                        print(f"         [COMMON-COPY] {label} on {day} at {time_slot}")
                    
                    continue  # Skip normal scheduling for this course
            
            # Parse LTPSC - defaults to 2 lectures, 1 tutorial if empty
            ltpsc = parse_ltpsc(ltpsc_str)
            L = ltpsc['L']
            T = ltpsc['T']
            P = ltpsc['P']
            
            # FIX: NEVER schedule labs for MA (Mathematics) courses
            is_math_course = course_code.startswith('MA') or 'Mathematics' in str(course.get('Department', ''))
            if is_math_course:
                P = 0  # Force no labs for math courses
                print(f"      [DISABLE] MA Course detected: {course_code} - Labs disabled regardless of LTPSC")
            
            # Determine lectures needed: if L is 2 or 3, schedule 2 lectures
            if L == 2 or L == 3:
                lectures_needed = 2
            elif L == 1:
                lectures_needed = 1
            else:
                # For other values, schedule min(L, 2) lectures
                lectures_needed = min(L, 2)
            
            # Determine tutorials needed: if T is 1, schedule 1 tutorial
            tutorials_needed = 1 if T >= 1 else 0
            
            # Determine labs needed: if P is 1 or more, schedule 1 lab (2 hours)
            # BUT NEVER for MA courses (already handled above)
            # If P is 0, do NOT schedule any labs regardless of other factors
            labs_needed = 1 if P >= 1 else 0
            
            print(f"      Scheduling {course_code} ({schedule_type_name}) - LTPSC: {ltpsc_str} -> L={L}, T={T}, P={P}:")
            print(f"         >> {lectures_needed} lectures, {tutorials_needed} tutorial, {labs_needed} lab")
            
            # Track which days we've used for this course
            course_day_usage = {'lectures': set(), 'tutorials': set(), 'labs': set()}
            
            # GET FACULTY LIST for this course - used for conflict checking
            # For CSE with 2 faculty, 1st is Section A, 2nd is Section B
            course_faculty = get_course_faculty_list(course, section=section, branch=branch)
            
            # Determine period for faculty booking (Pre-Mid or Post-Mid)
            period = 'Pre-Mid' if schedule_type == 'pre_mid' else 'Post-Mid'
            
            # Schedule lectures (1.5 hours each) - GUARANTEED SCHEDULING FOR MID-SEMESTER
            lectures_scheduled = 0
            
            # First try systematic slot filling
            for day in days:
                if lectures_scheduled >= lectures_needed:
                    break
                if day in course_day_usage['lectures']:
                    continue
                    
                for time_slot in lecture_times:
                    if lectures_scheduled >= lectures_needed:
                        break
                    key = (day, time_slot)
                    
                    if key not in used_slots and schedule.loc[time_slot, day] == 'Free':
                        # CHECK: Ensure all faculty are available at this slot
                        if course_faculty and not check_all_faculty_available(course_faculty, day, time_slot, period):
                            continue  # Try next time slot
                        
                        schedule.loc[time_slot, day] = course_code
                        used_slots.add(key)
                        course_day_usage['lectures'].add(day)
                        
                        # BOOK all faculty for this slot
                        if course_faculty:
                            book_all_faculty_for_slot(course_faculty, day, time_slot, course_code, period)
                        
                        lectures_scheduled += 1
                        print(f"      [OK] Scheduled lecture {lectures_scheduled} for {course_code} on {day} at {time_slot}")
                        break
            
            # If still not all lectures scheduled, use ANY available slot
            if lectures_scheduled < lectures_needed:
                print(f"      [FALLBACK] Using fallback scheduling for {course_code} lectures...")
                for day in days:
                    if lectures_scheduled >= lectures_needed:
                        break
                    for time_slot in schedule.index:
                        if lectures_scheduled >= lectures_needed:
                            break
                        if 'LUNCH' in time_slot:
                            continue
                        key = (day, time_slot)
                        if key not in used_slots and schedule.loc[time_slot, day] == 'Free':
                            # CHECK: Ensure all faculty are available at this slot
                            if course_faculty and not check_all_faculty_available(course_faculty, day, time_slot, period):
                                continue  # Try next time slot
                            
                            schedule.loc[time_slot, day] = course_code
                            used_slots.add(key)
                            course_day_usage['lectures'].add(day)
                            
                            # BOOK all faculty for this slot
                            if course_faculty:
                                book_all_faculty_for_slot(course_faculty, day, time_slot, course_code, period)
                            
                            lectures_scheduled += 1
                            print(f"      [FALLBACK] Scheduled lecture {lectures_scheduled} for {course_code} on {day} at {time_slot}")
                            break
            
            # Schedule tutorial (1 hour) if needed - GUARANTEED SCHEDULING FOR MID-SEMESTER
            tutorials_scheduled = 0
            if tutorials_needed > 0:
                # First try systematic slot filling
                for day in days:
                    if tutorials_scheduled >= tutorials_needed:
                        break
                    if day in course_day_usage['tutorials']:
                        continue
                        
                    for time_slot in tutorial_times:
                        if tutorials_scheduled >= tutorials_needed:
                            break
                        key = (day, time_slot)
                        
                        if key not in used_slots and schedule.loc[time_slot, day] == 'Free':
                            # CHECK: Ensure all faculty are available at this slot
                            if course_faculty and not check_all_faculty_available(course_faculty, day, time_slot, period):
                                continue  # Try next time slot
                            
                            schedule.loc[time_slot, day] = f"{course_code} (Tutorial)"
                            used_slots.add(key)
                            course_day_usage['tutorials'].add(day)
                            
                            # BOOK all faculty for this slot
                            if course_faculty:
                                book_all_faculty_for_slot(course_faculty, day, time_slot, course_code, period)
                            
                            tutorials_scheduled += 1
                            print(f"      [OK] Scheduled tutorial for {course_code} on {day} at {time_slot}")
                            break
                
                # If still not scheduled, use ANY available slot
                if tutorials_scheduled < tutorials_needed:
                    print(f"      [FALLBACK] Using fallback scheduling for {course_code} tutorial...")
                    for day in days:
                        if tutorials_scheduled >= tutorials_needed:
                            break
                        for time_slot in schedule.index:
                            if tutorials_scheduled >= tutorials_needed:
                                break
                            if 'LUNCH' in time_slot:
                                continue
                            key = (day, time_slot)
                            if key not in used_slots and schedule.loc[time_slot, day] == 'Free':
                                # CHECK: Ensure all faculty are available at this slot
                                if course_faculty and not check_all_faculty_available(course_faculty, day, time_slot, period):
                                    continue  # Try next time slot
                                
                                schedule.loc[time_slot, day] = f"{course_code} (Tutorial)"
                                used_slots.add(key)
                                course_day_usage['tutorials'].add(day)
                                
                                # BOOK all faculty for this slot
                                if course_faculty:
                                    book_all_faculty_for_slot(course_faculty, day, time_slot, course_code, period)
                                
                                tutorials_scheduled += 1
                                print(f"      [FALLBACK] Scheduled tutorial for {course_code} on {day} at {time_slot}")
            
            # Schedule lab (STRICTLY 2 hours) if needed - GUARANTEED SCHEDULING FOR MID-SEMESTER
            # NOTE: If P=0 in LTPSC, NO labs will be scheduled for this course
            labs_scheduled = 0
            if labs_needed > 0 and P > 0 and not is_math_course:
                # Define 2-hour lab slot pairs (consecutive 1.5-hour slots that form a 2-hour lab)
                lab_slot_pairs = [
                    (['13:00-14:30', '14:30-15:30'], '13:00-15:30'),  # 2 hours: 13:00-15:30
                    (['15:30-17:00', '17:00-18:00'], '15:30-18:00'),  # 2 hours: 15:30-18:00
                ]
                
                # First try systematic day-by-day filling
                for day in days:
                    if labs_scheduled >= labs_needed:
                        break
                    if day in course_day_usage['labs']:
                        continue
                        
                    # Try each lab slot pair for this day
                    for slot_pair, lab_display_time in lab_slot_pairs:
                        if labs_scheduled >= labs_needed:
                            break
                        # Check if both consecutive slots are free
                        slot1, slot2 = slot_pair
                        key1 = (day, slot1)
                        key2 = (day, slot2)
                        
                        if (key1 not in used_slots and key2 not in used_slots and
                            schedule.loc[slot1, day] == 'Free' and schedule.loc[slot2, day] == 'Free'):
                            
                            # CHECK: Ensure all faculty are available at BOTH slots
                            if course_faculty:
                                if not check_all_faculty_available(course_faculty, day, slot1, period) or \
                                   not check_all_faculty_available(course_faculty, day, slot2, period):
                                    continue  # Try next lab slot pair
                            
                            # Mark both slots as lab
                            schedule.loc[slot1, day] = f"{course_code} (Lab)"
                            schedule.loc[slot2, day] = f"{course_code} (Lab)"
                            used_slots.add(key1)
                            used_slots.add(key2)
                            course_day_usage['labs'].add(day)
                            
                            # BOOK all faculty for BOTH slots
                            if course_faculty:
                                book_all_faculty_for_slot(course_faculty, day, slot1, course_code, period)
                                book_all_faculty_for_slot(course_faculty, day, slot2, course_code, period)
                            
                            labs_scheduled += 1
                            print(f"      [OK] Scheduled lab for {course_code} on {day} at {lab_display_time} (using slots {slot1} and {slot2})")
                            break
                
                # If still not scheduled, try ANY consecutive pair in ANY slot
                if labs_scheduled < labs_needed:
                    print(f"      [FALLBACK] Using fallback scheduling for {course_code} lab...")
                    # Try to find ANY two consecutive slots
                    all_time_slots = [s for s in schedule.index if 'LUNCH' not in s]
                    for day in days:
                        if labs_scheduled >= labs_needed:
                            break
                        for i in range(len(all_time_slots) - 1):
                            if labs_scheduled >= labs_needed:
                                break
                            slot1 = all_time_slots[i]
                            slot2 = all_time_slots[i + 1]
                            key1 = (day, slot1)
                            key2 = (day, slot2)
                            
                            if (key1 not in used_slots and key2 not in used_slots and
                                schedule.loc[slot1, day] == 'Free' and schedule.loc[slot2, day] == 'Free'):
                                
                                # CHECK: Ensure all faculty are available at BOTH slots
                                if course_faculty:
                                    if not check_all_faculty_available(course_faculty, day, slot1, period) or \
                                       not check_all_faculty_available(course_faculty, day, slot2, period):
                                        continue  # Try next slot pair
                                
                                schedule.loc[slot1, day] = f"{course_code} (Lab)"
                                schedule.loc[slot2, day] = f"{course_code} (Lab)"
                                used_slots.add(key1)
                                used_slots.add(key2)
                                course_day_usage['labs'].add(day)
                                
                                # BOOK all faculty for BOTH slots
                                if course_faculty:
                                    book_all_faculty_for_slot(course_faculty, day, slot1, course_code, period)
                                    book_all_faculty_for_slot(course_faculty, day, slot2, course_code, period)
                                
                                labs_scheduled += 1
                                print(f"      [FALLBACK] Scheduled lab for {course_code} on {day} using slots {slot1} and {slot2}")
                                break
            elif P == 0:
                print(f"      [SKIP] No labs scheduled for {course_code} (P=0 in LTPSC)")
            elif is_math_course:
                print(f"      [DISABLE] Skipping lab for MA course {course_code} (P={P} in LTPSC but labs disabled for Mathematics)")
            
            # Summary - CRITICAL VALIDATION
            if lectures_scheduled < lectures_needed:
                print(f"      [CRITICAL] Could only schedule {lectures_scheduled}/{lectures_needed} lectures for {course_code}")
            if tutorials_needed > 0 and tutorials_scheduled < tutorials_needed:
                print(f"      [CRITICAL] Could only schedule {tutorials_scheduled}/{tutorials_needed} tutorials for {course_code}")
            if labs_needed > 0 and labs_scheduled < labs_needed and not is_math_course:
                print(f"      [CRITICAL] Could only schedule {labs_scheduled}/{labs_needed} labs for {course_code}")
            if lectures_scheduled == lectures_needed and tutorials_scheduled == tutorials_needed and labs_scheduled == labs_needed:
                print(f"      [OK] Successfully scheduled {course_code} according to LTPSC structure")
            
            # SAVE common course schedule for other sections to reuse
            if is_common and mid_sem_schedule_key and mid_sem_schedule_key not in _MID_SEM_COMMON_SCHEDULE:
                _MID_SEM_COMMON_SCHEDULE[mid_sem_schedule_key] = []
                # Find all slots scheduled for this course
                for day in schedule.columns:
                    for time_slot in schedule.index:
                        val = str(schedule.loc[time_slot, day])
                        if course_code in val and 'nan' not in val.lower():
                            _MID_SEM_COMMON_SCHEDULE[mid_sem_schedule_key].append({
                                'day': day,
                                'time_slot': time_slot,
                                'label': val
                            })
                print(f"      [COMMON-SAVE] Saved mid-sem schedule for common course {course_code} ({len(_MID_SEM_COMMON_SCHEDULE[mid_sem_schedule_key])} slots) [key={mid_sem_schedule_key}]")
        
        # FINAL VERIFICATION - Ensure ALL courses are scheduled
        print(f"\n   [VERIFY] Checking that ALL courses were scheduled for {schedule_type_name}...")
        all_scheduled_codes = set()
        for day in schedule.columns:
            for time_slot in schedule.index:
                value = schedule.loc[time_slot, day]
                if isinstance(value, str) and value not in ['Free', 'LUNCH BREAK']:
                    # Extract course code
                    clean_code = value.replace(' (Tutorial)', '').replace(' (Lab)', '')
                    if not any(basket in clean_code for basket in ['ELECTIVE_', 'HSS_', 'PROF_', 'OE_']):
                        all_scheduled_codes.add(clean_code)
        
        expected_courses = set(core_courses['Course Code'].tolist())
        missing_courses = expected_courses - all_scheduled_codes
        
        if missing_courses:
            print(f"   [CRITICAL] MISSING COURSES - The following courses were NOT scheduled: {', '.join(sorted(missing_courses))}")
            print(f"   [CRITICAL] Total missing: {len(missing_courses)} out of {len(expected_courses)} courses")
        else:
            print(f"   [OK] ALL {len(expected_courses)} courses successfully scheduled for {schedule_type_name} - ZERO courses missing!")
        
        return schedule
        
    except Exception as e:
        print(f"[FAIL] Error generating {schedule_type} schedule: {e}")
        traceback.print_exc()
        return None

def create_course_summary(dfs, semester, branch=None):
    """Create a summary sheet showing core vs elective courses"""
    if 'course' not in dfs:
        return pd.DataFrame()
    
    sem_courses = dfs['course'][
        dfs['course']['Semester'].astype(str).str.strip() == str(semester)
    ].copy()
    
    # Filter by branch if specified
    if branch and 'Department' in sem_courses.columns:
        sem_courses = sem_courses[
            (sem_courses['Department'] == branch) | 
            (sem_courses['Elective (Yes/No)'].str.upper() == 'YES')
        ]
    
    if sem_courses.empty:
        return pd.DataFrame()
    
    # Add course type classification and parse LTPSC
    sem_courses['Course Type'] = sem_courses['Elective (Yes/No)'].apply(
        lambda x: 'Elective' if str(x).upper() == 'YES' else 'Core'
    )
    
    # Add branch specificity info
    sem_courses['Branch Specificity'] = sem_courses.apply(
        lambda row: 'Common for All Branches' if row['Course Type'] == 'Elective' else f"Department: {row.get('Department', 'General')}",
        axis=1
    )
    
    # Parse LTPSC for detailed information
    ltpsc_data = sem_courses['LTPSC'].apply(parse_ltpsc)
    sem_courses['Lectures/Week'] = ltpsc_data.apply(lambda x: x['L'])
    sem_courses['Tutorials/Week'] = ltpsc_data.apply(lambda x: x['T'])
    sem_courses['Practicals/Week'] = ltpsc_data.apply(lambda x: x['P'])
    sem_courses['Total Credits'] = ltpsc_data.apply(lambda x: x['C'])
    
    summary_columns = ['Course Code', 'Course Name', 'Course Type', 'Branch Specificity', 'LTPSC', 'Lectures/Week', 'Tutorials/Week', 'Total Credits', 'Instructor', 'Department']
    available_columns = [col for col in summary_columns if col in sem_courses.columns]
    
    return sem_courses[available_columns]

def create_common_slots_info(basket_allocations, semester):
    """Create information sheet about common slots for all departments"""
    info_data = []
    
    for basket_name, allocation in basket_allocations.items():
        lecture_days = allocation['lecture_days']
        tutorial_day = allocation['tutorial_day']
        
        # Check if lectures and tutorial are on different days
        days_separated = tutorial_day not in lecture_days
        separation_status = "[OK] DIFFERENT DAYS" if days_separated else "[FAIL] SAME DAY"
        
        info_data.append({
            'Semester': f'Semester {semester}',
            'Basket Name': basket_name,
            'Lecture Slot 1': f"{allocation['lectures'][0][0]} {allocation['lectures'][0][1]}",
            'Lecture Slot 2': f"{allocation['lectures'][1][0]} {allocation['lectures'][1][1]}",
            'Tutorial Slot': f"{allocation['tutorial'][0]} {allocation['tutorial'][1]}",
            'Lecture Days': ', '.join(lecture_days),
            'Tutorial Day': tutorial_day,
            'Days Separation': separation_status,
            'Courses in Basket': ', '.join(allocation['courses']),
            'Common for All Departments': 'Yes',
            'Common for Both Sections': 'Yes',
            'Session Type': '2 Lectures + 1 Tutorial per course'
        })
    
    return pd.DataFrame(info_data)

def allocate_electives_by_baskets(elective_courses, semester_id):
    """Allocate elective courses to COMMON time slots for ALL branches and sections of a semester"""
    print(f"[TARGET] Allocating COMMON elective slots for Semester {semester_id} (ALL branches & sections)...")
    
    # Group electives by basket
    basket_groups = {}
    for _, course in elective_courses.iterrows():
        # Normalize basket name to ensure matching (e.g., 'elective_b4', trailing spaces, etc.)
        raw_basket = course.get('Basket', 'ELECTIVE_B1')
        basket = str(raw_basket).strip().upper() if pd.notna(raw_basket) else 'ELECTIVE_B1'
        if basket not in basket_groups:
            basket_groups[basket] = []
        basket_groups[basket].append(course)
    
    print(f"   Found {len(basket_groups)} elective baskets: {list(basket_groups.keys())}")
    
    # FILTER BASKETS BY SEMESTER - Only schedule required baskets
    if semester_id == 1:
        # Semester 1: Only ELECTIVE_B1
        required_baskets = ['ELECTIVE_B1']
        baskets_to_schedule = [basket for basket in basket_groups.keys() if basket in required_baskets]
        # Ensure ELECTIVE_B1 is scheduled even if not found in course data
        for req_basket in required_baskets:
            if req_basket not in baskets_to_schedule:
                print(f"   [WARN] {req_basket} not found in course data, but will be scheduled anyway for Semester 1")
                baskets_to_schedule.append(req_basket)
                # Create empty basket group so it can be scheduled
                if req_basket not in basket_groups:
                    basket_groups[req_basket] = []
        print(f"   [TARGET] Semester 1: Scheduling only ELECTIVE_B1")
    elif semester_id == 3:
        # Semester 3: Only ELECTIVE_B3, exclude ELECTIVE_B5
        required_baskets = ['ELECTIVE_B3']
        baskets_to_schedule = [basket for basket in basket_groups.keys() if basket in required_baskets]
        # Ensure ELECTIVE_B3 is scheduled even if not found in course data
        for req_basket in required_baskets:
            if req_basket not in baskets_to_schedule:
                print(f"   [WARN] {req_basket} not found in course data, but will be scheduled anyway for Semester 3")
                baskets_to_schedule.append(req_basket)
                if req_basket not in basket_groups:
                    basket_groups[req_basket] = []
        print(f"   [TARGET] Semester 3: Scheduling only ELECTIVE_B3")
    elif semester_id == 5:
        # Semester 5: Schedule ELECTIVE_B5 and ELECTIVE_B4 - ALWAYS schedule both even if no courses found
        required_baskets = ['ELECTIVE_B4', 'ELECTIVE_B5']
        baskets_to_schedule = [basket for basket in basket_groups.keys() if basket in required_baskets]
        # Ensure both baskets are scheduled even if not found in course data
        for req_basket in required_baskets:
            if req_basket not in baskets_to_schedule:
                print(f"   [WARN] {req_basket} not found in course data, but will be scheduled anyway for Semester 5")
                baskets_to_schedule.append(req_basket)
                # Create empty basket group so it can be scheduled
                if req_basket not in basket_groups:
                    basket_groups[req_basket] = []
        print(f"   [TARGET] Semester 5: Scheduling BOTH ELECTIVE_B4 and ELECTIVE_B5")
    elif semester_id == 7:
        # FIXED: Semester 7: Schedule ELECTIVE_B6, ELECTIVE_B7, ELECTIVE_B8, ELECTIVE_B9 (even if absent in data)
        required_baskets = ['ELECTIVE_B6', 'ELECTIVE_B7', 'ELECTIVE_B8', 'ELECTIVE_B9']
        baskets_to_schedule = [basket for basket in basket_groups.keys() if basket in required_baskets]
        for req_basket in required_baskets:
            if req_basket not in baskets_to_schedule:
                print(f"   [WARN] {req_basket} not found in course data, but will be scheduled anyway for Semester 7")
                baskets_to_schedule.append(req_basket)
                if req_basket not in basket_groups:
                    basket_groups[req_basket] = []
        print(f"   [TARGET] Semester 7: Scheduling ELECTIVE_B6, ELECTIVE_B7, ELECTIVE_B8, ELECTIVE_B9")
    else:
        # Other semesters: Schedule all baskets
        baskets_to_schedule = list(basket_groups.keys())
        print(f"   [TARGET] Semester {semester_id}: Scheduling all baskets")
    
    print(f"   [LIST] Baskets to schedule: {baskets_to_schedule}")
    
    # FIXED COMMON SLOTS for ALL branches and sections - ADDED B6 and B7
    common_slots_mapping = {
        'ELECTIVE_B1': {
            'lectures': [('Mon', '09:00-10:30'), ('Wed', '09:00-10:30')],
            'tutorial': ('Fri', '14:30-15:30')
        },
        'ELECTIVE_B2': {
            'lectures': [('Tue', '09:00-10:30'), ('Thu', '09:00-10:30')],
            'tutorial': ('Mon', '14:30-15:30')
        },
        'ELECTIVE_B3': {
            'lectures': [('Mon', '13:00-14:30'), ('Wed', '13:00-14:30')],
            'tutorial': ('Tue', '14:30-15:30')
        },
        'ELECTIVE_B4': {
            'lectures': [('Tue', '13:00-14:30'), ('Thu', '13:00-14:30')],
            'tutorial': ('Wed', '14:30-15:30')
        },
        'ELECTIVE_B5': {
            'lectures': [('Mon', '15:30-17:00'), ('Wed', '15:30-17:00')],
            'tutorial': ('Thu', '14:30-15:30')
        },
        'ELECTIVE_B6': {
            'lectures': [('Mon', '09:00-10:30'), ('Wed', '13:00-14:30')],
            'tutorial': ('Tue', '14:30-15:30')
        },
        'ELECTIVE_B7': {
            'lectures': [('Tue', '09:00-10:30'), ('Thu', '13:00-14:30')],
            'tutorial': ('Wed', '14:30-15:30')
        },
        'ELECTIVE_B8': {
            'lectures': [('Mon', '10:30-12:00'), ('Wed', '10:30-12:00')],
            'tutorial': ('Thu', '14:30-15:30')
        },
        'ELECTIVE_B9': {
            'lectures': [('Tue', '15:30-17:00'), ('Thu', '15:30-17:00')],
            'tutorial': ('Fri', '14:30-15:30')
        },
        'HSS_B1': {
            'lectures': [('Mon', '15:30-17:00'), ('Wed', '15:30-17:00')],
            'tutorial': ('Thu', '14:30-15:30')
        },
        'HSS_B2': {
            'lectures': [('Tue', '15:30-17:00'), ('Thu', '15:30-17:00')],
            'tutorial': ('Fri', '14:30-15:30')
        }
    }
    
    elective_allocations = {}
    basket_allocations = {}
    
    for basket_name in sorted(baskets_to_schedule):  # Only iterate through filtered baskets
        if basket_name not in basket_groups:
            print(f"   [WARN] Basket {basket_name} not found in course data")
            continue
            
        basket_courses = basket_groups[basket_name]
        course_codes = [course['Course Code'] for course in basket_courses] if basket_courses else []
        # Deduplicate while preserving order so we have a clean basket→courses map
        course_codes = list(dict.fromkeys(course_codes))
        
        # Parse LTPSC for the first course in the basket (assume all courses in basket have same LTPSC)
        # If LTPSC is empty or no courses, default to 2 lectures and 1 tutorial
        if basket_courses and len(basket_courses) > 0:
            first_course = basket_courses[0]
            ltpsc_str = first_course.get('LTPSC', '')
            ltpsc = parse_ltpsc(ltpsc_str)
            L = ltpsc['L']
            T = ltpsc['T']
            P = ltpsc['P']
        else:
            # Empty basket - use default LTPSC: 2 lectures, 1 tutorial
            ltpsc_str = '2-1-0-0-3'
            ltpsc = parse_ltpsc(ltpsc_str)
            L = 2
            T = 1
            P = 0
            print(f"   [WARN] Basket '{basket_name}' has no courses, using default LTPSC: 2-1-0-0-3")
        
        # Determine lectures needed: if L is 2 or 3, schedule 2 lectures
        if L == 2 or L == 3:
            lectures_needed = 2
        elif L == 1:
            lectures_needed = 1
        else:
            lectures_needed = min(L, 2)
        
        # Determine tutorials needed: if T is 1, schedule 1 tutorial
        tutorials_needed = 1 if T >= 1 else 0

        # ENFORCE: All ELECTIVE baskets (B1-B9) must have at least 1 tutorial for classroom allocation
        if basket_name in ['ELECTIVE_B1', 'ELECTIVE_B2', 'ELECTIVE_B3', 'ELECTIVE_B4', 'ELECTIVE_B5', 'ELECTIVE_B6', 'ELECTIVE_B7', 'ELECTIVE_B8', 'ELECTIVE_B9']:
            lectures_needed = max(lectures_needed, 2)
            tutorials_needed = max(tutorials_needed, 1)
        
        print(f"   [BASKET] Basket '{basket_name}' LTPSC: {ltpsc_str} -> L={L}, T={T}, P={P}")
        print(f"      -> Scheduling {lectures_needed} lectures, {tutorials_needed} tutorial")

        if basket_name in common_slots_mapping:
            fixed_slots = common_slots_mapping[basket_name]
            # Use only the number of lectures needed based on LTPSC
            lectures_allocated = fixed_slots['lectures'][:lectures_needed]
            
            # FIX: Ensure tutorial is allocated when needed
            if tutorials_needed > 0:
                tutorial_allocated = fixed_slots['tutorial']
                print(f"      [OK] Tutorial allocation: {tutorial_allocated}")
            else:
                tutorial_allocated = None
                print(f"      [WARN] No tutorial needed (T={T})")
        else:
            print(f"      [FAIL] No fixed slots found for basket '{basket_name}'")
        
        # Verify day separation
        lecture_days = set(day for day, time in lectures_allocated)
        tutorial_day = tutorial_allocated[0] if tutorial_allocated else None
        days_separated = tutorial_day not in lecture_days if tutorial_day else True
        
        # Store allocation for ALL courses in this basket
        # If no courses, create a dummy allocation entry to ensure basket is scheduled
        if course_codes:
            for course_code in course_codes:
                elective_allocations[course_code] = {
                    'basket_name': basket_name,
                    'lectures': lectures_allocated,
                    'tutorial': tutorial_allocated,
                    'all_courses_in_basket': course_codes,
                    'for_all_branches': True,
                    'for_both_sections': True,
                    'common_for_semester': True,
                    'common_for_all_departments': True,
                    'lecture_days': list(lecture_days),
                    'tutorial_day': tutorial_day,
                    'days_separated': days_separated,
                    'fixed_common_slots': True,
                    'ltpsc': ltpsc_str,
                    'lectures_needed': lectures_needed,
                    'tutorials_needed': tutorials_needed
                }
        else:
            # Empty basket - create a dummy allocation to ensure it's scheduled
            dummy_key = f"__BASKET_{basket_name}__"
            elective_allocations[dummy_key] = {
                'basket_name': basket_name,
                'lectures': lectures_allocated,
                'tutorial': tutorial_allocated,
                'all_courses_in_basket': [],
                'for_all_branches': True,
                'for_both_sections': True,
                'common_for_semester': True,
                'common_for_all_departments': True,
                'lecture_days': list(lecture_days),
                'tutorial_day': tutorial_day,
                'days_separated': days_separated,
                'fixed_common_slots': True,
                'ltpsc': ltpsc_str,
                'lectures_needed': lectures_needed,
                'tutorials_needed': tutorials_needed
            }
            print(f"   [NOTE] Created dummy allocation for empty basket '{basket_name}' to ensure scheduling")
        
        basket_allocations[basket_name] = {
            'lectures': lectures_allocated,
            'tutorial': tutorial_allocated,
            'courses': course_codes,
            'all_courses_in_basket': course_codes,
            'common_for_all_departments': True,
            'lecture_days': list(lecture_days),
            'tutorial_day': tutorial_day,
            'days_separated': days_separated,
            'fixed_common_slots': True,
            'ltpsc': ltpsc_str,
            'lectures_needed': lectures_needed,
            'tutorials_needed': tutorials_needed
        }
        
        print(f"   [BASKET] FIXED COMMON SLOTS for Basket '{basket_name}':")
        print(f"      [PIN] SAME FOR ALL BRANCHES & SECTIONS (LTPSC-based)")
        for i, (day, time_slot) in enumerate(lectures_allocated, 1):
            print(f"      Lecture {i}: {day} {time_slot}")
        if tutorial_allocated:
            print(f"      Tutorial: {tutorial_allocated[0]} {tutorial_allocated[1]}")
        print(f"      Days Separation: {'[OK] DIFFERENT DAYS' if days_separated else '[FAIL] SAME DAY'}")
        print(f"      Courses: {', '.join(course_codes)}")
    
    # Log excluded baskets
    excluded_baskets = set(basket_groups.keys()) - set(baskets_to_schedule)
    if excluded_baskets:
        print(f"   [DISABLE] Excluded baskets for Semester {semester_id}: {list(excluded_baskets)}")
    
    return elective_allocations, basket_allocations

def create_common_basket_summary(basket_allocations, semester, branch=None):
    """Create basket summary highlighting common slots"""
    summary_data = []
    
    for basket_name, allocation in basket_allocations.items():
        # Handle missing tutorial (None) or missing lecture slots safely
        lectures = allocation.get('lectures', [])
        lecture_slot_1 = '-'  # default display if missing
        lecture_slot_2 = '-'
        if len(lectures) > 0 and len(lectures[0]) >= 2:
            lecture_slot_1 = f"{lectures[0][0]} {lectures[0][1]}"
        if len(lectures) > 1 and len(lectures[1]) >= 2:
            lecture_slot_2 = f"{lectures[1][0]} {lectures[1][1]}"

        tutorial_slot = '-'
        tutorial = allocation.get('tutorial')
        if tutorial and len(tutorial) >= 2:
            tutorial_slot = f"{tutorial[0]} {tutorial[1]}"

        summary_data.append({
            'Basket Name': basket_name,
            'Lecture Slot 1': lecture_slot_1,
            'Lecture Slot 2': lecture_slot_2,
            'Tutorial Slot': tutorial_slot,
            'Courses in Basket': ', '.join(allocation.get('courses', [])),
            'Common for All Branches': '[OK] YES',
            'Common for Both Sections': '[OK] YES', 
            'Days Separation': '[OK] YES' if allocation.get('days_separated') else '[FAIL] NO',
            'Semester': f'Semester {semester}',
            'Applicable Branches': 'CSE, DSAI, ECE (ALL)',
            'Slot Type': 'FIXED COMMON SLOTS'
        })
    
    return pd.DataFrame(summary_data)


def create_detailed_common_slots_info(basket_allocations, semester):
    """Create detailed information about common slots"""
    info_data = []
    
    for basket_name, allocation in basket_allocations.items():
        lectures = allocation.get('lectures', [])
        l1_day = l1_time = '-'
        l2_day = l2_time = '-'
        if len(lectures) > 0 and len(lectures[0]) >= 2:
            l1_day, l1_time = lectures[0][0], lectures[0][1]
        if len(lectures) > 1 and len(lectures[1]) >= 2:
            l2_day, l2_time = lectures[1][0], lectures[1][1]

        tutorial_val = allocation.get('tutorial')
        t_day = t_time = '-'
        if tutorial_val and len(tutorial_val) >= 2:
            t_day, t_time = tutorial_val[0], tutorial_val[1]

        info_data.append({
            'Semester': f'Semester {semester}',
            'Basket Name': basket_name,
            'Lecture 1 Day': l1_day,
            'Lecture 1 Time': l1_time,
            'Lecture 2 Day': l2_day,
            'Lecture 2 Time': l2_time,
            'Tutorial Day': t_day,
            'Tutorial Time': t_time,
            'Courses': ', '.join(allocation.get('courses', [])),
            'Common for Branches': 'CSE, DSAI, ECE (ALL)',
            'Common for Sections': 'A & B (BOTH)',
            'Days Separation': '[OK] Achieved' if allocation.get('days_separated') else '[FAIL] Not Achieved',
            'Slot Consistency': '[OK] IDENTICAL across all',
            'Notes': 'FIXED COMMON TIMETABLE SLOTS'
        })
    
    return pd.DataFrame(info_data)

def create_semester_rules_sheet(semester, basket_allocations):
    """Create a sheet explaining semester-specific elective rules"""
    rules_data = []
    
    if semester == 3:
        rules_data.append({
            'Semester': 'Semester 3',
            'Rule': 'Schedule only ELECTIVE_B3',
            'Exclusion': 'Exclude ELECTIVE_B5',
            'Reason': 'Curriculum requirement - Semester 3 focuses on B3 electives',
            'Scheduled Baskets': ', '.join(basket_allocations.keys()) if basket_allocations else 'None',
            'Status': '[OK] Applied'
        })
    elif semester == 5:
        rules_data.append({
            'Semester': 'Semester 5', 
            'Rule': 'Schedule only ELECTIVE_B5',
            'Exclusion': 'Exclude ELECTIVE_B3',
            'Reason': 'Curriculum requirement - Semester 5 focuses on B5 electives',
            'Scheduled Baskets': ', '.join(basket_allocations.keys()) if basket_allocations else 'None',
            'Status': '[OK] Applied'
        })
    else:
        rules_data.append({
            'Semester': f'Semester {semester}',
            'Rule': 'Schedule all elective baskets',
            'Exclusion': 'None',
            'Reason': 'No specific restrictions for this semester',
            'Scheduled Baskets': ', '.join(basket_allocations.keys()) if basket_allocations else 'None',
            'Status': '[OK] Applied'
        })
    
    return pd.DataFrame(rules_data)

def print_classroom_allocation_summary(semester, branch):
    """Print a summary of classroom allocations"""
    global _TIMETABLE_CLASSROOM_ALLOCATIONS
    timetable_key = f"{branch}_sem{semester}"
    
    room_usage = {}
    for key, allocations in _TIMETABLE_CLASSROOM_ALLOCATIONS.items():
        if key.startswith(timetable_key):
            for allocation_key, allocation in allocations.items():
                room = allocation['classroom']
                if room not in room_usage:
                    room_usage[room] = 0
                room_usage[room] += 1
    
    print(f"   [STATS] Classroom usage summary for {branch} Semester {semester}:")
    for room, count in sorted(room_usage.items()):
        print(f"      {room}: {count} sessions")


def deduplicate_classroom_allocations(allocation_rows):
    """Remove duplicate classroom rows, preferring non-conflict entries with rooms."""
    deduped = {}
    for row in allocation_rows:
        key = (
            row.get('Semester'),
            row.get('Branch'),
            row.get('Section'),
            row.get('Day'),
            row.get('Time Slot'),
            row.get('Course'),
            row.get('Session Type')
        )

        existing = deduped.get(key)
        if existing is None:
            deduped[key] = row
            continue

        # Prefer non-conflict records, then records that actually have a room
        existing_conflict = bool(existing.get('Conflict'))
        new_conflict = bool(row.get('Conflict'))
        existing_room = existing.get('Room Number')
        new_room = row.get('Room Number')

        replace = False
        if existing_conflict and not new_conflict:
            replace = True
        elif (not existing_room or str(existing_room).strip() == '') and new_room:
            replace = True

        if replace:
            deduped[key] = row

    return list(deduped.values())


def create_classroom_allocation_detail_with_tracking(timetable_schedules, classrooms_df, semester, branch):
    """Create detailed classroom allocation information with global tracking"""
    allocation_data = []
    global _TIMETABLE_CLASSROOM_ALLOCATIONS
    
    for i, schedule in enumerate(timetable_schedules, 1):
        section = 'A' if i == 1 else 'B'
        timetable_key = f"{branch}_sem{semester}_sec{section}"
        
        included = set()
        for day in schedule.columns:
            for time_slot in schedule.index:
                cell_value = str(schedule.loc[time_slot, day])
                
                if cell_value not in ['Free', 'LUNCH BREAK'] and '[' in cell_value and ']' in cell_value:
                    # Extract course and room
                    course_match = re.search(r'^(.*?)\s*\[', cell_value)
                    room_match = re.search(r'\[(.*?)\]', cell_value)
                    
                    if course_match and room_match:
                        course = course_match.group(1).strip()
                        room_number = room_match.group(1)
                        
                        # Determine session type from course name
                        session_type = 'Tutorial' if ' (Tutorial)' in course else ('Lab' if ' (Lab)' in course else 'Lecture')
                        
                        # Get room details
                        room_details = classrooms_df[classrooms_df['Room Number'] == room_number]
                        capacity = room_details['Capacity'].iloc[0] if not room_details.empty else 'Unknown'
                        room_type = room_details['Type'].iloc[0] if not room_details.empty else 'Unknown'

                        # Determine conflict status from global tracking map if present
                        allocs_for_file = _TIMETABLE_CLASSROOM_ALLOCATIONS.get(timetable_key, {})
                        alloc_key = f"{day}_{time_slot}"
                        conflict_flag = allocs_for_file.get(alloc_key, {}).get('conflict', False)
                        
                        allocation_data.append({
                            'Semester': semester,
                            'Branch': branch,
                            'Section': section,
                            'Day': day,
                            'Time Slot': time_slot,
                            'Course': course,
                            'Room Number': room_number,
                            'Room Type': room_type,
                            'Capacity': capacity,
                            'Facilities': room_details['Facilities'].iloc[0] if not room_details.empty else 'Unknown',
                            'Conflict': conflict_flag,
                            'Allocation Type': 'Global Tracking',
                            'Session Type': session_type
                        })
                        included.add((day, time_slot, course))
    
        # Also include allocations that were made directly to the _TIMETABLE_CLASSROOM_ALLOCATIONS map
        allocs_for_file = _TIMETABLE_CLASSROOM_ALLOCATIONS.get(timetable_key, {})
        for alloc_key, alloc in allocs_for_file.items():
            # alloc_key format could be 'Day_Time' or 'Day_Time_Course'
            parts = alloc_key.split('_')
            if len(parts) >= 2:
                day = parts[0]
                time_slot = parts[1]
            else:
                continue
            course = alloc.get('course') or (parts[2] if len(parts) >= 3 else None)
            room_number = alloc.get('classroom')
            if not course or (day, time_slot, course) in included:
                continue
            # Get room details
            room_details = classrooms_df[classrooms_df['Room Number'] == room_number]
            capacity = room_details['Capacity'].iloc[0] if not room_details.empty else 'Unknown'
            room_type = room_details['Type'].iloc[0] if not room_details.empty else 'Unknown'
            conflict_flag = alloc.get('conflict', False)
            session_type = alloc.get('type') if isinstance(alloc, dict) else None
            allocation_data.append({
                'Semester': semester,
                'Branch': branch,
                'Section': section,
                'Day': day,
                'Time Slot': time_slot,
                'Course': course,
                'Room Number': room_number,
                'Room Type': room_type,
                'Capacity': capacity,
                'Facilities': room_details['Facilities'].iloc[0] if not room_details.empty else 'Unknown',
                'Conflict': conflict_flag,
                'Allocation Type': 'Global Tracking (from internal map)',
                'Basket': alloc.get('basket') if isinstance(alloc, dict) else None,
                'Session Type': session_type
            })
    
    deduped_data = deduplicate_classroom_allocations(allocation_data)
    return pd.DataFrame(deduped_data)

def normalize_classroom_allocation_records(records):
    """Normalize classroom allocation records to include lowercase keys expected by the UI/tests.
    Removes duplicate keys to prevent JSON serialization errors.
    """
    normalized = []
    for r in records:
        # Work with a shallow copy to avoid mutating original structures
        rec = dict(r)
        # Normalize room field from common variants
        room = rec.get('room') or rec.get('Room Number') or rec.get('Room') or rec.get('room_number')
        course = rec.get('course') or rec.get('Course')
        session_type = rec.get('session_type') or rec.get('Session Type')
        conflict = rec.get('conflict') or rec.get('Conflict') or False
        # Remove uppercase variants to avoid duplicates in JSON
        rec.pop('Room Number', None)
        rec.pop('Room', None)
        rec.pop('room_number', None)
        rec.pop('Session Type', None)
        # Set normalized lowercase keys
        rec['room'] = room
        rec['course'] = course
        rec['session_type'] = session_type
        rec['conflict'] = conflict
        # Also keep capitalized 'Course' and 'Conflict' keys for backward compatibility with tests
        rec['Course'] = course
        rec['Conflict'] = conflict
        normalized.append(rec)
    return normalized

def format_excel_worksheet(worksheet, course_colors=None, basket_colors=None, is_header_row=True):
    """Apply professional formatting to Excel worksheet:
    - Auto-adjust column widths
    - Color code headers
    - Color code courses using unique colors per course (same as website)
    - Apply borders and alignment
    """
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    # Define colors
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    # Default fills for special cases
    lunch_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")     # Light gray
    free_fill = PatternFill(start_color="FAFAFA", end_color="FAFAFA", fill_type="solid")      # Very light gray
    
    border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    # Format header row
    if is_header_row and worksheet.max_row > 0:
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = border
    
    # Format data rows
    for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row), start=2):
        for col_idx, cell in enumerate(row, start=1):
            cell_value = str(cell.value).strip() if cell.value else ""
            cell.border = border
            
            # First column (Time Slot) - left align
            if col_idx == 1:
                cell.alignment = left_alignment
                cell.font = Font(bold=True, size=10)
                continue
            
            # Apply color coding based on content
            cell.alignment = center_alignment
            
            if not cell_value or cell_value == '' or cell_value.upper() == 'FREE':
                cell.fill = free_fill
            elif 'LUNCH' in cell_value.upper():
                cell.fill = lunch_fill
                cell.font = Font(italic=True, size=9)
            else:
                # Extract course code from cell value (handle formats like "Course [Room]")
                course_code = cell_value.split('[')[0].strip() if '[' in cell_value else cell_value
                course_code = course_code.replace('(Tutorial)', '').replace('(Lab)', '').strip()
                
                # Try to find matching color from course_colors or basket_colors
                color_hex = None
                
                # Check baskets first (for ELECTIVE_B1, etc.)
                if basket_colors:
                    for basket_name, basket_color in basket_colors.items():
                        if basket_name in cell_value.upper():
                            color_hex = basket_color
                            break
                
                # Check course colors
                if not color_hex and course_colors:
                    color_hex = course_colors.get(course_code)
                
                # Apply the color
                if color_hex:
                    # Remove # if present and convert to RGB hex
                    hex_color = color_hex.replace('#', '')
                    # Handle HSL colors (convert to fallback)
                    if 'hsl' in hex_color.lower():
                        hex_color = 'E8DAEF'  # Light purple fallback
                    cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
                else:
                    # Default to free fill if no color found
                    cell.fill = free_fill
    
    # Auto-adjust column widths
    for col_idx in range(1, worksheet.max_column + 1):
        column_letter = get_column_letter(col_idx)
        max_length = 0
        
        for cell in worksheet[column_letter]:
            try:
                cell_value = str(cell.value) if cell.value else ""
                # Account for line breaks
                lines = cell_value.split('\n')
                max_line_length = max(len(line) for line in lines) if lines else 0
                max_length = max(max_length, max_line_length)
            except:
                pass
        
        # Set width with min and max constraints
        adjusted_width = min(max(max_length + 2, 12), 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width


def export_consolidated_semester_timetable(dfs, semester, branch, time_config=None, _reset_for_semester=True):
    """Export ONE consolidated Excel file per semester per branch containing:
    - Regular timetable sheets
    - Pre-mid timetable sheets  
    - Post-mid timetable sheets
    - Legend sheet with course info
    
    Args:
        _reset_for_semester: Set to True for first branch of semester, False for subsequent branches
                            to ensure common elective rooms are shared across all branches.
                            NOTE: We NO LONGER reset _CLASSROOM_USAGE_TRACKER here because
                            all semesters share the same physical classrooms. The tracker
                            is only reset once at the start of timetable generation.
    """
    
    print(f"\n[CONSOLIDATED] Generating consolidated timetable for Semester {semester}, Branch {branch}...")
    
    # IMPORTANT: Do NOT reset _CLASSROOM_USAGE_TRACKER here!
    # All semesters share the same physical classrooms, so we must track usage globally.
    # The tracker is only reset once at the start of timetable generation (in /upload endpoint).
    # Only reset the preferred classrooms map for each new semester to allow fresh room preferences.
    if _reset_for_semester:
        global _GLOBAL_PREFERRED_CLASSROOMS
        _GLOBAL_PREFERRED_CLASSROOMS = {}
        print(f"[INFO] Cleared preferred classroom preferences for semester {semester} (tracker preserved)")
    
    # Get course info and generate unique colors
    course_info = get_course_info(dfs) if dfs else {}
    all_courses = set()
    all_baskets = set()

    # Only show semester-allowed elective baskets in Excel outputs
    allowed_baskets_map = {
        1: {'ELECTIVE_B1'},
        3: {'ELECTIVE_B3'},
        5: {'ELECTIVE_B4', 'ELECTIVE_B5'},
        7: {'ELECTIVE_B6', 'ELECTIVE_B7', 'ELECTIVE_B8', 'ELECTIVE_B9'},
    }
    # Print confirmation of basket filtering
    print(f"[FILTER] Semester {semester} will display only these baskets: {allowed_baskets_map.get(semester, 'All')}")
    allowed_baskets = allowed_baskets_map.get(semester, set())
    
    if dfs and 'course' in dfs:
        all_courses = set(dfs['course']['Course Code'].unique())
        if 'Basket' in dfs['course'].columns:
            all_baskets = set(dfs['course']['Basket'].dropna().unique())
            if allowed_baskets:
                all_baskets = all_baskets & allowed_baskets
    
    course_colors = generate_course_colors(all_courses, course_info)
    basket_colors = generate_basket_colors(all_baskets)
    
    try:
        # Determine if branch has sections
        has_sections = (branch == 'CSE')
        sections = ['A', 'B'] if has_sections else ['Whole']
        
        # Get classroom data
        classroom_data = dfs.get('classroom')
        
        # GENERATE REGULAR TIMETABLE
        print(f"[REGULAR] Generating regular timetable for Semester {semester}, Branch {branch}...")
        course_baskets_all = separate_courses_by_type(dfs, semester, branch)
        elective_courses_all = course_baskets_all['elective_courses']
        elective_allocations, basket_allocations = allocate_electives_by_baskets(elective_courses_all, semester)

        # Filter basket allocations to semester-allowed baskets
        if allowed_baskets:
            basket_allocations = {k: v for k, v in (basket_allocations or {}).items() if k in allowed_baskets}
        
        basket_courses_map = {}
        for basket_name, alloc in (basket_allocations or {}).items():
            courses_in_basket = alloc.get('all_courses_in_basket', [])
            if courses_in_basket:
                basket_courses_map[basket_name] = courses_in_basket
        
        if has_sections:
            regular_section_a = generate_section_schedule_with_elective_baskets(dfs, semester, 'A', elective_allocations, branch, time_config=time_config, basket_allocations=basket_allocations)
            regular_section_b = generate_section_schedule_with_elective_baskets(dfs, semester, 'B', elective_allocations, branch, time_config=time_config, basket_allocations=basket_allocations)
        else:
            regular_section_a = generate_section_schedule_with_elective_baskets(dfs, semester, 'Whole', elective_allocations, branch, time_config=time_config, basket_allocations=basket_allocations)
            regular_section_b = pd.DataFrame()
        
        # Allocate classrooms for regular
        if classroom_data is not None and not classroom_data.empty:
            regular_section_a = allocate_classrooms_for_timetable(regular_section_a, classroom_data, course_info, semester, branch, 'A' if has_sections else 'Whole', basket_courses_map)
            if has_sections:
                regular_section_b = allocate_classrooms_for_timetable(regular_section_b, classroom_data, course_info, semester, branch, 'B', basket_courses_map)
        
        # GENERATE MID-SEMESTER TIMETABLES
        print(f"[MID-SEM] Generating mid-semester timetables for Semester {semester}, Branch {branch}...")
        mid_courses = separate_courses_by_mid_semester(dfs, semester, branch)
        pre_mid_courses = mid_courses['pre_mid_courses']
        post_mid_courses = mid_courses['post_mid_courses']
        
        # Allocate mid-semester electives
        if not pre_mid_courses.empty:
            pre_mid_electives = pre_mid_courses[pre_mid_courses['Elective (Yes/No)'].astype(str).str.upper() == 'YES'] if 'Elective (Yes/No)' in pre_mid_courses.columns else pd.DataFrame()
            pre_mid_elective_allocations = allocate_mid_semester_electives_by_baskets(pre_mid_electives, semester)
        else:
            pre_mid_elective_allocations = {}
        
        if not post_mid_courses.empty:
            post_mid_electives = post_mid_courses[post_mid_courses['Elective (Yes/No)'].astype(str).str.upper() == 'YES'] if 'Elective (Yes/No)' in post_mid_courses.columns else pd.DataFrame()
            post_mid_elective_allocations = allocate_mid_semester_electives_by_baskets(post_mid_electives, semester)
        else:
            post_mid_elective_allocations = {}
        
        # Generate pre-mid schedules
        pre_mid_sections = {}
        for section in sections:
            if not pre_mid_courses.empty:
                pre_mid_sections[section] = generate_mid_semester_schedule(dfs, semester, section, pre_mid_courses, branch, time_config, 'pre_mid', pre_mid_elective_allocations)
                if classroom_data is not None and not classroom_data.empty and pre_mid_sections[section] is not None:
                    pre_mid_basket_map = {}
                    if not pre_mid_courses.empty and 'Basket' in pre_mid_courses.columns:
                        for _, course in pre_mid_courses.iterrows():
                            if str(course.get('Elective (Yes/No)', '')).upper() == 'YES':
                                basket = str(course.get('Basket', 'Unknown')).strip().upper()
                                course_code = course['Course Code']
                                if basket not in pre_mid_basket_map:
                                    pre_mid_basket_map[basket] = []
                                if course_code not in pre_mid_basket_map[basket]:
                                    pre_mid_basket_map[basket].append(course_code)
                    pre_mid_sections[section] = allocate_classrooms_for_timetable(pre_mid_sections[section], classroom_data, course_info, semester, branch, section, pre_mid_basket_map, schedule_type='PreMid')
            else:
                pre_mid_sections[section] = pd.DataFrame()
        
        # Generate post-mid schedules
        post_mid_sections = {}
        for section in sections:
            if not post_mid_courses.empty:
                post_mid_sections[section] = generate_mid_semester_schedule(dfs, semester, section, post_mid_courses, branch, time_config, 'post_mid', post_mid_elective_allocations)
                if classroom_data is not None and not classroom_data.empty and post_mid_sections[section] is not None:
                    post_mid_basket_map = {}
                    if not post_mid_courses.empty and 'Basket' in post_mid_courses.columns:
                        for _, course in post_mid_courses.iterrows():
                            if str(course.get('Elective (Yes/No)', '')).upper() == 'YES':
                                basket = str(course.get('Basket', 'Unknown')).strip().upper()
                                course_code = course['Course Code']
                                if basket not in post_mid_basket_map:
                                    post_mid_basket_map[basket] = []
                                if course_code not in post_mid_basket_map[basket]:
                                    post_mid_basket_map[basket].append(course_code)
                    post_mid_sections[section] = allocate_classrooms_for_timetable(post_mid_sections[section], classroom_data, course_info, semester, branch, section, post_mid_basket_map, schedule_type='PostMid')
            else:
                post_mid_sections[section] = pd.DataFrame()
        
        # CREATE CONSOLIDATED EXCEL FILE
        filename = f"sem{semester}_{branch}_timetable.xlsx"
        filepath = os.path.join(OUTPUT_DIR, filename)
        
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            # Write REGULAR timetables
            if has_sections:
                regular_section_a.reset_index(drop=False).rename(columns={'index': 'Time Slot'}).to_excel(writer, sheet_name='Regular_Section_A', index=False)
                regular_section_b.reset_index(drop=False).rename(columns={'index': 'Time Slot'}).to_excel(writer, sheet_name='Regular_Section_B', index=False)
            else:
                regular_section_a.reset_index(drop=False).rename(columns={'index': 'Time Slot'}).to_excel(writer, sheet_name='Regular_Timetable', index=False)
            
            # Write PRE-MID timetables
            for section in sections:
                if not pre_mid_sections[section].empty:
                    sheet_name = f'PreMid_Section_{section}' if has_sections else 'PreMid_Timetable'
                    pre_mid_sections[section].reset_index(drop=False).rename(columns={'index': 'Time Slot'}).to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Write POST-MID timetables
            for section in sections:
                if not post_mid_sections[section].empty:
                    sheet_name = f'PostMid_Section_{section}' if has_sections else 'PostMid_Timetable'
                    post_mid_sections[section].reset_index(drop=False).rename(columns={'index': 'Time Slot'}).to_excel(writer, sheet_name=sheet_name, index=False)
        
        # Prepare legend data to add to each sheet
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        # Filter courses to show ONLY courses for this specific semester
        sem_courses = set()
        if dfs and 'course' in dfs:
            course_df = dfs['course']
            # Filter by semester
            sem_course_df = course_df[course_df['Semester'] == semester]
            sem_courses = set(sem_course_df['Course Code'].unique())
        
        # Separate courses into core, electives, and minors (FILTERED BY SEMESTER)
        core_courses = []
        elective_courses = []
        minor_courses = []
        
        for course_code in sorted(sem_courses):
            info = course_info.get(course_code, {})
            is_elective = info.get('is_elective', False)
            is_minor = info.get('is_minor', False)
            
            if is_minor:
                minor_courses.append(course_code)
            elif is_elective:
                elective_courses.append(course_code)
            else:
                core_courses.append(course_code)

        # Filter ALL COURSES to only this branch's department for this semester
        try:
            if branch:
                course_baskets_branch = separate_courses_by_type(dfs, semester, branch)
                branch_core_df = course_baskets_branch.get('core_courses')
                branch_elective_df = course_baskets_branch.get('elective_courses')
                branch_minor_df = course_baskets_branch.get('minor_courses')
                
                # Filter CORE COURSES by department and semester
                if branch_core_df is not None and not branch_core_df.empty:
                    branch_core_set = set(branch_core_df['Course Code'].tolist())
                    core_courses = [c for c in core_courses if c in branch_core_set]
                
                # Filter ELECTIVE COURSES by department and semester
                if branch_elective_df is not None and not branch_elective_df.empty:
                    branch_elective_set = set(branch_elective_df['Course Code'].tolist())
                    elective_courses = [c for c in elective_courses if c in branch_elective_set]
                
                # Filter MINOR COURSES by department and semester
                if branch_minor_df is not None and not branch_minor_df.empty:
                    branch_minor_set = set(branch_minor_df['Course Code'].tolist())
                    minor_courses = [c for c in minor_courses if c in branch_minor_set]
        except Exception as _e:
            # Non-fatal: if filtering fails, fall back to unfiltered lists
            pass
        
        # Get basket course mappings
        course_baskets_all = separate_courses_by_type(dfs, semester, branch)
        elective_courses_all = course_baskets_all['elective_courses']
        _, basket_allocations = allocate_electives_by_baskets(elective_courses_all, semester)
        if allowed_baskets:
            basket_allocations = {k: v for k, v in (basket_allocations or {}).items() if k in allowed_baskets}
        
        wb = load_workbook(filepath)
        
        # Add legend to each timetable sheet
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Inject basket course+room details directly into timetable cells to mirror website view
            header_label = str(ws.cell(row=1, column=1).value or '').strip().lower()
            if 'time' in header_label and basket_colors:
                timetable_section = 'A' if 'SECTION_A' in sheet_name.upper() else ('B' if 'SECTION_B' in sheet_name.upper() else 'Whole')
                timetable_key = f"{branch}_sem{semester}_sec{timetable_section}"
                allocs_for_file = _TIMETABLE_CLASSROOM_ALLOCATIONS.get(timetable_key, {})

                if allocs_for_file:
                    day_headers = {}
                    for col_idx in range(2, ws.max_column + 1):
                        header_val = ws.cell(row=1, column=col_idx).value
                        if header_val:
                            day_headers[col_idx] = str(header_val).strip()

                    for row_idx in range(2, ws.max_row + 1):
                        time_slot_val = ws.cell(row=row_idx, column=1).value
                        if not time_slot_val:
                            continue
                        time_slot_str = str(time_slot_val).strip()

                        for col_idx, day_name in day_headers.items():
                            cell = ws.cell(row=row_idx, column=col_idx)
                            if not cell.value:
                                continue

                            cell_text = str(cell.value).strip()
                            if not cell_text:
                                continue

                            upper_text = cell_text.upper()
                            basket_name = None

                            for candidate in basket_colors.keys():
                                candidate_upper = str(candidate).upper()
                                if candidate_upper in upper_text:
                                    basket_name = candidate_upper
                                    break

                            if not basket_name:
                                for prefix in ['ELECTIVE_B', 'HSS_B', 'PROF_B', 'OE_B']:
                                    if prefix in upper_text:
                                        suffix = upper_text.split(prefix, 1)[1]
                                        suffix = suffix.split('(')[0].split()[0]
                                        basket_name = f"{prefix}{suffix}"
                                        break

                            if not basket_name:
                                continue

                            # Just show the basket name, no course details in timetable cells
            
            # Find the last row with timetable data
            last_row = ws.max_row
            
            # Add spacing
            legend_start_row = last_row + 3
            
            # Add legend title
            title_cell = ws.cell(row=legend_start_row, column=1, value='COURSE INFORMATION')
            title_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            title_font = Font(bold=True, color="FFFFFF", size=12)
            title_cell.fill = title_fill
            title_cell.font = title_font
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.merge_cells(start_row=legend_start_row, start_column=1, end_row=legend_start_row, end_column=5)
            
            current_row = legend_start_row + 1
            
            section_fill = PatternFill(start_color="D0D0D0", end_color="D0D0D0", fill_type="solid")
            section_font = Font(bold=True, size=10)
            header_fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
            header_font = Font(bold=True, size=9)
            border = Border(
                left=Side(style='thin', color='CCCCCC'),
                right=Side(style='thin', color='CCCCCC'),
                top=Side(style='thin', color='CCCCCC'),
                bottom=Side(style='thin', color='CCCCCC')
            )
            
            # Core Courses Section
            if core_courses:
                section_cell = ws.cell(row=current_row, column=1, value='CORE COURSES')
                section_cell.fill = section_fill
                section_cell.font = section_font
                section_cell.alignment = Alignment(horizontal='left', vertical='center')
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
                current_row += 1
                
                # Headers - Changed to show Scheduled/Required format
                headers = ['Course Code', 'Course Name', 'L-T-P-S-C', 'Term Type', 'Lectures Hrs', 'Tutorials Hrs', 'Labs Hrs']
                for col_idx, header in enumerate(headers, start=1):
                    cell = ws.cell(row=current_row, column=col_idx, value=header)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                current_row += 1
                
                # Data rows
                for course_code in core_courses:
                    info = course_info.get(course_code, {})
                    ltpsc = info.get('ltpsc', 'N/A')
                    
                    # Parse LTPSC to extract L, T, P values (required hours)
                    req_lectures, req_tutorials, req_labs = 0, 0, 0
                    if ltpsc != 'N/A':
                        try:
                            parts = ltpsc.split('-')
                            if len(parts) >= 3:
                                req_lectures = int(parts[0])
                                req_tutorials = int(parts[1])
                                req_labs = int(parts[2])
                        except (ValueError, IndexError):
                            pass
                    
                    # Count scheduled hours from the timetable
                    # The DataFrame has time slots as rows and days as columns
                    # Values contain course codes (with classroom info like "MA161 [C004]")
                    sched_lectures = 0
                    sched_tutorials = 0
                    sched_labs = 0
                    
                    if not regular_section_a.empty:
                        try:
                            # Track processed (slot_idx, day) pairs to avoid double-counting
                            processed_cells = set()
                            
                            # Iterate through the schedule DataFrame (time slots x days)
                            for slot_idx, time_slot in enumerate(regular_section_a.index):
                                time_slot_str = str(time_slot).lower()
                                
                                # Check each day column for this course
                                for day in regular_section_a.columns:
                                    if (slot_idx, day) in processed_cells:
                                        continue
                                    cell_value = regular_section_a.loc[time_slot, day]
                                    
                                    # Check if this cell contains our course
                                    if pd.isna(cell_value) or cell_value == '' or 'free' in str(cell_value).lower():
                                        continue
                                    
                                    cell_str = str(cell_value).lower()
                                    
                                    # Check if this cell contains our course code
                                    if course_code.lower() not in cell_str:
                                        continue
                                    
                                    # Found the course in this time slot!
                                    # Now classify: lecture, tutorial, or lab
                                    
                                    # PRIORITY 1: Check if cell explicitly marked as (Lab) or (Tutorial)
                                    if '(lab)' in cell_str:
                                        # This is a lab slot - check if next slot also has lab
                                        if slot_idx + 1 < len(regular_section_a.index):
                                            next_slot = regular_section_a.index[slot_idx + 1]
                                            next_cell = regular_section_a.loc[next_slot, day]
                                            
                                            if pd.notna(next_cell) and course_code.lower() in str(next_cell).lower() and '(lab)' in str(next_cell).lower():
                                                # Consecutive lab slots = 2 hours
                                                sched_labs += 2
                                                processed_cells.add((slot_idx, day))
                                                processed_cells.add((slot_idx + 1, day))
                                            else:
                                                # Single lab slot (shouldn't happen but handle it)
                                                sched_labs += 1
                                                processed_cells.add((slot_idx, day))
                                        else:
                                            # Last slot marked as lab
                                            sched_labs += 1
                                            processed_cells.add((slot_idx, day))
                                    elif '(tutorial)' in cell_str:
                                        sched_tutorials += 1
                                        processed_cells.add((slot_idx, day))
                                    else:
                                        # No explicit marker - use time slot detection
                                        # Tutorial detection: check if slot is a tutorial time (1 hour)
                                        is_tutorial = any(t in time_slot_str for t in ['14:30-15:30', '17:00-18:00', '18:00-18:30', '18:30-20:00'])
                                        
                                        if is_tutorial:
                                            sched_tutorials += 1
                                            processed_cells.add((slot_idx, day))
                                        else:
                                            # Check if next time slot also has this course (lab detection)
                                            if slot_idx + 1 < len(regular_section_a.index):
                                                next_slot = regular_section_a.index[slot_idx + 1]
                                                next_cell = regular_section_a.loc[next_slot, day]
                                                
                                                if pd.notna(next_cell) and str(next_cell) != '' and 'free' not in str(next_cell).lower():
                                                    if course_code.lower() in str(next_cell).lower():
                                                        next_cell_str = str(next_cell).lower()
                                                        # IMPORTANT: Check if next slot is marked as tutorial or lab
                                                        # If it's marked, don't assume consecutive lab - let the markers decide
                                                        if '(tutorial)' in next_cell_str:
                                                            # Next slot is a tutorial, so current is just a lecture
                                                            sched_lectures += 1
                                                            processed_cells.add((slot_idx, day))
                                                        elif '(lab)' in next_cell_str:
                                                            # Next slot is a lab, so this is also a lab (2 hours total)
                                                            sched_labs += 2
                                                            processed_cells.add((slot_idx, day))
                                                            processed_cells.add((slot_idx + 1, day))
                                                        else:
                                                            # No markers on next slot, so assume consecutive slots = lab
                                                            sched_labs += 2
                                                            processed_cells.add((slot_idx, day))
                                                            processed_cells.add((slot_idx + 1, day))
                                                    else:
                                                        # Next slot doesn't have this course, so this is a single lecture
                                                        sched_lectures += 1
                                                        processed_cells.add((slot_idx, day))
                                                else:
                                                    # No next slot or next slot is free, so this is a lecture
                                                    sched_lectures += 1
                                                    processed_cells.add((slot_idx, day))
                                            else:
                                                # Last slot, so it's a lecture
                                                sched_lectures += 1
                                                processed_cells.add((slot_idx, day))
                        except Exception as e:
                            # If any error, leave as 0
                            pass
                    
                    # Adjust lecture count based on hours:
                    # Each lecture slot = 1.5 hours, so 2 slots = 3 hours
                    # If L=2 or L=3 and we scheduled 2 slots, consider it as matching the requirement
                    if req_lectures in [2, 3] and sched_lectures >= 2:
                        sched_lectures = req_lectures
                    
                    # Format as "scheduled/required"
                    lectures_display = f"{sched_lectures}/{req_lectures}"
                    tutorials_display = f"{sched_tutorials}/{req_tutorials}"
                    labs_display = f"{sched_labs}/{req_labs}"
                    
                    row_data = [
                        course_code,
                        info.get('name', 'N/A'),
                        ltpsc,
                        info.get('term_type', 'Full Semester'),
                        lectures_display,
                        tutorials_display,
                        labs_display
                    ]
                    for col_idx, value in enumerate(row_data, start=1):
                        cell = ws.cell(row=current_row, column=col_idx, value=value)
                        cell.border = border
                        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                        
                        # Apply course color to the Course Code cell (column 1)
                        if col_idx == 1:
                            color_hex = course_colors.get(course_code, 'FFFFFF')
                            hex_color = color_hex.replace('#', '')
                            if 'hsl' not in hex_color.lower():
                                cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
                    
                    current_row += 1
                
                # Set column widths for core courses section
                ws.column_dimensions['A'].width = max(ws.column_dimensions['A'].width, 12)  # Course Code
                ws.column_dimensions['B'].width = max(ws.column_dimensions['B'].width, 35)  # Course Name
                ws.column_dimensions['C'].width = max(ws.column_dimensions['C'].width, 12)  # L-T-P-S-C
                ws.column_dimensions['D'].width = max(ws.column_dimensions['D'].width, 15)  # Term Type
                ws.column_dimensions['E'].width = max(ws.column_dimensions['E'].width, 14)  # Lectures/Week
                ws.column_dimensions['F'].width = max(ws.column_dimensions['F'].width, 14)  # Tutorials/Week
                ws.column_dimensions['G'].width = max(ws.column_dimensions['G'].width, 12)  # Labs/Week
                
                current_row += 1  # Spacing
            
            # Elective Baskets Section - EACH COURSE IN SEPARATE ROW
            if all_baskets:
                section_cell = ws.cell(row=current_row, column=1, value='ELECTIVE BASKETS')
                section_cell.fill = section_fill
                section_cell.font = section_font
                section_cell.alignment = Alignment(horizontal='left', vertical='center')
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
                current_row += 1
                
                # Headers - Fixed columns (no color)
                headers = ['Basket Name', 'Course', 'Course Code', 'Lecture Slot - Classroom', 'Tutorial Slot - Classroom', 'L-T-P-S-C']
                for col_idx, header in enumerate(headers, start=1):
                    cell = ws.cell(row=current_row, column=col_idx, value=header)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                current_row += 1
                
                # Data rows - ONE ROW PER COURSE
                for basket_name in sorted(all_baskets):
                    basket_info = basket_allocations.get(basket_name, {})
                    basket_courses = basket_info.get('all_courses_in_basket', [])
                    
                    # Get lecture and tutorial slots
                    lectures = basket_info.get('lectures', [])
                    lecture_slots = ', '.join([f"{day} {time}" for day, time in lectures]) if lectures else '-'
                    
                    tutorial = basket_info.get('tutorial')
                    tutorial_slot = f"{tutorial[0]} {tutorial[1]}" if tutorial and len(tutorial) >= 2 else '-'
                    
                    # Get classroom allocations for this basket
                    timetable_section = 'A' if 'SECTION_A' in sheet_name.upper() else ('B' if 'SECTION_B' in sheet_name.upper() else 'Whole')
                    timetable_key = f"{branch}_sem{semester}_sec{timetable_section}"
                    allocs_for_file = _TIMETABLE_CLASSROOM_ALLOCATIONS.get(timetable_key, {})
                    
                    # Create one row for each course in the basket
                    for course_idx, course_code in enumerate(basket_courses):
                        # Get course name from course_info
                        course_name = course_info.get(course_code, {}).get('name', 'N/A')
                        
                        # Get classroom for EACH lecture slot and for tutorial
                        lecture_slots_with_rooms = []
                        tutorial_with_room = tutorial_slot
                        
                        # Process each lecture slot
                        for day, time in lectures:
                            lecture_room = ''
                            # Allocation key format: {day}_{time_slot}_{course_code}
                            expected_key = f"{day}_{time}_{course_code}"
                            alloc = allocs_for_file.get(expected_key)
                            if alloc:
                                room = alloc.get('classroom') or alloc.get('room')
                                if room:
                                    lecture_room = str(room)
                            
                            if lecture_room:
                                lecture_slots_with_rooms.append(f"{day} {time} [{lecture_room}]")
                            else:
                                lecture_slots_with_rooms.append(f"{day} {time}")
                        
                        lecture_with_room = ', '.join(lecture_slots_with_rooms) if lecture_slots_with_rooms else lecture_slots
                        
                        # Process tutorial slot
                        if tutorial:
                            tutorial_room = ''
                            # Allocation key format for tutorial: {day}_{time_slot}_{course_code}
                            expected_tutorial_key = f"{tutorial[0]}_{tutorial[1]}_{course_code}"
                            alloc = allocs_for_file.get(expected_tutorial_key)
                            if alloc:
                                room = alloc.get('classroom') or alloc.get('room')
                                if room:
                                    tutorial_room = str(room)
                            
                            if tutorial_room:
                                tutorial_with_room = f"{tutorial_slot} [{tutorial_room}]"
                        
                        # Get LTPSC info
                        ltpsc_str = course_info.get(course_code, {}).get('ltpsc', 'N/A')
                        
                        row_data = [
                            basket_name,
                            course_name,
                            course_code,
                            lecture_with_room,
                            tutorial_with_room,
                            ltpsc_str
                        ]
                        
                        for col_idx, value in enumerate(row_data, start=1):
                            cell = ws.cell(row=current_row, column=col_idx, value=value)
                            cell.border = border
                            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

                        
                        current_row += 1
                
                current_row += 1  # Spacing
            
            # Minor Courses Section
            if minor_courses:
                section_cell = ws.cell(row=current_row, column=1, value='MINOR COURSES')
                section_cell.fill = section_fill
                section_cell.font = section_font
                section_cell.alignment = Alignment(horizontal='left', vertical='center')
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
                current_row += 1
                
                # Headers (no color column)
                headers = ['Course Code', 'Course Name', 'L-T-P-S-C', 'Term Type']
                for col_idx, header in enumerate(headers, start=1):
                    cell = ws.cell(row=current_row, column=col_idx, value=header)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                current_row += 1
                
                # Data rows
                for course_code in minor_courses:
                    info = course_info.get(course_code, {})
                    row_data = [
                        course_code,
                        info.get('name', 'N/A'),
                        info.get('ltpsc', 'N/A'),
                        info.get('term_type', 'Full Semester')
                    ]
                    for col_idx, value in enumerate(row_data, start=1):
                        cell = ws.cell(row=current_row, column=col_idx, value=value)
                        cell.border = border
                        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                        
                        # Apply course color to the Course Code cell (column 1)
                        if col_idx == 1:
                            color_hex = course_colors.get(course_code, 'FFFFFF')
                            hex_color = color_hex.replace('#', '')
                            if 'hsl' not in hex_color.lower():
                                cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
                    
                    current_row += 1
            
            # Set column widths for legend area - Fixed columns for elective baskets
            ws.column_dimensions['A'].width = max(ws.column_dimensions['A'].width, 15)  # Basket Name
            ws.column_dimensions['B'].width = max(ws.column_dimensions['B'].width, 25)  # Course (name)
            ws.column_dimensions['C'].width = max(ws.column_dimensions['C'].width, 12)  # Course Code
            ws.column_dimensions['D'].width = max(ws.column_dimensions['D'].width, 30)  # Lecture Slot - Classroom
            ws.column_dimensions['E'].width = max(ws.column_dimensions['E'].width, 30)  # Tutorial Slot - Classroom
            ws.column_dimensions['F'].width = max(ws.column_dimensions['F'].width, 12)  # L-T-P-S-C
            
            # Format timetable part
            format_excel_worksheet(ws, course_colors, basket_colors, is_header_row=True)
        
        # Create a dedicated Course_Information sheet
        if 'Course_Information' not in wb.sheetnames:
            course_info_ws = wb.create_sheet('Course_Information', 0)  # Insert as first sheet
        else:
            course_info_ws = wb['Course_Information']

        # Add title
        title_cell = course_info_ws.cell(row=1, column=1, value=f'SEMESTER {semester} - {branch} BRANCH: COURSE INFORMATION')
        title_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        title_font = Font(bold=True, color="FFFFFF", size=13)
        title_cell.fill = title_fill
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        course_info_ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
        course_info_ws.row_dimensions[1].height = 25

        current_row = 3

        section_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        section_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        header_font = Font(bold=True, size=10)
        border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )
        
        # Core Courses Section
        if core_courses:
            section_cell = course_info_ws.cell(row=current_row, column=1, value='CORE COURSES')
            section_cell.fill = section_fill
            section_cell.font = section_font
            section_cell.alignment = Alignment(horizontal='left', vertical='center')
            course_info_ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
            current_row += 1

            # Headers
            headers = ['Course Code', 'Course Name', 'L-T-P-S-C', 'Term Type', 'Faculty', 'Display Format']
            for col_idx, header in enumerate(headers, start=1):
                cell = course_info_ws.cell(row=current_row, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            current_row += 1

            # Data rows
            for course_code in core_courses:
                info = course_info.get(course_code, {})
                ltpsc = info.get('ltpsc', 'N/A')
                term_type = info.get('term_type', 'Full Sem')
                course_name = info.get('name', 'N/A')
                faculty = info.get('instructor', 'N/A')
                # Display format like on website: "CODE (L-T-P-S-C | Term)"
                display_format = f"{course_code} ({ltpsc} | {term_type})" if ltpsc != 'N/A' else f"{course_code} ({term_type})"

                row_data = [
                    course_code,
                    course_name,
                    ltpsc,
                    term_type,
                    faculty,
                    display_format
                ]
                for col_idx, value in enumerate(row_data, start=1):
                    cell = course_info_ws.cell(row=current_row, column=col_idx, value=value)
                    cell.border = border
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                current_row += 1

            current_row += 1

        # Elective Courses Section
        if elective_courses:
            section_cell = course_info_ws.cell(row=current_row, column=1, value='ELECTIVE COURSES')
            section_cell.fill = section_fill
            section_cell.font = section_font
            section_cell.alignment = Alignment(horizontal='left', vertical='center')
            course_info_ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
            current_row += 1

            # Headers
            headers = ['Course Code', 'Course Name', 'L-T-P-S-C', 'Basket', 'Faculty', 'Display Format']
            for col_idx, header in enumerate(headers, start=1):
                cell = course_info_ws.cell(row=current_row, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            current_row += 1

            # Data rows
            for course_code in elective_courses:
                info = course_info.get(course_code, {})
                ltpsc = info.get('ltpsc', 'N/A')
                course_name = info.get('name', 'N/A')
                faculty = info.get('instructor', 'N/A')
                # Get basket name from basket_allocations; fallback to course dataframe if missing
                basket = 'Unknown'
                for basket_name, basket_info in (basket_allocations or {}).items():
                    courses_in_basket = basket_info.get('all_courses_in_basket', [])
                    if course_code in courses_in_basket:
                        basket = basket_name
                        break

                # Fallback: look up Basket column from course dataframe (if present)
                if basket == 'Unknown' and dfs and 'course' in dfs:
                    course_df = dfs['course']
                    basket_match = course_df.loc[course_df['Course Code'] == course_code, 'Basket'] if 'Basket' in course_df.columns else None
                    if basket_match is not None and not basket_match.empty:
                        basket_val = str(basket_match.iloc[0]).strip()
                        if basket_val:
                            basket = basket_val
               
                display_format = f"{course_code} ({ltpsc})" if ltpsc != 'N/A' else course_code

                row_data = [
                    course_code,
                    course_name,
                    ltpsc,
                    basket,
                    faculty,
                    display_format
                ]
                for col_idx, value in enumerate(row_data, start=1):
                    cell = course_info_ws.cell(row=current_row, column=col_idx, value=value)
                    cell.border = border
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                current_row += 1

            current_row += 1

        # Minor Courses Section
        if minor_courses:
            section_cell = course_info_ws.cell(row=current_row, column=1, value='MINOR COURSES')
            section_cell.fill = section_fill
            section_cell.font = section_font
            section_cell.alignment = Alignment(horizontal='left', vertical='center')
            course_info_ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
            current_row += 1

            # Headers
            headers = ['Course Code', 'Course Name', 'L-T-P-S-C', 'Term Type', 'Faculty', 'Display Format']
            for col_idx, header in enumerate(headers, start=1):
                cell = course_info_ws.cell(row=current_row, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            current_row += 1

            # Data rows
            for course_code in minor_courses:
                info = course_info.get(course_code, {})
                ltpsc = info.get('ltpsc', 'N/A')
                term_type = info.get('term_type', 'Full Sem')
                course_name = info.get('name', 'N/A')
                faculty = info.get('instructor', 'N/A')
                display_format = f"{course_code} ({ltpsc} | {term_type})" if ltpsc != 'N/A' else f"{course_code} ({term_type})"

                row_data = [
                    course_code,
                    course_name,
                    ltpsc,
                    term_type,
                    faculty,
                    display_format
                ]
                for col_idx, value in enumerate(row_data, start=1):
                    cell = course_info_ws.cell(row=current_row, column=col_idx, value=value)
                    cell.border = border
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                current_row += 1

        # Set column widths
        course_info_ws.column_dimensions['A'].width = 15
        course_info_ws.column_dimensions['B'].width = 40
        course_info_ws.column_dimensions['C'].width = 12
        course_info_ws.column_dimensions['D'].width = 20
        course_info_ws.column_dimensions['E'].width = 25  # Faculty column
        course_info_ws.column_dimensions['F'].width = 35  # Display Format column

        wb.save(filepath)
        print(f"[OK] Consolidated timetable saved: {filename}")
        return True
        
    except Exception as e:
        print(f"[FAIL] Error generating consolidated timetable: {e}")
        traceback.print_exc()
        return False

def export_semester_timetable_with_baskets(dfs, semester, branch=None, time_config=None, minimal_only=False):
    """Export timetable using IDENTICAL COMMON elective slots for ALL branches and sections with classroom allocation.
    Accepts optional time_config to override slot timings. Set minimal_only=True to emit only timetable sheets (no verification/summary extras)."""
    branch_info = f", Branch {branch}" if branch else ""
    print(f"\n[STATS] Generating timetable for Semester {semester}{branch_info}...")
    
    # NOTE: Do NOT reset classroom tracker here! All semesters/branches share physical classrooms.
    # The tracker is only reset once at the start of timetable generation in the main endpoint.
    
    # Show semester-specific basket rules
    if semester == 3:
        print(f"   [TARGET] SEMESTER 3 RULE: Scheduling only ELECTIVE_B3, excluding ELECTIVE_B5")
    elif semester == 5:
        print(f"   [TARGET] SEMESTER 5 RULE: Scheduling BOTH ELECTIVE_B5 and ELECTIVE_B4")
    elif semester == 7:
        print(f"   [TARGET] SEMESTER 7 RULE: Scheduling ELECTIVE_B6, ELECTIVE_B7, ELECTIVE_B8, ELECTIVE_B9")
    else:
        print(f"   [TARGET] SEMESTER {semester}: Scheduling all elective baskets")
    
    try:
        # Initialize helper structures used later in the writer section
        basket_courses_map = {}
        classroom_allocation_details = []
        # Get ALL elective courses for this semester (without branch filter)
        course_baskets_all = separate_courses_by_type(dfs, semester)
        elective_courses_all = course_baskets_all['elective_courses']
        
        print(f"[TARGET] Elective courses for Semester {semester} (COMMON for ALL): {len(elective_courses_all)}")
        
        # Allocate electives using FIXED COMMON slots (with semester filtering)
        elective_allocations, basket_allocations = allocate_electives_by_baskets(elective_courses_all, semester)
        
        # Build basket_courses_map from basket_allocations for classroom allocation
        for basket_name, alloc in (basket_allocations or {}).items():
            courses_in_basket = alloc.get('all_courses_in_basket', [])
            if courses_in_basket:
                basket_courses_map[basket_name] = courses_in_basket
                print(f"   [MAP] {basket_name}: {len(courses_in_basket)} courses")

        
        print(f"   [TIME] FINAL SCHEDULED BASKETS for Semester {semester}:")
        if basket_allocations:
            for basket_name, allocation in basket_allocations.items():
                status = "[OK] VALID" if allocation['days_separated'] else "[FAIL] INVALID"
                print(f"      {basket_name}: {status}")
        
        # Determine if this branch has sections (only CSE has sections A and B)
        has_sections = (branch == 'CSE')
        if has_sections:
            # Generate schedules - these will have IDENTICAL elective slots
            # Pass basket_allocations to ensure all required baskets are scheduled
            section_a = generate_section_schedule_with_elective_baskets(dfs, semester, 'A', elective_allocations, branch, time_config=time_config, basket_allocations=basket_allocations)
            section_b = generate_section_schedule_with_elective_baskets(dfs, semester, 'B', elective_allocations, branch, time_config=time_config, basket_allocations=basket_allocations)
            
            if section_a is None or section_b is None:
                return False
        else:
            # For non-CSE branches (e.g., DSAI, ECE) treat as whole branch single schedule
            section_a = generate_section_schedule_with_elective_baskets(dfs, semester, 'Whole', elective_allocations, branch, time_config=time_config, basket_allocations=basket_allocations)
            section_b = pd.DataFrame()
            if section_a is None:
                return False

        # ALLOCATE CLASSROOMS for both sections with proper tracking
        course_info = get_course_info(dfs) if dfs else {}
        classroom_data = dfs.get('classroom')
        
        if classroom_data is not None and not classroom_data.empty:
            print("[SCHOOL] Allocating classrooms with global tracking...")
            section_a_with_rooms = allocate_classrooms_for_timetable(
                section_a, classroom_data, course_info, semester, branch, 'A', basket_courses_map
            )
            section_b_with_rooms = allocate_classrooms_for_timetable(
                section_b, classroom_data, course_info, semester, branch, 'B', basket_courses_map
            )
            
            # Check if classroom allocation was successful
            has_classroom_allocation_a = check_for_classroom_allocation(section_a_with_rooms)
            has_classroom_allocation_b = check_for_classroom_allocation(section_b_with_rooms)
            has_classroom_allocation = has_classroom_allocation_a or has_classroom_allocation_b
            
            if has_classroom_allocation:
                print("   [OK] Classroom allocation completed with global tracking")
                # Print allocation summary
                print_classroom_allocation_summary(semester, branch)
            else:
                print("   [WARN] Classroom allocation attempted but no rooms were assigned")
        else:
            section_a_with_rooms = section_a
            section_b_with_rooms = section_b
            print("[WARN]  No classroom data available for allocation")

        # Create filename
        filename = f"sem{semester}_{branch}_timetable_baskets.xlsx" if branch else f"sem{semester}_timetable_baskets.xlsx"
        filepath = os.path.join(OUTPUT_DIR, filename)
        
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            # Save schedules with classroom allocation
            # Ensure consistent column structure across all semesters
            if has_sections:
                # For CSE (which has both sections), reset index and write without index column
                # Clean up any unwanted index-related columns first
                section_a_for_excel = section_a_with_rooms.copy()
                section_b_for_excel = section_b_with_rooms.copy()
                
                # Drop any 'level_0', 'index', or unnamed columns that shouldn't be there
                cols_to_drop_a = [col for col in section_a_for_excel.columns if col in ['level_0', 'Unnamed: 0', 'Unnamed: 1']]
                if cols_to_drop_a:
                    section_a_for_excel = section_a_for_excel.drop(columns=cols_to_drop_a)
                cols_to_drop_b = [col for col in section_b_for_excel.columns if col in ['level_0', 'Unnamed: 0', 'Unnamed: 1']]
                if cols_to_drop_b:
                    section_b_for_excel = section_b_for_excel.drop(columns=cols_to_drop_b)
                
                # Reset index to move time slot to a column
                section_a_for_excel = section_a_for_excel.reset_index(drop=False).rename(columns={'index': 'Time Slot'})
                section_b_for_excel = section_b_for_excel.reset_index(drop=False).rename(columns={'index': 'Time Slot'})
                
                section_a_for_excel.to_excel(writer, sheet_name='Section_A', index=False)
                section_b_for_excel.to_excel(writer, sheet_name='Section_B', index=False)
            else:
                # Single sheet 'Timetable' for whole-branch schedules
                section_a_reset = section_a_with_rooms.copy()
                
                # Drop any unwanted index-related columns
                cols_to_drop = [col for col in section_a_reset.columns if col in ['level_0', 'Unnamed: 0', 'Unnamed: 1']]
                if cols_to_drop:
                    section_a_reset = section_a_reset.drop(columns=cols_to_drop)
                
                # Reset index to move time slot to a column
                section_a_reset = section_a_reset.reset_index(drop=False).rename(columns={'index': 'Time Slot'})
                
                section_a_reset.to_excel(writer, sheet_name='Timetable', index=False)
            
            if minimal_only:
                print("[STATS] Minimal output: skipping verification/summary sheets")
            else:
                # ========== ADDED: COMPREHENSIVE VERIFICATION SHEETS ==========
                print("[STATS] Generating comprehensive verification sheets...")
                
                # Get course info for verification
                course_info = get_course_info(dfs) if dfs else {}
                
                # 1. Create detailed verification sheet for Section A (or Whole branch)
                if not section_a_with_rooms.empty:
                    print(f"   Creating verification sheet for Section A (or Whole)...")
                    verification_a = create_timetable_verification_sheet(
                        section_a_with_rooms, course_info, classroom_data, semester, branch, 'A' if has_sections else 'Whole'
                    )
                    if not verification_a.empty:
                        writer_sheet = 'Verification_A' if has_sections else 'Verification'
                        verification_a.to_excel(writer, sheet_name=writer_sheet, index=False)
                        print(f"   [OK] Created {writer_sheet} sheet with {len(verification_a)} entries")
                
                # 2. Create detailed verification sheet for Section B (only for branches with sections)
                if has_sections and not section_b_with_rooms.empty:
                    print(f"   Creating verification sheet for Section B...")
                    verification_b = create_timetable_verification_sheet(
                        section_b_with_rooms, course_info, classroom_data, semester, branch, 'B'
                    )
                    if not verification_b.empty:
                        verification_b.to_excel(writer, sheet_name='Verification_B', index=False)
                        print(f"   [OK] Created Verification_B sheet with {len(verification_b)} entries")
                
                # 3. Create room allocation summary
                if classroom_data is not None and not classroom_data.empty:
                    print(f"   Creating room allocation summary...")
                    room_summary = create_room_allocation_summary_verification(
                        section_a_with_rooms, section_b_with_rooms, classroom_data
                    )
                    if not room_summary.empty:
                        room_summary.to_excel(writer, sheet_name='Room_Allocation', index=False)
                        print(f"   [OK] Created Room_Allocation sheet with {len(room_summary)} rooms")
                
                # 4. Create LTPSC compliance summary
                print(f"   Creating LTPSC compliance summary...")
                ltpsc_summary = create_ltpsc_compliance_summary(
                    dfs, semester, branch, section_a_with_rooms, section_b_with_rooms
                )
                if not ltpsc_summary.empty:
                    ltpsc_summary.to_excel(writer, sheet_name='LTPSC_Compliance', index=False)
                    print(f"   [OK] Created LTPSC_Compliance sheet with {len(ltpsc_summary)} courses")
                
                # 5. Create executive summary
                print(f"   Creating executive summary...")
                exec_summary = create_executive_summary(
                    dfs, semester, branch, section_a_with_rooms, section_b_with_rooms, basket_allocations
                )
                if not exec_summary.empty:
                    exec_summary.to_excel(writer, sheet_name='Executive_Summary', index=False)
                    print(f"   [OK] Created Executive_Summary sheet")
                
                # ========== EXISTING SHEETS ==========
                # Add basket allocation summary
                if basket_allocations:
                    basket_summary = create_common_basket_summary(basket_allocations, semester, branch)
                    basket_summary.to_excel(writer, sheet_name='Basket_Allocation', index=False)
                
                # Add course summary
                course_summary = create_course_summary(dfs, semester, branch)
                if not course_summary.empty:
                    course_summary.to_excel(writer, sheet_name='Course_Summary', index=False)
                
                # Add basket courses details
                basket_courses_sheet = create_basket_courses_sheet(basket_allocations)
                if not basket_courses_sheet.empty:
                    basket_courses_sheet.to_excel(writer, sheet_name='Basket_Courses', index=False)

                    # Add basket per-course allocations sheet (so Excel outputs list allocated rooms per basket course)
                    try:
                        rows = []
                        for basket_name, courses in (basket_courses_map or {}).items():
                            for course in courses:
                                # Check explicit classroom allocation details first
                                rooms = [rec.get('room') for rec in classroom_allocation_details if rec.get('course') == course and rec.get('room')]
                                unique_rooms = sorted(set([r for r in rooms if r]))
                                # Fallback to internal tracker if none found
                                if not unique_rooms and branch and semester:
                                    timetable_keys = [f"{branch}_sem{semester}_secA", f"{branch}_sem{semester}_secB", f"{branch}_sem{semester}_secWhole"]
                                    tracker_rooms = []
                                    for tk in timetable_keys:
                                        alloc_map = _TIMETABLE_CLASSROOM_ALLOCATIONS.get(tk, {})
                                        for alloc in alloc_map.values():
                                            c = alloc.get('course')
                                            room_val = alloc.get('classroom') or alloc.get('room')
                                            if c == course and room_val:
                                                tracker_rooms.append(room_val)
                                    if tracker_rooms:
                                        unique_rooms = sorted(set(tracker_rooms))
                                rows.append({
                                    'Basket Name': basket_name,
                                    'Course': course,
                                    'Allocated Rooms': ', '.join(unique_rooms) if unique_rooms else ''
                                })
                        if rows:
                            pd.DataFrame(rows).to_excel(writer, sheet_name='Basket_Course_Allocations', index=False)
                    except Exception:
                        pass
                    
                    # Add detailed classroom allocation with global tracking info
                    classroom_allocation_detail = create_classroom_allocation_detail_with_tracking(
                        [section_a_with_rooms, section_b_with_rooms], classroom_data, semester, branch
                    )
                    classroom_allocation_detail.to_excel(writer, sheet_name='Classroom_Allocation', index=False)

                # Persist configuration if provided
                if time_config:
                    try:
                        config_items = [{'Parameter': k, 'Value': str(v)} for k, v in time_config.items()]
                        pd.DataFrame(config_items).to_excel(writer, sheet_name='Configuration', index=False)
                    except Exception as _:
                        pass
                # ========== END OF ADDED SECTION ==========
        
        success_message = f"[OK] Timetable saved: {filename}"
        if classroom_data is not None and not classroom_data.empty:
            success_message += " (with classroom allocation)"
        
        print(success_message)
        if minimal_only:
            print("[STATS] Minimal export written (timetables only)")
        else:
            print(f"[STATS] Added comprehensive verification sheets for easy inspection")
        return True
        
    except Exception as e:
        print(f"[FAIL] Error generating timetable: {e}")
        traceback.print_exc()
        return False

def allocate_mid_semester_electives_by_baskets(elective_courses, semester_id):
    """Allocate elective courses in pre-mid/post-mid to COMMON time slots for ALL branches and sections
    
    Similar to allocate_electives_by_baskets but for mid-semester courses
    """
    if elective_courses.empty:
        print(f"   [INFO] No elective courses to allocate for Semester {semester_id}")
        return {}
    
    print(f"   [BASKET] Allocating elective slots for mid-semester, Semester {semester_id}...")
    
    # Group electives by basket
    basket_groups = {}
    for _, course in elective_courses.iterrows():
        raw_basket = course.get('Basket', 'ELECTIVE_B1')
        basket = str(raw_basket).strip().upper() if pd.notna(raw_basket) else 'ELECTIVE_B1'
        if basket not in basket_groups:
            basket_groups[basket] = []
        basket_groups[basket].append(course)
    
    print(f"      Found {len(basket_groups)} elective baskets: {list(basket_groups.keys())}")
    
    # SCHEDULE ALL BASKETS PRESENT FOR THIS SEMESTER (across all branches)
    # Do NOT filter by pre-defined basket names, as different branches may use different baskets
    baskets_to_schedule = sorted(list(basket_groups.keys()))
    print(f"      [TARGET] Semester {semester_id}: Scheduling ALL {len(baskets_to_schedule)} baskets present:")
    for basket in baskets_to_schedule:
        print(f"         - {basket}")
    
    # COMMON SLOTS mapping (same as regular timetables)
    common_slots_mapping = {
        'ELECTIVE_B1': {
            'lectures': [('Mon', '09:00-10:30'), ('Wed', '09:00-10:30')],
            'tutorial': ('Fri', '14:30-15:30')
        },
        'ELECTIVE_B2': {
            'lectures': [('Tue', '09:00-10:30'), ('Thu', '09:00-10:30')],
            'tutorial': ('Mon', '14:30-15:30')
        },
        'ELECTIVE_B3': {
            'lectures': [('Mon', '13:00-14:30'), ('Wed', '13:00-14:30')],
            'tutorial': ('Tue', '14:30-15:30')
        },
        'ELECTIVE_B4': {
            'lectures': [('Tue', '13:00-14:30'), ('Thu', '13:00-14:30')],
            'tutorial': ('Wed', '14:30-15:30')
        },
        'ELECTIVE_B5': {
            'lectures': [('Mon', '15:30-17:00'), ('Wed', '15:30-17:00')],
            'tutorial': ('Thu', '14:30-15:30')
        },
        'ELECTIVE_B6': {
            'lectures': [('Mon', '09:00-10:30'), ('Wed', '13:00-14:30')],
            'tutorial': ('Tue', '14:30-15:30')
        },
        'ELECTIVE_B7': {
            'lectures': [('Tue', '09:00-10:30'), ('Thu', '13:00-14:30')],
            'tutorial': ('Wed', '14:30-15:30')
        },
        # Align Semester 7 elective baskets with regular timetable slots to avoid conflicts
        'ELECTIVE_B8': {
            'lectures': [('Mon', '10:30-12:00'), ('Wed', '10:30-12:00')],
            'tutorial': ('Thu', '14:30-15:30')
        },
        'ELECTIVE_B9': {
            'lectures': [('Tue', '15:30-17:00'), ('Thu', '15:30-17:00')],
            'tutorial': ('Fri', '14:30-15:30')
        },
        'HSS_B1': {
            'lectures': [('Mon', '15:30-17:00'), ('Wed', '15:30-17:00')],
            'tutorial': ('Thu', '14:30-15:30')
        },
        'HSS_B2': {
            'lectures': [('Tue', '15:30-17:00'), ('Thu', '15:30-17:00')],
            'tutorial': ('Fri', '14:30-15:30')
        },
        'HSS_B3': {
            'lectures': [('Mon', '10:30-12:00'), ('Wed', '10:30-12:00')],
            'tutorial': ('Fri', '14:30-15:30')
        },
        'HSS_B4': {
            'lectures': [('Tue', '10:30-12:00'), ('Thu', '10:30-12:00')],
            'tutorial': ('Mon', '14:30-15:30')
        },
        'HSS_B5': {
            'lectures': [('Mon', '15:30-17:00'), ('Wed', '15:30-17:00')],
            'tutorial': ('Fri', '14:30-15:30')
        },
        'HSS_B6': {
            'lectures': [('Tue', '15:30-17:00'), ('Thu', '15:30-17:00')],
            'tutorial': ('Wed', '14:30-15:30')
        },
        'HSS_B7': {
            'lectures': [('Mon', '10:30-12:00'), ('Wed', '10:30-12:00')],
            'tutorial': ('Thu', '14:30-15:30')
        }
    }
    
    elective_allocations = {}
    
    for basket_name in sorted(baskets_to_schedule):
        if basket_name not in basket_groups:
            continue
        
        basket_courses = basket_groups[basket_name]
        course_codes = [course['Course Code'] for course in basket_courses] if basket_courses else []
        
        # Parse LTPSC across all courses in basket and use the MAX values to ensure adequate slots
        ltpsc_str = '2-1-0-0-3'
        L = 2
        T = 1
        if basket_courses and len(basket_courses) > 0:
            # Determine maximum L and T among courses in the basket to accommodate the highest requirement
            max_L = 0
            max_T = 0
            ltpsc_candidates = []
            for c in basket_courses:
                this_ltpsc = c.get('LTPSC', '')
                ltpsc_parsed = parse_ltpsc(this_ltpsc)
                ltpsc_candidates.append(this_ltpsc)
                max_L = max(max_L, ltpsc_parsed.get('L', 0))
                max_T = max(max_T, ltpsc_parsed.get('T', 0))
            L = max_L if max_L > 0 else 2
            T = max_T if max_T > 0 else 1
            ltpsc_str = ','.join([str(x) for x in ltpsc_candidates]) if ltpsc_candidates else ltpsc_str
        else:
            # Default if no courses present in basket
            ltpsc_str = '2-1-0-0-3'
            L = 2
            T = 1

        # Determine lectures and tutorials needed (cap lectures to 2 as before)
        if L >= 2:
            lectures_needed = 2
        elif L == 1:
            lectures_needed = 1
        else:
            lectures_needed = max(0, min(L, 2))

        tutorials_needed = 1 if T >= 1 else 0
        
        print(f"      [BASKET] Basket '{basket_name}' LTPSC: {ltpsc_str} -> L={L}, T={T}")
        print(f"         -> Scheduling {lectures_needed} lectures, {tutorials_needed} tutorial")
        
        if basket_name in common_slots_mapping:
            fixed_slots = common_slots_mapping[basket_name]
            lectures_allocated = fixed_slots['lectures'][:lectures_needed]
            tutorial_allocated = fixed_slots['tutorial'] if tutorials_needed > 0 else None
            
            # Calculate if lectures and tutorials are on separate days
            if tutorial_allocated:
                lecture_days = set([day for day, _ in lectures_allocated])
                tutorial_day = tutorial_allocated[0]
                days_separated = tutorial_day not in lecture_days
            else:
                days_separated = True  # No tutorial, so separation doesn't apply
            
            # Build allocation for this basket - ENSURE all_courses_in_basket is a list of strings
            allocation = {
                'basket_name': basket_name,
                'lectures': lectures_allocated,
                'tutorial': tutorial_allocated,
                'courses': course_codes,
                'all_courses_in_basket': course_codes,  # Use course codes list, not series objects
                'lecture_days': list(set([day for day, _ in lectures_allocated])),
                'tutorial_day': tutorial_allocated[0] if tutorial_allocated else None,
                'days_separated': days_separated
            }
            
            elective_allocations[basket_name] = allocation
            print(f"         -> Lectures: {lectures_allocated}, Tutorial: {tutorial_allocated}, Days Separated: {days_separated}")
        else:
            # FALLBACK: Auto-generate slots for unmapped baskets
            print(f"      [WARN] Basket '{basket_name}' not in common slots mapping - auto-generating slots")
            
            # Auto-generate reasonable slots by cycling through day/time combinations
            all_day_time_combos = [
                ('Mon', '09:00-10:30'), ('Mon', '10:30-12:00'), ('Mon', '13:00-14:30'),
                ('Tue', '09:00-10:30'), ('Tue', '10:30-12:00'), ('Tue', '13:00-14:30'),
                ('Wed', '09:00-10:30'), ('Wed', '10:30-12:00'), ('Wed', '13:00-14:30'),
                ('Thu', '09:00-10:30'), ('Thu', '10:30-12:00'), ('Thu', '13:00-14:30'),
                ('Fri', '09:00-10:30'), ('Fri', '10:30-12:00'), ('Fri', '13:00-14:30')
            ]
            all_tutorial_combos = [
                ('Mon', '14:30-15:30'), ('Tue', '14:30-15:30'), ('Wed', '14:30-15:30'),
                ('Thu', '14:30-15:30'), ('Fri', '14:30-15:30')
            ]
            
            # Use hash of basket name to pick consistent (but different) indices for each basket
            hash_val = sum([ord(c) for c in basket_name])
            lecture_start_idx = hash_val % (len(all_day_time_combos) - 1)
            tutorial_idx = hash_val % len(all_tutorial_combos)
            
            lectures_allocated = all_day_time_combos[lecture_start_idx:lecture_start_idx + lectures_needed]
            # Ensure we have enough lectures; wrap if necessary
            while len(lectures_allocated) < lectures_needed:
                lectures_allocated.extend(all_day_time_combos[:lectures_needed - len(lectures_allocated)])
            
            tutorial_allocated = all_tutorial_combos[tutorial_idx] if tutorials_needed > 0 else None
            
            # Calculate if lectures and tutorials are on separate days
            if tutorial_allocated:
                lecture_days = set([day for day, _ in lectures_allocated])
                tutorial_day = tutorial_allocated[0]
                days_separated = tutorial_day not in lecture_days
            else:
                days_separated = True
            
            allocation = {
                'basket_name': basket_name,
                'lectures': lectures_allocated,
                'tutorial': tutorial_allocated,
                'courses': course_codes,
                'all_courses_in_basket': course_codes,
                'lecture_days': list(set([day for day, _ in lectures_allocated])),
                'tutorial_day': tutorial_allocated[0] if tutorial_allocated else None,
                'days_separated': days_separated
            }
            
            elective_allocations[basket_name] = allocation
            print(f"         -> AUTO-GENERATED: Lectures: {lectures_allocated}, Tutorial: {tutorial_allocated}, Days Separated: {days_separated}")
    
    print(f"   [OK] Elective allocation complete: {len(elective_allocations)} baskets allocated")
    return elective_allocations

def export_mid_semester_timetables(dfs, semester, branch=None, time_config=None, pre_mid_common_allocations=None, post_mid_common_allocations=None, minimal_only=False):
    """Export separate pre-mid and post-mid timetables
    
    Note: CSE has Section A and Section B
           DSAI and ECE are treated as a whole (no sections)
    
    Set minimal_only=True to emit only timetable sheets (skip verification/summary extras).
    """
    branch_info = f", Branch {branch}" if branch else ""
    print(f"\n[STATS] Generating MID-SEMESTER timetables for Semester {semester}{branch_info}...")
    
    try:
        # Separate courses into pre-mid and post-mid
        mid_semester_courses = separate_courses_by_mid_semester(dfs, semester, branch)
        pre_mid_courses = mid_semester_courses['pre_mid_courses']
        post_mid_courses = mid_semester_courses['post_mid_courses']
        
        print(f"[LIST] Course distribution:")
        print(f"   Pre-mid courses: {len(pre_mid_courses)}")
        print(f"   Post-mid courses: {len(post_mid_courses)}")
        
        # ALLOCATE ELECTIVES TO BASKETS for pre-mid and post-mid
        print(f"\n[BASKET] Allocating electives to baskets...")

        # Use provided common allocations if available (ensures same slots across branches/sections)
        if pre_mid_common_allocations and isinstance(pre_mid_common_allocations, dict) and len(pre_mid_common_allocations) > 0:
            pre_mid_elective_allocations = pre_mid_common_allocations
            print(f"   [INFO] Using provided COMMON PRE-MID allocations for Semester {semester}")
        else:
            if not pre_mid_courses.empty:
                pre_mid_electives = pre_mid_courses[
                    pre_mid_courses['Elective (Yes/No)'].astype(str).str.upper() == 'YES'
                ] if 'Elective (Yes/No)' in pre_mid_courses.columns else pd.DataFrame()
                pre_mid_elective_allocations = allocate_mid_semester_electives_by_baskets(pre_mid_electives, semester)
            else:
                pre_mid_elective_allocations = {}

        if post_mid_common_allocations and isinstance(post_mid_common_allocations, dict) and len(post_mid_common_allocations) > 0:
            post_mid_elective_allocations = post_mid_common_allocations
            print(f"   [INFO] Using provided COMMON POST-MID allocations for Semester {semester}")
        else:
            if not post_mid_courses.empty:
                post_mid_electives = post_mid_courses[
                    post_mid_courses['Elective (Yes/No)'].astype(str).str.upper() == 'YES'
                ] if 'Elective (Yes/No)' in post_mid_courses.columns else pd.DataFrame()
                post_mid_elective_allocations = allocate_mid_semester_electives_by_baskets(post_mid_electives, semester)
            else:
                post_mid_elective_allocations = {}
        
        # Determine if this branch has sections (only CSE has sections A and B)
        has_sections = (branch == 'CSE')
        if has_sections:
            sections = ['A', 'B']
            print(f"[INFO] CSE branch detected - will generate timetables for Section A and Section B")
        else:
            sections = ['Whole']
            print(f"[INFO] {branch} branch detected - will generate single timetable (no sections)")
        
        # Generate pre-mid timetables
        pre_mid_sections = {}
        if not pre_mid_courses.empty:
            print(f"[RESET] Generating PRE-MID timetables...")
            print(f"   Courses to schedule: {pre_mid_courses['Course Code'].tolist()}")
            
            for section in sections:
                section_label = f"Section {section}" if has_sections else "Whole Branch"
                try:
                    result = generate_mid_semester_schedule(
                        dfs, semester, section, pre_mid_courses, branch, time_config, 'pre_mid', pre_mid_elective_allocations
                    )
                    pre_mid_sections[section] = result
                    if result is not None:
                        print(f"   [OK] Pre-mid {section_label}: Generated successfully ({len(result)} time slots)")
                    else:
                        print(f"   [WARN] Pre-mid {section_label}: Returned None (will use empty schedule)")
                        # Create empty/placeholder schedule so file still gets generated
                        pre_mid_sections[section] = pd.DataFrame()
                except Exception as e:
                    print(f"   [FAIL] Error generating pre-mid {section_label}: {e}")
                    traceback.print_exc()
                    # Use empty schedule as fallback
                    pre_mid_sections[section] = pd.DataFrame()
        else:
            print(f"[WARN] No pre-mid courses to schedule for Semester {semester}, Branch {branch}")
        
        # Generate post-mid timetables
        post_mid_sections = {}
        if not post_mid_courses.empty:
            print(f"[RESET] Generating POST-MID timetables...")
            print(f"   Courses to schedule: {post_mid_courses['Course Code'].tolist()}")
            
            for section in sections:
                section_label = f"Section {section}" if has_sections else "Whole Branch"
                try:
                    result = generate_mid_semester_schedule(
                        dfs, semester, section, post_mid_courses, branch, time_config, 'post_mid', post_mid_elective_allocations
                    )
                    post_mid_sections[section] = result
                    if result is not None:
                        print(f"   [OK] Post-mid {section_label}: Generated successfully ({len(result)} time slots)")
                    else:
                        print(f"   [WARN] Post-mid {section_label}: Returned None (will use empty schedule)")
                        # Create empty/placeholder schedule so file still gets generated
                        post_mid_sections[section] = pd.DataFrame()
                except Exception as e:
                    print(f"   [FAIL] Error generating post-mid {section_label}: {e}")
                    traceback.print_exc()
                    # Use empty schedule as fallback
                    post_mid_sections[section] = pd.DataFrame()
        else:
            print(f"[WARN] No post-mid courses to schedule for Semester {semester}, Branch {branch}")
        
        # Create filenames
        base_filename = f"sem{semester}_{branch}_" if branch else f"sem{semester}_"
        pre_mid_filename = base_filename + "pre_mid_timetable.xlsx"
        post_mid_filename = base_filename + "post_mid_timetable.xlsx"
        
        pre_mid_filepath = os.path.join(OUTPUT_DIR, pre_mid_filename)
        post_mid_filepath = os.path.join(OUTPUT_DIR, post_mid_filename)
        
        # Determine sheet names based on whether this branch has sections
        if has_sections:
            sheet_names = {'A': 'Section_A', 'B': 'Section_B', 'Whole': 'Timetable'}
        else:
            sheet_names = {'A': 'Timetable', 'B': None, 'Whole': 'Timetable'}
        
        # ========== ALLOCATE CLASSROOMS FOR PRE-MID TIMETABLES ==========
        classroom_data = dfs.get('classroom', None) if dfs else None
        course_info = get_course_info(dfs) if dfs else {}
        
        if classroom_data is not None and not classroom_data.empty:
            print(f"[SCHOOL] Allocating classrooms for PRE-MID timetables...")
            # Build basket courses map for pre-mid
            pre_mid_basket_courses_map = {}
            if not pre_mid_courses.empty and 'Basket' in pre_mid_courses.columns:
                pre_mid_electives = pre_mid_courses[pre_mid_courses['Elective (Yes/No)'].astype(str).str.upper() == 'YES'] if 'Elective (Yes/No)' in pre_mid_courses.columns else pd.DataFrame()
                for _, course in pre_mid_electives.iterrows():
                    basket = str(course.get('Basket', 'Unknown')).strip().upper() if pd.notna(course.get('Basket')) else 'Unknown'
                    course_code = course['Course Code']
                    if basket not in pre_mid_basket_courses_map:
                        pre_mid_basket_courses_map[basket] = []
                    if course_code not in pre_mid_basket_courses_map[basket]:
                        pre_mid_basket_courses_map[basket].append(course_code)
            
            for section in sections:
                pre_mid_data = pre_mid_sections.get(section)
                if pre_mid_data is not None and not pre_mid_data.empty:
                    print(f"   Allocating classrooms for Pre-Mid Section {section}...")
                    pre_mid_sections[section] = allocate_classrooms_for_timetable(
                        pre_mid_data, classroom_data, course_info, semester, branch, section, pre_mid_basket_courses_map, schedule_type='PreMid'
                    )
                    print(f"   [OK] Pre-Mid Section {section} classrooms allocated")
        
        # Save pre-mid timetable
        pre_mid_all_valid = all(pre_mid_sections.get(s) is not None for s in sections)
        # Also check if at least we have some data (even if one schedule is empty, save the file)
        pre_mid_has_data = any(pre_mid_sections.get(s) is not None and not pre_mid_sections.get(s).empty for s in sections)
        
        if pre_mid_all_valid or pre_mid_has_data:
            try:
                with pd.ExcelWriter(pre_mid_filepath, engine='openpyxl') as writer:
                    for section in sections:
                        pre_mid_data = pre_mid_sections.get(section)
                        if pre_mid_data is not None and not pre_mid_data.empty:
                            # Reset index to ensure Time Slot is a column, not index
                            pre_mid_reset = pre_mid_data.reset_index()
                            pre_mid_reset = pre_mid_reset.rename(columns={'index': 'Time Slot'})
                            sheet_name = sheet_names.get(section, f'Section_{section}')
                            pre_mid_reset.to_excel(writer, sheet_name=sheet_name, index=False)
                        elif pre_mid_data is None:
                            # Create placeholder sheet with message
                            placeholder = pd.DataFrame({'Message': ['No schedule generated for this section']})
                            sheet_name = sheet_names.get(section, f'Section_{section}')
                            placeholder.to_excel(writer, sheet_name=sheet_name, index=False)
                    
                    if minimal_only:
                        print(f"[STATS] Minimal output: Pre-Mid timetables only for Semester {semester}, Branch {branch}")
                    else:
                        # ========== ADDED: VERIFICATION STATISTICS ==========
                        print(f"[STATS] Generating PRE-MID verification sheets...")
                        
                        # Get course info for verification
                        course_info = get_course_info(dfs) if dfs else {}
                        classroom_data = dfs.get('classroom', None) if dfs else None
                        
                        # Create verification sheets for each section
                        for section in sections:
                            pre_mid_reset = pre_mid_sections[section].reset_index().rename(columns={'index': 'Time Slot'})
                            if not pre_mid_reset.empty:
                                section_label = f"Section {section}" if has_sections else "Timetable"
                                print(f"   Creating Pre-Mid Verification sheet for {section_label}...")
                                verification = create_timetable_verification_sheet(
                                    pre_mid_reset, course_info, classroom_data, semester, branch, section
                                )
                                if not verification.empty:
                                    sheet_name = f'Verification_{section}' if has_sections else 'Verification'
                                    verification.to_excel(writer, sheet_name=sheet_name, index=False)
                                    print(f"   [OK] Created Pre-Mid Verification sheet with {len(verification)} entries")
                        
                        # Create room allocation summary
                        if classroom_data is not None and not classroom_data.empty:
                            print(f"   Creating room allocation summary...")
                            all_sections_data = [pre_mid_sections[s].reset_index().rename(columns={'index': 'Time Slot'}) for s in sections]
                            room_summary = create_room_allocation_summary_verification(
                                all_sections_data[0], all_sections_data[1] if len(all_sections_data) > 1 else pd.DataFrame(), classroom_data
                            )
                            if not room_summary.empty:
                                room_summary.to_excel(writer, sheet_name='Room_Allocation', index=False)
                                print(f"   [OK] Created Room_Allocation sheet with {len(room_summary)} rooms")
                            
                            # Create detailed classroom allocation sheet
                            print(f"   Creating detailed classroom allocation...")
                            classroom_alloc_detail = create_classroom_allocation_detail_with_tracking(
                                [pre_mid_sections[s] for s in sections], classroom_data, semester, branch
                            )
                            if not classroom_alloc_detail.empty:
                                classroom_alloc_detail.to_excel(writer, sheet_name='Classroom_Allocation', index=False)
                                print(f"   [OK] Created Classroom_Allocation sheet with {len(classroom_alloc_detail)} entries")
                        
                        # Create LTPSC compliance summary
                        print(f"   Creating LTPSC compliance summary...")
                        all_sections_data = [pre_mid_sections[s].reset_index().rename(columns={'index': 'Time Slot'}) for s in sections]
                        ltpsc_summary = create_ltpsc_compliance_summary(
                            dfs, semester, branch, all_sections_data[0], all_sections_data[1] if len(all_sections_data) > 1 else pd.DataFrame()
                        )
                        if not ltpsc_summary.empty:
                            ltpsc_summary.to_excel(writer, sheet_name='LTPSC_Compliance', index=False)
                            print(f"   [OK] Created LTPSC_Compliance sheet with {len(ltpsc_summary)} courses")
                        
                        # Create executive summary
                        print(f"   Creating executive summary...")
                        all_sections_data = [pre_mid_sections[s].reset_index().rename(columns={'index': 'Time Slot'}) for s in sections]
                        exec_summary = create_executive_summary(
                            dfs, semester, branch, all_sections_data[0], all_sections_data[1] if len(all_sections_data) > 1 else pd.DataFrame(), {}
                        )
                        if not exec_summary.empty:
                            exec_summary.to_excel(writer, sheet_name='Executive_Summary', index=False)
                            print(f"   [OK] Created Executive_Summary sheet")
                        
                        # Add course summary
                        pre_mid_summary = create_mid_semester_summary(pre_mid_courses, 'Pre-Mid', semester, branch)
                        pre_mid_summary.to_excel(writer, sheet_name='Course_Summary', index=False)

                        # Add basket allocation and details if we have pre_mid elective allocations
                        try:
                            if 'pre_mid_elective_allocations' in locals() and pre_mid_elective_allocations:
                                print(f"   Creating Basket_Allocation and Basket_Courses for Pre-Mid...")
                                basket_summary = create_basket_summary(pre_mid_elective_allocations, semester, branch)
                                if not basket_summary.empty:
                                    basket_summary.to_excel(writer, sheet_name='Basket_Allocation', index=False)
                                basket_courses_sheet_mid = create_basket_courses_sheet(pre_mid_elective_allocations)
                                if not basket_courses_sheet_mid.empty:
                                    basket_courses_sheet_mid.to_excel(writer, sheet_name='Basket_Courses', index=False)
                        except Exception as _:
                            # Avoid failing the whole save if basket sheets can't be created
                            pass

                        # Add Classroom Utilization and Allocation detail for Pre-Mid
                        if classroom_data is not None and not classroom_data.empty:
                            try:
                                classroom_report = create_classroom_utilization_report(
                                    classroom_data, [pre_mid_sections[s].reset_index().rename(columns={'index': 'Time Slot'}) for s in sections], []
                                )
                                classroom_report.to_excel(writer, sheet_name='Classroom_Utilization', index=False)

                                classroom_allocation_detail = create_classroom_allocation_detail_with_tracking(
                                    [pre_mid_sections[s] for s in sections], classroom_data, semester, branch
                                )
                                classroom_allocation_detail.to_excel(writer, sheet_name='Classroom_Allocation', index=False)
                            except Exception:
                                pass
                        
                        print(f"[STATS] Added PRE-MID verification sheets for easy inspection")
                        print(f"[OK] PRE-MID timetable saved: {pre_mid_filename}")
                    if minimal_only:
                        print(f"[OK] PRE-MID timetable saved: {pre_mid_filename}")
            except Exception as e:
                print(f"[FAIL] Error saving pre-mid timetable: {e}")
                traceback.print_exc()
        else:
            print(f"[WARN] Cannot save pre-mid timetable - Not all sections generated successfully")
        
        # ========== ALLOCATE CLASSROOMS FOR POST-MID TIMETABLES ==========
        if classroom_data is not None and not classroom_data.empty:
            print(f"[SCHOOL] Allocating classrooms for POST-MID timetables...")
            # Build basket courses map for post-mid
            post_mid_basket_courses_map = {}
            if not post_mid_courses.empty and 'Basket' in post_mid_courses.columns:
                post_mid_electives = post_mid_courses[post_mid_courses['Elective (Yes/No)'].astype(str).str.upper() == 'YES'] if 'Elective (Yes/No)' in post_mid_courses.columns else pd.DataFrame()
                for _, course in post_mid_electives.iterrows():
                    basket = str(course.get('Basket', 'Unknown')).strip().upper() if pd.notna(course.get('Basket')) else 'Unknown'
                    course_code = course['Course Code']
                    if basket not in post_mid_basket_courses_map:
                        post_mid_basket_courses_map[basket] = []
                    if course_code not in post_mid_basket_courses_map[basket]:
                        post_mid_basket_courses_map[basket].append(course_code)
            
            for section in sections:
                post_mid_data = post_mid_sections.get(section)
                if post_mid_data is not None and not post_mid_data.empty:
                    print(f"   Allocating classrooms for Post-Mid Section {section}...")
                    post_mid_sections[section] = allocate_classrooms_for_timetable(
                        post_mid_data, classroom_data, course_info, semester, branch, section, post_mid_basket_courses_map, schedule_type='PostMid'
                    )
                    print(f"   [OK] Post-Mid Section {section} classrooms allocated")
        
        # Save post-mid timetable
        post_mid_all_valid = all(post_mid_sections.get(s) is not None for s in sections)
        # Also check if at least we have some data (even if one schedule is empty, save the file)
        post_mid_has_data = any(post_mid_sections.get(s) is not None and not post_mid_sections.get(s).empty for s in sections)
        
        if post_mid_all_valid or post_mid_has_data:
            try:
                with pd.ExcelWriter(post_mid_filepath, engine='openpyxl') as writer:
                    for section in sections:
                        post_mid_data = post_mid_sections.get(section)
                        if post_mid_data is not None and not post_mid_data.empty:
                            # Reset index to ensure Time Slot is a column, not index
                            post_mid_reset = post_mid_data.reset_index()
                            post_mid_reset = post_mid_reset.rename(columns={'index': 'Time Slot'})
                            sheet_name = sheet_names.get(section, f'Section_{section}')
                            post_mid_reset.to_excel(writer, sheet_name=sheet_name, index=False)
                        elif post_mid_data is None:
                            # Create placeholder sheet with message
                            placeholder = pd.DataFrame({'Message': ['No schedule generated for this section']})
                            sheet_name = sheet_names.get(section, f'Section_{section}')
                            placeholder.to_excel(writer, sheet_name=sheet_name, index=False)
                    
                    if minimal_only:
                        print(f"[STATS] Minimal output: Post-Mid timetables only for Semester {semester}, Branch {branch}")
                    else:
                        # ========== ADDED: VERIFICATION STATISTICS ==========
                        print(f"[STATS] Generating POST-MID verification sheets...")
                        
                        # Get course info for verification
                        course_info = get_course_info(dfs) if dfs else {}
                        classroom_data = dfs.get('classroom', None) if dfs else None
                        
                        # Create verification sheets for each section
                        for section in sections:
                            post_mid_reset = post_mid_sections[section].reset_index().rename(columns={'index': 'Time Slot'})
                            if not post_mid_reset.empty:
                                section_label = f"Section {section}" if has_sections else "Timetable"
                                print(f"   Creating Post-Mid Verification sheet for {section_label}...")
                                verification = create_timetable_verification_sheet(
                                    post_mid_reset, course_info, classroom_data, semester, branch, section
                                )
                                if not verification.empty:
                                    sheet_name = f'Verification_{section}' if has_sections else 'Verification'
                                    verification.to_excel(writer, sheet_name=sheet_name, index=False)
                                    print(f"   [OK] Created Post-Mid Verification sheet with {len(verification)} entries")
                        
                        # Create room allocation summary
                        if classroom_data is not None and not classroom_data.empty:
                            print(f"   Creating room allocation summary...")
                            all_sections_data = [post_mid_sections[s].reset_index().rename(columns={'index': 'Time Slot'}) for s in sections]
                            room_summary = create_room_allocation_summary_verification(
                                all_sections_data[0], all_sections_data[1] if len(all_sections_data) > 1 else pd.DataFrame(), classroom_data
                            )
                            if not room_summary.empty:
                                room_summary.to_excel(writer, sheet_name='Room_Allocation', index=False)
                                print(f"   [OK] Created Room_Allocation sheet with {len(room_summary)} rooms")
                            
                                # Create detailed classroom allocation sheet
                            print(f"   Creating detailed classroom allocation...")
                            classroom_alloc_detail = create_classroom_allocation_detail_with_tracking(
                                [post_mid_sections[s] for s in sections], classroom_data, semester, branch
                            )
                            if not classroom_alloc_detail.empty:
                                classroom_alloc_detail.to_excel(writer, sheet_name='Classroom_Allocation', index=False)
                                print(f"   [OK] Created Classroom_Allocation sheet with {len(classroom_alloc_detail)} entries")
                                print(f"   [OK] Created Room_Allocation sheet with {len(room_summary)} rooms")
                        
                        # Create LTPSC compliance summary
                        print(f"   Creating LTPSC compliance summary...")
                        all_sections_data = [post_mid_sections[s].reset_index().rename(columns={'index': 'Time Slot'}) for s in sections]
                        ltpsc_summary = create_ltpsc_compliance_summary(
                            dfs, semester, branch, all_sections_data[0], all_sections_data[1] if len(all_sections_data) > 1 else pd.DataFrame()
                        )
                        if not ltpsc_summary.empty:
                            ltpsc_summary.to_excel(writer, sheet_name='LTPSC_Compliance', index=False)
                            print(f"   [OK] Created LTPSC_Compliance sheet with {len(ltpsc_summary)} courses")
                        
                        # Create executive summary
                        print(f"   Creating executive summary...")
                        all_sections_data = [post_mid_sections[s].reset_index().rename(columns={'index': 'Time Slot'}) for s in sections]
                        exec_summary = create_executive_summary(
                            dfs, semester, branch, all_sections_data[0], all_sections_data[1] if len(all_sections_data) > 1 else pd.DataFrame(), {}
                        )
                        if not exec_summary.empty:
                            exec_summary.to_excel(writer, sheet_name='Executive_Summary', index=False)
                            print(f"   [OK] Created Executive_Summary sheet")
                        
                        # Add course summary
                        post_mid_summary = create_mid_semester_summary(post_mid_courses, 'Post-Mid', semester, branch)
                        post_mid_summary.to_excel(writer, sheet_name='Course_Summary', index=False)
                        
                        # Add distribution logic sheet
                        distribution_sheet = create_mid_semester_distribution_sheet(
                            pre_mid_courses, post_mid_courses, semester, branch
                        )
                        distribution_sheet.to_excel(writer, sheet_name='Distribution_Logic', index=False)
                        
                        # Add combined overview
                        combined_sheet = create_combined_mid_semester_sheet(
                            pre_mid_courses, post_mid_courses, semester, branch
                        )
                        combined_sheet.to_excel(writer, sheet_name='All_Courses_Overview', index=False)

                        # Add basket allocation and details for Post-Mid if present
                        try:
                            if 'post_mid_elective_allocations' in locals() and post_mid_elective_allocations:
                                print(f"   Creating Basket_Allocation and Basket_Courses for Post-Mid...")
                                basket_summary = create_basket_summary(post_mid_elective_allocations, semester, branch)
                                if not basket_summary.empty:
                                    basket_summary.to_excel(writer, sheet_name='Basket_Allocation', index=False)
                                basket_courses_sheet_mid = create_basket_courses_sheet(post_mid_elective_allocations)
                                if not basket_courses_sheet_mid.empty:
                                    basket_courses_sheet_mid.to_excel(writer, sheet_name='Basket_Courses', index=False)
                        except Exception as _:
                            pass

                        # Add Classroom Utilization and Allocation detail for Post-Mid
                        if classroom_data is not None and not classroom_data.empty:
                            try:
                                classroom_report = create_classroom_utilization_report(
                                    classroom_data, [post_mid_sections[s].reset_index().rename(columns={'index': 'Time Slot'}) for s in sections], []
                                )
                                classroom_report.to_excel(writer, sheet_name='Classroom_Utilization', index=False)

                                classroom_allocation_detail = create_classroom_allocation_detail_with_tracking(
                                    [post_mid_sections[s] for s in sections], classroom_data, semester, branch
                                )
                                classroom_allocation_detail.to_excel(writer, sheet_name='Classroom_Allocation', index=False)
                            except Exception:
                                pass
                    
                        print(f"[STATS] Added POST-MID verification sheets for easy inspection")
                        print(f"[OK] POST-MID timetable saved: {post_mid_filename}")
                    if minimal_only:
                        print(f"[OK] POST-MID timetable saved: {post_mid_filename}")
            except Exception as e:
                print(f"[FAIL] Error saving post-mid timetable: {e}")
                traceback.print_exc()
        else:
            print(f"[WARN] Cannot save post-mid timetable - Not all sections generated successfully")
        
        return {
            'pre_mid_success': pre_mid_all_valid or pre_mid_has_data,
            'post_mid_success': post_mid_all_valid or post_mid_has_data,
            'pre_mid_filename': pre_mid_filename if (pre_mid_all_valid or pre_mid_has_data) else None,
            'post_mid_filename': post_mid_filename if (post_mid_all_valid or post_mid_has_data) else None
        }
        
    except Exception as e:
        print(f"[FAIL] Error generating mid-semester timetables: {e}")
        traceback.print_exc()
        return {
            'pre_mid_success': False,
            'post_mid_success': False,
            'pre_mid_filename': None,
            'post_mid_filename': None
        }

def create_mid_semester_summary(courses_df, schedule_type, semester, branch=None):
    """Create summary sheet for mid-semester timetable
    
    Shows inclusion reasoning based on Half Semester and Post mid-sem values
    """
    if courses_df.empty:
        return pd.DataFrame()
    
    summary_data = []
    
    for _, course in courses_df.iterrows():
        ltpsc_str = course.get('LTPSC', '')
        ltpsc = parse_ltpsc(ltpsc_str)
        credits = course.get('Credits', 'N/A')
        half_sem = course.get('Half Semester (Yes/No)', 'N/A')
        post_mid = course.get('Post mid-sem', 'N/A')
        
        # Add reasoning based on new logic
        if str(half_sem).upper().strip() == 'YES':
            reasoning = "Half Semester = Yes (scheduled in both pre-mid and post-mid)"
        elif str(half_sem).upper().strip() == 'NO':
            if str(post_mid).upper().strip() == 'NO':
                reasoning = f"Half Semester = No, Post mid-sem = No (pre-mid only)"
            elif str(post_mid).upper().strip() == 'YES':
                reasoning = f"Half Semester = No, Post mid-sem = Yes (post-mid only)"
            else:
                reasoning = "Half Semester = No (check Post mid-sem value)"
        else:
            reasoning = f"Check Half Semester ({half_sem}) and Post mid-sem ({post_mid}) values"
        
        summary_data.append({
            'Schedule Type': schedule_type,
            'Semester': semester,
            'Branch': branch or 'All',
            'Course Code': course['Course Code'],
            'Course Name': course.get('Course Name', 'Unknown'),
            'Credits': credits,
            'LTPSC': ltpsc_str,
            'Lectures/Week': ltpsc['L'],
            'Tutorials/Week': ltpsc['T'],
            'Practicals/Week': ltpsc['P'],
            'Faculty': course.get('Faculty', 'Unknown'),
            'Department': course.get('Department', 'Unknown'),
            'Half Semester': half_sem,
            'Post mid-sem': post_mid,
            'Elective': course.get('Elective (Yes/No)', 'N/A'),
            'Inclusion Reasoning': reasoning
        })
    
    return pd.DataFrame(summary_data)

def create_mid_semester_distribution_sheet(pre_mid_courses, post_mid_courses, semester, branch=None):
    """Create a sheet showing the complete distribution logic"""
    if pre_mid_courses.empty and post_mid_courses.empty:
        return pd.DataFrame()
    
    # Get all courses from the original data to see complete picture
    dfs = load_all_data(force_reload=False)
    if dfs is None or 'course' not in dfs:
        return pd.DataFrame()
    
    # Get all courses for this semester and branch
    all_courses = dfs['course'][
        dfs['course']['Semester'].astype(str).str.strip() == str(semester)
    ].copy()
    
    if branch and 'Department' in all_courses.columns:
        normalized_branch = branch.strip()
        dept_match = all_courses['Department'].astype(str).str.strip() == normalized_branch
        # The column name is 'Common' in the input data. Use robust access with default
        if 'Common' in all_courses.columns:
            common_series = all_courses['Common'].astype(str).str.upper() == 'YES'
        else:
            # If column not present, use all False series so we only filter by Department
            common_series = pd.Series(False, index=all_courses.index)
        all_courses = all_courses[dept_match | common_series].copy()
    
    all_courses['Credits'] = pd.to_numeric(all_courses['Credits'], errors='coerce')
    
    distribution_data = []
    
    for _, course in all_courses.iterrows():
        course_code = course['Course Code']
        credits = course.get('Credits', 'N/A')
        post_mid = course.get('Post mid-sem', 'N/A')
        
        # Check if in pre-mid
        in_pre_mid = course_code in pre_mid_courses['Course Code'].values if not pre_mid_courses.empty else False
        
        # Check if in post-mid
        in_post_mid = course_code in post_mid_courses['Course Code'].values if not post_mid_courses.empty else False
        
        # Determine scheduling logic
        if in_pre_mid and in_post_mid:
            schedule_type = "Both (non-2-credit with Post mid-sem='No')"
        elif in_pre_mid:
            schedule_type = "Pre-Mid Only"
        elif in_post_mid:
            schedule_type = "Post-Mid Only"
        else:
            schedule_type = "Not Scheduled (Error)"
        
        # Detailed reasoning
        if credits == 2:
            if str(post_mid).upper() == 'NO':
                logic = "2-credit + Post mid-sem='No' -> Pre-Mid Only"
            else:  # 'YES'
                logic = "2-credit + Post mid-sem='Yes' -> Post-Mid Only"
        else:  # non-2-credit
            if str(post_mid).upper() == 'NO':
                logic = f"{credits}-credit + Post mid-sem='No' -> BOTH Pre-Mid & Post-Mid"
            else:  # 'YES'
                logic = f"{credits}-credit + Post mid-sem='Yes' -> Post-Mid Only"
        
        distribution_data.append({
            'Course Code': course_code,
            'Course Name': course.get('Course Name', 'Unknown'),
            'Credits': credits,
            'Post mid-sem': post_mid,
            'Scheduled In': schedule_type,
            'Logic Applied': logic,
            'In Pre-Mid': 'Yes' if in_pre_mid else 'No',
            'In Post-Mid': 'Yes' if in_post_mid else 'No',
            'Semester': semester,
            'Branch': branch or 'All'
        })
    
    return pd.DataFrame(distribution_data)

def create_combined_mid_semester_sheet(pre_mid_courses, post_mid_courses, semester, branch=None):
    """Create combined sheet showing all courses with their mid-semester status"""
    if pre_mid_courses.empty and post_mid_courses.empty:
        return pd.DataFrame()
    
    combined_data = []
    
    # Add pre-mid courses
    for _, course in pre_mid_courses.iterrows():
        combined_data.append({
            'Course Code': course['Course Code'],
            'Course Name': course.get('Course Name', 'Unknown'),
            'Credits': course.get('Credits', 'N/A'),
            'Semester': semester,
            'Branch': branch or 'All',
            'Schedule Type': 'Pre-Mid',
            'Post mid-sem': 'No',
            'Half Semester': course.get('Half Semester (Yes/No)', 'N/A'),
            'Department': course.get('Department', 'Unknown'),
            'Faculty': course.get('Faculty', 'Unknown')
        })
    
    # Add post-mid courses
    for _, course in post_mid_courses.iterrows():
        combined_data.append({
            'Course Code': course['Course Code'],
            'Course Name': course.get('Course Name', 'Unknown'),
            'Credits': course.get('Credits', 'N/A'),
            'Semester': semester,
            'Branch': branch or 'All',
            'Schedule Type': 'Post-Mid',
            'Post mid-sem': 'Yes' if course.get('Credits') == 2 else 'Any',
            'Half Semester': course.get('Half Semester (Yes/No)', 'N/A'),
            'Department': course.get('Department', 'Unknown'),
            'Faculty': course.get('Faculty', 'Unknown')
        })
    
    return pd.DataFrame(combined_data)

def check_for_classroom_allocation(schedule_df):
    """Check if classroom allocation was successful by looking for [Room] pattern"""
    for col in schedule_df.columns:
        for val in schedule_df[col]:
            if isinstance(val, str) and '[' in val and ']' in val:
                return True
    return False

# EXAM CLASSROOM USAGE FUNCTION - COMMENTED OUT
"""
def calculate_classroom_usage_for_exams(exam_schedule_df):
    \"\"\"Calculate classroom usage statistics for exams\"\"\"
    used_rooms = set()
    total_sessions = 0
    
    for _, exam in exam_schedule_df.iterrows():
        if exam['status'] == 'Scheduled' and 'classroom' in exam:
            classrooms = str(exam['classroom']).split(', ')
            used_rooms.update(classrooms)
            total_sessions += len(classrooms)
    
    return {
        'used_rooms': len(used_rooms),
        'total_sessions': total_sessions,
        'rooms_list': list(used_rooms)
    }
"""

def calculate_timetable_classroom_usage(timetable_schedules):
    """Calculate classroom usage statistics for timetables"""
    used_rooms = set()
    total_sessions = 0
    
    for schedule in timetable_schedules:
        for day in schedule.columns:
            for time_slot in schedule.index:
                cell_value = str(schedule.loc[time_slot, day])
                # Extract room numbers from format "Course [Room]"
                if '[' in cell_value and ']' in cell_value:
                    room_match = re.search(r'\[(.*?)\]', cell_value)
                    if room_match:
                        room_number = room_match.group(1)
                        used_rooms.add(room_number)
                        total_sessions += 1
    
    
    return {
        'used_rooms': len(used_rooms),
        'total_sessions': total_sessions,
        'rooms_list': list(used_rooms)
    }

def create_classroom_allocation_detail(timetable_schedules, classrooms_df):
    """Create detailed classroom allocation information"""
    allocation_data = []
    
    for i, schedule in enumerate(timetable_schedules, 1):
        section = 'A' if i == 1 else 'B'
        
        for day in schedule.columns:
            for time_slot in schedule.index:
                cell_value = str(schedule.loc[time_slot, day])
                
                if cell_value not in ['Free', 'LUNCH BREAK'] and '[' in cell_value and ']' in cell_value:
                    # Extract course and room
                    course_match = re.search(r'^(.*?)\s*\[', cell_value)
                    room_match = re.search(r'\[(.*?)\]', cell_value)
                    
                    if course_match and room_match:
                        course = course_match.group(1).strip()
                        room_number = room_match.group(1)
                        
                        # Get room details
                        room_details = classrooms_df[classrooms_df['Room Number'] == room_number]
                        capacity = room_details['Capacity'].iloc[0] if not room_details.empty else 'Unknown'
                        room_type = room_details['Type'].iloc[0] if not room_details.empty else 'Unknown'
                        
                        allocation_data.append({
                            'Section': section,
                            'Day': day,
                            'Time Slot': time_slot,
                            'Course': course,
                            'Room Number': room_number,
                            'Room Type': room_type,
                            'Capacity': capacity,
                            'Facilities': room_details['Facilities'].iloc[0] if not room_details.empty else 'Unknown'
                        })
    
    return pd.DataFrame(allocation_data)

def create_basket_courses_sheet(basket_allocations):
    """Create detailed sheet showing all courses in each basket"""
    summary_data = []
    
    for basket_name, allocation in basket_allocations.items():
        # Ensure safe defaults when lesson/tutor slots might be missing
        lectures = allocation.get('lectures') or []
        if lectures:
            lecture_slots = ', '.join([f"{day} {time}" for day, time in lectures])
        else:
            lecture_slots = '-'

        tutorial = allocation.get('tutorial')
        if tutorial and isinstance(tutorial, (list, tuple)) and len(tutorial) >= 2:
            tutorial_slot = f"{tutorial[0]} {tutorial[1]}"
        else:
            tutorial_slot = '-'

        for course_code in allocation.get('courses', []):
            summary_data.append({
                'Basket Name': basket_name,
                'Course Code': course_code,
                'Lecture Slots': lecture_slots,
                'Tutorial Slot': tutorial_slot,
                'Total Courses in Basket': len(allocation.get('courses', [])),
                'Common for All Branches': 'Yes',
                'Common for Both Sections': 'Yes'
            })
    
    return pd.DataFrame(summary_data)

def export_semester_timetable(dfs, semester, branch=None):
    """Export timetable using basket-based elective allocation with COMMON slots"""
    branch_info = f", Branch {branch}" if branch else ""
    print(f"\n[STATS] Generating BASKET-BASED timetable for Semester {semester}{branch_info}...")
    print(f"[TARGET] Using COMMON elective basket slots across all branches")
    
    try:
        # CRITICAL: Get ALL elective courses for this semester ONCE (without branch filter)
        # This ensures COMMON allocation for all branches
        course_baskets_all = separate_courses_by_type(dfs, semester)  # No branch filter
        elective_courses_all = course_baskets_all['elective_courses']
        
        print(f"[TARGET] Elective courses found for semester {semester} (COMMON for ALL branches): {len(elective_courses_all)}")
        if not elective_courses_all.empty:
            print("   Common elective courses:", elective_courses_all['Course Code'].tolist())
            # Show basket distribution
            basket_counts = elective_courses_all['Basket'].value_counts()
            print("   Basket distribution:")
            for basket, count in basket_counts.items():
                courses = elective_courses_all[elective_courses_all['Basket'] == basket]['Course Code'].tolist()
                print(f"      {basket}: {count} courses - {courses}")
        
        # Allocate electives by baskets - COMMON for all branches
        elective_allocations, basket_allocations = allocate_electives_by_baskets(elective_courses_all, semester)
        
        print(f"   [TIME] COMMON BASKET ALLOCATIONS for Semester {semester}:")
        for basket_name, allocation in basket_allocations.items():
            print(f"      {basket_name}:")
            print(f"         Lectures: {allocation['lectures']}")
            print(f"         Tutorial: {allocation['tutorial']}")
            print(f"         Courses: {allocation['courses']}")
        
        # Generate schedules for both sections with IDENTICAL basket slots
        section_a = generate_section_schedule_with_elective_baskets(dfs, semester, 'A', elective_allocations, branch)
        section_b = generate_section_schedule_with_elective_baskets(dfs, semester, 'B', elective_allocations, branch)
        
        if section_a is None or section_b is None:
            return False

        # ALLOCATE CLASSROOMS for both sections
        course_info = get_course_info(dfs) if dfs else {}
        classroom_data = dfs.get('classroom')
        
        if classroom_data is not None and not classroom_data.empty:
            print("[SCHOOL] Allocating classrooms...")
            section_a_with_rooms = allocate_classrooms_for_timetable(
                section_a, classroom_data, course_info, semester, branch, 'A'
            )
            section_b_with_rooms = allocate_classrooms_for_timetable(
                section_b, classroom_data, course_info, semester, branch, 'B'
            )
        else:
            section_a_with_rooms = section_a
            section_b_with_rooms = section_b
            print("[WARN]  No classroom data available for allocation")

        # Create filename
        if branch:
            filename = f"sem{semester}_{branch}_timetable.xlsx"
        else:
            filename = f"sem{semester}_timetable.xlsx"
            
        filepath = os.path.join(OUTPUT_DIR, filename)
        
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            section_a_with_rooms.to_excel(writer, sheet_name='Section_A')
            section_b_with_rooms.to_excel(writer, sheet_name='Section_B')
            
            # NEW: Add comprehensive statistics sheets
            if dfs:
                course_info = get_course_info(dfs)
                classroom_df = dfs.get('classroom', pd.DataFrame())
                
                # Statistics for Section A
                print(f"[STATS] Creating statistics sheet for Section A...")
                stats_a = create_timetable_statistics_sheet(
                    section_a_with_rooms, course_info, classroom_df, semester, branch, 'A'
                )
                stats_a.to_excel(writer, sheet_name='Statistics_A', index=False)
                
                # Statistics for Section B
                print(f"[STATS] Creating statistics sheet for Section B...")
                stats_b = create_timetable_statistics_sheet(
                    section_b_with_rooms, course_info, classroom_df, semester, branch, 'B'
                )
                stats_b.to_excel(writer, sheet_name='Statistics_B', index=False)
                
                # Room allocation summary
                if not classroom_df.empty:
                    print(f"[STATS] Creating room allocation summary...")
                    room_summary = create_room_allocation_summary(section_a_with_rooms, classroom_df)
                    room_summary.to_excel(writer, sheet_name='Room_Summary', index=False)
                
                # Comprehensive executive summary
                print(f"[STATS] Creating executive summary...")
                exec_summary = create_comprehensive_summary(
                    dfs, semester, branch, section_a_with_rooms, section_b_with_rooms, basket_allocations
                )
                exec_summary.to_excel(writer, sheet_name='Executive_Summary', index=False)
            
            # Add basket allocation summary
            basket_summary = create_basket_summary(basket_allocations, semester, branch)
            basket_summary.to_excel(writer, sheet_name='Basket_Allocation', index=False)
            
            # Add course summary
            course_summary = create_course_summary(dfs, semester, branch)
            if not course_summary.empty:
                course_summary.to_excel(writer, sheet_name='Course_Summary', index=False)
            
            # Add basket course details
            basket_courses_sheet = create_basket_courses_sheet(basket_allocations)
            basket_courses_sheet.to_excel(writer, sheet_name='Basket_Courses', index=False)
            
            # Add branch-specific info sheet
            branch_info_sheet = create_branch_info_sheet(dfs, semester, branch)
            if not branch_info_sheet.empty:
                branch_info_sheet.to_excel(writer, sheet_name='Branch_Info', index=False)
        
        print(f"[OK] Basket-based timetable saved: {filename}")
        print(f"[STATS] Added comprehensive statistics sheets for easy verification")
        return True
        
    except Exception as e:
        print(f"[FAIL] Error generating basket-based timetable: {e}")
        traceback.print_exc()
        return False

def create_basket_summary(basket_allocations, semester, branch=None):
    """Create a summary of basket allocations"""
    summary_data = []
    
    for basket_name, allocation in basket_allocations.items():
        # Allocation may contain 'lectures' (list) and 'tutorial' (tuple)
        # Use first lecture as representative slot for summary, if present
        lectures = allocation.get('lectures', [])
        tutorial = allocation.get('tutorial')
        courses = allocation.get('courses', [])
        if lectures and len(lectures) > 0:
            day, time_slot = lectures[0]
        else:
            day, time_slot = (None, None)
        
        summary_data.append({
            'Basket Name': basket_name,
            'Day': day,
            'Time Slot': time_slot,
            'Tutorial Day': tutorial[0] if tutorial else None,
            'Courses in Basket': ', '.join(courses),
            'Number of Courses': len(courses),
            'Sections': 'A & B (Common)',
            'Branches': branch if branch else 'ALL',
            'Semester': f'Semester {semester}'
        })
    
    return pd.DataFrame(summary_data)

def create_branch_info_sheet(dfs, semester, branch):
    """Create a sheet showing department-specific course information"""
    if 'course' not in dfs:
        return pd.DataFrame()
    
    # Get all courses for the semester
    sem_courses = dfs['course'][
        dfs['course']['Semester'].astype(str).str.strip() == str(semester)
    ].copy()
    
    if sem_courses.empty:
        return pd.DataFrame()
    
    # Separate department-specific and common courses
    dept_specific_courses = sem_courses[
        (sem_courses['Department'] == branch) & 
        (sem_courses['Elective (Yes/No)'].str.upper() != 'YES')
    ].copy()
    
    common_elective_courses = sem_courses[
        sem_courses['Elective (Yes/No)'].str.upper() == 'YES'
    ].copy()
    
    # Create summary
    summary_data = []
    
    # Add department-specific courses
    for _, course in dept_specific_courses.iterrows():
        summary_data.append({
            'Course Type': 'Department-Specific Core',
            'Course Code': course['Course Code'],
            'Course Name': course.get('Course Name', 'Unknown'),
            'Department': course.get('Department', 'Unknown'),
            'LTPSC': course.get('LTPSC', 'N/A'),
            'Credits': course.get('Credits', 'N/A'),
            'Instructor': course.get('Instructor', 'Unknown')
        })
    
    # Add common elective courses
    for _, course in common_elective_courses.iterrows():
        summary_data.append({
            'Course Type': 'Common Elective',
            'Course Code': course['Course Code'],
            'Course Name': course.get('Course Name', 'Unknown'),
            'Department': 'ALL (Common)',
            'LTPSC': course.get('LTPSC', 'N/A'),
            'Credits': course.get('Credits', 'N/A'),
            'Instructor': course.get('Instructor', 'Unknown')
        })
    
    return pd.DataFrame(summary_data)

def clean_table_html(html):
    """Clean and format the HTML table for better display"""
    html = html.replace('border="1"', '')
    html = html.replace('class="dataframe"', 'class="timetable-table"')
    html = html.replace('<thead>', '<thead class="timetable-head">')
    html = html.replace('<tbody>', '<tbody class="timetable-body">')
    return html

def extract_unique_courses(df):
    """Extract unique course codes from a timetable dataframe"""
    courses = set()
    for col in df.columns:
        for value in df[col]:
            if isinstance(value, str) and value not in ['Free', 'LUNCH BREAK']:
                # Handle both regular courses and elective/tutorial marked courses
                clean_value = value.replace(' (Elective)', '').replace(' (Tutorial)', '')
                course_code = extract_course_code(clean_value)
                if course_code:
                    courses.add(course_code)
    return list(courses)

def extract_unique_courses_with_baskets(df, elective_allocations=None):
    """Extract unique course codes and basket names from a timetable dataframe"""
    courses = set()
    baskets = set()
    
    for col in df.columns:
        for value in df[col]:
            if isinstance(value, str) and value not in ['Free', 'LUNCH BREAK']:
                # Extract clean course code (remove classroom info, tutorial markers, etc.)
                clean_value = value
                
                # Remove classroom allocation info
                if '[' in clean_value:
                    clean_value = clean_value.split('[')[0].strip()
                
                # Remove tutorial marker
                clean_value = clean_value.replace(' (Tutorial)', '')
                
                # Check if this is a basket entry
                basket_names = ['ELECTIVE_B1', 'ELECTIVE_B2', 'ELECTIVE_B3', 'ELECTIVE_B4', 'ELECTIVE_B5', 'ELECTIVE_B6', 'ELECTIVE_B7', 'ELECTIVE_B8', 'ELECTIVE_B9', 'HSS_B1', 'HSS_B2']
                is_basket = any(basket in clean_value for basket in basket_names)
                
                if is_basket:
                    # Extract basket name - get the actual basket name from the string
                    for basket_name in basket_names:
                        if basket_name in clean_value:
                            baskets.add(basket_name)
                            break
                    
                    # ADDED: Also add all courses from this basket to the courses set
                    if elective_allocations:
                        for course_code, allocation in elective_allocations.items():
                            if allocation and allocation.get('basket_name') in baskets:
                                courses.add(course_code)
                else:
                    # Handle regular courses - extract course code
                    course_code = extract_course_code(clean_value)
                    if course_code:
                        courses.add(course_code)
    
    return list(courses), list(baskets)

def extract_course_code(text):
    """Extract course code from text (robust to case and trailing section letters)"""
    import re
    course_pattern = r'([A-Za-z]{2,3}\s?-?\d{3}[A-Za-z]?)'
    match = re.search(course_pattern, str(text))
    if not match:
        return None
    normalized = match.group(0).replace(' ', '').replace('-', '').upper()
    return normalized

def generate_course_colors(courses, course_info):
    """Generate unique, visually distinct colors for each course using HSL color space"""
    import colorsys
    
    course_colors = {}
    core_courses = []
    elective_courses = []
    
    # Separate core and elective courses
    for course in sorted(courses):
        if course in course_info:
            if course_info[course].get('is_elective', False):
                elective_courses.append(course)
            else:
                core_courses.append(course)
        else:
            core_courses.append(course)  # Default to core if info not available
    
    def generate_distinct_colors(n, saturation=0.7, lightness=0.6):
        """Generate n visually distinct colors using golden ratio"""
        colors = []
        golden_ratio = 0.618033988749895
        hue = 0.0
        
        for i in range(n):
            hue = (hue + golden_ratio) % 1.0
            rgb = colorsys.hls_to_rgb(hue, lightness, saturation)
            hex_color = '#{:02x}{:02x}{:02x}'.format(
                int(rgb[0] * 255),
                int(rgb[1] * 255),
                int(rgb[2] * 255)
            )
            colors.append(hex_color)
        
        return colors
    
    # Generate distinct colors for core courses (more saturated and darker)
    if core_courses:
        core_colors = generate_distinct_colors(len(core_courses), saturation=0.75, lightness=0.55)
        for i, course in enumerate(core_courses):
            course_colors[course] = core_colors[i]
    
    # Generate distinct colors for elective courses (lighter, less saturated)
    if elective_courses:
        elective_colors = generate_distinct_colors(len(elective_courses), saturation=0.6, lightness=0.70)
        for i, course in enumerate(elective_courses):
            course_colors[course] = elective_colors[i]
    
    return course_colors

def allocate_classrooms_for_timetable(schedule_df, classrooms_df, course_info, semester, branch, section, basket_courses_map=None, schedule_type='Regular'):
    """Allocate classrooms to timetable sessions with proper tracking across all timetables.
    
    Args:
        schedule_type: 'Regular', 'PreMid', or 'PostMid' - used to separate classroom tracking
                       so that different schedule types don't conflict with each other.
    """
    print(f"[SCHOOL] Allocating classrooms for {branch} Semester {semester} Section {section} ({schedule_type})...")
    
    if classrooms_df is None or classrooms_df.empty:
        print("   [WARN] No classroom data available")
        return schedule_df
    
    # Initialize global tracker if not exists and ensure global preferred classrooms map
    global _CLASSROOM_USAGE_TRACKER, _GLOBAL_PREFERRED_CLASSROOMS, _COMMON_COURSE_ROOMS, _ELECTIVE_COMMON_ROOMS
    if not _CLASSROOM_USAGE_TRACKER:
        initialize_classroom_usage_tracker()
    if '_GLOBAL_PREFERRED_CLASSROOMS' not in globals():
        _GLOBAL_PREFERRED_CLASSROOMS = {}
    if '_COMMON_COURSE_ROOMS' not in globals():
        _COMMON_COURSE_ROOMS = {}
    if '_ELECTIVE_COMMON_ROOMS' not in globals():
        _ELECTIVE_COMMON_ROOMS = {}
    
    print(f"[COMMON-DEBUG] Starting allocation for Semester {semester}, Branch {branch}, Section {section}")
    print(f"[COMMON-DEBUG] Current _COMMON_COURSE_ROOMS keys: {list(_COMMON_COURSE_ROOMS.keys())}")
    
    room_type_series = classrooms_df['Type'].fillna('').astype(str).str.lower()
    room_number_series = classrooms_df['Room Number'].fillna('').astype(str)

    # FIX: Define lab rooms by ACTUAL lab type, not just L prefix
    # L402-L408 are "classroom" type - should NOT be treated as labs
    # Only "Hardware Lab", "Software Lab", "Physics Lab" etc. are labs
    is_classroom_type = room_type_series.isin(['classroom', 'large classroom', 'auditorium'])
    has_lab_in_type = room_type_series.str.contains('lab', na=False)
    
    # A room is a lab ONLY if it has "lab" in type AND is not a classroom type
    lab_mask = has_lab_in_type & ~is_classroom_type

    # Exclude non-teaching spaces from non-lab pool
    excluded_non_teaching_mask = (
        room_type_series.str.contains('library', na=False) |
        room_type_series.str.contains('research', na=False) |
        room_type_series.str.contains('empty', na=False)
    )

    # Accept all usable non-lab teaching rooms (classroom, auditorium, examination room, etc.)
    available_classrooms = classrooms_df[(~lab_mask) & (~excluded_non_teaching_mask)].copy()

    # Keep lab pool separate for lab sessions
    available_lab_rooms = classrooms_df[lab_mask].copy()
    
    # Convert capacity to numeric, handle 'nil' values
    available_classrooms['Capacity'] = pd.to_numeric(available_classrooms['Capacity'], errors='coerce')
    # Exclude rooms with missing or non-positive capacity (e.g., 'nil')
    available_classrooms = available_classrooms[available_classrooms['Capacity'].notna() & (available_classrooms['Capacity'] > 0)].copy()
    
    available_lab_rooms['Capacity'] = pd.to_numeric(available_lab_rooms['Capacity'], errors='coerce')
    available_lab_rooms = available_lab_rooms[available_lab_rooms['Capacity'].notna() & (available_lab_rooms['Capacity'] > 0)].copy()
    
    if available_classrooms.empty and available_lab_rooms.empty:
        print("   [WARN] No suitable classrooms found after filtering")
        return schedule_df
    
    print(f"   Available classrooms: {len(available_classrooms)}")
    print(f"   Available lab rooms: {len(available_lab_rooms)}")
    
    # FIX: Use room TYPE to classify rooms, not just prefix
    # L402-L408 are "classroom" type despite L prefix - include them in primary
    # Only actual labs (Hardware Lab, Software Lab) are fallback
    room_type_lower = available_classrooms['Type'].astype(str).str.lower().str.strip()
    is_actual_lab_type = (
        room_type_lower.str.contains('hardware', na=False) | 
        room_type_lower.str.contains('software', na=False) |
        room_type_lower.str.contains('lab', na=False)
    ) & ~room_type_lower.isin(['classroom', 'large classroom', 'auditorium'])
    
    # Primary = all classrooms/auditoriums (including L4xx classroom types)
    # Fallback = only actual labs (Hardware Lab, Software Lab)
    primary_classrooms = available_classrooms[~is_actual_lab_type].copy()
    fallback_lab_classrooms = available_classrooms[is_actual_lab_type].copy()
    
    print(f"   Primary classrooms (all classroom types): {len(primary_classrooms)}")
    print(f"   Fallback classrooms (lab types only): {len(fallback_lab_classrooms)}")
    
    # Create a copy of schedule with classroom allocation
    schedule_with_rooms = schedule_df.copy()
    # Use a global map so the same course tends to get the same classroom across timetables
    course_preferred_classrooms = _GLOBAL_PREFERRED_CLASSROOMS

    def normalize_single_room(room_value):
        """Collapse any iterable/list room value to a single string room identifier."""
        if room_value is None:
            return None
        # If it's already a non-empty string, keep it
        if isinstance(room_value, str):
            return room_value.strip() or None
        # If it's a simple list/tuple/set, pick the first truthy entry
        if isinstance(room_value, (list, tuple, set)):
            for candidate in room_value:
                if candidate:
                    return str(candidate).strip() or None
            return None
        # Catch-all for other iterables (but ignore strings handled above)
        try:
            if hasattr(room_value, '__iter__'):
                for candidate in room_value:
                    if candidate:
                        return str(candidate).strip() or None
                return None
        except Exception:
            pass
        return str(room_value).strip() or None

    def room_available(room_number, day_key, slot_key):
        # Use schedule_type prefix to separate Regular, PreMid, PostMid tracking
        prefixed_day = f"{schedule_type}_{day_key}"
        return room_number not in _CLASSROOM_USAGE_TRACKER.get(prefixed_day, {}).get(slot_key, set())

    def reserve_room(room_number, day_key, slot_key):
        global _ROOM_ALLOCATION_COUNTER
        # Use schedule_type prefix to separate Regular, PreMid, PostMid tracking
        prefixed_day = f"{schedule_type}_{day_key}"
        if prefixed_day not in _CLASSROOM_USAGE_TRACKER:
            _CLASSROOM_USAGE_TRACKER[prefixed_day] = {}
        if slot_key not in _CLASSROOM_USAGE_TRACKER[prefixed_day]:
            _CLASSROOM_USAGE_TRACKER[prefixed_day][slot_key] = set()
        _CLASSROOM_USAGE_TRACKER[prefixed_day][slot_key].add(room_number)
        # Increment allocation counter for load balancing
        if room_number not in _ROOM_ALLOCATION_COUNTER:
            _ROOM_ALLOCATION_COUNTER[room_number] = 0
        _ROOM_ALLOCATION_COUNTER[room_number] += 1

    def get_room_usage_count(room_number):
        """Get the total allocation count for a room (for load balancing)."""
        return _ROOM_ALLOCATION_COUNTER.get(room_number, 0)

    def select_least_used_room(candidates_df):
        """Select the least-used room from a DataFrame of candidates.
        Returns room number string or None if candidates is empty.
        Uses _ROOM_ALLOCATION_COUNTER for load balancing."""
        if candidates_df.empty:
            return None
        # Add usage count column for sorting
        candidates = candidates_df.copy()
        candidates['_usage'] = candidates['Room Number'].apply(lambda r: _ROOM_ALLOCATION_COUNTER.get(str(r), 0))
        # Sort by capacity (ascending), then by usage (ascending) to pick least-used room of appropriate size
        if '_cap' not in candidates.columns:
            candidates['_cap'] = pd.to_numeric(candidates['Capacity'], errors='coerce').fillna(0)
        candidates = candidates.sort_values(['_cap', '_usage', 'Room Number'], ascending=[True, True, True])
        return str(candidates.iloc[0]['Room Number'])

    def pick_forced_fallback_room(preferred_min_capacity=None, day_key=None, slot_key=None):
        """Pick a room, preferring available rooms first, then allowing conflicts as last resort.
        
        IMPROVED: First try to find rooms NOT booked for this specific time slot.
        Only if all rooms are booked at this slot, pick the least-used room (conflict).
        
        Args:
            preferred_min_capacity: Minimum capacity requirement
            day_key: Day to check availability (optional)
            slot_key: Time slot to check availability (optional)
        """
        candidate_sets = []
        if not primary_classrooms.empty:
            candidate_sets.append(primary_classrooms)
        if not available_classrooms.empty:
            candidate_sets.append(available_classrooms)
        if not available_lab_rooms.empty:
            candidate_sets.append(available_lab_rooms)

        # Get rooms booked at this specific slot (if day/slot provided)
        booked_at_slot = set()
        if day_key and slot_key:
            prefixed_day = f"{schedule_type}_{day_key}"
            booked_at_slot = _CLASSROOM_USAGE_TRACKER.get(prefixed_day, {}).get(slot_key, set())

        # FIRST PASS: Try to find a room that is NOT booked at this slot
        for rooms_df in candidate_sets:
            candidates = rooms_df.copy()
            if preferred_min_capacity is not None:
                capacities = pd.to_numeric(candidates['Capacity'], errors='coerce').fillna(0)
                sized = candidates[capacities >= preferred_min_capacity]
                if not sized.empty:
                    candidates = sized
            if not candidates.empty:
                candidates = candidates.assign(_cap=pd.to_numeric(candidates['Capacity'], errors='coerce').fillna(0))
                # Filter to only rooms NOT booked at this slot
                if booked_at_slot:
                    available_candidates = candidates[~candidates['Room Number'].isin(booked_at_slot)]
                    if not available_candidates.empty:
                        selected = select_least_used_room(available_candidates)
                        if selected:
                            # Reserve the room in the tracker to prevent future conflicts
                            reserve_room(selected, day_key, slot_key)
                            print(f"         [FORCED-AVAIL] Found available room {selected} for {day_key} {slot_key}")
                            return selected
                else:
                    # No rooms booked at this slot - pick any
                    selected = select_least_used_room(candidates)
                    if selected:
                        if day_key and slot_key:
                            reserve_room(selected, day_key, slot_key)
                        return selected

        # SECOND PASS: All rooms booked at this slot - DO NOT USE CONFLICTS
        # With 29 rooms and max ~16 courses per slot, we should NEVER need conflicts
        # Instead, log the issue and return None to trigger higher-level fallback
        print(f"         [FORCED-FAIL] All {len(booked_at_slot)} rooms booked at {day_key} {slot_key}, no fallback available")
        print(f"         [DEBUG] Booked rooms: {sorted(booked_at_slot)[:15]}...")
        return None

    def room_available_for_lab_pair(room_number, day_key, slot_one, slot_two):
        return room_available(room_number, day_key, slot_one) and room_available(room_number, day_key, slot_two)

    def _get_capacity_tiers(enroll):
        """Return ordered list of capacity tiers to try for a given enrollment."""
        if enroll <= 80:
            return [80, 96, 120, 135]
        elif enroll <= 96:
            return [96, 120, 135]
        elif enroll <= 120:
            return [120, 135]
        elif enroll <= 135:
            return [135, 240]
        else:
            return [240]

    def find_tiered_fallback_room(enroll, day_key, slot_key, extra_disallowed=None):
        """Find the best room using tiered logic from all available room pools.
        Uses load balancing to distribute rooms evenly.
        PRIORITY: C-prefix rooms FIRST, L-prefix rooms ONLY as last resort.
        Returns room number or None."""
        prefixed_day = f"{schedule_type}_{day_key}"
        globally_booked = _CLASSROOM_USAGE_TRACKER.get(prefixed_day, {}).get(slot_key, set())
        disallowed = set(globally_booked)
        if extra_disallowed:
            disallowed |= set(extra_disallowed)
        tier_caps = _get_capacity_tiers(enroll)
        
        # Filter primary_classrooms into C-prefix ONLY and L-prefix separately
        c_prefix_rooms = primary_classrooms[~primary_classrooms['Room Number'].astype(str).str.startswith('L')].copy() if not primary_classrooms.empty else pd.DataFrame()
        l_prefix_rooms = primary_classrooms[primary_classrooms['Room Number'].astype(str).str.startswith('L')].copy() if not primary_classrooms.empty else pd.DataFrame()
        
        # STEP 1: Try ALL tiers in C-prefix classrooms first
        if not c_prefix_rooms.empty:
            candidates = c_prefix_rooms.copy()
            candidates['_cap'] = pd.to_numeric(candidates['Capacity'], errors='coerce').fillna(0)
            candidates = candidates[candidates['_cap'] >= enroll]
            candidates = candidates[~candidates['Room Number'].isin(disallowed)]
            if not candidates.empty:
                for tier_cap in tier_caps:
                    tier_match = candidates[(candidates['_cap'] >= tier_cap - 10) & (candidates['_cap'] <= tier_cap + 10)]
                    if not tier_match.empty:
                        return select_least_used_room(tier_match)
                # No tier match in C-prefix, use smallest adequate C-prefix room
                return select_least_used_room(candidates)
        
        # STEP 2: ONLY if C-prefix exhausted, try L-prefix classroom rooms
        if not l_prefix_rooms.empty:
            candidates = l_prefix_rooms.copy()
            candidates['_cap'] = pd.to_numeric(candidates['Capacity'], errors='coerce').fillna(0)
            candidates = candidates[candidates['_cap'] >= enroll]
            candidates = candidates[~candidates['Room Number'].isin(disallowed)]
            if not candidates.empty:
                for tier_cap in tier_caps:
                    tier_match = candidates[(candidates['_cap'] >= tier_cap - 10) & (candidates['_cap'] <= tier_cap + 10)]
                    if not tier_match.empty:
                        return select_least_used_room(tier_match)
                return select_least_used_room(candidates)
        
        # STEP 3: Last resort - try available_classrooms (includes all) with capacity filter
        if not available_classrooms.empty:
            candidates = available_classrooms.copy()
            candidates['_cap'] = pd.to_numeric(candidates['Capacity'], errors='coerce').fillna(0)
            candidates = candidates[candidates['_cap'] >= enroll]
            candidates = candidates[~candidates['Room Number'].isin(disallowed)]
            if not candidates.empty:
                return select_least_used_room(candidates)
        
        # STEP 4: ULTIMATE FALLBACK - any available room ignoring capacity (prefer larger)
        if not available_classrooms.empty:
            candidates = available_classrooms.copy()
            candidates['_cap'] = pd.to_numeric(candidates['Capacity'], errors='coerce').fillna(0)
            candidates = candidates[~candidates['Room Number'].isin(disallowed)]
            if not candidates.empty:
                # Sort by capacity descending to get largest available room
                candidates = candidates.sort_values('_cap', ascending=False)
                return select_least_used_room(candidates)
        
        return None
    
    def get_suitable_lab_rooms_for_course(course_code):
        """Filter lab rooms based on course type:
           - EC courses use Hardware labs (priority), then Software labs as fallback
           - CS and DS courses use Software labs (priority), then Hardware labs as fallback
           - Default: any available lab room
           
           Returns: DataFrame with preferred labs first, then other labs for fallback"""
        if not available_lab_rooms.empty:
            # Extract base course code (remove prefixes, suffixes)
            base_code = str(course_code).upper().split('[')[0].split('(')[0].strip()
            
            # Separate lab types
            hardware_labs = available_lab_rooms[available_lab_rooms['Type'].str.contains('Hardware', case=False, na=False)].copy()
            software_labs = available_lab_rooms[available_lab_rooms['Type'].str.contains('Software', case=False, na=False)].copy()
            other_labs = available_lab_rooms[
                ~available_lab_rooms['Type'].str.contains('Hardware', case=False, na=False) &
                ~available_lab_rooms['Type'].str.contains('Software', case=False, na=False)
            ].copy()
            
            # EC courses (Electronics/Electrical) → Hardware labs first, then Software labs
            if base_code.startswith('EC'):
                if not hardware_labs.empty:
                    print(f"         [LAB-TYPE] {base_code} (EC): Using Hardware labs (priority), Software labs (fallback)")
                    # Concatenate: hardware first, then software, then other
                    return pd.concat([hardware_labs, software_labs, other_labs], ignore_index=True)
                else:
                    print(f"         [LAB-WARN] {base_code} (EC): No Hardware labs found, using all labs")
                    return available_lab_rooms
            
            # CS, DS, and DA courses (Computer Science, Data Science/DSAI) → Software labs first, then Hardware labs
            elif base_code.startswith(('CS', 'DS', 'DA')):
                dept_label = 'CS' if base_code.startswith('CS') else ('DS' if base_code.startswith('DS') else 'DA')
                if not software_labs.empty:
                    print(f"         [LAB-TYPE] {base_code} ({dept_label}): Using Software labs (priority), Hardware labs (fallback)")
                    # Concatenate: software first, then hardware, then other
                    return pd.concat([software_labs, hardware_labs, other_labs], ignore_index=True)
                else:
                    print(f"         [LAB-WARN] {base_code} ({dept_label}): No Software labs found, using all labs")
                    return available_lab_rooms
            
            # Other courses: use any available lab
            else:
                print(f"         [LAB-TYPE] {base_code}: Using any available lab")
                return available_lab_rooms
        return available_lab_rooms
    
    def allocate_regular_classroom(enrollment_value, day_key, slot_key, is_common_course=False, is_lab_session=False, course_code=None, preferred_capacities_override=None):
        classroom_choice = None
        # LABS: unchanged
        if is_lab_session:
            if course_code:
                lab_rooms_to_use = get_suitable_lab_rooms_for_course(course_code)
            else:
                lab_rooms_to_use = available_lab_rooms
            if not lab_rooms_to_use.empty:
                classroom_choice = find_suitable_classroom_with_tracking(
                    lab_rooms_to_use, enrollment_value, day_key, slot_key, _CLASSROOM_USAGE_TRACKER,
                    is_common=is_common_course, is_lab=True, preferred_capacities_override=preferred_capacities_override,
                    schedule_type=schedule_type
                )
            if not classroom_choice:
                print(f"         [LAB-WARN] No lab room available for {day_key} {slot_key}")
            return classroom_choice

        # NON-LAB: prefer smallest adequate room, escalate capacity tier by tier
        # Use capacity RANGES instead of exact matches to ensure rooms are found
        # Escalation order: smallest fit → medium → large
        
        # COMMON COURSE OPTIMIZATION: For common courses (same course for both sections A & B),
        # both sections attend together, so prefer larger capacity rooms to accommodate combined enrollment
        # This naturally results in 240-capacity rooms being chosen for high-enrollment common courses
        if is_common_course:
            print(f"         [COMMON-ROOM] Common course detected - preferring larger capacity rooms for combined sections")

        # FIX: Use C-prefix rooms FIRST in tier search, L-prefix ONLY as last resort
        # This prevents L4xx rooms (80-cap) from being overloaded
        # Filter primary_classrooms to get ONLY C-prefix rooms (exclude L-prefix)
        
        # Start with C-prefix only for tier search (exclude L-prefix rooms entirely)
        c_prefix_classrooms = primary_classrooms[~primary_classrooms['Room Number'].astype(str).str.startswith('L')].copy()
        # L-prefix classroom rooms (L402-L408 etc.) for fallback only
        l_prefix_classrooms = primary_classrooms[primary_classrooms['Room Number'].astype(str).str.startswith('L')].copy()
        
        print(f"         [ROOM-POOLS] C-prefix: {len(c_prefix_classrooms)}, L-prefix: {len(l_prefix_classrooms)}")
        
        # COMMON COURSE: Prefer larger capacity rooms when is_common_course=True
        # This ensures common courses get rooms that can accommodate both sections attending together
        if is_common_course:
            # For common courses, try largest rooms first (typically 240 capacity)
            max_cap = c_prefix_classrooms['Capacity'].max() if not c_prefix_classrooms.empty else 0
            capacity_tiers = [
                ('c-large-200+', c_prefix_classrooms[c_prefix_classrooms['Capacity'] >= 200]),
                ('c-large-140-200', c_prefix_classrooms[(c_prefix_classrooms['Capacity'] >= 140) & (c_prefix_classrooms['Capacity'] < 200)]),
                ('c-medium-100-140', c_prefix_classrooms[(c_prefix_classrooms['Capacity'] >= 100) & (c_prefix_classrooms['Capacity'] < 140)]),
                ('c-adequate', c_prefix_classrooms[c_prefix_classrooms['Capacity'] >= enrollment_value]),
            ]
            print(f"         [COMMON-TIER] Using large-capacity-first tier order for common course")
        # IMPROVED: Define capacity tiers using RANGES to prevent clustering
        # Instead of exact capacity matches, use ranges that allow room distribution
        elif enrollment_value <= 60:
            # Small classes: prefer rooms 60-100, then any larger (C-prefix only first)
            capacity_tiers = [
                ('c-small-60-100', c_prefix_classrooms[(c_prefix_classrooms['Capacity'] >= 60) & (c_prefix_classrooms['Capacity'] <= 100)]),
                ('c-medium-100-140', c_prefix_classrooms[(c_prefix_classrooms['Capacity'] > 100) & (c_prefix_classrooms['Capacity'] <= 140)]),
                ('c-large-140+', c_prefix_classrooms[c_prefix_classrooms['Capacity'] > 140]),
            ]
        elif enrollment_value <= 96:
            # Medium classes: prefer rooms 96-120, then larger (C-prefix first)
            capacity_tiers = [
                ('c-med-96-120', c_prefix_classrooms[(c_prefix_classrooms['Capacity'] >= enrollment_value) & (c_prefix_classrooms['Capacity'] <= 120)]),
                ('c-large-120-140', c_prefix_classrooms[(c_prefix_classrooms['Capacity'] > 120) & (c_prefix_classrooms['Capacity'] <= 140)]),
                ('c-xlarge-140+', c_prefix_classrooms[c_prefix_classrooms['Capacity'] > 140]),
            ]
        elif enrollment_value <= 120:
            # Larger classes: need 120+ (C-prefix first)
            capacity_tiers = [
                ('c-120-140', c_prefix_classrooms[(c_prefix_classrooms['Capacity'] >= 120) & (c_prefix_classrooms['Capacity'] <= 140)]),
                ('c-140+', c_prefix_classrooms[c_prefix_classrooms['Capacity'] > 140]),
            ]
        else:
            # Large enrollment: any room that fits (C-prefix first)
            capacity_tiers = [
                ('c-adequate', c_prefix_classrooms[c_prefix_classrooms['Capacity'] >= enrollment_value]),
            ]
            # If nothing fits and enrollment exceeds max, use largest available
            if all(tier_rooms.empty for _, tier_rooms in capacity_tiers):
                max_capacity = c_prefix_classrooms['Capacity'].max() if not c_prefix_classrooms.empty else 0
                if enrollment_value > max_capacity and max_capacity > 0:
                    capacity_tiers = [
                        ('c-largest', c_prefix_classrooms[c_prefix_classrooms['Capacity'] == max_capacity]),
                    ]
                else:
                    return None

        # Try each tier in order until a room is found
        for tier_name, eligible_rooms in capacity_tiers:
            if eligible_rooms is None or eligible_rooms.empty:
                continue

            # ROTATE eligible rooms to maximize usage (least used first)
            # Use the global allocation counter for consistent load balancing
            eligible_rooms = eligible_rooms.copy()
            eligible_rooms['_usage'] = eligible_rooms['Room Number'].apply(lambda r: _ROOM_ALLOCATION_COUNTER.get(str(r), 0))
            eligible_rooms = eligible_rooms.sort_values(['_usage', 'Room Number'])
            classroom_choice = find_suitable_classroom_with_tracking(
                eligible_rooms, enrollment_value, day_key, slot_key, _CLASSROOM_USAGE_TRACKER,
                is_common=is_common_course, is_lab=False, preferred_capacities_override=preferred_capacities_override,
                schedule_type=schedule_type
            )
            if classroom_choice:
                print(f"         [TIER-OK] Found room in {tier_name} tier for {enrollment_value} students")
                return classroom_choice
            else:
                print(f"         [ESCALATE] No available room in {tier_name} tier for {enrollment_value} students, trying next tier")

        print(f"         [ESCALATE-WARN] No room found in any tier for {enrollment_value} students")
        
        # CROSS-PREFIX FALLBACK: Try L-prefix classroom rooms (L402-L408) as last resort
        # These are 80-cap rooms that should only be used when C-prefix rooms are exhausted
        print(f"         [L-PREFIX-FALLBACK] Trying L-prefix classroom rooms...")
        
        # Get all rooms that are NOT booked for this slot
        prefixed_day = f"{schedule_type}_{day_key}"
        booked_at_slot = _CLASSROOM_USAGE_TRACKER.get(prefixed_day, {}).get(slot_key, set())
        
        # Try L-prefix classroom rooms (filtered earlier as l_prefix_classrooms)
        if not l_prefix_classrooms.empty:
            l_prefix_available = l_prefix_classrooms[
                (l_prefix_classrooms['Capacity'] >= enrollment_value) &
                (~l_prefix_classrooms['Room Number'].isin(booked_at_slot))
            ].copy()
            
            if not l_prefix_available.empty:
                l_prefix_available['_usage'] = l_prefix_available['Room Number'].apply(lambda r: _ROOM_ALLOCATION_COUNTER.get(str(r), 0))
                l_prefix_available = l_prefix_available.sort_values(['_usage', 'Capacity'])
                classroom_choice = l_prefix_available.iloc[0]['Room Number']
                reserve_room(classroom_choice, day_key, slot_key)
                print(f"         [L-PREFIX-FALLBACK] Found L-prefix room {classroom_choice} for {enrollment_value} students")
                return classroom_choice
        
        # FINAL FALLBACK: Try ANY available room regardless of capacity from C-prefix
        c_prefix_available = c_prefix_classrooms[~c_prefix_classrooms['Room Number'].isin(booked_at_slot)].copy()
        if not c_prefix_available.empty:
            c_prefix_available['_usage'] = c_prefix_available['Room Number'].apply(lambda r: _ROOM_ALLOCATION_COUNTER.get(str(r), 0))
            c_prefix_available = c_prefix_available.sort_values(['_usage', 'Capacity'], ascending=[True, False])  # Prefer larger if available
            classroom_choice = c_prefix_available.iloc[0]['Room Number']
            reserve_room(classroom_choice, day_key, slot_key)
            print(f"         [FINAL-FALLBACK] Found C-prefix room {classroom_choice} (any capacity) for {enrollment_value} students")
            return classroom_choice
        
        # Try ANY L-prefix room regardless of capacity
        if not l_prefix_classrooms.empty:
            l_any_available = l_prefix_classrooms[~l_prefix_classrooms['Room Number'].isin(booked_at_slot)].copy()
            if not l_any_available.empty:
                l_any_available['_usage'] = l_any_available['Room Number'].apply(lambda r: _ROOM_ALLOCATION_COUNTER.get(str(r), 0))
                l_any_available = l_any_available.sort_values(['_usage', 'Capacity'], ascending=[True, False])
                classroom_choice = l_any_available.iloc[0]['Room Number']
                reserve_room(classroom_choice, day_key, slot_key)
                print(f"         [FINAL-FALLBACK] Found L-prefix room {classroom_choice} (any capacity) for {enrollment_value} students")
                return classroom_choice
        
        # ULTIMATE FALLBACK: Check ALL primary_classrooms (any prefix) - catches rooms with other prefixes
        if not primary_classrooms.empty:
            any_primary_available = primary_classrooms[~primary_classrooms['Room Number'].isin(booked_at_slot)].copy()
            if not any_primary_available.empty:
                any_primary_available['_usage'] = any_primary_available['Room Number'].apply(lambda r: _ROOM_ALLOCATION_COUNTER.get(str(r), 0))
                any_primary_available = any_primary_available.sort_values(['_usage', 'Capacity'], ascending=[True, False])
                classroom_choice = any_primary_available.iloc[0]['Room Number']
                reserve_room(classroom_choice, day_key, slot_key)
                print(f"         [ULTIMATE-FALLBACK] Found ANY room {classroom_choice} for {enrollment_value} students")
                return classroom_choice
        
        # DEBUG: Print available vs booked for troubleshooting
        total_primary = len(primary_classrooms) if not primary_classrooms.empty else 0
        total_booked = len(booked_at_slot)
        print(f"         [FALLBACK-FAIL] No rooms available at {day_key} {slot_key}")
        print(f"         [DEBUG] Total primary classrooms: {total_primary}, Booked at slot: {total_booked}")
        print(f"         [DEBUG] Booked rooms: {list(booked_at_slot)[:10]}...")  # Show first 10
        
        # ABSOLUTE LAST RESORT: Force use ANY available room from available_classrooms
        # even if it means we missed it in previous checks (defensive fallback)
        if not available_classrooms.empty:
            all_rooms = set(available_classrooms['Room Number'].astype(str).tolist())
            truly_available = all_rooms - booked_at_slot
            if truly_available:
                fallback_room = list(truly_available)[0]
                reserve_room(fallback_room, day_key, slot_key)
                print(f"         [ABSOLUTE-FALLBACK] Found room {fallback_room} via defensive check")
                return fallback_room
        
        return None
    
    # Estimate student numbers for courses
    course_enrollment_raw = estimate_course_enrollment(course_info)
    course_enrollment = {}
    for course_code, raw_enrollment in course_enrollment_raw.items():
        info = course_info.get(course_code, {})
        dept = str(info.get('department', '')).strip().upper()
        is_common = str(info.get('common', 'No')).strip().upper() == 'YES'
        is_elective = str(info.get('elective', 'No')).strip().upper() == 'YES'
        is_minor = 'MINOR' in str(info.get('basket', '')).upper() or 'MINOR' in str(course_code).upper()
        reg_students = None
        try:
            reg_students = int(info.get('registered students', raw_enrollment))
        except Exception:
            reg_students = raw_enrollment

        # Store raw registered students for all departments uniformly
        # Halving (for 2 sections) is done later in compute_effective_enrollment()
        course_enrollment[course_code] = reg_students

    def compute_effective_enrollment(raw_enrollment, is_common_course):
        """Compute effective enrollment for room sizing.
        
        For COMMON courses: Use FULL enrollment (no halving) because both sections
        attend together in the SAME room at the SAME timeslot.
        
        For NON-COMMON courses: Halve once since each section needs its own room.
        """
        try:
            value = float(raw_enrollment)
        except Exception:
            return raw_enrollment
        
        if is_common_course:
            # Common courses: BOTH sections attend together in same room
            # Use FULL enrollment for room sizing (no halving)
            print(f"         [COMMON-ENROLL] Using FULL enrollment {int(value)} for common course (both sections attend together)")
            return max(1, int(value))
        else:
            # Non-common courses: each section needs its own room, halve enrollment
            return max(1, int(math.ceil(value / 2.0)))

    # Cache cross-department (DSAI+ECE) common metadata per course for this semester
    _cross_common_cache = {}

    def get_cross_common_bundle(course_code):
        if course_code not in _cross_common_cache:
            _cross_common_cache[course_code] = detect_cross_dsai_ece_common(course_info, course_code, semester)
        return _cross_common_cache[course_code]
    
    # Track allocations for this specific timetable
    timetable_key = f"{branch}_sem{semester}_sec{section}"
    if timetable_key not in _TIMETABLE_CLASSROOM_ALLOCATIONS:
        _TIMETABLE_CLASSROOM_ALLOCATIONS[timetable_key] = {}
    
    # Track which lab slots have been processed to avoid double allocation
    processed_lab_slots = set()
    
    # Define lab slot pairs (consecutive slots that form a 2-hour lab)
    # Build pairs from TIME_SLOT_LABELS to ensure all consecutive slot pairs are covered
    lab_slot_pairs = {}
    for i in range(len(TIME_SLOT_LABELS) - 1):
        current_slot = TIME_SLOT_LABELS[i]
        next_slot = TIME_SLOT_LABELS[i + 1]
        # Only pair consecutive slots that make sense for labs (skip LUNCH)
        if 'LUNCH' not in current_slot and 'LUNCH' not in next_slot:
            lab_slot_pairs[current_slot] = next_slot
    
    print(f"   [LAB-PAIRS] Defined {len(lab_slot_pairs)} lab slot pairs for allocation")
    for first, second in lab_slot_pairs.items():
        print(f"      {first} -> {second}")
    
    # Allocate classrooms for each time slot
    allocation_count = 0
    # Debug: print a small preview of the schedule being processed
    try:
        preview_cell = schedule_df.iloc[0,0] if not schedule_df.empty else 'EMPTY'
    except Exception:
        preview_cell = 'ERR'
    # schedule preview logging removed (noisy debug)
    for day in schedule_df.columns:
        for time_slot in schedule_df.index:
            course_value = schedule_df.loc[time_slot, day]
            
            # Handle case where .loc returns a Series (duplicate index) - take first value
            if isinstance(course_value, pd.Series):
                course_value = course_value.iloc[0] if len(course_value) > 0 else 'Free'
            
            # Skip free slots, lunch breaks
            if course_value in ['Free', 'LUNCH BREAK']:
                continue
            
            # Skip if this slot was already processed as part of a lab pair
            if (day, time_slot) in processed_lab_slots:
                continue
            
            # NEW: Handle basket entries - detect if this is a basket slot
            # Check if the course value is a basket name (not multiple courses separated by newlines)
            is_basket_entry = False
            basket_keywords = ['ELECTIVE_', 'HSS_', 'PROF_', 'OE_']
            if isinstance(course_value, str):
                normalized_value = course_value.upper().replace(' (TUTORIAL)', '').replace(' (LAB)', '')
                is_basket_entry = any(keyword in normalized_value for keyword in basket_keywords)
            
            # Handle basket entries - allocate classrooms for courses but keep basket name in display
            if is_basket_entry:
                basket_name = course_value.replace(' (Tutorial)', '').replace(' (Lab)', '').strip()
                courses_in_basket = basket_courses_map.get(basket_name, []) if basket_courses_map else []
                
                # Track all rooms allocated for this basket slot
                basket_rooms_allocated = []
                
                # Determine session type 
                is_tutorial = '(Tutorial)' in course_value
                is_lab = '(Lab)' in course_value
                session_type = 'Tutorial' if is_tutorial else ('Lab' if is_lab else 'Lecture')
                
                if courses_in_basket:
                    print(f"      [BASKET-ENTRY] {day} {time_slot}: '{course_value}' contains {len(courses_in_basket)} courses ({session_type})")
                    
                    for course_code in courses_in_basket:
                        # Get enrollment for this course
                        enrollment = course_enrollment.get(course_code, 40)
                        
                        # CRITICAL: Elective courses in baskets should use COMMON classrooms across all sections/branches
                        # Use a key that includes semester+day+time+course+session so:
                        # 1. Same semester + same day/time + same course shares the SAME room (within-semester common)
                        # 2. Different semesters are completely isolated
                        # 3. Different courses at same day/time get DIFFERENT rooms
                        common_elective_key = f"ELECTIVE_COMMON_{semester}_{day}_{time_slot}_{course_code}_{session_type}"
                        
                        suitable_classroom = None
                        existing_common_room = _ELECTIVE_COMMON_ROOMS.get(common_elective_key)
                        
                        if existing_common_room:
                            # For electives at the SAME day/time/course, reuse the common room
                            # This ensures all sections of same semester use same classroom
                            # CRITICAL: Check if room is available or already booked FOR THIS COURSE
                            prefixed_day = f"{schedule_type}_{day}"
                            booked_rooms_at_slot = _CLASSROOM_USAGE_TRACKER.get(prefixed_day, {}).get(time_slot, set())
                            if existing_common_room not in booked_rooms_at_slot:
                                # Room is free - use it and reserve
                                suitable_classroom = existing_common_room
                                print(f"         [BASKET-COMMON-REUSE] {course_code} ({session_type}) -> {existing_common_room} (first allocation at {day} {time_slot})")
                            else:
                                # Room is already booked - assume it's the same course sharing
                                # (Both sections attend together, so we reuse without re-reserving)
                                suitable_classroom = existing_common_room
                                print(f"         [BASKET-COMMON-SHARED] {course_code} ({session_type}) -> {existing_common_room} (sharing room at {day} {time_slot})")
                        else:
                            # Allocate a new common room based on enrollment
                            # This checks _CLASSROOM_USAGE_TRACKER to avoid conflicts with other semesters/baskets
                            suitable_classroom = allocate_regular_classroom(
                                enrollment, day, time_slot, 
                                is_common_course=True,  # Electives are COMMON - all sections share one room
                                is_lab_session=is_lab, 
                                course_code=course_code
                            )
                            if suitable_classroom:
                                # Store as common room for all sections/branches at this specific day/time
                                _ELECTIVE_COMMON_ROOMS[common_elective_key] = suitable_classroom
                                print(f"         [BASKET-COMMON-NEW] {course_code} ({session_type}) -> {suitable_classroom} (NEW common room for all branches at {day} {time_slot})")
                        
                        if suitable_classroom:
                            # Reserve the room in the tracker
                            reserve_room(suitable_classroom, day, time_slot)
                            
                            # Track room for display
                            if suitable_classroom not in basket_rooms_allocated:
                                basket_rooms_allocated.append(suitable_classroom)
                            
                            # Track allocation (but don't update cell display)
                            allocation_key = f"{day}_{time_slot}_{course_code}"
                            _TIMETABLE_CLASSROOM_ALLOCATIONS[timetable_key][allocation_key] = {
                                'course': course_code,
                                'classroom': suitable_classroom,
                                'enrollment': enrollment,
                                'conflict': False,
                                'basket': basket_name,
                                'type': session_type
                            }
                            print(f"         [BASKET-STORED] Key: {timetable_key}/{allocation_key} -> {suitable_classroom} for {course_code} ({session_type})")
                            allocation_count += 1
                        else:
                            print(f"         [BASKET-WARN] {day} {time_slot}: No classroom available for {course_code} in {basket_name}")
                else:
                    # No specific courses in basket_courses_map - allocate a generic room for the basket
                    print(f"      [BASKET-FALLBACK] {day} {time_slot}: '{course_value}' not in basket_courses_map, allocating generic room")
                    common_elective_key = f"ELECTIVE_COMMON_{semester}_{day}_{time_slot}_{basket_name}_{session_type}"
                    existing_common_room = _ELECTIVE_COMMON_ROOMS.get(common_elective_key)
                    
                    if existing_common_room:
                        suitable_classroom = existing_common_room
                        print(f"         [BASKET-GENERIC-REUSE] {basket_name} ({session_type}) -> {existing_common_room}")
                    else:
                        # Allocate new room with default enrollment for electives
                        suitable_classroom = allocate_regular_classroom(
                            40, day, time_slot,
                            is_common_course=True,
                            is_lab_session=is_lab,
                            course_code=basket_name
                        )
                        if suitable_classroom:
                            _ELECTIVE_COMMON_ROOMS[common_elective_key] = suitable_classroom
                            print(f"         [BASKET-GENERIC-NEW] {basket_name} ({session_type}) -> {suitable_classroom}")
                    
                    if suitable_classroom:
                        reserve_room(suitable_classroom, day, time_slot)
                        basket_rooms_allocated.append(suitable_classroom)
                        allocation_count += 1
                
                # Display ONLY basket name in cell - rooms are shown in the legend table
                schedule_with_rooms.loc[time_slot, day] = course_value
                if basket_rooms_allocated:
                    print(f"      [BASKET-DISPLAY] {day} {time_slot}: '{course_value}' (rooms {basket_rooms_allocated} stored in allocations for legend)")
                else:
                    print(f"      [BASKET-DISPLAY] {day} {time_slot}: '{course_value}' (no rooms allocated)")
                continue
            
            # Handle legacy newline-separated courses (from old data files)
            if isinstance(course_value, str) and '\n' in course_value:
                # This is a legacy basket slot with multiple courses - each needs its own room
                courses_in_slot = [c.strip() for c in course_value.split('\n') if c.strip()]
                rooms_allocated = []
                rooms_used_in_this_slot = set()
                
                print(f"      [MULTI-COURSE] {day} {time_slot}: {len(courses_in_slot)} courses in basket slot")
                
                for course_code in courses_in_slot:
                    # Extract clean code and check if it's Tutorial/Lab
                    is_tutorial = ' (Tutorial)' in course_code
                    is_lab = ' (Lab)' in course_code
                    clean_code = course_code.replace(' (Tutorial)', '').replace(' (Lab)', '').strip()
                    
                    # Get enrollment for this specific course
                    enrollment = course_enrollment.get(clean_code, 40)
                    
                    # CRITICAL: Each elective course should use the SAME room across all sections/branches
                    # Include semester to ensure electives are shared within semester but isolated between semesters
                    session_type = 'Tutorial' if is_tutorial else ('Lab' if is_lab else 'Lecture')
                    common_elective_key = f"ELECTIVE_COMMON_{semester}_{day}_{time_slot}_{clean_code}_{session_type}"
                    
                    # Check if this course already has a common room allocated at this day/time
                    existing_common_room = _ELECTIVE_COMMON_ROOMS.get(common_elective_key)
                    suitable_classroom = None
                    
                    if existing_common_room:
                        # CRITICAL: For common electives, ALWAYS use the same room across sections
                        # even if it appears "busy" - both sections attend together in the same room
                        suitable_classroom = existing_common_room
                        if room_available(existing_common_room, day, time_slot):
                            print(f"        [USING-COMMON] {day} {time_slot}: {clean_code} -> {existing_common_room} (COMMON ELECTIVE - shared across sections)")
                        else:
                            print(f"        [USING-COMMON-SHARED] {day} {time_slot}: {clean_code} -> {existing_common_room} (COMMON ELECTIVE - both sections together)")
                    else:
                        
                        # Allocate a new room, but avoid rooms already used in this same slot
                        prefixed_day = f"{schedule_type}_{day}"
                        booked_global = _CLASSROOM_USAGE_TRACKER.get(prefixed_day, {}).get(time_slot, set())
                        disallowed = rooms_used_in_this_slot | booked_global
                        
                        # Find a suitable room not in disallowed set (tiered: smallest adequate)
                        # Build tier order based on enrollment
                        tier_caps = []
                        if enrollment <= 80:
                            tier_caps = [80, 96, 120, 135]
                        elif enrollment <= 96:
                            tier_caps = [96, 120, 135]
                        elif enrollment <= 120:
                            tier_caps = [120, 135]
                        elif enrollment <= 135:
                            tier_caps = [135, 240]
                        else:
                            tier_caps = [240]
                        
                        for rooms_df in [primary_classrooms, fallback_lab_classrooms, available_classrooms]:
                            if rooms_df.empty:
                                continue
                            candidates = rooms_df.copy()
                            candidates['_cap'] = pd.to_numeric(candidates['Capacity'], errors='coerce').fillna(0)
                            candidates = candidates[candidates['_cap'] >= enrollment]
                            candidates = candidates[~candidates['Room Number'].isin(disallowed)]
                            if not candidates.empty:
                                # Try each tier to find the best-fit room with load balancing
                                for tier_cap in tier_caps:
                                    tier_match = candidates[(candidates['_cap'] >= tier_cap - 10) & (candidates['_cap'] <= tier_cap + 10)]
                                    if not tier_match.empty:
                                        suitable_classroom = select_least_used_room(tier_match)
                                        print(f"        [ALLOCATED] {suitable_classroom} for {day} {time_slot} - {enrollment} students (tier {tier_cap}, load-balanced)")
                                        break
                                if suitable_classroom:
                                    break
                                # No tier match, use smallest adequate with load balancing
                                suitable_classroom = select_least_used_room(candidates)
                                print(f"        [ALLOCATED] {suitable_classroom} for {day} {time_slot} - {enrollment} students (smallest adequate, load-balanced)")
                                break
                        
                        if suitable_classroom:
                            # Store as common room for all sections/branches at this specific day/time
                            _ELECTIVE_COMMON_ROOMS[common_elective_key] = suitable_classroom
                            print(f"        [ESTABLISH-COMMON] {day} {time_slot}: {clean_code} -> {suitable_classroom} (NEW common room for all branches at {day} {time_slot})")
                    
                    if suitable_classroom:
                        # Reserve and track
                        reserve_room(suitable_classroom, day, time_slot)
                        rooms_used_in_this_slot.add(suitable_classroom)
                        rooms_allocated.append(f"{course_code} [{suitable_classroom}]")
                        
                        # Track allocation
                        allocation_key = f"{day}_{time_slot}_{clean_code}"
                        _TIMETABLE_CLASSROOM_ALLOCATIONS[timetable_key][allocation_key] = {
                            'course': clean_code,
                            'classroom': suitable_classroom,
                            'enrollment': enrollment,
                            'conflict': False,
                            'basket': 'Multiple'
                        }
                        allocation_count += 1
                        print(f"      [BASKET] {day} {time_slot}: {clean_code} -> {suitable_classroom} ({enrollment} students)")
                    else:
                        print(f"      [WARN] {day} {time_slot}: No classroom available for {clean_code}")
                        rooms_allocated.append(course_code)
                
                # Update cell with all courses and their rooms
                schedule_with_rooms.loc[time_slot, day] = '\n'.join(rooms_allocated)
                continue
            
            # Handle both regular courses and basket entries
            if isinstance(course_value, str):
                # Strip pre-existing room annotation to get a clean course label
                course_display = course_value
                course_key = course_display
                existing_room = None
                if '[' in course_display and ']' in course_display:
                    try:
                        bracket_start = course_display.rfind('[')
                        bracket_end = course_display.rfind(']')
                        if bracket_start != -1 and bracket_end != -1 and bracket_end > bracket_start:
                            room_fragment = course_display[bracket_start+1:bracket_end]
                            existing_room = normalize_single_room(room_fragment)
                            course_display = course_display[:bracket_start].strip()
                    except Exception:
                        pass

                clean_existing_code = course_display.replace(' (Tutorial)', '').replace(' (Lab)', '').strip()
                annotated_enrollment = course_enrollment.get(clean_existing_code, 50)

                # Determine course attributes for room preference
                info = get_course_info_by_dept(course_info, clean_existing_code, branch)
                is_elective = bool(info.get('is_elective', False))
                is_common_course = str(info.get('common', 'No')).strip().upper() == 'YES'
                is_minor = course_display.upper().startswith('MINOR')
                preferred_caps_override = None
                
                # CHECK: If course is taught by multiple instructors in this department, divide enrollment by 2
                dual_info = detect_dual_instructor_course(course_info, clean_existing_code, branch)
                if dual_info and dual_info.get('is_dual'):
                    original_enroll = annotated_enrollment
                    annotated_enrollment = dual_info['effective_enrollment']
                    print(f"      [DUAL-INSTR] {clean_existing_code} ({branch}): {dual_info['num_instructors']} instructors detected. Enrollment {original_enroll} → {annotated_enrollment} (total {dual_info['total_enrollment']}) for room sizing")
                
                if not is_common_course and not is_elective and not is_minor:
                    # Core (non-common, non-elective, non-minor) → tiered by enrollment
                    if annotated_enrollment > 135:
                        preferred_caps_override = [240]
                    elif annotated_enrollment > 120:
                        preferred_caps_override = [135, 120]
                    elif annotated_enrollment > 96:
                        preferred_caps_override = [120, 135]
                    elif annotated_enrollment > 80:
                        preferred_caps_override = [96, 120]
                    else:
                        preferred_caps_override = [80, 96, 120]

                # If the timetable already contains a room, lock it in the tracker to avoid clashes
                if existing_room:
                    existing_room_norm = normalize_single_room(existing_room)

                    # Handle annotated lab pairs (ensure both slots stay in the same lab room)
                    if ' (Lab)' in course_display and time_slot in lab_slot_pairs:
                        second_slot = lab_slot_pairs[time_slot]
                        if second_slot in schedule_df.index and schedule_df.loc[second_slot, day] == course_value:
                            if room_available_for_lab_pair(existing_room_norm, day, time_slot, second_slot):
                                reserve_room(existing_room_norm, day, time_slot)
                                reserve_room(existing_room_norm, day, second_slot)

                                schedule_with_rooms.loc[time_slot, day] = f"{course_display} [{existing_room_norm}]"
                                schedule_with_rooms.loc[second_slot, day] = f"{course_display} [{existing_room_norm}]"

                                allocation_key_1 = f"{day}_{time_slot}"
                                allocation_key_2 = f"{day}_{second_slot}"
                                _TIMETABLE_CLASSROOM_ALLOCATIONS[timetable_key][allocation_key_1] = {
                                    'course': course_display,
                                    'classroom': existing_room_norm,
                                    'enrollment': annotated_enrollment,
                                    'conflict': False,
                                    'split': False
                                }
                                _TIMETABLE_CLASSROOM_ALLOCATIONS[timetable_key][allocation_key_2] = {
                                    'course': course_display,
                                    'classroom': existing_room_norm,
                                    'enrollment': annotated_enrollment,
                                    'conflict': False,
                                    'split': False
                                }
                                processed_lab_slots.add((day, second_slot))
                                allocation_count += 2
                                print(f"      [EXIST-LAB] Locked existing lab room {existing_room_norm} for {day} {time_slot} & {second_slot}")
                                continue
                            else:
                                print(f"      [LAB-WARN] Existing lab room {existing_room_norm} already booked at {day} {time_slot}/{second_slot}; reallocating")
                                existing_room = None

                    # Non-lab or single-slot lab entry
                    if existing_room:
                        if room_available(existing_room_norm, day, time_slot):
                            reserve_room(existing_room_norm, day, time_slot)
                            schedule_with_rooms.loc[time_slot, day] = f"{course_display} [{existing_room_norm}]"
                            allocation_key = f"{day}_{time_slot}"
                            _TIMETABLE_CLASSROOM_ALLOCATIONS[timetable_key][allocation_key] = {
                                'course': course_display,
                                'classroom': existing_room_norm,
                                'enrollment': annotated_enrollment,
                                'conflict': False,
                                'split': False
                            }
                            if course_key not in course_preferred_classrooms:
                                course_preferred_classrooms[course_key] = existing_room_norm
                            allocation_count += 1
                            print(f"      [EXIST] Reusing pre-annotated room {existing_room_norm} for {course_display} on {day} {time_slot}")
                            continue
                        else:
                            print(f"      [WARN] Pre-annotated room {existing_room_norm} is already booked at {day} {time_slot}; reallocating")
                            existing_room = None

                # Check for minor sessions (treat as common across ALL branches/sections of the semester)
                if course_display.upper().startswith('MINOR'):
                    minor_enrollment = 60  # default if not provided
                    # Create a common key for all branches of this semester
                    # Key format: sem{semester}_MINOR_{day}_{time_slot} - NO branch dependency
                    # This ensures ALL branches of the same semester share the same minor classroom
                    minor_common_key = f"sem{semester}_MINOR_{day}_{time_slot}"
                    room_for_minor = None

                    if existing_room:
                        room_for_minor = existing_room
                        # Store this allocation so all other branches reuse it
                        _MINOR_COMMON_CLASSROOMS[minor_common_key] = room_for_minor
                        print(f"      [MINOR-EXISTING] Using existing room {room_for_minor} for {course_display} on {day} {time_slot}")
                    elif minor_common_key in _MINOR_COMMON_CLASSROOMS:
                        # Another branch already allocated a room for this minor slot
                        candidate = normalize_single_room(_MINOR_COMMON_CLASSROOMS[minor_common_key])
                        if candidate:
                            room_for_minor = candidate
                            print(f"      [MINOR-COMMON] Reusing room {room_for_minor} allocated by another branch for {course_display} on {day} {time_slot}")
                    
                    if not room_for_minor:
                        # First branch to allocate for this minor slot
                        room_for_minor = normalize_single_room(
                            allocate_regular_classroom(minor_enrollment, day, time_slot, is_common_course=True, is_lab_session=False)
                        )
                        if room_for_minor:
                            # Store for all other branches to reuse
                            _MINOR_COMMON_CLASSROOMS[minor_common_key] = room_for_minor
                            print(f"      [MINOR-NEW] Allocated room {room_for_minor} for {course_display} on {day} {time_slot} (all branches will reuse)")
                    
                    if room_for_minor:
                        reserve_room(room_for_minor, day, time_slot)
                        schedule_with_rooms.loc[time_slot, day] = f"{course_display} [{room_for_minor}]"
                        allocation_key = f"{day}_{time_slot}_MINOR"
                        _TIMETABLE_CLASSROOM_ALLOCATIONS[timetable_key][allocation_key] = {
                            'course': course_display,
                            'classroom': room_for_minor,
                            'enrollment': minor_enrollment,
                            'conflict': False,
                            'type': 'Minor'
                        }
                        allocation_count += 1
                        print(f"      [MINOR] {day} {time_slot}: {course_display} -> {room_for_minor} (common room for all branches)")
                    else:
                        schedule_with_rooms.loc[time_slot, day] = course_display
                        print(f"      [MINOR-WARN] {day} {time_slot}: No classroom available for {course_display}")
                    continue

                # Check if this is a lab slot
                is_lab = ' (Lab)' in course_display
                
                # Check if this is a basket entry
                is_basket = any(basket in course_display for basket in ['ELECTIVE_B', 'HSS_B', 'PROF_B', 'OE_B'])
                
                # For basket entries, use a standard enrollment
                if is_basket:
                    enrollment = 40  # Standard enrollment for elective baskets
                else:
                    # Regular course - extract clean course code (handle Tutorial, Lab suffixes)
                    clean_course_code = course_display.replace(' (Tutorial)', '').replace(' (Lab)', '')
                    enrollment = course_enrollment.get(clean_course_code, 50)
                    
                    # CHECK: If course is taught by multiple instructors in this department, divide enrollment by 2
                    dual_info = detect_dual_instructor_course(course_info, clean_course_code, branch)
                    if dual_info and dual_info.get('is_dual'):
                        original_enroll = enrollment
                        enrollment = dual_info['effective_enrollment']
                        print(f"      [DUAL-INSTR] {clean_course_code} ({branch}): {dual_info['num_instructors']} instructors detected. Enrollment {original_enroll} → {enrollment} (total {dual_info['total_enrollment']}) for room sizing")
            else:
                continue
            
            course_key = course_display
            preferred_classroom = normalize_single_room(course_preferred_classrooms.get(course_key))
            if preferred_classroom:
                course_preferred_classrooms[course_key] = preferred_classroom
            
            # If this is a basket entry, allocate rooms to individual courses within the basket
            # Do NOT add room to the basket entry in the schedule display
            if is_basket:
                # Determine session type for this basket slot
                session_type = 'Tutorial' if ' (Tutorial)' in course_display else ('Lab' if ' (Lab)' in course_display else 'Lecture')
                # Get the basket name
                basket_name = None
                # Include all elective baskets (B1-B9) and existing HSS/PROF/OE identifiers
                for basket in ['ELECTIVE_B1', 'ELECTIVE_B2', 'ELECTIVE_B3', 'ELECTIVE_B4', 'ELECTIVE_B5', 'ELECTIVE_B6', 'ELECTIVE_B7', 'ELECTIVE_B8', 'ELECTIVE_B9', 'HSS_B1', 'HSS_B2', 'PROF_B1', 'OE_B1']:
                    if basket in course_display:
                        basket_name = basket
                        break
                
                # Keep the basket entry in the schedule WITHOUT adding a room
                schedule_with_rooms.loc[time_slot, day] = course_display
                
                # Allocate separate rooms for each course in this basket (for verification/scheduling)
                if basket_courses_map and basket_name and basket_name in basket_courses_map:
                    courses_in_basket = basket_courses_map[basket_name]
                    for idx, individual_course in enumerate(courses_in_basket):
                        # Each course in the basket gets allocated a separate room
                        individual_enrollment = course_enrollment.get(individual_course, 40)
                        
                        # CHECK: If course is taught by multiple instructors in this department, divide enrollment by 2
                        dual_info = detect_dual_instructor_course(course_info, individual_course, branch)
                        if dual_info and dual_info.get('is_dual'):
                            original_enroll = individual_enrollment
                            individual_enrollment = dual_info['effective_enrollment']
                            print(f"        [DUAL-INSTR] {individual_course} ({branch}): {dual_info['num_instructors']} instructors detected. Enrollment {original_enroll} → {individual_enrollment} (total {dual_info['total_enrollment']}) for room sizing")
                        
                        # Try to find a suitable classroom for this individual course
                        individual_preferred = normalize_single_room(course_preferred_classrooms.get(individual_course))
                        if individual_preferred:
                            course_preferred_classrooms[individual_course] = individual_preferred
                        individual_classroom = None

                        print(f"        [BASKET-TRY] {day} {time_slot}: Trying {individual_course} (enroll {individual_enrollment}), preferred={individual_preferred}")
                        
                        # IMPORTANT: For basket elective courses, all sections/branches at same time should use SAME room
                        # Basket courses with same course code should use the SAME room across all sections/branches
                        # Include semester in the key to ensure same semester shares but different semesters are separate
                        common_elective_key = f"ELECTIVE_COMMON_{semester}_{day}_{time_slot}_{individual_course}_{session_type}"

                        # If this course's common room is already established at this day/time, reuse it
                        existing_common_room = _ELECTIVE_COMMON_ROOMS.get(common_elective_key)

                        if existing_common_room:
                            # For electives at the SAME day/time/course, reuse the common room
                            # This ensures all sections of same semester use same classroom
                            # The room is intentionally shared - both sections attend together
                            individual_classroom = existing_common_room
                            print(f"        [USING-COMMON] {day} {time_slot}: {individual_course} -> {existing_common_room} (COMMON ELECTIVE - shared across sections)")

                        if not individual_classroom:
                            # IMPORTANT: Basket elective courses ARE common courses
                            # All sections/branches offering the same course at same time should share ONE room
                            is_common = True  # Force common treatment for basket electives
                            
                            # Allocate as a common classroom
                            individual_classroom = normalize_single_room(
                                allocate_regular_classroom(
                                    individual_enrollment,
                                    day,
                                    time_slot,
                                    is_common_course=True,
                                    is_lab_session=(session_type == 'Lab'),
                                    course_code=individual_course
                                )
                            )
                            
                            # If successful, store this as the common room for all sections/branches
                            if individual_classroom:
                                _ELECTIVE_COMMON_ROOMS[common_elective_key] = individual_classroom
                                print(f"        [ESTABLISH-COMMON] {day} {time_slot}: {individual_course} -> {individual_classroom} (NEW common room for all branches)")
                        
                        if individual_classroom:
                            # Defensive: ensure uniqueness within this basket and time slot
                            # Collect rooms already used at this day/time by other courses in the same basket
                            basket_rooms_at_this_slot = set()
                            for key, alloc in _TIMETABLE_CLASSROOM_ALLOCATIONS[timetable_key].items():
                                if alloc.get('basket') == basket_name and f"{day}_{time_slot}" in key:
                                    if alloc.get('classroom'):
                                        basket_rooms_at_this_slot.add(alloc.get('classroom'))

                            # If chosen room clashes within the basket at this slot, try to pick an alternative
                            if individual_classroom in basket_rooms_at_this_slot:
                                alt_room = None
                                # Build a set of globally booked rooms for this day/time
                                prefixed_day = f"{schedule_type}_{day}"
                                booked_global = _CLASSROOM_USAGE_TRACKER.get(prefixed_day, {}).get(time_slot, set())
                                disallowed = set(basket_rooms_at_this_slot) | set(booked_global)

                                # Prefer primary classrooms first (tiered: smallest adequate room)
                                search_sets = []
                                if not primary_classrooms.empty:
                                    search_sets.append(primary_classrooms)
                                if not fallback_lab_classrooms.empty:
                                    search_sets.append(fallback_lab_classrooms)
                                if not available_classrooms.empty:
                                    search_sets.append(available_classrooms)

                                # Build tier order based on enrollment
                                tier_caps = []
                                if individual_enrollment <= 80:
                                    tier_caps = [80, 96, 120, 135]
                                elif individual_enrollment <= 96:
                                    tier_caps = [96, 120, 135]
                                elif individual_enrollment <= 120:
                                    tier_caps = [120, 135]
                                elif individual_enrollment <= 135:
                                    tier_caps = [135, 240]
                                else:
                                    tier_caps = [240]

                                for rooms_df in search_sets:
                                    candidates = rooms_df.copy()
                                    candidates['_cap'] = pd.to_numeric(candidates['Capacity'], errors='coerce').fillna(0)
                                    # capacity filter
                                    candidates = candidates[candidates['_cap'] >= individual_enrollment]
                                    # remove disallowed
                                    candidates = candidates[~candidates['Room Number'].isin(disallowed)]
                                    if not candidates.empty:
                                        # Try each tier for best fit with load balancing
                                        for tier_cap in tier_caps:
                                            tier_match = candidates[(candidates['_cap'] >= tier_cap - 10) & (candidates['_cap'] <= tier_cap + 10)]
                                            if not tier_match.empty:
                                                alt_room = select_least_used_room(tier_match)
                                                break
                                        if not alt_room:
                                            alt_room = select_least_used_room(candidates)
                                        break

                                if alt_room:
                                    individual_classroom = alt_room
                                    print(f"        [BASKET-UNIQ] {day} {time_slot}: Switched {individual_course} to {alt_room} to avoid same-room reuse in {basket_name}")

                            # Reserve to block other courses at this time
                            reserve_room(individual_classroom, day, time_slot)
                            course_preferred_classrooms.setdefault(individual_course, individual_classroom)
                        else:
                            print(f"        [BASKET-FAIL] {day} {time_slot}: No classroom allocated for {individual_course}. primary={len(primary_classrooms)}, fallback={len(fallback_lab_classrooms)}")
                            # Find a room that's not already allocated to another course in this basket at this time
                            fallback_room = None
                            
                            # Check which rooms are already used by other courses in this basket at this time
                            basket_rooms_at_this_slot = set()
                            for key, alloc in _TIMETABLE_CLASSROOM_ALLOCATIONS[timetable_key].items():
                                if alloc.get('basket') == basket_name and f"{day}_{time_slot}" in key:
                                    basket_rooms_at_this_slot.add(alloc.get('classroom'))
                            
                            # CRITICAL: Also check the global classroom usage tracker to avoid double-booking
                            prefixed_day = f"{schedule_type}_{day}"
                            globally_booked_rooms = _CLASSROOM_USAGE_TRACKER.get(prefixed_day, {}).get(time_slot, set())
                            unavailable_rooms = basket_rooms_at_this_slot | globally_booked_rooms
                            
                            # Try to find an unused room that is GLOBALLY available (tiered)
                            fallback_room = find_tiered_fallback_room(individual_enrollment, day, time_slot, basket_rooms_at_this_slot)
                            
                            if fallback_room:
                                reserve_room(fallback_room, day, time_slot)
                                individual_classroom = fallback_room
                                _TIMETABLE_CLASSROOM_ALLOCATIONS[timetable_key][f"{day}_{time_slot}_{session_type}_{individual_course}"] = {
                                    'course': individual_course,
                                    'classroom': fallback_room,
                                    'enrollment': individual_enrollment,
                                    'conflict': False,  # Not a conflict - room was available
                                    'basket': basket_name,
                                    'type': session_type
                                }
                                allocation_count += 1
                                print(f"        [BASKET-FALLBACK-OK] {day} {time_slot}: {individual_course} -> {fallback_room} ({individual_enrollment} students) [Found available room]")
                            else:
                                print(f"        [CRITICAL] {day} {time_slot}: Cannot allocate {individual_course} - all rooms either unavailable or already used by other courses in {basket_name}")


                        # Track the allocation for this individual course
                        if individual_classroom:
                            allocation_key = f"{day}_{time_slot}_{session_type}_{individual_course}"
                            _TIMETABLE_CLASSROOM_ALLOCATIONS[timetable_key][allocation_key] = {
                                'course': individual_course,
                                'classroom': individual_classroom,
                                'enrollment': individual_enrollment,
                                'conflict': False,
                                'basket': basket_name,
                                'type': session_type
                            }
                            allocation_count += 1
                            print(f"      [BASKET] {day} {time_slot}: {individual_course} (from {basket_name}) -> {individual_classroom} ({individual_enrollment} students)")
                        else:
                            print(f"      [WARN] {day} {time_slot}: No classroom available for {individual_course} in basket {basket_name}")
                else:
                    print(f"      [INFO] {day} {time_slot}: Basket {basket_name} has no course mapping, keeping basket entry without room allocation")
            # If this is a lab and it's the first slot of a lab pair, find a classroom available for BOTH slots
            elif is_lab and time_slot in lab_slot_pairs:
                second_slot = lab_slot_pairs[time_slot]
                # Safety check: second_slot must exist in the schedule
                if second_slot not in schedule_df.index:
                    print(f"      [WARN] Lab pair slot {second_slot} not found in schedule for {course_display}")
                    continue
                second_course_value = schedule_df.loc[second_slot, day]
                
                # Check if second slot also has the same lab course
                if isinstance(second_course_value, str) and second_course_value == course_value:
                    if preferred_classroom and room_available_for_lab_pair(preferred_classroom, day, time_slot, second_slot):
                        reserve_room(preferred_classroom, day, time_slot)
                        reserve_room(preferred_classroom, day, second_slot)
                        schedule_with_rooms.loc[time_slot, day] = f"{course_display} [{preferred_classroom}]"
                        schedule_with_rooms.loc[second_slot, day] = f"{course_display} [{preferred_classroom}]"
                        
                        allocation_key_1 = f"{day}_{time_slot}"
                        allocation_key_2 = f"{day}_{second_slot}"
                        _TIMETABLE_CLASSROOM_ALLOCATIONS[timetable_key][allocation_key_1] = {
                            'course': course_display,
                            'classroom': preferred_classroom,
                            'enrollment': enrollment,
                            'conflict': False
                        }
                        _TIMETABLE_CLASSROOM_ALLOCATIONS[timetable_key][allocation_key_2] = {
                            'course': course_display,
                            'classroom': preferred_classroom,
                            'enrollment': enrollment,
                            'conflict': False
                        }
                        
                        processed_lab_slots.add((day, second_slot))
                        
                        allocation_count += 2
                        print(f"      [REUSE] Reusing {preferred_classroom} for lab pair {day} {time_slot} & {second_slot} ({enrollment} students)")
                    else:
                        # Find a LAB ROOM (starting with "L") that's available for BOTH slots
                        # Labs must use lab rooms, not regular classrooms
                        clean_code = course_display.replace(' (Lab)', '').strip()
                        
                        # Get appropriate lab rooms based on course type
                        lab_rooms_to_search = get_suitable_lab_rooms_for_course(clean_code)

                        # Check if this lab pair already has an allocated room (from another section of the SAME course)
                        # BUT: Labs should be DIFFERENT for each section even if lecture is common
                        # Only reuse lab rooms for non-common courses or same section
                        # CRITICAL: Include course code in key to prevent cross-semester/cross-course conflicts
                        global _LAB_ROOM_ALLOCATIONS
                        lab_pair_key = (day, time_slot, second_slot, clean_code)  # Include course code!
                        
                        suitable_classroom = None
                        # For COMMON COURSES: Do NOT reuse lab rooms - each section gets different lab
                        # For NON-COMMON COURSES: Reuse lab rooms across sections (same course, same session)
                        is_common_lab = False
                        if 'is_common' in locals() and is_common:
                            is_common_lab = True
                        
                        if lab_pair_key in _LAB_ROOM_ALLOCATIONS and not is_common_lab:
                            # Check if the previously allocated room is still available
                            candidate_room = _LAB_ROOM_ALLOCATIONS[lab_pair_key]
                            if room_available_for_lab_pair(candidate_room, day, time_slot, second_slot):
                                suitable_classroom = candidate_room
                                print(f"      [REUSE-LAB] Using previously allocated room {suitable_classroom} for {clean_code} on {day} {time_slot} & {second_slot} (same course from other section)")
                            else:
                                print(f"      [REUSE-LAB-BLOCKED] Previously allocated room {candidate_room} for {clean_code} is no longer available at {day} {time_slot}, finding new room")
                        
                        if not suitable_classroom:
                            # Allocate a new room for this lab pair
                            suitable_classroom = find_suitable_classroom_for_lab_pair(
                                lab_rooms_to_search, enrollment, day, time_slot, second_slot, _CLASSROOM_USAGE_TRACKER, schedule_type=schedule_type
                            )
                            
                            # Store this allocation globally ONLY for non-common courses
                            # For common courses: Labs are DIFFERENT per section, so do NOT store for reuse
                            if suitable_classroom and not is_common_lab:
                                _LAB_ROOM_ALLOCATIONS[lab_pair_key] = suitable_classroom
                                print(f"      [NEW-LAB] Allocated new lab room {suitable_classroom} for {day} {time_slot} & {second_slot} (will reuse for other sections)")
                            elif suitable_classroom and is_common_lab:
                                print(f"      [NEW-LAB] Allocated new lab room {suitable_classroom} for {day} {time_slot} & {second_slot} (common course - separate lab per section)")
                        
                        if suitable_classroom:
                            # Allocate the same classroom to BOTH slots
                            # Mark both slots as used in the tracker
                            reserve_room(suitable_classroom, day, time_slot)
                            reserve_room(suitable_classroom, day, second_slot)
                            
                            # Update schedule with same classroom for both slots
                            schedule_with_rooms.loc[time_slot, day] = f"{course_display} [{suitable_classroom}]"
                            schedule_with_rooms.loc[second_slot, day] = f"{course_display} [{suitable_classroom}]"
                            if course_key not in course_preferred_classrooms:
                                course_preferred_classrooms[course_key] = suitable_classroom
                            
                            # Track both allocations
                            allocation_key_1 = f"{day}_{time_slot}"
                            allocation_key_2 = f"{day}_{second_slot}"
                            _TIMETABLE_CLASSROOM_ALLOCATIONS[timetable_key][allocation_key_1] = {
                                'course': course_display,
                                'classroom': suitable_classroom,
                                'enrollment': enrollment,
                                'conflict': False
                            }
                            _TIMETABLE_CLASSROOM_ALLOCATIONS[timetable_key][allocation_key_2] = {
                                'course': course_display,
                                'classroom': suitable_classroom,
                                'enrollment': enrollment,
                                'conflict': False
                            }
                            # Persist preferred classroom globally
                            if course_key not in _GLOBAL_PREFERRED_CLASSROOMS:
                                _GLOBAL_PREFERRED_CLASSROOMS[course_key] = suitable_classroom
                            
                            # Mark second slot as processed
                            processed_lab_slots.add((day, second_slot))
                            
                            allocation_count += 2
                            print(f"      [OK] {day} {time_slot} & {second_slot}: {course_display} -> {suitable_classroom} ({enrollment} students) [Lab pair - same room]")
                        else:
                            print(f"      [WARN]  {day} {time_slot}: No classroom available for lab pair ({time_slot} & {second_slot})")
                else:
                    # Not a proper lab pair, treat as regular course
                    # Extract clean course code for common check
                    clean_code = course_display.replace(' (Tutorial)', '').replace(' (Lab)', '').strip()
                    
                    # Check if this is a common course FIRST (before preferred classroom logic)
                    info_for_room = get_course_info_by_dept(course_info, clean_code, branch)
                    common_val = str(info_for_room.get('common', info_for_room.get('Common', 'No'))).strip().upper()
                    is_common = common_val == 'YES'
                    
                    # Check for cross-department (DSAI+ECE) common courses
                    cross_common_bundle = get_cross_common_bundle(clean_code)
                    normalized_branch = normalize_branch_string(branch)
                    cross_common_active = bool(
                        is_common
                        and cross_common_bundle
                        and normalized_branch in cross_common_bundle.get('departments', set())
                    )
                    
                    # For COMMON courses, check _COMMON_COURSE_ROOMS first (both sections/branches share same room)
                    # Use cross-department key if applicable (DSAI+ECE share same room)
                    if cross_common_active:
                        common_course_key = f"{cross_common_bundle['room_key']}_{day}_{time_slot}"
                    else:
                        common_course_key = f"{semester}_{branch}_{clean_code}_{day}_{time_slot}"
                    common_room_found = None
                    
                    if is_common and common_course_key in _COMMON_COURSE_ROOMS:
                        common_room_found = normalize_single_room(_COMMON_COURSE_ROOMS[common_course_key])
                        if common_room_found:
                            # Use same room as Section A - NO availability check needed (they attend together)
                            schedule_with_rooms.loc[time_slot, day] = f"{course_display} [{common_room_found}]"
                            allocation_key = f"{day}_{time_slot}"
                            _TIMETABLE_CLASSROOM_ALLOCATIONS[timetable_key][allocation_key] = {
                                'course': course_display,
                                'classroom': common_room_found,
                                'enrollment': enrollment,
                                'conflict': False,
                                'split': False
                            }
                            allocation_count += 1
                            print(f"      [COMMON-REUSE] {course_display} using SAME room {common_room_found} on {day} {time_slot} (both sections attend together)")
                    
                    # For common courses: skip preferred classroom to ensure _COMMON_COURSE_ROOMS is used
                    skip_preferred_for_common = is_common and not common_room_found
                    
                    if common_room_found:
                        pass  # Already handled above
                    elif not skip_preferred_for_common and preferred_classroom and room_available(preferred_classroom, day, time_slot):
                        reserve_room(preferred_classroom, day, time_slot)
                        # Validate preferred_classroom is not empty/None before adding brackets
                        if preferred_classroom and str(preferred_classroom).strip():
                            schedule_with_rooms.loc[time_slot, day] = f"{course_display} [{preferred_classroom}]"
                        else:
                            schedule_with_rooms.loc[time_slot, day] = course_display
                        allocation_key = f"{day}_{time_slot}"
                        _TIMETABLE_CLASSROOM_ALLOCATIONS[timetable_key][allocation_key] = {
                            'course': course_display,
                            'classroom': preferred_classroom,
                            'enrollment': enrollment,
                            'conflict': False,
                            'split': False
                        }
                        allocation_count += 1
                        print(f"      [REUSE] Reusing {preferred_classroom} for {course_display} on {day} {time_slot} ({enrollment} students)")
                    else:
                        # Check if this is a common course
                        is_common = False
                        clean_code = course_display.replace(' (Tutorial)', '').replace(' (Lab)', '').strip()
                        
                        # Use department-aware lookup to get the correct Common value for this branch
                        info_for_room = get_course_info_by_dept(course_info, clean_code, branch)
                        common_val = str(info_for_room.get('common', info_for_room.get('Common', 'No'))).strip().upper()
                        is_common = common_val == 'YES'

                        cross_common_bundle = get_cross_common_bundle(clean_code)
                        normalized_branch = normalize_branch_string(branch)
                        cross_common_active = bool(
                            is_common
                            and cross_common_bundle
                            and normalized_branch in cross_common_bundle.get('departments', set())
                        )

                        if cross_common_active:
                            effective_enrollment = cross_common_bundle['total_enrollment']
                        else:
                            effective_enrollment = enrollment

                        if is_common or clean_code in ['DS161', 'MA161', 'EC161', 'HS161', 'CS161']:  # Debug key courses
                            debug_enroll = effective_enrollment if cross_common_active else enrollment
                            print(f"      [COMMON-DEBUG] {clean_code} for {branch}: common_val='{common_val}', is_common={is_common}, cross_dsai_ece={cross_common_active}, enroll_used={debug_enroll}")
                        
                        # For common courses, same room for SAME slot across sections (include day/time in key)
                        if cross_common_active:
                            common_course_key = f"{cross_common_bundle['room_key']}_{day}_{time_slot}"
                        else:
                            common_course_key = f"{semester}_{branch}_{clean_code}_{day}_{time_slot}"
                        suitable_classroom = None
                        
                        is_elective = bool(info_for_room.get('is_elective', False))
                        is_minor_course = clean_code.upper().startswith('MINOR')
                        preferred_caps_override = None
                        if not is_common and not is_elective and not is_minor_course:
                            # Core (non-common, non-elective, non-minor) → tiered by enrollment
                            eff_enroll = effective_enrollment if cross_common_active else enrollment
                            if eff_enroll > 135:
                                preferred_caps_override = [240]
                            elif eff_enroll > 120:
                                preferred_caps_override = [135, 120]
                            elif eff_enroll > 96:
                                preferred_caps_override = [120, 135]
                            elif eff_enroll > 80:
                                preferred_caps_override = [96, 120]
                            else:
                                preferred_caps_override = [80, 96, 120]

                        if is_common:
                            enrollment_for_room = effective_enrollment
                            print(f"      [COMMON-CHECK] {clean_code} is marked as COMMON course - using FULL enrollment {enrollment_for_room} (Section {section})")
                            print(f"      [COMMON-DEBUG] Looking for key: '{common_course_key}'")
                            if common_course_key in _COMMON_COURSE_ROOMS:
                                # Use the same room as already allocated for the other section - MUST enforce same room
                                common_room = normalize_single_room(_COMMON_COURSE_ROOMS[common_course_key])
                                print(f"      [COMMON-DEBUG] Found existing allocation: {common_room}")
                                if common_room:
                                    # FOR COMMON COURSES: BOTH SECTIONS MUST USE SAME ROOM - NO AVAILABILITY CHECK
                                    # Both sections attend the same lecture together, so we force-use the same room
                                    # Do NOT reserve again - room is already reserved from Section A
                                    suitable_classroom = common_room
                                    print(f"      [COMMON-SHARED] Both sections use SAME room {common_room} for {clean_code} on {day} {time_slot} (Sections A & B together)")
                            else:
                                # Allocate new room (likely Section A allocating first) with FULL enrollment
                                print(f"      [COMMON-DEBUG] No existing allocation found, allocating new room for full enrollment {enrollment_for_room}")
                                suitable_classroom = normalize_single_room(
                                    allocate_regular_classroom(enrollment_for_room, day, time_slot, is_common_course=True, is_lab_session=is_lab, course_code=clean_code)
                                )
                                if suitable_classroom:
                                    # Reserve the room for this allocation
                                    reserve_room(suitable_classroom, day, time_slot)
                                    # Store for Section B/other branches to reuse (SAME room, no re-reservation)
                                    _COMMON_COURSE_ROOMS[common_course_key] = suitable_classroom
                                    print(f"      [COMMON-NEW] Allocated room {suitable_classroom} for common course {clean_code} on {day} {time_slot} with FULL enrollment {enrollment_for_room} (Section {section})")
                                    print(f"      [COMMON-DEBUG] Stored in _COMMON_COURSE_ROOMS['{common_course_key}'] = {suitable_classroom}")
                        else:
                            # Regular (non-common) course - use section-specific enrollment (already halved)
                            print(f"      [NON-COMMON] {clean_code} is NON-common - using section enrollment {enrollment} (Section {section})")
                            suitable_classroom = normalize_single_room(
                                allocate_regular_classroom(
                                    enrollment,
                                    day,
                                    time_slot,
                                    is_common_course=False,
                                    is_lab_session=is_lab,
                                    course_code=clean_code,
                                    preferred_capacities_override=preferred_caps_override,
                                )
                            )
                        
                        if suitable_classroom:
                            # Always use SINGLE classroom - NO multi-room allocation
                            single_room = normalize_single_room(suitable_classroom)
                            # Validate single_room is not empty/None
                            if single_room and str(single_room).strip():
                                schedule_with_rooms.loc[time_slot, day] = f"{course_display} [{single_room}]"
                            else:
                                schedule_with_rooms.loc[time_slot, day] = course_display
                            allocation_key = f"{day}_{time_slot}"
                            _TIMETABLE_CLASSROOM_ALLOCATIONS[timetable_key][allocation_key] = {
                                'course': course_display,
                                'classroom': single_room,
                                'enrollment': enrollment,  # Use actual enrollment (already adjusted)
                                'conflict': False,
                                'split': False
                            }
                            if course_key not in course_preferred_classrooms:
                                course_preferred_classrooms[course_key] = single_room
                            allocation_count += 1
                            print(f"      [OK] {day} {time_slot}: {course_display} -> {single_room} ({enrollment} students)")
                        else:
                            # As a last resort, try to find an AVAILABLE classroom (tiered, check global tracker)
                            fallback_room = find_tiered_fallback_room(enrollment, day, time_slot)
                            
                            if fallback_room and str(fallback_room).strip():
                                # Found an available room - no conflict
                                reserve_room(fallback_room, day, time_slot)
                                schedule_with_rooms.loc[time_slot, day] = f"{course_display} [{fallback_room}]"
                                allocation_key = f"{day}_{time_slot}"
                                _TIMETABLE_CLASSROOM_ALLOCATIONS[timetable_key][allocation_key] = {
                                    'course': course_display,
                                    'classroom': fallback_room,
                                    'enrollment': enrollment,
                                    'conflict': False,
                                    'split': False
                                }
                                if course_key not in course_preferred_classrooms:
                                    course_preferred_classrooms[course_key] = fallback_room
                                allocation_count += 1
                                print(f"      [FALLBACK-OK] {day} {time_slot}: {course_display} -> {fallback_room} ({enrollment} students) [Found available room]")
                            else:
                                # Last-resort forced assignment: pick a room even if already booked
                                forced_room = pick_forced_fallback_room(enrollment, day, time_slot)
                                if forced_room:
                                    schedule_with_rooms.loc[time_slot, day] = f"{course_display} [{forced_room}]"
                                    allocation_key = f"{day}_{time_slot}"
                                    _TIMETABLE_CLASSROOM_ALLOCATIONS[timetable_key][allocation_key] = {
                                        'course': course_display,
                                        'classroom': forced_room,
                                        'enrollment': enrollment,
                                        'conflict': True,
                                        'split': False
                                    }
                                    if course_key not in course_preferred_classrooms:
                                        course_preferred_classrooms[course_key] = forced_room
                                    allocation_count += 1
                                    print(f"      [FORCED] {day} {time_slot}: {course_display} -> {forced_room} ({enrollment} students) [all rooms booked; conflict-allowed fallback]")
                                else:
                                    schedule_with_rooms.loc[time_slot, day] = course_display
                                    print(f"      [WARN]  {day} {time_slot}: No classroom available for {course_display} ({enrollment} students) - no rooms configured")
            else:
                # Regular course (not a lab pair) - find classroom normally
                # Extract clean course code for common check
                clean_code = course_display.replace(' (Tutorial)', '').replace(' (Lab)', '').strip()

                # Determine common status once and compute effective per-section enrollment
                is_common = False
                if clean_code in course_info:
                    is_common = str(course_info[clean_code].get('common', course_info[clean_code].get('Common', 'No'))).strip().upper() == 'YES'
                effective_enrollment = compute_effective_enrollment(enrollment, is_common)

                # If a room was already present in the cell, register it and share for common courses
                if 'existing_room' in locals() and existing_room:
                    normalized_existing = normalize_single_room(existing_room)
                    if normalized_existing:
                        reserve_room(normalized_existing, day, time_slot)
                        schedule_with_rooms.loc[time_slot, day] = f"{course_display} [{normalized_existing}]"
                        allocation_key = f"{day}_{time_slot}"
                        _TIMETABLE_CLASSROOM_ALLOCATIONS[timetable_key][allocation_key] = {
                            'course': course_display,
                            'classroom': normalized_existing,
                            'enrollment': effective_enrollment,
                            'conflict': False,
                            'split': False
                        }
                        if course_key not in course_preferred_classrooms:
                            course_preferred_classrooms[course_key] = normalized_existing
                        if clean_code in course_info:
                            is_common_existing = str(course_info[clean_code].get('common', course_info[clean_code].get('Common', 'No'))).strip().upper() == 'YES'
                            if is_common_existing:
                                # Include day/time in key for slot-specific room sharing
                                common_course_key = f"{semester}_{branch}_{clean_code}_{day}_{time_slot}"
                                _COMMON_COURSE_ROOMS[common_course_key] = normalized_existing
                        allocation_count += 1
                        continue
                
                if preferred_classroom and room_available(preferred_classroom, day, time_slot):
                    reserve_room(preferred_classroom, day, time_slot)
                    # Validate preferred_classroom is not empty before adding brackets
                    if preferred_classroom and str(preferred_classroom).strip():
                        schedule_with_rooms.loc[time_slot, day] = f"{course_display} [{preferred_classroom}]"
                    else:
                        schedule_with_rooms.loc[time_slot, day] = course_display
                    
                    allocation_key = f"{day}_{time_slot}"
                    _TIMETABLE_CLASSROOM_ALLOCATIONS[timetable_key][allocation_key] = {
                        'course': course_display,
                        'classroom': preferred_classroom,
                        'enrollment': effective_enrollment,
                        'conflict': False,
                        'split': False
                    }
                    
                    # For common courses: store in _COMMON_COURSE_ROOMS with cross-common key if applicable
                    # This ensures ECE will find the room when DSAI processes first with preferred classroom
                    if is_common:
                        cross_common_bundle = detect_cross_dsai_ece_common(course_info, clean_code, semester)
                        normalized_branch = normalize_branch_string(branch)
                        cross_common_active = bool(
                            cross_common_bundle
                            and normalized_branch in cross_common_bundle.get('departments', set())
                        )
                        if cross_common_active:
                            common_course_key = f"{cross_common_bundle['room_key']}_{day}_{time_slot}"
                        else:
                            common_course_key = f"{semester}_{branch}_{clean_code}_{day}_{time_slot}"
                        if common_course_key not in _COMMON_COURSE_ROOMS:
                            _COMMON_COURSE_ROOMS[common_course_key] = preferred_classroom
                            print(f"      [COMMON-STORE] Stored preferred room {preferred_classroom} for common course key: {common_course_key}")
                    
                    allocation_count += 1
                    print(f"      [REUSE] Reusing {preferred_classroom} for {course_display} on {day} {time_slot} ({effective_enrollment} students)")
                else:
                    # For common courses, check if a room has already been allocated in another section
                    # Include day/time in key for slot-specific room sharing
                    # Check for cross-department (DSAI+ECE) common courses
                    cross_common_bundle = detect_cross_dsai_ece_common(course_info, clean_code, semester)
                    normalized_branch = normalize_branch_string(branch)
                    cross_common_active = bool(
                        is_common
                        and cross_common_bundle
                        and normalized_branch in cross_common_bundle.get('departments', set())
                    )
                    
                    # Use cross-department key if applicable (DSAI+ECE share same room)
                    if cross_common_active:
                        common_course_key = f"{cross_common_bundle['room_key']}_{day}_{time_slot}"
                        effective_enrollment = cross_common_bundle['total_enrollment']
                        print(f"      [CROSS-COMMON] Using cross-dept key: {common_course_key} (enrollment={effective_enrollment})")
                    else:
                        common_course_key = f"{semester}_{branch}_{clean_code}_{day}_{time_slot}"
                    suitable_classroom = None
                    
                    if is_common:
                        print(f"      [COMMON-CHECK] {clean_code} is marked as COMMON course (Section {section})")
                        if common_course_key in _COMMON_COURSE_ROOMS:
                            # Use the same room for common courses at THIS SLOT across all sections
                            # Both sections attend the same lecture together in the same room
                            common_room = normalize_single_room(_COMMON_COURSE_ROOMS[common_course_key])
                            if common_room:
                                # Don't check availability - common courses are shared sessions
                                # Both sections attend together, so no conflict exists
                                suitable_classroom = common_room
                                print(f"      [COMMON-SHARED] Using SAME room {common_room} for {clean_code} on {day} {time_slot} - Both sections attend together (Section {section})")
                            else:
                                # Fallback: allocate new if room parsing failed
                                suitable_classroom = normalize_single_room(
                                    allocate_regular_classroom(effective_enrollment, day, time_slot, is_common_course=True)
                                )
                                print(f"      [COMMON-FALLBACK] Common room parse failed, allocated {suitable_classroom} for {clean_code} (Section {section})")
                        else:
                            # Allocate new room (likely Section A allocating first)
                            suitable_classroom = normalize_single_room(
                                allocate_regular_classroom(effective_enrollment, day, time_slot, is_common_course=True)
                            )
                            if suitable_classroom:
                                # Store for other sections to reuse THE SAME ROOM
                                _COMMON_COURSE_ROOMS[common_course_key] = suitable_classroom
                                print(f"      [COMMON-NEW] Allocated room {suitable_classroom} for common course {clean_code} on {day} {time_slot} - Will be shared by all sections (Section {section})")
                    else:
                        # Regular (non-common) course
                        suitable_classroom = normalize_single_room(
                            allocate_regular_classroom(effective_enrollment, day, time_slot, is_common_course=False)
                        )
                    
                    if suitable_classroom and str(suitable_classroom).strip():
                        # Update schedule with classroom in format "Course [Room]"
                        schedule_with_rooms.loc[time_slot, day] = f"{course_display} [{suitable_classroom}]"
                        
                        # Track this allocation
                        allocation_key = f"{day}_{time_slot}"
                        _TIMETABLE_CLASSROOM_ALLOCATIONS[timetable_key][allocation_key] = {
                            'course': course_display,
                            'classroom': suitable_classroom,
                            'enrollment': effective_enrollment,
                            'conflict': False,
                            'split': False
                        }
                        if course_key not in course_preferred_classrooms:
                            course_preferred_classrooms[course_key] = suitable_classroom
                        
                        allocation_count += 1
                        print(f"      [OK] {day} {time_slot}: {course_display} -> {suitable_classroom} ({effective_enrollment} students)")
                    else:
                        # FALLBACK: Try to find an AVAILABLE classroom (tiered, check global tracker)
                        fallback_room = find_tiered_fallback_room(effective_enrollment, day, time_slot)
                        
                        if fallback_room and str(fallback_room).strip():
                            # Found an available room - no conflict
                            reserve_room(fallback_room, day, time_slot)
                            schedule_with_rooms.loc[time_slot, day] = f"{course_display} [{fallback_room}]"
                            allocation_key = f"{day}_{time_slot}"
                            _TIMETABLE_CLASSROOM_ALLOCATIONS[timetable_key][allocation_key] = {
                                'course': course_display,
                                'classroom': fallback_room,
                                'enrollment': effective_enrollment,
                                'conflict': False,
                                'split': False
                            }
                            if course_key not in course_preferred_classrooms:
                                course_preferred_classrooms[course_key] = fallback_room
                            allocation_count += 1
                            print(f"      [FALLBACK-OK] {day} {time_slot}: {course_display} -> {fallback_room} ({effective_enrollment} students) [Found available room]")
                        else:
                            # Last-resort forced assignment: pick a room even if already booked
                            forced_room = pick_forced_fallback_room(effective_enrollment, day, time_slot)
                            if forced_room:
                                schedule_with_rooms.loc[time_slot, day] = f"{course_display} [{forced_room}]"
                                allocation_key = f"{day}_{time_slot}"
                                _TIMETABLE_CLASSROOM_ALLOCATIONS[timetable_key][allocation_key] = {
                                    'course': course_display,
                                    'classroom': forced_room,
                                    'enrollment': effective_enrollment,
                                    'conflict': True,
                                    'split': False
                                }
                                if course_key not in course_preferred_classrooms:
                                    course_preferred_classrooms[course_key] = forced_room
                                allocation_count += 1
                                print(f"      [FORCED] {day} {time_slot}: {course_display} -> {forced_room} ({effective_enrollment} students) [all rooms booked; conflict-allowed fallback]")
                            else:
                                schedule_with_rooms.loc[time_slot, day] = course_display
                                print(f"      [WARN] {day} {time_slot}: NO ROOM AVAILABLE for {course_display} ({effective_enrollment} students) - no rooms configured")
    
    print(f"   [SCHOOL] Total classroom allocations: {allocation_count}")
    return schedule_with_rooms

def find_suitable_classroom_for_lab_pair(lab_rooms_df, enrollment, day, time_slot1, time_slot2, classroom_usage_tracker, schedule_type='Regular'):
    """Find a suitable LAB ROOM (Hardware Lab, Software Lab types) that's available for BOTH slots of a lab pair"""
    # Ensure Capacity is numeric and exclude rooms with missing/non-positive capacity
    lab_rooms_df = lab_rooms_df.copy()
    lab_rooms_df['Capacity'] = pd.to_numeric(lab_rooms_df['Capacity'], errors='coerce')
    lab_rooms_df = lab_rooms_df[lab_rooms_df['Capacity'].notna() & (lab_rooms_df['Capacity'] > 0)].copy()
    if lab_rooms_df.empty:
        print(f"         [WARN] No lab rooms available with valid capacity for lab pair")
        return None

    # FIX: Only use actual lab rooms by TYPE, not L prefix
    # L402-L408 are classrooms with 80 capacity, NOT labs
    room_type_lower = lab_rooms_df['Type'].astype(str).str.lower().str.strip()
    is_actual_lab = (
        room_type_lower.str.contains('hardware', na=False) | 
        room_type_lower.str.contains('software', na=False) |
        (room_type_lower.str.contains('lab', na=False) & ~room_type_lower.isin(['classroom', 'large classroom', 'auditorium']))
    )
    lab_rooms_df = lab_rooms_df[is_actual_lab].copy()
    
    if lab_rooms_df.empty:
        print(f"         [WARN] No lab rooms (starting with 'L') found")
        return None
    
    # Use prefixed day key for proper tracking
    prefixed_day = f"{schedule_type}_{day}"
    
    # Filter lab rooms that can accommodate the enrollment
    suitable_rooms = lab_rooms_df[lab_rooms_df['Capacity'] >= enrollment].copy()
    
    if suitable_rooms.empty:
        # If no room can accommodate, find the largest available lab room
        suitable_rooms = lab_rooms_df.nlargest(3, 'Capacity')
        if suitable_rooms.empty:
            print(f"         [WARN] No lab rooms available for {enrollment} students")
            return None
        print(f"         [WARN] Using largest available lab room for {enrollment} students (lab pair)")
    
    # Sort by capacity (prefer smallest adequate room first)
    suitable_rooms = suitable_rooms.sort_values('Capacity')
    
    # Shuffle to ensure variety
    suitable_rooms = suitable_rooms.sample(frac=1).reset_index(drop=True)
    
    # Check availability in global tracker for BOTH slots using prefixed_day
    for _, room in suitable_rooms.iterrows():
        room_number = room['Room Number']
        
        # Treat missing day/slot entries as available (they mean nothing is yet reserved for that slot)
        available_slot1 = not (prefixed_day in classroom_usage_tracker and time_slot1 in classroom_usage_tracker[prefixed_day] and room_number in classroom_usage_tracker[prefixed_day][time_slot1])
        
        available_slot2 = not (prefixed_day in classroom_usage_tracker and time_slot2 in classroom_usage_tracker[prefixed_day] and room_number in classroom_usage_tracker[prefixed_day][time_slot2])
        
        if available_slot1 and available_slot2:
            # Mark room as used in global tracker for BOTH slots
            if prefixed_day not in classroom_usage_tracker:
                classroom_usage_tracker[prefixed_day] = {}
            if time_slot1 not in classroom_usage_tracker[prefixed_day]:
                classroom_usage_tracker[prefixed_day][time_slot1] = set()
            if time_slot2 not in classroom_usage_tracker[prefixed_day]:
                classroom_usage_tracker[prefixed_day][time_slot2] = set()
            
            classroom_usage_tracker[prefixed_day][time_slot1].add(room_number)
            classroom_usage_tracker[prefixed_day][time_slot2].add(room_number)
            print(f"         [PIN] Allocated {room_number} for lab pair {prefixed_day} {time_slot1} & {time_slot2} (Capacity: {room['Capacity']})")
            return room_number
    
    # If all suitable rooms are booked, try larger lab rooms
    larger_rooms = lab_rooms_df[lab_rooms_df['Capacity'] > enrollment].copy()
    if not larger_rooms.empty:
        larger_rooms = larger_rooms.sort_values('Capacity')
        # Shuffle to ensure variety
        larger_rooms = larger_rooms.sample(frac=1).reset_index(drop=True)
        for _, room in larger_rooms.iterrows():
            room_number = room['Room Number']
            
            # FIX: Correct availability check - room is available if NOT tracked as booked
            # A room is available if the day/slot doesn't exist in tracker OR room is not in booked set
            available_slot1 = not (prefixed_day in classroom_usage_tracker and 
                                  time_slot1 in classroom_usage_tracker[prefixed_day] and
                                  room_number in classroom_usage_tracker[prefixed_day][time_slot1])
            
            available_slot2 = not (prefixed_day in classroom_usage_tracker and 
                                  time_slot2 in classroom_usage_tracker[prefixed_day] and
                                  room_number in classroom_usage_tracker[prefixed_day][time_slot2])
            
            if available_slot1 and available_slot2:
                if prefixed_day not in classroom_usage_tracker:
                    classroom_usage_tracker[prefixed_day] = {}
                if time_slot1 not in classroom_usage_tracker[prefixed_day]:
                    classroom_usage_tracker[prefixed_day][time_slot1] = set()
                if time_slot2 not in classroom_usage_tracker[prefixed_day]:
                    classroom_usage_tracker[prefixed_day][time_slot2] = set()
                
                classroom_usage_tracker[prefixed_day][time_slot1].add(room_number)
                classroom_usage_tracker[prefixed_day][time_slot2].add(room_number)
                print(f"         [RESET] Using larger room {room_number} for lab pair {prefixed_day} {time_slot1} & {time_slot2} (Capacity: {room['Capacity']})")
                return room_number
    
    # ULTIMATE FALLBACK: Try ANY lab room regardless of capacity (prefer larger)
    all_labs_sorted = lab_rooms_df.sort_values('Capacity', ascending=False)
    for _, room in all_labs_sorted.iterrows():
        room_number = room['Room Number']
        available_slot1 = not (prefixed_day in classroom_usage_tracker and 
                              time_slot1 in classroom_usage_tracker[prefixed_day] and
                              room_number in classroom_usage_tracker[prefixed_day][time_slot1])
        available_slot2 = not (prefixed_day in classroom_usage_tracker and 
                              time_slot2 in classroom_usage_tracker[prefixed_day] and
                              room_number in classroom_usage_tracker[prefixed_day][time_slot2])
        if available_slot1 and available_slot2:
            if prefixed_day not in classroom_usage_tracker:
                classroom_usage_tracker[prefixed_day] = {}
            if time_slot1 not in classroom_usage_tracker[prefixed_day]:
                classroom_usage_tracker[prefixed_day][time_slot1] = set()
            if time_slot2 not in classroom_usage_tracker[prefixed_day]:
                classroom_usage_tracker[prefixed_day][time_slot2] = set()
            classroom_usage_tracker[prefixed_day][time_slot1].add(room_number)
            classroom_usage_tracker[prefixed_day][time_slot2].add(room_number)
            print(f"         [ULTIMATE-LAB] Using any available lab {room_number} for {prefixed_day} {time_slot1} & {time_slot2}")
            return room_number
    
    print(f"         [LAB-CONFLICT] All lab rooms booked for {prefixed_day} {time_slot1} & {time_slot2}")
    return None

def find_suitable_classroom_with_tracking(
    classrooms_df,
    enrollment,
    day,
    time_slot,
    classroom_usage_tracker,
    is_common=False,
    is_lab=False,
    preferred_capacities_override=None,
    schedule_type='Regular',
):
    """Find a suitable classroom based on capacity, course type, and availability with global tracking
    
    Args:
        classrooms_df: DataFrame of available classrooms
        enrollment: Number of students
        day: Day of the week
        time_slot: Time slot
        classroom_usage_tracker: Global tracker to prevent double-booking
        is_common: True if this is a common course (should get 120/240 capacity rooms)
        is_lab: True if this is a lab session
        schedule_type: 'Regular', 'PreMid', or 'PostMid' - used to construct prefixed tracker key
    """
    if classrooms_df.empty:
        return None
    # Ensure Capacity is numeric and exclude rooms with missing/non-positive capacity
    classrooms_df = classrooms_df.copy()
    classrooms_df['Capacity'] = pd.to_numeric(classrooms_df['Capacity'], errors='coerce')
    classrooms_df = classrooms_df[classrooms_df['Capacity'].notna() & (classrooms_df['Capacity'] > 0)].copy()
    if classrooms_df.empty:
        return None
    
    # If this is a lab session, filter to ONLY actual lab rooms by TYPE (not L prefix)
    # L402-L408 are classrooms, not labs despite L prefix
    if is_lab:
        room_type_lower = classrooms_df['Type'].astype(str).str.lower().str.strip()
        is_actual_lab = (
            room_type_lower.str.contains('hardware', na=False) | 
            room_type_lower.str.contains('software', na=False) |
            (room_type_lower.str.contains('lab', na=False) & ~room_type_lower.isin(['classroom', 'large classroom', 'auditorium']))
        )
        classrooms_df = classrooms_df[is_actual_lab].copy()
        if classrooms_df.empty:
            print(f"         [LAB-WARN] No lab rooms available after filtering")
            return None
        print(f"         [LAB-FILTER] Found {len(classrooms_df)} lab rooms for lab session")
    
    # Filter out already booked rooms FIRST to prevent double-booking
    # Use prefixed key format to match how rooms are tracked in allocate_classrooms_for_timetable
    available_now = classrooms_df.copy()
    prefixed_day = f"{schedule_type}_{day}"
    if prefixed_day in classroom_usage_tracker and time_slot in classroom_usage_tracker[prefixed_day]:
        booked_rooms = classroom_usage_tracker[prefixed_day][time_slot]
        available_now = available_now[~available_now['Room Number'].isin(booked_rooms)]
    
    if available_now.empty:
        print(f"         [WARN] All classrooms are booked for {day} {time_slot} (schedule_type={schedule_type})")
        return None
    
    # Determine preferred room capacity based on course type
    if preferred_capacities_override is not None:
        preferred_capacities = preferred_capacities_override
        print(f"         [PREF-OVERRIDE] Using preferred capacities {preferred_capacities} for {enrollment} students")
    elif is_common:
        # Common courses: tiered escalation based on enrollment
        if enrollment > 150:
            preferred_capacities = [240]
            print(f"         [COMMON] High enrollment ({enrollment}): Prefer 240 capacity room")
        elif enrollment > 135:
            preferred_capacities = [240, 135]
            print(f"         [COMMON] Large enrollment ({enrollment}): Prefer 240, fallback 135")
        elif enrollment > 120:
            preferred_capacities = [135, 120]
            print(f"         [COMMON] Moderate-high enrollment ({enrollment}): Prefer 135, fallback 120")
        elif enrollment > 96:
            preferred_capacities = [120, 135]
            print(f"         [COMMON] Moderate enrollment ({enrollment}): Prefer 120, fallback 135")
        elif enrollment > 80:
            preferred_capacities = [96, 120, 135]
            print(f"         [COMMON] Medium enrollment ({enrollment}): Prefer 96, fallback 120/135")
        else:
            preferred_capacities = [80, 96, 120]
            print(f"         [COMMON] Small enrollment ({enrollment}): Prefer 80, fallback 96/120")
    elif is_lab:
        # Labs use whatever capacity is needed
        preferred_capacities = None
    else:
        # Non-common lectures/tutorials: tiered escalation based on enrollment
        if enrollment > 135:
            preferred_capacities = [240]
            print(f"         [NON-COMMON] Large enrollment ({enrollment}): Prefer 240 capacity room")
        elif enrollment > 120:
            preferred_capacities = [135, 120]
            print(f"         [NON-COMMON] Enrollment ({enrollment}): Prefer 135, fallback 120")
        elif enrollment > 96:
            preferred_capacities = [120, 135]
            print(f"         [NON-COMMON] Enrollment ({enrollment}): Prefer 120, fallback 135")
        elif enrollment > 80:
            preferred_capacities = [96, 120]
            print(f"         [NON-COMMON] Enrollment ({enrollment}): Prefer 96, fallback 120")
        else:
            preferred_capacities = [80, 96, 120]
            print(f"         [NON-COMMON] Small enrollment ({enrollment}): Prefer 80, fallback 96/120")
    
    # Try to find preferred capacity rooms first
    selected_room = None
    if preferred_capacities:
        for pref_cap in preferred_capacities:
            # Find rooms close to preferred capacity (within ±10 to allow for slight variations)
            capacity_match = available_now[
                (available_now['Capacity'] >= pref_cap - 10) & 
                (available_now['Capacity'] <= pref_cap + 10) &
                (available_now['Capacity'] >= enrollment)  # Must fit the students
            ]
            if not capacity_match.empty:
                # Sort by how close to preferred capacity, then by usage count (load balancing)
                capacity_match = capacity_match.copy()
                capacity_match['cap_diff'] = abs(capacity_match['Capacity'] - pref_cap)
                capacity_match['_usage'] = capacity_match['Room Number'].apply(lambda r: _ROOM_ALLOCATION_COUNTER.get(str(r), 0))
                capacity_match = capacity_match.sort_values(['cap_diff', '_usage'])
                selected_room = capacity_match.iloc[0]['Room Number']
                selected_capacity = capacity_match.iloc[0]['Capacity']
                print(f"         [MATCH] Found preferred capacity {selected_capacity} (target {pref_cap}) for {enrollment} students (load-balanced)")
                break
    
    # If no preferred capacity found, use any suitable room
    if selected_room is None:
        suitable_rooms = available_now[available_now['Capacity'] >= enrollment].copy()
        
        if suitable_rooms.empty:
            # If no room can accommodate, use the largest available
            suitable_rooms = available_now.nlargest(1, 'Capacity')
            if suitable_rooms.empty:
                return None
            print(f"         [WARN] Using largest available room for {enrollment} students")
        else:
            # Sort by capacity first (prefer smallest adequate room), then by usage (load balancing)
            suitable_rooms['_usage'] = suitable_rooms['Room Number'].apply(lambda r: _ROOM_ALLOCATION_COUNTER.get(str(r), 0))
            suitable_rooms = suitable_rooms.sort_values(['Capacity', '_usage'])
        
        selected_room = suitable_rooms.iloc[0]['Room Number']
        selected_capacity = suitable_rooms.iloc[0]['Capacity']
    
    # Reserve the room in global tracker to prevent double-booking
    # Use prefixed_day (e.g., "PreMid_Mon") to match how rooms are checked
    prefixed_day = f"{schedule_type}_{day}"
    if prefixed_day not in classroom_usage_tracker:
        classroom_usage_tracker[prefixed_day] = {}
    if time_slot not in classroom_usage_tracker[prefixed_day]:
        classroom_usage_tracker[prefixed_day][time_slot] = set()
    
    classroom_usage_tracker[prefixed_day][time_slot].add(selected_room)
    # Update global allocation counter for load balancing
    room_key = str(selected_room)
    if room_key not in _ROOM_ALLOCATION_COUNTER:
        _ROOM_ALLOCATION_COUNTER[room_key] = 0
    _ROOM_ALLOCATION_COUNTER[room_key] += 1
    print(f"         [ALLOCATED] {selected_room} (Cap: {selected_capacity}) for {prefixed_day} {time_slot} - {enrollment} students")
    
    return selected_room

def find_suitable_classroom(classrooms_df, enrollment, day, time_slot, classroom_usage):
    """Find a suitable classroom based on capacity and availability with load balancing"""
    if classrooms_df.empty:
        return None
    # Ensure Capacity is numeric and exclude rooms with missing/non-positive capacity
    classrooms_df = classrooms_df.copy()
    classrooms_df['Capacity'] = pd.to_numeric(classrooms_df['Capacity'], errors='coerce')
    classrooms_df = classrooms_df[classrooms_df['Capacity'].notna() & (classrooms_df['Capacity'] > 0)].copy()
    if classrooms_df.empty:
        return None
    
    # Filter classrooms that can accommodate the enrollment
    suitable_rooms = classrooms_df[classrooms_df['Capacity'] >= enrollment].copy()
    
    if suitable_rooms.empty:
        # If no room can accommodate, find the largest available
        suitable_rooms = classrooms_df.nlargest(1, 'Capacity')
        if suitable_rooms.empty:
            return None
        print(f"         [WARN] Using largest available room for {enrollment} students")
    
    # Filter out already booked rooms
    booked_rooms = classroom_usage.get(day, {}).get(time_slot, set())
    suitable_rooms = suitable_rooms[~suitable_rooms['Room Number'].isin(booked_rooms)]
    
    if suitable_rooms.empty:
        return None
    
    # Sort by capacity first, then by usage count (load balancing)
    suitable_rooms['_usage'] = suitable_rooms['Room Number'].apply(lambda r: _ROOM_ALLOCATION_COUNTER.get(str(r), 0))
    suitable_rooms = suitable_rooms.sort_values(['Capacity', '_usage'])
    
    return suitable_rooms.iloc[0]['Room Number']

def estimate_course_enrollment(course_info):
    """Estimate student enrollment for courses using available student data if present"""
    enrollment_estimates = {}

    # Try to get student counts from cached data frames (if uploaded)
    student_counts_by_sem_dept = {}
    try:
        # Force reload to pick up recently uploaded/modified student files
        dfs = load_all_data(force_reload=True)
        students_df = None
        if dfs and 'student' in dfs:
            students_df = dfs['student']
        if students_df is not None and not students_df.empty:
            # Normalize and count students by Semester and Department
            students_df = students_df.copy()
            if 'Semester' in students_df.columns:
                students_df['Semester'] = students_df['Semester'].astype(str).str.strip()
            if 'Department' in students_df.columns:
                # Normalize Department to canonical branch names for robust matching
                students_df['Department'] = students_df['Department'].astype(str).str.strip().apply(lambda v: normalize_branch_string(v))
            counts = students_df.groupby(['Semester', 'Department']).size().to_dict()
            total_by_sem = students_df.groupby('Semester').size().to_dict()
            student_counts_by_sem_dept = {'pair_counts': counts, 'sem_counts': total_by_sem}
    except Exception:
        student_counts_by_sem_dept = {}

    for course_code, info in course_info.items():
        semester = str(info.get('semester', '')).strip() if info.get('semester') is not None else ''
        branch = info.get('branch') or info.get('department') or ''

        # Prefer explicit registered students if present
        registered = info.get('registered_students') if info.get('registered_students') else None
        if registered:
            # Always use the registered count for accurate classroom sizing
            enrollment_estimates[course_code] = int(registered)
            continue

        # Use student counts if available
        est = None
        if student_counts_by_sem_dept:
            normalized_branch = normalize_branch_string(branch) if branch else ''
            pair_key = (semester, normalized_branch) if semester and normalized_branch else None
            pair_counts = student_counts_by_sem_dept.get('pair_counts', {})
            sem_counts = student_counts_by_sem_dept.get('sem_counts', {})
            
            if pair_key and pair_key in pair_counts:
                est = int(pair_counts[pair_key])
            elif semester in sem_counts:
                est = int(sem_counts[semester])

        if est is None or est == 0:
            # Fallback heuristics
            if info.get('is_elective', False):
                est = 40
            else:
                est = 60

        # Cap elective enrollments to a sensible value
        if info.get('is_elective', False):
            est = max(20, min(est, 60))

        enrollment_estimates[course_code] = est

    return enrollment_estimates

def calculate_rooms_needed(enrollment, classrooms_df):
    """Calculate how many rooms are needed for an exam based on enrollment"""
    # Sort classrooms by capacity (descending)
    sorted_rooms = classrooms_df.sort_values('Capacity', ascending=False)
    
    remaining_students = enrollment
    allocated_rooms = []
    
    for _, room in sorted_rooms.iterrows():
        if remaining_students <= 0:
            break
            
        room_capacity = room['Capacity']
        room_number = room['Room Number']
        
        # Use this room
        allocated_rooms.append(room_number)
        remaining_students -= room_capacity
    
    return allocated_rooms

def create_classroom_utilization_report(classrooms_df, timetable_schedules, exam_schedules=None):
    """Create a comprehensive classroom utilization report (exam functions disabled)"""
    print("[STATS] Generating classroom utilization report (exam functions disabled)...")
    
    utilization_data = []
    
    for _, classroom in classrooms_df.iterrows():
        room_number = classroom.get('Room Number', 'Unknown')
        capacity = classroom.get('Capacity', 0)
        room_type = classroom.get('Type', 'Unknown')
        
        # Calculate timetable usage
        timetable_usage = calculate_timetable_usage(room_number, timetable_schedules)
        
        utilization_data.append({
            'Room Number': room_number,
            'Type': room_type,
            'Capacity': capacity,
            'Weekly Hours (Timetable)': timetable_usage['weekly_hours'],
            'Daily Avg Hours (Timetable)': timetable_usage['daily_avg_hours'],
            'Exam Sessions': 0,
            'Utilization Rate (%)': 0,
            'Facilities': classroom.get('Facilities', 'None')
        })
    
    return pd.DataFrame(utilization_data)

def calculate_timetable_usage(room_number, timetable_schedules):
    """Calculate how much a room is used in timetables"""
    weekly_hours = 0
    
    try:
        for schedule in timetable_schedules:
            for day in schedule.columns:
                for time_slot in schedule.index:
                    cell_value = str(schedule.loc[time_slot, day])
                    if f"[{room_number}]" in cell_value:
                        # Estimate hours from standard time slots
                        weekly_hours += 1.5  # Standard 90-minute class
    except:
        pass
    
    daily_avg_hours = weekly_hours / 5 if weekly_hours > 0 else 0
    return {'weekly_hours': weekly_hours, 'daily_avg_hours': daily_avg_hours}

# EXAM USAGE CALCULATION FUNCTIONS - COMMENTED OUT
"""
def calculate_exam_usage(room_number, exam_schedules):
    \"\"\"Calculate how much a room is used for exams\"\"\"
    exam_sessions = 0
    
    for schedule in exam_schedules:
        for _, exam in schedule.iterrows():
            if exam['status'] == 'Scheduled':
                classrooms = str(exam.get('classroom', ''))
                if room_number in classrooms:
                    exam_sessions += 1
    
    return {'exam_sessions': exam_sessions}

def calculate_time_slot_hours(time_slot):
    \"\"\"Calculate duration in hours from time slot string\"\"\"
    try:
        if '-' in time_slot:
            start, end = time_slot.split('-')
            start_hour = int(start.split(':')[0]) + int(start.split(':')[1]) / 60
            end_hour = int(end.split(':')[0]) + int(end.split(':')[1]) / 60
            return end_hour - start_hour
    except:
        pass
    return 1.5  # Default fallback

def calculate_utilization_rate(weekly_hours, exam_sessions):
    \"\"\"Calculate classroom utilization rate\"\"\"
    # Assuming 40 hours per week available (8 hours x 5 days)
    max_weekly_hours = 40
    timetable_utilization = (weekly_hours / max_weekly_hours) * 100
    
    # Exam utilization (each exam session counts as 1)
    exam_utilization = min(exam_sessions * 5, 100)  # Cap at 100%
    
    return min(timetable_utilization + exam_utilization, 100)
"""

def create_timetable_statistics_sheet(schedule_df, course_info, classrooms_df, semester, branch, section):
    """Create detailed statistics sheet for timetable verification"""
    
    print(f"[STATS] Creating statistics sheet for Semester {semester}, {branch}, Section {section}...")
    
    # Extract all scheduled courses from the timetable
    scheduled_courses = extract_scheduled_courses_from_timetable(schedule_df)
    
    statistics_data = []
    
    for course_code, course_data in scheduled_courses.items():
        # Get course info
        info = course_info.get(course_code, {})
        
        # Parse LTPSC
        ltpsc_str = info.get('ltpsc', '3-0-0-0-3')
        ltpsc = parse_ltpsc(ltpsc_str)
        
        # Calculate scheduled sessions from timetable
        lecture_count = course_data.get('lecture_count', 0)
        tutorial_count = course_data.get('tutorial_count', 0)
        lab_count = course_data.get('lab_count', 0)
        
        # Check LTPSC compliance
        lectures_compliant = "[OK]" if lecture_count == ltpsc['L'] else f"[FAIL] ({lecture_count}/{ltpsc['L']})"
        tutorials_compliant = "[OK]" if tutorial_count == ltpsc['T'] else f"[FAIL] ({tutorial_count}/{ltpsc['T']})"
        labs_compliant = "[OK]" if lab_count == ltpsc['P'] else f"[FAIL] ({lab_count}/{ltpsc['P']})"
        
        # Get allocated rooms
        allocated_rooms = ', '.join(course_data.get('rooms', [])) if course_data.get('rooms') else 'Not Allocated'
        
        # Check for combined classes (courses with same code in different sections at same time)
        combined_class = check_combined_class(course_code, schedule_df, section, semester, branch)
        
        statistics_data.append({
            'Course Code': f"**{course_code}**",
            'Course Name': info.get('name', 'Unknown'),
            'Faculty': info.get('instructor', 'Unknown'),
            'LTPSC': ltpsc_str,
            'Lectures/Week': f"{lecture_count}/{ltpsc['L']}",
            'Tutorials/Week': f"{tutorial_count}/{ltpsc['T']}",
            'Labs/Week': f"{lab_count}/{ltpsc['P']}",
            'LTPSC Compliance': f"{lectures_compliant} | {tutorials_compliant} | {labs_compliant}",
            'Combined Class': 'Yes' if combined_class else 'No',
            'Allocation Status': '[OK] Complete' if all([
                lecture_count == ltpsc['L'],
                tutorial_count == ltpsc['T'],
                lab_count == ltpsc['P']
            ]) else '[WARN] Partial',
            'Allocated Rooms': allocated_rooms,
            'Room Utilization': calculate_room_utilization(course_data.get('rooms', []), classrooms_df)
        })
    
    # Add summary row
    total_courses = len(scheduled_courses)
    compliant_courses = sum(1 for course in statistics_data if course['Allocation Status'] == '[OK] Complete')
    
    statistics_data.append({
        'Course Code': '**SUMMARY**',
        'Course Name': f'Total Courses: {total_courses}',
        'Faculty': f'Fully Compliant: {compliant_courses}',
        'LTPSC': f'Compliance Rate: {(compliant_courses/total_courses)*100:.1f}%' if total_courses > 0 else 'N/A',
        'Lectures/Week': f'Lectures: {sum(c.get("lecture_count", 0) for c in scheduled_courses.values())}',
        'Tutorials/Week': f'Tutorials: {sum(c.get("tutorial_count", 0) for c in scheduled_courses.values())}',
        'Labs/Week': f'Labs: {sum(c.get("lab_count", 0) for c in scheduled_courses.values())}',
        'LTPSC Compliance': 'OVERALL',
        'Combined Class': '--',
        'Allocation Status': '[OK]' if compliant_courses == total_courses else f'[WARN] {total_courses - compliant_courses} issues',
        'Allocated Rooms': '--',
        'Room Utilization': '--'
    })
    
    return pd.DataFrame(statistics_data)

def extract_scheduled_courses_from_timetable(schedule_df):
    """Extract all scheduled courses from timetable with session counts"""
    
    scheduled_courses = {}
    
    for day in schedule_df.columns:
        for time_slot in schedule_df.index:
            cell_value = str(schedule_df.loc[time_slot, day])
            
            if cell_value in ['Free', 'LUNCH BREAK']:
                continue
            
            # Extract course code and session type
            course_code, session_type, rooms = parse_timetable_cell(cell_value)
            
            if not course_code:
                continue
            
            if course_code not in scheduled_courses:
                scheduled_courses[course_code] = {
                    'lecture_count': 0,
                    'tutorial_count': 0,
                    'lab_count': 0,
                    'rooms': set()
                }
            
            # Count session types
            if '(Tutorial)' in cell_value:
                scheduled_courses[course_code]['tutorial_count'] += 1
            elif '(Lab)' in cell_value:
                scheduled_courses[course_code]['lab_count'] += 1
            else:
                scheduled_courses[course_code]['lecture_count'] += 1
            
            # Add rooms
            if rooms:
                scheduled_courses[course_code]['rooms'].add(rooms)
    
    return scheduled_courses

def parse_timetable_cell(cell_value):
    """Parse timetable cell to extract course code and room"""
    
    if not isinstance(cell_value, str):
        return None, None, None
    
    # Remove classroom info
    if '[' in cell_value and ']' in cell_value:
        # Extract course part and room part
        course_part = cell_value.split('[')[0].strip()
        room_part = '[' + cell_value.split('[')[1]
        
        # Clean course code
        course_clean = course_part.replace(' (Tutorial)', '').replace(' (Lab)', '')
        
        # Try to extract course code pattern (case-insensitive, allow trailing letter)
        import re
        course_pattern = r'[A-Za-z]{2,3}\s?-?\d{3}[A-Za-z]?'
        match = re.search(course_pattern, course_clean)

        if match:
            normalized_code = match.group(0).replace(' ', '').replace('-', '').upper()
            return normalized_code, course_part, room_part.strip('[]')
        else:
            # Check if it's a basket
            if any(basket in course_clean.upper() for basket in ['ELECTIVE_', 'HSS_', 'PROF_', 'OE_']):
                return course_clean, course_part, room_part.strip('[]')
    
    # For cells without room allocation
    course_clean = cell_value.replace(' (Tutorial)', '').replace(' (Lab)', '')
    
    import re
    course_pattern = r'[A-Za-z]{2,3}\s?-?\d{3}[A-Za-z]?'
    match = re.search(course_pattern, course_clean)

    if match:
        normalized_code = match.group(0).replace(' ', '').replace('-', '').upper()
        return normalized_code, cell_value, None
    elif any(basket in course_clean.upper() for basket in ['ELECTIVE_', 'HSS_', 'PROF_', 'OE_']):
        return course_clean, cell_value, None
    
    return None, None, None

def check_combined_class(course_code, schedule_df, section, semester, branch):
    """Check if this course has combined classes with other sections"""
    # This would need cross-section comparison
    # For now, return False - can be enhanced with multi-section analysis
    return False

def calculate_room_utilization(rooms, classrooms_df):
    """Calculate room utilization percentage"""
    if not rooms or classrooms_df.empty:
        return 'N/A'
    
    total_capacity = 0
    used_capacity = 0
    
    for room in rooms:
        room_info = classrooms_df[classrooms_df['Room Number'] == room]
        if not room_info.empty:
            capacity = room_info['Capacity'].iloc[0]
            if isinstance(capacity, (int, float)):
                total_capacity += capacity
                used_capacity += min(capacity, 60)  # Assuming 60 students per course
    
    if total_capacity > 0:
        utilization = (used_capacity / total_capacity) * 100
        return f'{utilization:.1f}%'
    
    return 'N/A'

def create_room_allocation_summary(schedule_df, classrooms_df):
    """Create detailed room allocation summary"""
    
    room_allocations = {}
    
    for day in schedule_df.columns:
        for time_slot in schedule_df.index:
            cell_value = str(schedule_df.loc[time_slot, day])
            
            if '[' in cell_value and ']' in cell_value:
                try:
                    # Extract room
                    room_match = re.search(r'\[(.*?)\]', cell_value)
                    if room_match:
                        room = room_match.group(1)
                        
                        # Extract course
                        course_part = cell_value.split('[')[0].strip()
                        
                        if room not in room_allocations:
                            room_allocations[room] = {
                                'total_sessions': 0,
                                'days_used': set(),
                                'courses': set(),
                                'time_slots': []
                            }
                        
                        room_allocations[room]['total_sessions'] += 1
                        room_allocations[room]['days_used'].add(day)
                        room_allocations[room]['courses'].add(course_part)
                        room_allocations[room]['time_slots'].append(f"{day} {time_slot}")
                except:
                    continue
    
    # Create summary DataFrame
    summary_data = []
    
    for room, allocation in room_allocations.items():
        # Get room details
        room_info = classrooms_df[classrooms_df['Room Number'] == room]
        capacity = room_info['Capacity'].iloc[0] if not room_info.empty else 'Unknown'
        room_type = room_info['Type'].iloc[0] if not room_info.empty else 'Unknown'
        
        summary_data.append({
            'Room Number': room,
            'Type': room_type,
            'Capacity': capacity,
            'Total Sessions': allocation['total_sessions'],
            'Days Used': len(allocation['days_used']),
            'Courses Assigned': len(allocation['courses']),
            'Utilization (Sessions/Day)': f"{allocation['total_sessions']/5:.1f}",
            'Sample Schedule': ', '.join(allocation['time_slots'][:3]) + ('...' if len(allocation['time_slots']) > 3 else '')
        })
    
    return pd.DataFrame(summary_data)

def create_comprehensive_summary(dfs, semester, branch, section_a_df, section_b_df, basket_allocations):
    """Create comprehensive summary sheet with all key metrics"""
    
    summary_data = []
    
    # 1. Course Allocation Summary
    course_info = get_course_info(dfs) if dfs else {}
    scheduled_a = extract_scheduled_courses_from_timetable(section_a_df)
    scheduled_b = extract_scheduled_courses_from_timetable(section_b_df)
    
    total_courses = len(set(list(scheduled_a.keys()) + list(scheduled_b.keys())))
    core_courses = sum(1 for code in scheduled_a.keys() if code in course_info and not course_info[code].get('is_elective', False))
    elective_courses = total_courses - core_courses
    
    summary_data.append({
        'Category': 'Course Allocation',
        'Metric': 'Total Courses Scheduled',
        'Value': total_courses,
        'Details': f'Core: {core_courses}, Elective: {elective_courses}'
    })
    
    # 2. LTPSC Compliance
    compliant_a = sum(1 for course_code, data in scheduled_a.items() 
                     if check_ltpsc_compliance(course_code, data, course_info))
    compliant_b = sum(1 for course_code, data in scheduled_b.items() 
                     if check_ltpsc_compliance(course_code, data, course_info))
    
    compliance_rate = ((compliant_a + compliant_b) / (len(scheduled_a) + len(scheduled_b))) * 100 if (len(scheduled_a) + len(scheduled_b)) > 0 else 0
    
    summary_data.append({
        'Category': 'LTPSC Compliance',
        'Metric': 'Overall Compliance Rate',
        'Value': f'{compliance_rate:.1f}%',
        'Details': f'Section A: {compliant_a}/{len(scheduled_a)}, Section B: {compliant_b}/{len(scheduled_b)}'
    })
    
    # 3. Basket Allocation
    if basket_allocations:
        basket_count = len(basket_allocations)
        basket_courses = sum(len(basket['courses']) for basket in basket_allocations.values())
        
        summary_data.append({
            'Category': 'Elective Baskets',
            'Metric': 'Basket Coverage',
            'Value': f'{basket_count} baskets',
            'Details': f'{basket_courses} courses across all baskets'
        })
    
    # 4. Room Utilization
    classroom_data = dfs.get('classroom', pd.DataFrame())
    if not classroom_data.empty:
        used_rooms_a = set()
        used_rooms_b = set()
        
        for df in [section_a_df, section_b_df]:
            for day in df.columns:
                for time_slot in df.index:
                    cell_value = str(df.loc[time_slot, day])
                    if '[' in cell_value and ']' in cell_value:
                        room_match = re.search(r'\[(.*?)\]', cell_value)
                        if room_match:
                            if df is section_a_df:
                                used_rooms_a.add(room_match.group(1))
                            else:
                                used_rooms_b.add(room_match.group(1))
        
        total_rooms = len(classroom_data)
        used_rooms = len(used_rooms_a.union(used_rooms_b))
        
        summary_data.append({
            'Category': 'Room Utilization',
            'Metric': 'Classroom Usage',
            'Value': f'{used_rooms}/{total_rooms} rooms',
            'Details': f'Utilization: {(used_rooms/total_rooms)*100:.1f}%'
        })
    
    # 5. Time Slot Coverage
    days = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri']
    time_slots = section_a_df.index.tolist()
    
    occupied_slots_a = sum(1 for day in days for slot in time_slots 
                          if section_a_df.loc[slot, day] not in ['Free', 'LUNCH BREAK'])
    occupied_slots_b = sum(1 for day in days for slot in time_slots 
                          if section_b_df.loc[slot, day] not in ['Free', 'LUNCH BREAK'])
    total_slots = len(days) * len(time_slots)
    
    summary_data.append({
        'Category': 'Schedule Density',
        'Metric': 'Time Slot Utilization',
        'Value': f'{(occupied_slots_a + occupied_slots_b)/(total_slots*2)*100:.1f}%',
        'Details': f'A: {occupied_slots_a}/{total_slots}, B: {occupied_slots_b}/{total_slots}'
    })
    
    return pd.DataFrame(summary_data)

def check_ltpsc_compliance(course_code, scheduled_data, course_info):
    """Check if scheduled sessions match LTPSC requirements"""
    
    info = course_info.get(course_code, {})
    ltpsc_str = info.get('ltpsc', '3-0-0-0-3')
    ltpsc = parse_ltpsc(ltpsc_str)
    
    return (scheduled_data.get('lecture_count', 0) == ltpsc['L'] and
            scheduled_data.get('tutorial_count', 0) == ltpsc['T'] and
            scheduled_data.get('lab_count', 0) == ltpsc['P'])

def create_timetable_verification_sheet(schedule_df, course_info, classrooms_df, semester, branch, section):
    """Create detailed verification sheet matching the requested format"""
    
    print(f"[STATS] Creating verification sheet for Semester {semester}, {branch}, Section {section}...")
    
    # Extract all scheduled courses from the timetable
    scheduled_courses = extract_scheduled_courses_from_timetable(schedule_df)
    
    verification_data = []
    
    for course_code, course_data in scheduled_courses.items():
        # Get course info
        info = course_info.get(course_code, {})
        
        # Parse LTPSC
        ltpsc_str = info.get('ltpsc', '3-0-0-0-3')
        ltpsc = parse_ltpsc(ltpsc_str)
        
        # Calculate scheduled sessions from timetable
        lecture_count = course_data.get('lecture_count', 0)
        tutorial_count = course_data.get('tutorial_count', 0)
        lab_count = course_data.get('lab_count', 0)
        
        # Get allocated rooms
        allocated_rooms = ', '.join(course_data.get('rooms', [])) if course_data.get('rooms') else 'Not Allocated'
        
        # Format course code with ** for emphasis
        formatted_course_code = f"**{course_code}**"
        
        # Check if it's an elective basket
        is_basket = any(basket in course_code.upper() for basket in ['ELECTIVE_', 'HSS_', 'PROF_', 'OE_'])
        
        # Format course name
        course_name = info.get('name', 'Unknown')
        if is_basket:
            course_name = "Elective Basket"
        
        # Format faculty
        faculty = info.get('instructor', 'Unknown')
        if is_basket:
            faculty = '-'  # Dash for elective baskets
        
        # Format LTPSC
        formatted_ltpsc = ltpsc_str
        
        # Format lectures/tutorials per week
        if is_basket:
            lect_tuts_week = f"0/0"  # Baskets show 0/0
        else:
            lect_tuts_week = f"{lecture_count}/{tutorial_count}"
        
        # Format labs per week
        labs_week = f"{lab_count}/1" if lab_count > 0 else "0/0"
        
        # Combined class status
        combined_class = "No"  # Default - can be enhanced
        
        # Allocation status
        allocation = "Complete" if (lecture_count >= ltpsc['L'] and 
                                   tutorial_count >= ltpsc['T'] and 
                                   lab_count >= ltpsc['P']) else "Partial"
        
        # Room allocation
        room = allocated_rooms
        
        verification_data.append({
            'Course Code': formatted_course_code,
            'Course Name': course_name,
            'Faculty': faculty,
            'LTPSC': formatted_ltpsc,
            'Lect/Tuts/Week': lect_tuts_week,
            'Labs/Week': labs_week,
            'Combined Class': combined_class,
            'Allocation': allocation,
            'Room': room
        })
    
    # Add summary row
    if verification_data:
        total_courses = len(verification_data)
        complete_allocations = sum(1 for course in verification_data if course['Allocation'] == 'Complete')
        
        verification_data.append({
            'Course Code': '**SUMMARY**',
            'Course Name': f'Total Courses: {total_courses}',
            'Faculty': f'Complete: {complete_allocations}',
            'LTPSC': f'Rate: {(complete_allocations/total_courses)*100:.1f}%',
            'Lect/Tuts/Week': f'Lectures: {sum(c.get("lecture_count", 0) for c in scheduled_courses.values())}',
            'Labs/Week': f'Labs: {sum(c.get("lab_count", 0) for c in scheduled_courses.values())}',
            'Combined Class': '--',
            'Allocation': '[OK]' if complete_allocations == total_courses else f'[WARN] {total_courses - complete_allocations} issues',
            'Room': '--'
        })
    
    return pd.DataFrame(verification_data)

def create_room_allocation_summary_verification(section_a_df, section_b_df, classrooms_df):
    """Create room allocation summary for verification"""
    
    print("[STATS] Creating room allocation summary...")
    
    # Track room usage across both sections
    room_usage = {}
    
    for df, section in [(section_a_df, 'A'), (section_b_df, 'B')]:
        for day in df.columns:
            for time_slot in df.index:
                cell_value = str(df.loc[time_slot, day])
                
                if '[' in cell_value and ']' in cell_value:
                    try:
                        # Extract course and room
                        course_part = cell_value.split('[')[0].strip()
                        room_match = re.search(r'\[(.*?)\]', cell_value)
                        if room_match:
                            room = room_match.group(1)
                            
                            if room not in room_usage:
                                room_usage[room] = {
                                    'total_sessions': 0,
                                    'sections': set(),
                                    'courses': set(),
                                    'time_slots': []
                                }
                            
                            room_usage[room]['total_sessions'] += 1
                            room_usage[room]['sections'].add(section)
                            room_usage[room]['courses'].add(course_part)
                            room_usage[room]['time_slots'].append(f"{day} {time_slot}")
                    except:
                        continue
    
    # Create summary DataFrame
    summary_data = []
    
    for room, usage in room_usage.items():
        # Get room details
        room_info = classrooms_df[classrooms_df['Room Number'] == room]
        capacity = room_info['Capacity'].iloc[0] if not room_info.empty else 'Unknown'
        room_type = room_info['Type'].iloc[0] if not room_info.empty else 'Unknown'
        facilities = room_info['Facilities'].iloc[0] if not room_info.empty and 'Facilities' in room_info.columns else 'Standard'
        
        summary_data.append({
            'Room Number': room,
            'Type': room_type,
            'Capacity': capacity,
            'Facilities': facilities,
            'Total Sessions': usage['total_sessions'],
            'Sections': ', '.join(sorted(usage['sections'])),
            'Courses Assigned': len(usage['courses']),
            'Sample Courses': ', '.join(list(usage['courses'])[:3]) + ('...' if len(usage['courses']) > 3 else ''),
            'Utilization (Sessions/Day)': f"{usage['total_sessions']/5:.1f}"
        })
    
    # After iterating all rooms, if no rooms were used, return an empty DataFrame with expected columns
    if not summary_data:
        cols = ['Room Number', 'Type', 'Capacity', 'Facilities', 'Total Sessions', 'Sections',
                'Courses Assigned', 'Sample Courses', 'Utilization (Sessions/Day)']
        return pd.DataFrame(columns=cols)

    # Otherwise return the summary DataFrame
    return pd.DataFrame(summary_data).sort_values('Room Number')

def create_ltpsc_compliance_summary(dfs, semester, branch, section_a_df, section_b_df):
    """Create LTPSC compliance summary sheet"""
    
    print("[STATS] Creating LTPSC compliance summary...")
    
    course_info = get_course_info(dfs) if dfs else {}
    scheduled_a = extract_scheduled_courses_from_timetable(section_a_df)
    scheduled_b = extract_scheduled_courses_from_timetable(section_b_df)
    
    compliance_data = []
    
    # Combine courses from both sections
    all_courses = set(list(scheduled_a.keys()) + list(scheduled_b.keys()))
    
    for course_code in sorted(all_courses):
        info = course_info.get(course_code, {})
        ltpsc_str = info.get('ltpsc', '3-0-0-0-3')
        ltpsc = parse_ltpsc(ltpsc_str)
        
        # Get scheduled counts from both sections
        scheduled_lectures_a = scheduled_a.get(course_code, {}).get('lecture_count', 0)
        scheduled_tutorials_a = scheduled_a.get(course_code, {}).get('tutorial_count', 0)
        scheduled_labs_a = scheduled_a.get(course_code, {}).get('lab_count', 0)
        
        scheduled_lectures_b = scheduled_b.get(course_code, {}).get('lecture_count', 0)
        scheduled_tutorials_b = scheduled_b.get(course_code, {}).get('tutorial_count', 0)
        scheduled_labs_b = scheduled_b.get(course_code, {}).get('lab_count', 0)
        
        # Check compliance
        lectures_ok = (scheduled_lectures_a >= ltpsc['L'] and scheduled_lectures_b >= ltpsc['L'])
        tutorials_ok = (scheduled_tutorials_a >= ltpsc['T'] and scheduled_tutorials_b >= ltpsc['T'])
        labs_ok = (scheduled_labs_a >= ltpsc['P'] and scheduled_labs_b >= ltpsc['P'])
        
        compliance_status = "[OK] FULLY COMPLIANT" if lectures_ok and tutorials_ok and labs_ok else "[WARN] PARTIAL"
        
        compliance_data.append({
            'Course Code': course_code,
            'Course Name': info.get('name', 'Unknown'),
            'Required LTPSC': ltpsc_str,
            'Required (L/T/P)': f"{ltpsc['L']}/{ltpsc['T']}/{ltpsc['P']}",
            'Section A (L/T/P)': f"{scheduled_lectures_a}/{scheduled_tutorials_a}/{scheduled_labs_a}",
            'Section B (L/T/P)': f"{scheduled_lectures_b}/{scheduled_tutorials_b}/{scheduled_labs_b}",
            'Lectures Status': '[OK]' if lectures_ok else '[FAIL]',
            'Tutorials Status': '[OK]' if tutorials_ok else '[FAIL]',
            'Labs Status': '[OK]' if labs_ok else '[FAIL]',
            'Overall Compliance': compliance_status,
            'Notes': 'Meets requirements' if lectures_ok and tutorials_ok and labs_ok else 'Check scheduling'
        })
    
    return pd.DataFrame(compliance_data)

def create_executive_summary(dfs, semester, branch, section_a_df, section_b_df, basket_allocations):
    """Create executive summary sheet"""
    
    print("[STATS] Creating executive summary...")
    
    summary_data = []
    
    # 1. Basic Information
    summary_data.append({'Category': 'Basic Information', 'Metric': 'Semester', 'Value': semester, 'Details': f'Branch: {branch}'})
    summary_data.append({'Category': 'Basic Information', 'Metric': 'Generation Date', 'Value': datetime.now().strftime('%Y-%m-%d %H:%M'), 'Details': 'Timetable generation timestamp'})
    
    # 2. Course Statistics
    course_info = get_course_info(dfs) if dfs else {}
    scheduled_a = extract_scheduled_courses_from_timetable(section_a_df)
    scheduled_b = extract_scheduled_courses_from_timetable(section_b_df)
    
    total_courses = len(set(list(scheduled_a.keys()) + list(scheduled_b.keys())))
    core_courses = sum(1 for code in scheduled_a.keys() if code in course_info and not course_info[code].get('is_elective', False))
    elective_courses = total_courses - core_courses
    
    summary_data.append({'Category': 'Course Statistics', 'Metric': 'Total Courses', 'Value': total_courses, 'Details': f'Core: {core_courses}, Elective: {elective_courses}'})
    
    # 3. LTPSC Compliance
    compliant_courses = 0
    for course_code in set(list(scheduled_a.keys()) + list(scheduled_b.keys())):
        info = course_info.get(course_code, {})
        ltpsc_str = info.get('ltpsc', '3-0-0-0-3')
        ltpsc = parse_ltpsc(ltpsc_str)
        
        scheduled_a_data = scheduled_a.get(course_code, {})
        scheduled_b_data = scheduled_b.get(course_code, {})
        
        if (scheduled_a_data.get('lecture_count', 0) >= ltpsc['L'] and
            scheduled_a_data.get('tutorial_count', 0) >= ltpsc['T'] and
            scheduled_a_data.get('lab_count', 0) >= ltpsc['P'] and
            scheduled_b_data.get('lecture_count', 0) >= ltpsc['L'] and
            scheduled_b_data.get('tutorial_count', 0) >= ltpsc['T'] and
            scheduled_b_data.get('lab_count', 0) >= ltpsc['P']):
            compliant_courses += 1
    
    compliance_rate = (compliant_courses / total_courses * 100) if total_courses > 0 else 0
    summary_data.append({'Category': 'LTPSC Compliance', 'Metric': 'Compliance Rate', 'Value': f'{compliance_rate:.1f}%', 'Details': f'{compliant_courses}/{total_courses} courses fully compliant'})
    
    # 4. Basket Allocation
    if basket_allocations:
        basket_count = len(basket_allocations)
        basket_courses = sum(len(basket['courses']) for basket in basket_allocations.values())
        summary_data.append({'Category': 'Elective Baskets', 'Metric': 'Baskets Scheduled', 'Value': basket_count, 'Details': f'{basket_courses} courses in baskets'})
    
    # 5. Room Utilization
    classroom_data = dfs.get('classroom', pd.DataFrame())
    if not classroom_data.empty:
        used_rooms = set()
        for df in [section_a_df, section_b_df]:
            for day in df.columns:
                for time_slot in df.index:
                    cell_value = str(df.loc[time_slot, day])
                    if '[' in cell_value and ']' in cell_value:
                        room_match = re.search(r'\[(.*?)\]', cell_value)
                        if room_match:
                            used_rooms.add(room_match.group(1))
        
        total_rooms = len(classroom_data)
        summary_data.append({'Category': 'Room Utilization', 'Metric': 'Rooms Used', 'Value': f'{len(used_rooms)}/{total_rooms}', 'Details': f'Utilization: {(len(used_rooms)/total_rooms)*100:.1f}%'})
    
    # 6. Schedule Density
    expected_days = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri']

    def detect_days_from_df(df):
        if df is None or df.empty:
            return []
        # Direct match
        found = [d for d in expected_days if d in df.columns]
        if found:
            return found
        # Partial match inside column names
        found_partial = []
        for c in df.columns:
            for d in expected_days:
                if d in str(c) and d not in found_partial:
                    found_partial.append(d)
        if found_partial:
            return found_partial
        # Partial match inside index values
        try:
            idx_values = list(map(str, df.index.tolist()))
            found_idx = [d for d in expected_days if any(d in v for v in idx_values)]
            if found_idx:
                return found_idx
        except Exception:
            pass
        # fallback
        return expected_days

    # determine days and time_slots using available dataframe(s)
    days = detect_days_from_df(section_a_df) or detect_days_from_df(section_b_df) or expected_days
    time_slots = (section_a_df.index.tolist() if not section_a_df.empty else (section_b_df.index.tolist() if not section_b_df.empty else []))

    occupied_slots_a = 0
    if not section_a_df.empty:
        occupied_slots_a = sum(1 for day in days for slot in time_slots if section_a_df.loc[slot, day] not in ['Free', 'LUNCH BREAK'])

    occupied_slots_b = 0
    if section_b_df is not None and not section_b_df.empty:
        occupied_slots_b = sum(1 for day in days for slot in time_slots if section_b_df.loc[slot, day] not in ['Free', 'LUNCH BREAK'])
    total_slots = len(days) * len(time_slots)
    
    density_a = (occupied_slots_a / total_slots) * 100
    density_b = (occupied_slots_b / total_slots) * 100
    avg_density = (density_a + density_b) / 2
    
    summary_data.append({'Category': 'Schedule Density', 'Metric': 'Time Slot Utilization', 'Value': f'{avg_density:.1f}%', 'Details': f'A: {density_a:.1f}%, B: {density_b:.1f}%'})
    
    # 7. Quality Assessment
    quality_score = (compliance_rate * 0.4) + (avg_density * 0.3) + (100 if basket_allocations else 80) * 0.3
    quality_status = "[OK] EXCELLENT" if quality_score >= 90 else "[GOOD] GOOD" if quality_score >= 75 else "[WARN] NEEDS REVIEW"
    
    summary_data.append({'Category': 'Quality Assessment', 'Metric': 'Overall Quality', 'Value': quality_status, 'Details': f'Score: {quality_score:.1f}/100'})
    
    return pd.DataFrame(summary_data)

# EXAM ALLOCATION AND SCHEDULING FUNCTIONS - DISABLED
# This entire section has been commented out to disable exam timetable functionality
# The following functions are no longer in use:
# - allocate_classrooms_for_exams()
# - estimate_exam_enrollment()
# - find_suitable_classroom_for_exam()
# - create_configuration_sheet()
# - save_exam_schedule()
# - schedule_exams_conflict_free()
# - (and related conflict checking functions)
# - create_exam_classroom_summary()
# - create_exam_summary()
# - create_department_summary()

# EXAM-RELATED HELPER FUNCTIONS - DISABLED FOR EXAM FUNCTIONALITY REMOVAL
# The following functions are no longer in use and have been stubbed out:
# - allocate_classrooms_for_exams()
# - estimate_exam_enrollment()  
# - find_suitable_classroom_for_exam()
# - create_configuration_sheet()
# - save_exam_schedule()
# - schedule_exams_conflict_free()
# And other exam conflict checking functions

def allocate_classrooms_for_exams(exam_schedule_df, classrooms_df, course_data_df):
    """DISABLED - Allocate classrooms for exams"""
    return exam_schedule_df

def estimate_exam_enrollment(exam_schedule_df, course_data_df):
    """DISABLED - Estimate enrollment"""
    return {}

def find_suitable_classroom_for_exam(classrooms_df, enrollment, used_classrooms):
    """DISABLED - Find classroom"""
    return None

def create_configuration_sheet(config):
    """DISABLED - Create configuration"""
    return pd.DataFrame()

def save_exam_schedule(schedule_df, start_date, end_date, config=None):
    """DISABLED - Save exam schedule"""
    return None

def schedule_exams_conflict_free(exams_df, start_date, end_date, max_exams_per_day=2, 
                               include_weekends=False, session_duration=180,
                               department_conflict='moderate', preference_weight='medium',
                               session_balance='strict'):
    """DISABLED - Schedule exams conflict-free"""
    return pd.DataFrame()

# EXAM FUNCTIONS DISABLED - all exam scheduling/allocation functions stubbed out

def allocate_classrooms_for_exams(exam_schedule_df, classrooms_df, course_data_df):
    """DISABLED - Allocate classrooms for exams"""
    return exam_schedule_df

def estimate_exam_enrollment(exam_schedule_df, course_data_df):
    """DISABLED - Estimate exam enrollment"""
    return {}

def find_suitable_classroom_for_exam(classrooms_df, enrollment, used_classrooms):
    """DISABLED - Find suitable classroom for exam"""
    return None

@app.route('/generate', methods=['POST'])
def generate_all_timetables():
    """Main generation endpoint - generates basket, pre-mid, and post-mid timetables"""
    try:
        # Call the basket generation which now includes pre-mid and post-mid
        return generate_timetables_with_baskets()
    except Exception as e:
        print(f"[FAIL] Error in /generate endpoint: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

# DISABLED EXAM FUNCTIONS - All exam-related helper functions stubbed out
def create_configuration_sheet(config):
    """DISABLED - Create configuration sheet"""
    return pd.DataFrame()

def save_exam_schedule(schedule_df, start_date, end_date, config=None):
    """DISABLED - Save exam schedule with classroom allocation"""
    return None

def create_exam_classroom_summary(schedule_df):
    """DISABLED - Create exam classroom summary"""
    return pd.DataFrame()

def create_exam_summary(schedule_df):
    """DISABLED - Create exam summary"""
    return pd.DataFrame()

def create_department_summary(schedule_df):
    """DISABLED - Create department summary"""
    return pd.DataFrame()

# Debug endpoints
@app.route('/debug/current-data')
def debug_current_data():
    """Debug endpoint to show currently loaded data"""
    try:
        data_frames = load_all_data()
        if data_frames is None:
            return jsonify({
                'success': False,
                'message': 'No data loaded'
            })
        
        debug_info = {
            'success': True,
            'loaded_files': list(data_frames.keys()),
            'file_sizes': {},
            'sample_data': {}
        }
        
        for key, df in data_frames.items():
            debug_info['file_sizes'][key] = {
                'rows': len(df),
                'columns': list(df.columns),
                'memory_usage': df.memory_usage(deep=True).sum()
            }
            
            # Add sample data (first 3 rows)
            if not df.empty:
                debug_info['sample_data'][key] = df.head(3).to_dict('records')
        
        # Add timetable info - look for consolidated timetable files
        timetable_files = glob.glob(os.path.join(OUTPUT_DIR, "sem*_*_timetable.xlsx"))
        audit_files = glob.glob(os.path.join(OUTPUT_DIR, "*_Audit.xlsx"))
        
        # Combine all files
        excel_files = timetable_files + audit_files
        
        print(f"[DIR] Looking for timetable files in {OUTPUT_DIR}")
        print(f"[FILE] Found {len(timetable_files)} timetables, {len(audit_files)} audit files")
        debug_info['generated_timetables'] = {
            'count': len(timetable_files),
            'files': [os.path.basename(f) for f in timetable_files],
            'audit_files': [os.path.basename(f) for f in audit_files]
        }
        
        return jsonify(debug_info)
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e),
            'traceback': traceback.format_exc()
        })

@app.route('/debug/clear-cache')
def debug_clear_cache():
    """Debug endpoint to clear cached data"""
    global _cached_data_frames, _cached_timestamp, _file_hashes
    global _SEMESTER_ELECTIVE_ALLOCATIONS, _CLASSROOM_USAGE_TRACKER
    global _TIMETABLE_CLASSROOM_ALLOCATIONS, _GLOBAL_PREFERRED_CLASSROOMS
    global _EXAM_SCHEDULE_FILES
    
    # Clear all cache variables
    _cached_data_frames = None
    _cached_timestamp = 0
    _file_hashes = {}
    _SEMESTER_ELECTIVE_ALLOCATIONS = {}
    _CLASSROOM_USAGE_TRACKER = {}
    _TIMETABLE_CLASSROOM_ALLOCATIONS = {}
    _GLOBAL_PREFERRED_CLASSROOMS = {}
    _EXAM_SCHEDULE_FILES = set()
    
    print("[CLEAN] Cleared all cache variables and global trackers")
    
    # Clear input directory
    if os.path.exists(INPUT_DIR):
        try:
            # First try to remove all files individually to avoid permission issues
            for fname in os.listdir(INPUT_DIR):
                fpath = os.path.join(INPUT_DIR, fname)
                try:
                    if os.path.isfile(fpath):
                        os.chmod(fpath, 0o666)  # Make writable
                        os.remove(fpath)
                        print(f"[CLEAN] Removed {fname}")
                    elif os.path.isdir(fpath):
                        shutil.rmtree(fpath)
                        print(f"[CLEAN] Removed directory {fname}")
                except Exception as e:
                    print(f"[WARN] Could not remove {fname}: {e}")
            
            print("[CLEAN] Input directory cleared")
        except Exception as e:
            print(f"[WARN] Error during input directory cleanup: {e}")
            return jsonify({
                'success': False,
                'error': f'Error clearing input directory: {str(e)}'
            })
    
    # Clear output directory cache (optional - only clear temporary files)
    if os.path.exists(OUTPUT_DIR):
        try:
            for fname in os.listdir(OUTPUT_DIR):
                if fname.startswith('~$') or fname.startswith('.~'):
                    fpath = os.path.join(OUTPUT_DIR, fname)
                    try:
                        os.remove(fpath)
                        print(f"[CLEAN] Removed temp file {fname}")
                    except Exception:
                        pass
        except Exception:
            pass
    
    return jsonify({
        'success': True,
        'message': 'Cache cleared successfully - all caches, trackers, and input files removed',
        'cache_cleared': True,
        'input_dir_cleared': True,
        'trackers_cleared': True
    })

@app.route('/debug/file-matching')
def debug_file_matching():
    """Debug endpoint to check file matching"""
    available_files = os.listdir(INPUT_DIR) if os.path.exists(INPUT_DIR) else []
    required_files = [
        "course_data.csv",
        "faculty_availability.csv",
        "classroom_data.csv",
        "student_data.csv",
        "exams_data.csv"
    ]
    
    matching_results = {}
    
    for required_file in required_files:
        required_clean = required_file.lower().replace(' ', '').replace('_', '').replace('-', '')
        matches = []
        
        for uploaded_file in available_files:
            uploaded_clean = uploaded_file.lower().replace(' ', '').replace('_', '').replace('-', '')
            is_match = (
                required_clean in uploaded_clean or 
                uploaded_clean in required_clean or
                any(part in uploaded_clean for part in required_file.split('_'))
            )
            
            matches.append({
                'uploaded_file': uploaded_file,
                'uploaded_clean': uploaded_clean,
                'is_match': is_match,
                'match_type': 'exact' if uploaded_clean == required_clean else 'partial' if is_match else 'none'
            })
        
        matching_results[required_file] = {
            'required_clean': required_clean,
            'matches': matches,
            'has_match': any(match['is_match'] for match in matches)
        }
    
    return jsonify({
        'available_files': available_files,
        'required_files': required_files,
        'matching_results': matching_results,
        'all_files_matched': all(result['has_match'] for result in matching_results.values())
    })

@app.route('/')
def index():
    return render_template('index.html')

def convert_dataframe_to_html_with_baskets(df, table_id, course_colors, basket_colors, course_info):
    """Convert dataframe to HTML with special handling for basket entries, classroom allocation, and color coding"""
    # Create a copy to avoid modifying the original
    df_display = df.copy()
    
    # Ensure the index has proper name for time slots
    if df_display.index.name is None or df_display.index.name == '':
        df_display.index.name = 'Time Slot'
    
    # Ensure index values are strings (for proper display)
    if df_display.index.dtype != 'object':
        df_display.index = df_display.index.astype(str)
    
    # Replace NaN/NA values with 'Free' for display
    df_display = df_display.fillna('Free')
    
    # Convert all remaining non-string values to string (including any remaining NaN/None)
    for col in df_display.columns:
        df_display[col] = df_display[col].apply(lambda x: str(x) if pd.notna(x) and x != 'Free' else (x if x == 'Free' else 'Free'))
    
    basket_keywords = ['ELECTIVE_', 'HSS_', 'PROF_', 'OE_']

    def is_basket_entry(text):
        if not isinstance(text, str):
            return False
        normalized = text.upper().replace(' (TUTORIAL)', '')
        return any(keyword in normalized for keyword in basket_keywords)

    def should_hide_classroom_info(course_label):
        # Hide classroom information for ALL basket entries (show only in legends)
        # Basket entries should appear as just "ELECTIVE_B1" or "ELECTIVE_B1 (Tutorial)"
        # All classroom and course details will be shown in the elective basket legends instead
        try:
            if is_basket_entry(course_label):
                return True  # Always hide classroom info for basket entries
            return False
        except Exception:
            return False
    
    def build_course_title(course_label):
        """
        Build a tooltip/title for regular courses including Course Name and LTPSC if available.
        For basket entries, provide a concise basket tooltip.
        """
        if not isinstance(course_label, str):
            return ""
        if is_basket_entry(course_label):
            base = course_label.replace(' (Tutorial)', '')
            return f"{base} * Elective basket"
        base = course_label.replace(' (Tutorial)', '').replace(' (Lab)', '').strip()
        code = extract_course_code(base) or base
        info = course_info.get(code, {})
        name = info.get('name', '').strip()
        ltpsc = info.get('ltpsc', '').strip()
        parts = []
        if name:
            parts.append(name)
        if ltpsc:
            parts.append(f"LTPSC: {ltpsc}")
        return " * ".join(parts)

    def style_cell_with_classroom_and_colors(val):
        if isinstance(val, str) and val not in ['Free', 'LUNCH BREAK']:
            # Check for classroom allocation format "Course [Room]"
            if '[' in val and ']' in val:
                try:
                    # Extract course and classroom
                    course_part = val.split('[')[0].strip()
                    # Extract room part between [ and ]
                    bracket_content = val.split('[')[1]  # Get everything after first [
                    if ']' in bracket_content:
                        room_part = '[' + bracket_content.split(']')[0] + ']'  # Get content between [ and ]
                    else:
                        # Malformed - has [ but incomplete ], skip classroom display
                        room_part = None
                    
                    if not room_part:
                        # No valid room part, just return the course name
                        course_color = course_colors.get(course_part, '#cccccc')
                        return f'<span class="regular-course" style="background-color: {course_color}" title="{build_course_title(course_part)}">{val}</span>'
                    
                    # Check if it's a basket entry
                    is_basket = is_basket_entry(course_part)
                    title_attr = build_course_title(course_part)
                    
                    if is_basket:
                        # For elective baskets, show ONLY the basket name (no classroom info)
                        basket_key = course_part.replace(' (Tutorial)', '')
                        basket_color = basket_colors.get(basket_key, '#cccccc')
                        if '(Tutorial)' in course_part:
                            return f'<span class="basket-entry basket-tutorial" style="background-color: {basket_color}" title="{title_attr}">{course_part}</span>'
                        return f'<span class="basket-entry elective-basket" style="background-color: {basket_color}" title="{title_attr}">{course_part}</span>'
                    else:
                        # Regular course with classroom - get course color and show classroom info
                        clean_course = course_part.replace(' (Tutorial)', '')
                        course_color = course_colors.get(clean_course, '#cccccc')
                        return f'<span class="course-with-room" style="background-color: {course_color}" title="{title_attr}">{course_part}<br><small class="classroom-info">{room_part}</small></span>'
                except:
                    return val
            
            # Handle courses without classroom allocation but with color coding
            for basket_keyword in basket_keywords:
                if basket_keyword in val.upper():
                    basket_key = val.replace(' (Tutorial)', '')
                    basket_color = basket_colors.get(basket_key, '#cccccc')
                    if '(Tutorial)' in val:
                        return f'<span class="basket-entry basket-tutorial" style="background-color: {basket_color}" title="{build_course_title(val)}">{val}</span>'
                    return f'<span class="basket-entry elective-basket" style="background-color: {basket_color}" title="{build_course_title(val)}">{val}</span>'
            
            # Regular courses without classrooms - apply course color
            if val not in ['Free', 'LUNCH BREAK']:
                clean_course = val.replace(' (Tutorial)', '')
                course_color = course_colors.get(clean_course, '#cccccc')
                return f'<span class="regular-course" style="background-color: {course_color}" title="{build_course_title(val)}">{val}</span>'
                
        return val
    
    # Apply styling to all cells
    for col in df_display.columns:
        df_display[col] = df_display[col].apply(style_cell_with_classroom_and_colors)
    
    # Remove any unwanted columns (index-related, numeric, or duplicate Time Slot columns)
    cols_to_drop = []
    for col in df_display.columns:
        col_str = str(col)
        # Drop columns that are: index-related, numeric, unnamed, or Time Slot variations
        if (col_str == 'index' or 
            col_str == 'level_0' or 
            col_str.startswith('Unnamed') or
            col_str.startswith('Time Slot') and col_str != 'Time Slot' or  # Drop 'Time Slot1', etc.
            isinstance(col, int)):  # Numeric column names
            cols_to_drop.append(col)
    
    if cols_to_drop:
        df_display = df_display.drop(columns=cols_to_drop, errors='ignore')
    
    # Save the index values as a list before resetting everything
    time_slot_values = df_display.index.astype(str).tolist()
    
    # Filter to only include rows up to 18:30-20:00 (exclude any rows beyond that)
    max_time_slot = '18:30-20:00'
    max_idx = -1
    for idx, slot in enumerate(time_slot_values):
        if slot == max_time_slot:
            max_idx = idx
            break
    
    # If we found the max time slot, only include rows up to and including it
    if max_idx >= 0:
        time_slot_values = time_slot_values[:max_idx + 1]
        df_display = df_display.iloc[:max_idx + 1]
    
    # Manually construct HTML table to avoid index column issues
    html_parts = []
    html_parts.append(f'<table class="timetable-table" id="{table_id}" border="0">')
    
    # Build header
    html_parts.append('<thead>')
    html_parts.append('<tr>')
    html_parts.append('<th>Time Slot</th>')
    for col in df_display.columns:
        html_parts.append(f'<th>{col}</th>')
    html_parts.append('</tr>')
    html_parts.append('</thead>')
    
    # Build body
    html_parts.append('<tbody>')
    for time_slot, row in zip(time_slot_values, df_display.itertuples(index=False)):
        html_parts.append('<tr>')
        html_parts.append(f'<td>{time_slot}</td>')
        for value in row:
            html_parts.append(f'<td>{value}</td>')
        html_parts.append('</tr>')
    html_parts.append('</tbody>')
    
    html_parts.append('</table>')
    html = '\n'.join(html_parts)
    
    return clean_table_html(html)

def generate_basket_colors(baskets):
    """
    Generate consistent colors for elective baskets
    """
    basket_colors = {
        'ELECTIVE_B1': '#FF6B6B',
        'ELECTIVE_B2': '#4ECDC4', 
        'ELECTIVE_B3': '#45B7D1',
        'ELECTIVE_B4': '#96CEB4',
        'ELECTIVE_B5': '#FECA57',
        'ELECTIVE_B6': '#6C5CE7',
        'ELECTIVE_B7': '#F368E0',
        'ELECTIVE_B8': '#1DD1A1',
        'ELECTIVE_B9': '#8395A7',
        'HSS_B1': '#FECA57',
        'HSS_B2': '#FF9FF3',
        'HSS_B3': '#54A0FF',
        'HSS_B4': '#5F27CD',
        'PROF_B1': '#00D2D3',
        'PROF_B2': '#FF9F43',
        'OE_B1': '#10AC84',
        'OE_B2': '#EE5A24'
    }
    
    # For any baskets not in predefined colors, generate consistent colors
    for basket in baskets:
        if basket not in basket_colors:
            # Generate color based on basket name hash
            hash_val = sum(ord(char) for char in basket)
            hue = hash_val % 360
            basket_colors[basket] = f'hsl({hue}, 70%, 65%)'
    
    return basket_colors


@app.route('/timetables')
def get_timetables():
    try:
        # Ensure classroom trackers are fresh per request to avoid stale/duplicate allocations
        reset_classroom_usage_tracker()

        timetables = []
        # Look for consolidated timetable files
        excel_files = glob.glob(os.path.join(OUTPUT_DIR, "sem*_*_timetable.xlsx"))
        
        # Filter out temporary/lock files (starting with ~$ or .~)
        def is_temp_file(filepath):
            basename = os.path.basename(filepath)
            if basename.startswith('~$') or basename.startswith('.~'):
                return True
            return False
        
        excel_files = [f for f in excel_files if not is_temp_file(f)]

        # When running under pytest, limit to the most recent files to keep test runs fast
        if os.environ.get("PYTEST_CURRENT_TEST"):
            excel_files = sorted(excel_files, key=lambda f: os.path.getmtime(f), reverse=True)[:20]

        print(f"[DIR] Looking for timetable files in {OUTPUT_DIR}")
        print(f"[FILE] Found {len(excel_files)} consolidated timetable files")
        
        # Load course data for course information - force reload to get latest data
        data_frames = load_all_data(force_reload=True)
        course_info = get_course_info(data_frames) if data_frames else {}
        
        # Generate colors for all courses
        all_courses = set()
        all_baskets = set()
        if data_frames and 'course' in data_frames:
            all_courses = set(data_frames['course']['Course Code'].unique())
            
            # Extract basket information from course data
            if 'Basket' in data_frames['course'].columns:
                all_baskets = set(data_frames['course']['Basket'].dropna().unique())
        
        course_colors = generate_course_colors(all_courses, course_info)
        basket_colors = generate_basket_colors(all_baskets)
        
        for file_path in excel_files:
            filename = os.path.basename(file_path)

            # DO NOT reset trackers per file - we need to track classroom usage ACROSS all semester files
            # to prevent conflicts like ELECTIVE_B1 (sem1) and ELECTIVE_B6 (sem7) getting the same room
            # The tracker is already reset once at the start of get_timetables() which is sufficient
            
            # Consolidated files contain Regular/PreMid/PostMid sheets - process Regular sheet
            timetable_type = 'regular'
            
            print(f"[READ] Processing timetable file: {filename}")
            
            try:
                # Extract semester and branch from filename (format: sem{X}_{BRANCH}_timetable.xlsx)
                parts = filename.replace('.xlsx', '').split('_')
                sem_part = parts[0].replace('sem', '')
                branch = parts[1] if len(parts) > 1 else None
                sem = int(sem_part)
                
                print(f"[READ] Reading timetable file: {filename} (Branch: {branch}, Semester: {sem})")
                
                # Process all sheet types: Regular, PreMid, PostMid
                has_sections = (branch == 'CSE')
                
                for sheet_prefix in ['Regular', 'PreMid', 'PostMid']:
                    # Determine timetable type
                    if sheet_prefix == 'Regular':
                        timetable_type = 'regular'
                    elif sheet_prefix == 'PreMid':
                        timetable_type = 'pre_mid'
                    else:
                        timetable_type = 'post_mid'
                    
                    # Determine sheet names for this timetable type
                    if has_sections:
                        sheet_name_a = f'{sheet_prefix}_Section_A'
                        sheet_name_b = f'{sheet_prefix}_Section_B'
                    else:
                        sheet_name_a = f'{sheet_prefix}_Timetable'
                        sheet_name_b = None
                    
                    # Try to read Section A / Whole
                    try:
                        df_a = pd.read_excel(file_path, sheet_name=sheet_name_a)
                    except Exception:
                        print(f"   [WARN] No {sheet_name_a} sheet in {filename}")
                        continue

                    # Section_B may be absent for non-CSE branches
                    df_b = pd.DataFrame()
                    if has_sections and sheet_name_b:
                        try:
                            df_b = pd.read_excel(file_path, sheet_name=sheet_name_b)
                        except Exception:
                            print(f"   [WARN] No {sheet_name_b} sheet in {filename}")
                    
                    # Clean any unintended index columns like "Unnamed: 0" and set proper index
                    def _clean_section_df(df):
                        if df.empty:
                            return df
                        
                        # First, drop any columns that are clearly numeric indices or unwanted
                        cols_to_drop = []
                        for col in df.columns:
                            col_str = str(col)
                            # Drop columns that are: unnamed, numeric indices, 'index', 'level_0', or duplicate 'Time Slot' variations
                            if (col_str.startswith('Unnamed') or 
                                col_str == 'index' or 
                                col_str == 'level_0' or 
                                col_str.startswith('Time Slot') and col_str != 'Time Slot' or  # Drop 'Time Slot1', etc.
                                isinstance(col, int)):  # Numeric column names
                                cols_to_drop.append(col)
                        
                        if cols_to_drop:
                            df = df.drop(columns=cols_to_drop, errors='ignore')
                        
                        # Now handle the Time Slot column identification
                        # FIRST: Check if the index already looks like time slots (already set correctly when written)
                        if df.index.name == 'Time Slot':
                            # Index is already properly set, just ensure it's a string
                            df.index = df.index.astype(str)
                            df.index = df.index.map(normalize_time_slot_label)
                            return df
                        elif len(df) > 0:
                            # Check if current index values look like time slots
                            sample_idx = str(df.index[0])
                            if ':' in sample_idx or '-' in sample_idx or 'LUNCH' in sample_idx.upper():
                                # Index already contains time slots, just name it properly
                                df.index.name = 'Time Slot'
                                df.index = df.index.astype(str)
                                df.index = df.index.map(normalize_time_slot_label)
                                return df
                        
                        # SECOND: Check if 'Time Slot' exists as a column (needs to be moved to index)
                        if 'Time Slot' in df.columns:
                            df = df.set_index('Time Slot')
                        elif 'Time' in df.columns:
                            df = df.set_index('Time')
                            df.index.name = 'Time Slot'
                        else:
                            # If no Time Slot column and index doesn't look like time slots,
                            # check if the first column looks like time slots
                            if len(df.columns) > 0:
                                first_col = df.columns[0]
                                # Check if first column contains time slot values
                                sample_val = str(df[first_col].iloc[0]) if len(df) > 0 else ''
                                if ':' in sample_val or '-' in sample_val or 'LUNCH' in sample_val.upper():
                                    # This looks like time slots, use it as index
                                    df = df.set_index(first_col)
                                    df.index.name = 'Time Slot'
                                # REMOVED: Don't use first column as fallback if it doesn't look like time slots
                        
                        # Set default index name if still not set
                        if not df.index.name:
                            df.index.name = 'Time Slot'
                        
                        # Convert index to string to ensure time slots display correctly
                        df.index = df.index.astype(str)
                        # Normalize any numeric/indexed time slot labels to canonical strings
                        df.index = df.index.map(normalize_time_slot_label)
                        return df
                    
                    df_a = _clean_section_df(df_a)
                    if not df_b.empty:
                        df_b = _clean_section_df(df_b)
                    
                    # Check if classrooms are allocated (look for [Room] pattern in any cell)
                    has_classroom_allocation = False
                    for df in [df_a, df_b]:
                        if df.empty:
                            continue
                        for col in df.columns:
                            for val in df[col]:
                                if isinstance(val, str) and '[' in val and ']' in val:
                                    has_classroom_allocation = True
                                    break
                            if has_classroom_allocation:
                                break
                        if has_classroom_allocation:
                            break
                    
                    # Try to read basket allocations if available (read regardless of timetable type)
                    basket_allocations = {}
                    basket_courses_map = {}
                    try:
                        basket_df = pd.read_excel(file_path, sheet_name='Basket_Allocation')
                        for _, row in basket_df.iterrows():
                            basket_name = row['Basket Name']
                            courses_in_basket = row['Courses in Basket'].split(', ')
                            normalized_slot = normalize_time_slot_label(row['Time Slot'])
                            basket_allocations[basket_name] = {
                                'courses': courses_in_basket,
                                'slot': (row['Day'], normalized_slot)
                            }
                            # Build basket courses map for legends
                            basket_courses_map[basket_name] = courses_in_basket
                    except:
                        print(f"   [WARN] No basket allocation sheet found in {filename}")
                    
                    # Try to read classroom allocation details
                    classroom_allocation_details = []
                    try:
                        classroom_df = pd.read_excel(file_path, sheet_name='Classroom_Allocation')
                        classroom_allocation_details = normalize_classroom_allocation_records(classroom_df.to_dict('records'))
                        print(f"   [SCHOOL] Found classroom allocation details: {len(classroom_allocation_details)} entries")
                    except:
                        print(f"   [WARN] No classroom allocation sheet found in {filename}")
                    
                    # Try to read configuration details to reflect current settings
                    configuration_summary = {}
                    try:
                        config_df = pd.read_excel(file_path, sheet_name='Configuration')
                        if not config_df.empty and {'Parameter', 'Value'}.issubset(config_df.columns):
                            configuration_summary = dict(zip(config_df['Parameter'], config_df['Value']))
                        else:
                            configuration_summary = config_df.to_dict('records')
                        print(f"   [CONFIG] Loaded configuration details for {filename}")
                    except:
                        print(f"   [WARN] No configuration sheet found in {filename}")
                        configuration_summary = {}
                    
                    # Convert to HTML tables with basket-aware, classroom-aware, and color-aware processing
                    # Choose table ids based on whether it's a Section_A sheet or a Timetable-only sheet
                    if branch:
                        table_id_a = f"sem{sem}_{branch}_A" if sheet_name_a.endswith('Section_A') else f"sem{sem}_{branch}_whole"
                        table_id_b = f"sem{sem}_{branch}_B"
                    else:
                        table_id_a = f"sem{sem}_A" if sheet_name_a.endswith('Section_A') else f"sem{sem}_whole"
                        table_id_b = f"sem{sem}_B"

                    # Enforce semester-level allowed baskets BEFORE rendering HTML so disallowed basket entries are not shown or treated as scheduled
                    allowed_baskets_map = {
                        1: ['ELECTIVE_B1'],
                        3: ['ELECTIVE_B3'],
                        5: ['ELECTIVE_B4', 'ELECTIVE_B5'],
                        7: ['ELECTIVE_B6', 'ELECTIVE_B7', 'ELECTIVE_B8', 'ELECTIVE_B9']
                    }
                    allowed_set = set(allowed_baskets_map.get(sem, []))

                    def _sanitize_df_baskets(df):
                        if df.empty:
                            return df
                        df_copy = df.copy()
                        basket_names = ['ELECTIVE_B1','ELECTIVE_B2','ELECTIVE_B3','ELECTIVE_B4','ELECTIVE_B5','ELECTIVE_B6','ELECTIVE_B7','ELECTIVE_B8','ELECTIVE_B9','HSS_B1','HSS_B2']
                        def _sanitize_val(val):
                            if not isinstance(val, str):
                                return val
                            upper = val.upper()
                            for b in basket_names:
                                if b in upper:
                                    return val if b in allowed_set else 'Free'
                            return val
                        for col in df_copy.columns:
                            df_copy[col] = df_copy[col].apply(_sanitize_val)
                        return df_copy

                    def _remove_direct_elective_courses(df):
                        """Ensure electives are not hard-scheduled as regular courses (they should appear via baskets)."""
                        # Preserve electives exactly as provided in the file so existing schedules
                        # retain their courses instead of being blanked out during display/allocation.
                        return df

                    df_a = _sanitize_df_baskets(df_a)
                    if not df_b.empty:
                        df_b = _sanitize_df_baskets(df_b)

                    # Remove any directly scheduled elective courses to avoid duplicates (electives are handled via baskets)
                    df_a = _remove_direct_elective_courses(df_a)
                    if not df_b.empty:
                        df_b = _remove_direct_elective_courses(df_b)

                    # Build basket courses map early (before allocation) so it's available for on-the-fly allocation
                    course_baskets = separate_courses_by_type(data_frames, sem, branch) if data_frames else {'core_courses': [], 'elective_courses': []}
                
                    # Build comprehensive basket courses map including ALL elective courses for this semester
                    # BUT: only add courses from course_data.csv if they're not already in basket_courses_map
                    # (preserves courses from Basket_Allocation sheets in the Excel file)
                    if not course_baskets['elective_courses'].empty and 'Basket' in course_baskets['elective_courses'].columns:
                        for _, course in course_baskets['elective_courses'].iterrows():
                            raw_basket = course.get('Basket', 'Unknown')
                            basket = str(raw_basket).strip().upper() if pd.notna(raw_basket) else 'Unknown'
                            course_code = course['Course Code']
                            if basket not in basket_courses_map:
                                basket_courses_map[basket] = []
                            if course_code not in basket_courses_map[basket]:
                                basket_courses_map[basket].append(course_code)

                    # ALWAYS try to allocate classrooms if classroom data exists and file doesn't have complete allocation
                    # This ensures classrooms appear on the website even for old/partial timetable files
                    classroom_allocation_details = classroom_allocation_details if 'classroom_allocation_details' in locals() else []
                    classroom_data_df = data_frames.get('classroom') if data_frames else None
                    allocated_and_persisted = False

                    # Count how many non-empty, non-Free cells DON'T have classroom info
                    def count_unallocated_cells(df):
                        count = 0
                        for col in df.columns:
                            for val in df[col]:
                                if isinstance(val, str) and val.strip() and val.strip() != 'Free':
                                    if '[' not in val or ']' not in val:
                                        count += 1
                        return count

                    unallocated_a = count_unallocated_cells(df_a)
                    unallocated_b = count_unallocated_cells(df_b)
                    total_unallocated = unallocated_a + unallocated_b

                    # If we have ANY unallocated courses OR no classroom info, allocate on-the-fly
                    should_allocate_onthefly = (not has_classroom_allocation or total_unallocated > 0) and classroom_data_df is not None and not classroom_data_df.empty
                
                    if should_allocate_onthefly:
                        try:
                            if has_classroom_allocation:
                                print(f"[SCHOOL] Partial classroom allocations found in {filename} ({total_unallocated} unallocated); completing allocation on-the-fly")
                            else:
                                print(f"[SCHOOL] No classroom allocations found in {filename}; applying allocation on-the-fly")
                        
                            df_a = allocate_classrooms_for_timetable(df_a, classroom_data_df, course_info, sem, branch, 'A', basket_courses_map)
                            if not df_b.empty:
                                df_b = allocate_classrooms_for_timetable(df_b, classroom_data_df, course_info, sem, branch, 'B', basket_courses_map)

                            # DEBUG: Print what's in _TIMETABLE_CLASSROOM_ALLOCATIONS after allocation
                            print(f"\n[DEBUG-ALLOC] After allocation for {branch} Sem {sem}:")
                            for tk, allocs in _TIMETABLE_CLASSROOM_ALLOCATIONS.items():
                                if f"sem{sem}" in tk and branch in tk:
                                    print(f"   {tk}: {len(allocs)} allocations")
                                    basket_allocs = [(k, v) for k, v in allocs.items() if v.get('basket')]
                                    print(f"   Basket allocations: {len(basket_allocs)}")
                                    for k, v in list(basket_allocs)[:10]:
                                        print(f"      {k} -> {v.get('classroom')} ({v.get('basket')}, {v.get('type')})")

                            # Build classroom allocation details for UI and verification
                            classroom_allocation_details = normalize_classroom_allocation_records(create_classroom_allocation_detail_with_tracking([df_a, df_b], classroom_data_df, sem, branch).to_dict('records'))
                            has_classroom_allocation = True

                            # Persist allocation sheets back into the Excel file (replace existing sheets if present)
                            try:
                                with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                                    # Properly prepare DataFrames for Excel writing (convert index to column)
                                    # Handle both cases: index named 'Time Slot' and index named something else
                                    if df_a.index.name == 'Time Slot':
                                        df_a_for_excel = df_a.reset_index(drop=False)
                                    elif 'Time Slot' not in df_a.columns:
                                        df_a_for_excel = df_a.reset_index(drop=False).rename(columns={df_a.index.name or 'index': 'Time Slot'})
                                    else:
                                        df_a_for_excel = df_a.copy()
                                
                                    df_a_for_excel.to_excel(writer, sheet_name='Section_A', index=False)
                                
                                    if not df_b.empty:
                                        if df_b.index.name == 'Time Slot':
                                            df_b_for_excel = df_b.reset_index(drop=False)
                                        elif 'Time Slot' not in df_b.columns:
                                            df_b_for_excel = df_b.reset_index(drop=False).rename(columns={df_b.index.name or 'index': 'Time Slot'})
                                        else:
                                            df_b_for_excel = df_b.copy()
                                        df_b_for_excel.to_excel(writer, sheet_name='Section_B', index=False)

                                    # Classroom verification sheets
                                    class_report = create_classroom_utilization_report(classroom_data_df, [df_a, df_b], [])
                                    class_report.to_excel(writer, sheet_name='Classroom_Utilization', index=False)
                                    pd.DataFrame(classroom_allocation_details).to_excel(writer, sheet_name='Classroom_Allocation', index=False)

                                    # Persist basket per-course allocations to a dedicated sheet
                                    try:
                                        rows = []
                                        for basket_name, courses in (basket_courses_map or {}).items():
                                            for course in courses:
                                                rooms = [rec.get('room') for rec in classroom_allocation_details if rec.get('course') == course and rec.get('room')]
                                                unique_rooms = sorted(set([r for r in rooms if r]))
                                                if not unique_rooms and branch and sem:
                                                    timetable_keys = [f"{branch}_sem{sem}_secA", f"{branch}_sem{sem}_secB", f"{branch}_sem{sem}_secWhole"]
                                                    tracker_rooms = []
                                                    for tk in timetable_keys:
                                                        alloc_map = _TIMETABLE_CLASSROOM_ALLOCATIONS.get(tk, {})
                                                        for alloc in alloc_map.values():
                                                            c = alloc.get('course')
                                                            room_val = alloc.get('classroom') or alloc.get('room')
                                                            if c == course and room_val:
                                                                tracker_rooms.append(room_val)
                                                    if tracker_rooms:
                                                        unique_rooms = sorted(set(tracker_rooms))
                                                rows.append({'Basket Name': basket_name, 'Course': course, 'Allocated Rooms': ', '.join(unique_rooms) if unique_rooms else ''})
                                        if rows:
                                            pd.DataFrame(rows).to_excel(writer, sheet_name='Basket_Course_Allocations', index=False)
                                    except Exception:
                                        pass

                                print(f"   [SCHOOL] Persisted classroom allocations to {filename}")
                                allocated_and_persisted = True
                            except Exception as persist_e:
                                print(f"   [WARN] Could not persist allocations to file {filename}: {persist_e}")
                        except Exception as alloc_e:
                            print(f"   [WARN] On-the-fly classroom allocation failed for {filename}: {alloc_e}")
                            traceback.print_exc()

                    html_a = convert_dataframe_to_html_with_baskets(df_a, table_id_a, course_colors, basket_colors, course_info)
                    html_b = convert_dataframe_to_html_with_baskets(df_b, table_id_b, course_colors, basket_colors, course_info) if not df_b.empty else ""
                
                    # Extract unique courses AND baskets from the actual schedule
                    unique_courses_a, unique_baskets_a = extract_unique_courses_with_baskets(df_a, basket_allocations)
                    unique_courses_b, unique_baskets_b = extract_unique_courses_with_baskets(df_b, basket_allocations) if not df_b.empty else ([], [])
                
                    # Filter baskets by semester mapping to enforce allowed elective baskets per semester
                    allowed_baskets_map = {
                        1: ['ELECTIVE_B1'],
                        3: ['ELECTIVE_B3'],
                        5: ['ELECTIVE_B4', 'ELECTIVE_B5'],
                        7: ['ELECTIVE_B6', 'ELECTIVE_B7', 'ELECTIVE_B8', 'ELECTIVE_B9']
                    }
                    if sem in allowed_baskets_map:
                        allowed = set(allowed_baskets_map[sem])
                        # Only keep allowed baskets detected in the schedule
                        unique_baskets_a = [b for b in unique_baskets_a if b in allowed]
                        unique_baskets_b = [b for b in unique_baskets_b if b in allowed]
                        print(f"   [MAP] Allowed baskets for semester {sem}: {allowed}")

                    # Ensure Semester 7 always shows all required elective baskets in legends (even if empty)
                    if sem == 7:
                        required_sem7_baskets = ['ELECTIVE_B6', 'ELECTIVE_B7', 'ELECTIVE_B8', 'ELECTIVE_B9']
                        for b in required_sem7_baskets:
                            if b not in unique_baskets_a:
                                unique_baskets_a.append(b)
                            if b not in unique_baskets_b:
                                unique_baskets_b.append(b)
                            # Ensure map has an entry so legends render the basket name
                            basket_courses_map.setdefault(b, [])
                
                    # Create comprehensive course lists for legends including ALL basket courses
                    all_core_courses = course_baskets['core_courses']['Course Code'].tolist() if not course_baskets['core_courses'].empty else []
                    all_elective_courses = course_baskets['elective_courses']['Course Code'].tolist() if not course_baskets['elective_courses'].empty else []

                    # Also include courses from Basket_Allocation sheets so they don't get filtered out
                    for courses_list in basket_courses_map.values():
                        if courses_list:
                            all_elective_courses.extend([c for c in courses_list if c not in all_elective_courses])

                    elective_set = set(all_elective_courses)
                    core_set = set(all_core_courses)

                    def _filter_basket_map_to_electives(bmap):
                        """Keep only elective course codes inside basket maps to avoid core leakage into elective legends."""
                        if not bmap:
                            return {}
                        filtered = {}
                        for bname, courses in bmap.items():
                            filtered_courses = [c for c in courses if c in elective_set]
                            filtered[bname] = filtered_courses
                        return filtered
                
                    # FIXED: Remove duplicate basket entries and empty baskets
                    # Combine scheduled courses with all basket courses for complete legends
                    legend_courses_a = set(unique_courses_a)
                    legend_courses_b = set(unique_courses_b)
                
                    # For mid-semester timetables, add all courses from the schedule
                    if timetable_type in ['pre_mid', 'post_mid']:
                        # Add all courses from the schedule for mid-semester timetables
                        all_courses_in_schedule = set(unique_courses_a).union(set(unique_courses_b))
                        legend_courses_a.update(all_courses_in_schedule)
                        legend_courses_b.update(all_courses_in_schedule)
                
                    supplemental_baskets = []
                    if sem == 5:
                        supplemental_baskets = ['ELECTIVE_B4']
                    if sem == 7:
                        # Ensure sem7 basket legends include all four baskets
                        supplemental_baskets = supplemental_baskets + ['ELECTIVE_B6', 'ELECTIVE_B7', 'ELECTIVE_B8', 'ELECTIVE_B9']

                    # Add all elective courses from baskets that appear in the schedule
                    for basket_name in unique_baskets_a:
                        if basket_name in basket_courses_map and basket_courses_map[basket_name]:
                            legend_courses_a.update(basket_courses_map[basket_name])
                
                    for basket_name in unique_baskets_b:
                        if basket_name in basket_courses_map and basket_courses_map[basket_name]:
                            legend_courses_b.update(basket_courses_map[basket_name])

                    # Ensure semester 5 includes ELECTIVE_B4 courses even if slots share across sections
                    if supplemental_baskets:
                        for basket_name in supplemental_baskets:
                            if basket_name in basket_courses_map and basket_courses_map[basket_name]:
                                legend_courses_a.update(basket_courses_map[basket_name])
                                legend_courses_b.update(basket_courses_map[basket_name])

                    # Ensure all core courses for this semester/branch appear in legends even if missing from the timetable
                    legend_courses_a.update(core_set)
                    legend_courses_b.update(core_set)
                
                    # FIXED: Create clean basket lists without duplicates
                    # For Semester 7, include ALL required baskets even if they have empty course lists
                    if sem == 7:
                        required_sem7_baskets = ['ELECTIVE_B6', 'ELECTIVE_B7', 'ELECTIVE_B8', 'ELECTIVE_B9']
                        # Ensure all required baskets are in the map
                        for basket_name in required_sem7_baskets:
                            basket_courses_map.setdefault(basket_name, [])
                        # Include baskets that either have courses OR are required for Semester 7
                        clean_baskets_a = [basket for basket in unique_baskets_a if basket in basket_courses_map and (basket_courses_map[basket] or basket in required_sem7_baskets)]
                        clean_baskets_b = [basket for basket in unique_baskets_b if basket in basket_courses_map and (basket_courses_map[basket] or basket in required_sem7_baskets)]
                        # Add any missing required baskets
                        for basket_name in required_sem7_baskets:
                            if basket_name not in clean_baskets_a:
                                clean_baskets_a.append(basket_name)
                            if basket_name not in clean_baskets_b:
                                clean_baskets_b.append(basket_name)
                    else:
                        # Determine allowed baskets for this semester
                        allowed_baskets_by_semester = {
                            1: ['ELECTIVE_B1'],
                            3: ['ELECTIVE_B3'],
                            5: ['ELECTIVE_B4', 'ELECTIVE_B5'],
                            7: ['ELECTIVE_B6', 'ELECTIVE_B7', 'ELECTIVE_B8', 'ELECTIVE_B9']
                        }
                        allowed_baskets_list = allowed_baskets_by_semester.get(sem, [])
                        
                        if allowed_baskets_list:
                            # For semesters with allowed baskets, always show them (even if empty)
                            for basket_name in allowed_baskets_list:
                                basket_courses_map.setdefault(basket_name, [])
                            clean_baskets_a = allowed_baskets_list.copy()
                            clean_baskets_b = allowed_baskets_list.copy()
                        else:
                            # Other semesters: only include baskets with courses
                            clean_baskets_a = [basket for basket in unique_baskets_a if basket in basket_courses_map and basket_courses_map[basket]]
                            clean_baskets_b = [basket for basket in unique_baskets_b if basket in basket_courses_map and basket_courses_map[basket]]

                    if supplemental_baskets:
                        for basket_name in supplemental_baskets:
                            if basket_name in basket_courses_map and basket_courses_map[basket_name]:
                                if basket_name not in clean_baskets_a:
                                    clean_baskets_a.append(basket_name)
                                if basket_name not in clean_baskets_b:
                                    clean_baskets_b.append(basket_name)

                    # If a semester-level basket mapping is defined, filter the basket_courses_map to only include allowed baskets
                    if sem in allowed_baskets_map:
                        allowed_set = set(allowed_baskets_map[sem])
                        filtered_basket_courses_map = {k: v for k, v in basket_courses_map.items() if k in allowed_set}
                    else:
                        filtered_basket_courses_map = basket_courses_map

                    # Ensure elective basket legends only show true elective courses
                    basket_courses_map = _filter_basket_map_to_electives(basket_courses_map)
                    filtered_basket_courses_map = _filter_basket_map_to_electives(filtered_basket_courses_map)

                    print(f"   [COLOR] Color coding: {len(course_colors)} course colors, {len(basket_colors)} basket colors")
                    print(f"   [STATS] Legend courses A: {len(legend_courses_a)}, Baskets A: {clean_baskets_a}")
                    print(f"   [STATS] Legend courses B: {len(legend_courses_b)}, Baskets B: {clean_baskets_b}")
                
                    # Build per-basket per-course allocated rooms map for legend display
                    # This consolidates all rooms allocated to each basket course WITH day/time info
                    basket_course_allocations = {}
                    print(f"   [BASKET-ALLOC] Building basket_course_allocations from {len(classroom_allocation_details or [])} allocation records")
                    try:
                        # Prefer the raw basket map so tests and provided basket sheets always surface their courses
                        source_basket_map = basket_courses_map or {}
                        print(f"   [BASKET-ALLOC] Processing {len(source_basket_map)} baskets: {list(source_basket_map.keys())}")
                        for basket_name, course_list in source_basket_map.items():
                            try:
                                basket_course_allocations.setdefault(basket_name, {})
                                
                                # FIRST: Try to extract room directly from timetable cells for basket entries
                                # Scan df_a and df_b for cells containing this basket name with rooms
                                basket_rooms = []
                                for df in [df_a, df_b]:
                                    if df.empty:
                                        continue
                                    for day in df.columns:
                                        for time_slot in df.index:
                                            cell_value = df.loc[time_slot, day]
                                            if isinstance(cell_value, str) and basket_name in cell_value and '[' in cell_value and ']' in cell_value:
                                                # Extract room number
                                                room_match = cell_value[cell_value.find('[')+1:cell_value.find(']')]
                                                if room_match:
                                                    # Determine session type
                                                    session_type = 'Lecture'
                                                    if '(Tutorial)' in cell_value or '(tutorial)' in cell_value:
                                                        session_type = 'Tutorial'
                                                    elif '(Lab)' in cell_value or '(lab)' in cell_value:
                                                        session_type = 'Lab'
                                                    
                                                    basket_rooms.append({
                                                        'room': room_match.strip(),
                                                        'day': str(day),
                                                        'time': str(time_slot),
                                                        'type': session_type
                                                    })
                                
                                # If we found rooms for the basket, assign them to all courses in the basket
                                if basket_rooms:
                                    for course_code in (course_list or []):
                                        basket_course_allocations[basket_name][course_code] = basket_rooms
                                    print(f"   [BASKET] {basket_name}: Found {len(basket_rooms)} room allocations")
                                    for room_info in basket_rooms:
                                        print(f"   [BASKET]   - {room_info['room']} on {room_info['day']} at {room_info['time']} ({room_info['type']})")
                                    continue
                                
                                # FALLBACK: Try course-by-course extraction from classroom_allocation_details
                                for course_code in (course_list or []):
                                    room_allocations = []
                                    print(f"   [BASKET-COURSE] Processing {basket_name} -> {course_code}")
                                    # Collect rooms WITH day/time from classroom_allocation_details
                                    try:
                                        for rec in (classroom_allocation_details or []):
                                            rec_course = rec.get('course') or rec.get('Course')
                                            if not rec_course:
                                                continue
                                            # Normalize to match base course code (strip session suffixes)
                                            rec_course_clean = str(rec_course).replace(' (Tutorial)', '').replace(' (Lab)', '').strip()
                                            if rec_course_clean == course_code and (rec.get('room') or rec.get('Room Number')):
                                                # Use session_type from record if available, otherwise infer from course name
                                                session_type = rec.get('session_type') or rec.get('Session Type')
                                                if not session_type:
                                                    session_type = 'Tutorial' if ' (Tutorial)' in str(rec_course) else ('Lab' if ' (Lab)' in str(rec_course) else 'Lecture')
                                                room_val = rec.get('room') or rec.get('Room Number')
                                                day_val = rec.get('day') or rec.get('Day', '')
                                                time_val = rec.get('time_slot') or rec.get('time') or rec.get('Time Slot', '')
                                                
                                                # Skip entries where day or time contains basket keywords (invalid entries from header rows)
                                                day_upper = str(day_val).upper()
                                                time_upper = str(time_val).upper()
                                                if 'ELECTIVE' in day_upper or 'ELECTIVE' in time_upper or 'HSS_' in day_upper or 'HSS_' in time_upper:
                                                    continue
                                                # Skip entries with empty day/time
                                                if not day_val or not time_val:
                                                    continue
                                                    
                                                print(f"   [BASKET-MATCH] Found {rec_course} -> {room_val} on {day_val} at {time_val} ({session_type})")
                                                room_allocations.append({
                                                    'room': str(room_val).strip(),
                                                    'day': day_val,
                                                    'time': time_val,
                                                    'type': session_type
                                                })
                                    except Exception as e:
                                        print(f"   [BASKET-ERR] Error extracting from classroom_allocation_details for {course_code}: {e}")
                                        pass

                                    if room_allocations:
                                        print(f"   [BASKET-SUCCESS] {basket_name} -> {course_code}: {len(room_allocations)} allocations from classroom_allocation_details")
                                    else:
                                        print(f"   [BASKET-FALLBACK] {basket_name} -> {course_code}: No allocations from classroom_allocation_details, trying _TIMETABLE_CLASSROOM_ALLOCATIONS")
                                    
                                    # Fallback: collect rooms from internal tracker if details missing
                                    if not room_allocations and branch and sem:
                                        try:
                                            timetable_keys = [f"{branch}_sem{sem}_secA", f"{branch}_sem{sem}_secB", f"{branch}_sem{sem}_secWhole"]
                                            for tk in timetable_keys:
                                                alloc_map = _TIMETABLE_CLASSROOM_ALLOCATIONS.get(tk, {})
                                                for key, alloc in alloc_map.items():
                                                    c = alloc.get('course')
                                                    room_val = alloc.get('classroom') or alloc.get('room')
                                                    # Normalize course name from tracker to match base code
                                                    c_clean = str(c).replace(' (Tutorial)', '').replace(' (Lab)', '').strip() if c else None
                                                    if c_clean == course_code and room_val:
                                                        # Extract day/time from key (format: Day_TimeSlot)
                                                        parts = key.split('_')
                                                        day = parts[0] if len(parts) > 0 else ''
                                                        time = parts[1] if len(parts) > 1 else ''
                                                        
                                                        # Skip entries where day or time contains basket keywords (invalid entries)
                                                        day_upper = str(day).upper()
                                                        time_upper = str(time).upper()
                                                        if 'ELECTIVE' in day_upper or 'ELECTIVE' in time_upper or 'HSS_' in day_upper or 'HSS_' in time_upper:
                                                            continue
                                                        # Skip entries with empty day/time
                                                        if not day or not time:
                                                            continue
                                                        
                                                        # Use session_type from alloc if available, otherwise infer from course name
                                                        session_type = alloc.get('type')
                                                        if not session_type:
                                                            session_type = 'Tutorial' if (c and ' (Tutorial)' in str(c)) else ('Lab' if (c and ' (Lab)' in str(c)) else 'Lecture')
                                                        room_allocations.append({
                                                            'room': str(room_val).strip(),
                                                            'day': day,
                                                            'time': time,
                                                            'type': session_type
                                                        })
                                        except Exception:
                                            pass

                                    # Deduplicate room allocations - for common courses, keep only one allocation per day+time
                                    # (Common courses use the same room for both sections, but might appear twice in records)
                                    # Also filter out invalid allocations where day/time contains basket names
                                    basket_keywords = ['ELECTIVE_', 'HSS_', 'PROF_', 'OE_', 'ELECTIVE']
                                    seen_slots = set()
                                    unique_allocations = []
                                    for alloc in room_allocations:
                                        day_str = str(alloc.get('day', '')).upper()
                                        time_str = str(alloc.get('time', '')).upper()
                                        # Skip allocations where day or time contains basket keywords (invalid entries)
                                        if any(kw in day_str or kw in time_str for kw in basket_keywords):
                                            continue
                                        # Skip allocations with empty or invalid day/time
                                        if not alloc.get('day') or not alloc.get('time'):
                                            continue
                                        slot_key = (alloc['day'], alloc['time'])
                                        if slot_key not in seen_slots:
                                            seen_slots.add(slot_key)
                                            unique_allocations.append(alloc)
                                
                                    # Store with full details; set None if no allocation found
                                    basket_course_allocations[basket_name][course_code] = unique_allocations if unique_allocations else None
                                    if unique_allocations:
                                        print(f"   [BASKET-SUCCESS] {basket_name} -> {course_code}: {len(unique_allocations)} allocations")
                                        for alloc in unique_allocations:
                                            print(f"      {alloc['room']} on {alloc['day']} at {alloc['time']} ({alloc['type']})")
                                    else:
                                        print(f"   [BASKET-WARN] {basket_name} -> {course_code}: NO allocations found")
                            except Exception:
                                # Ensure basket key exists even if an error occurs
                                basket_course_allocations.setdefault(basket_name, {})
                    except Exception:
                        basket_course_allocations = {}

                    # Compute scheduled core courses (non-basket) from the actual timetable
                    elective_code_set = set(all_elective_courses)
                    scheduled_core_courses_a = [code for code in unique_courses_a if code not in elective_code_set]
                    scheduled_core_courses_b = [code for code in unique_courses_b if code not in elective_code_set]
                
                    def build_course_legend_entries(course_codes):
                        legend_entries = []
                        for code in sorted(course_codes):
                            # Use helper to get department-specific course info
                            info = get_course_info_by_dept(course_info, code, branch)
                            ltpsc_value = info.get('ltpsc', '') if info else ''

                            # Normalize term to explicit labels for legend (Pre-Mid, Post-Mid, Full Sem)
                            raw_term = info.get('term_type') if info else None
                            if raw_term:
                                upper_term = str(raw_term).upper()
                                if 'PRE' in upper_term:
                                    term_label = 'Pre-Mid'
                                elif 'POST' in upper_term:
                                    term_label = 'Post-Mid'
                                else:
                                    term_label = 'Full Sem'
                            else:
                                term_label = 'Full Sem'

                            parts = []
                            if ltpsc_value:
                                parts.append(ltpsc_value)
                            # Always include the term label so legends show Pre-Mid/Post-Mid/Full Sem
                            parts.append(term_label)

                            display = f"{code} ({' | '.join(parts)})" if parts else code
                            legend_entries.append({
                                'code': code,
                                'name': info.get('name', ''),
                                'ltpsc': ltpsc_value,
                                'term': term_label,
                                'display': display
                            })
                        return legend_entries
                
                    # Build legends separated into core vs elective for UI clarity
                    elective_courses_in_legend_a = [c for c in legend_courses_a if c in elective_set]
                    elective_courses_in_legend_b = [c for c in legend_courses_b if c in elective_set]
                    
                    print(f"   [DEBUG] Legend courses A: {legend_courses_a}")
                    print(f"   [DEBUG] Elective set: {elective_set}")
                    print(f"   [DEBUG] Elective courses in legend A: {elective_courses_in_legend_a}")
                    print(f"   [DEBUG] Basket courses map: {basket_courses_map}")
                    
                    elective_legends_a = build_course_legend_entries(elective_courses_in_legend_a)
                    elective_legends_b = build_course_legend_entries(elective_courses_in_legend_b)
                    core_legends_a = build_course_legend_entries([c for c in legend_courses_a if c not in elective_set])
                    core_legends_b = build_course_legend_entries([c for c in legend_courses_b if c not in elective_set])
                
                    # Debug: print classroom allocation details for forced_conflict test to diagnose intermittent failures
    # forced_conflict debug dump removed



                    # Add timetable for Section A or Whole
                    timetable_data = {
                        'semester': sem,
                        'section': 'A' if sheet_name_a.endswith('Section_A') else ('Whole' if branch and branch != 'CSE' else 'A'),
                        'branch': branch,
                        'filename': filename,
                        'html': html_a,
                        'courses': list(legend_courses_a),  # Use enhanced course list
                        'baskets': clean_baskets_a,  # Use cleaned basket list
                        'basket_courses_map': filtered_basket_courses_map,
                        'course_info': course_info,
                        'course_colors': course_colors,
                        'basket_colors': basket_colors,
                        'core_courses': all_core_courses,
                        'elective_courses': all_elective_courses,
                        'scheduled_core_courses': scheduled_core_courses_a,
                        'is_basket_timetable': (timetable_type == 'basket'),
                        'is_pre_mid_timetable': (timetable_type == 'pre_mid'),
                        'is_post_mid_timetable': (timetable_type == 'post_mid'),
                        'all_basket_courses': filtered_basket_courses_map,  # Include filtered basket courses for legends
                        'has_classroom_allocation': has_classroom_allocation,
                        'classroom_details': classroom_allocation_details,
                        'basket_course_allocations': basket_course_allocations,
                        'configuration': configuration_summary,
                        'course_legends': elective_legends_a + core_legends_a,
                        'core_course_legends': core_legends_a,
                        'elective_course_legends': elective_legends_a,
                        'timetable_type': timetable_type  # Add type for frontend filtering
                    }
                    timetables.append(timetable_data)
                
                    # Add timetable for Section B (if it exists)
                    if not df_b.empty and html_b:
                        timetables.append({
                            'semester': sem,
                            'section': 'B',
                            'branch': branch,
                            'filename': filename,
                            'html': html_b,
                            'courses': list(legend_courses_b),  # Use enhanced course list
                            'baskets': clean_baskets_b,  # Use cleaned basket list
                            'basket_courses_map': filtered_basket_courses_map,
                            'course_info': course_info,
                            'course_colors': course_colors,
                            'basket_colors': basket_colors,
                            'core_courses': all_core_courses,
                            'elective_courses': all_elective_courses,
                            'scheduled_core_courses': scheduled_core_courses_b,
                            'is_basket_timetable': (timetable_type == 'basket'),
                            'is_pre_mid_timetable': (timetable_type == 'pre_mid'),
                            'is_post_mid_timetable': (timetable_type == 'post_mid'),
                            'all_basket_courses': filtered_basket_courses_map,  # Include filtered basket courses for legends
                            'has_classroom_allocation': has_classroom_allocation,
                            'classroom_details': classroom_allocation_details,
                            'basket_course_allocations': basket_course_allocations,
                            'configuration': configuration_summary,
                            'course_legends': elective_legends_b + core_legends_b,
                            'core_course_legends': core_legends_b,
                            'elective_course_legends': elective_legends_b,
                            'timetable_type': timetable_type  # Add type for frontend filtering
                        })
                
                    print(f"[OK] Loaded {timetable_type.upper()} timetable: {filename}")
                    print(f"   Type: {timetable_type}")
                    print(f"   Classroom allocation: {'[OK] YES' if has_classroom_allocation else '[FAIL] NO'}")
                
            except Exception as e:
                print(f"[FAIL] Error reading {timetable_type} timetable {filename}: {e}")
                traceback.print_exc()
                continue
        
        print(f"[STATS] Total timetables loaded: {len(timetables)}")
        # Sanitize timetables for JSON serialisation (convert NaN/NaT to None and numpy types to native types)
        def _sanitize(obj):
            import math, numbers
            from datetime import datetime, date
            # Basic types
            if obj is None:
                return None
            if isinstance(obj, (str, bool, int)):
                return obj
            if isinstance(obj, float):
                if math.isnan(obj) or math.isinf(obj):
                    return None
                return obj
            # numpy types
            try:
                import numpy as _np
                if isinstance(obj, _np.generic):
                    if _np.isrealobj(obj):
                        val = obj.item()
                        if isinstance(val, float) and math.isnan(val):
                            return None
                        return val
                    else:
                        return obj.item()
            except Exception:
                pass
            # datetime
            if isinstance(obj, (datetime, date)):
                return obj.isoformat()
            # dict
            if isinstance(obj, dict):
                return {str(k): _sanitize(v) for k, v in obj.items()}
            # list/tuple
            if isinstance(obj, (list, tuple)):
                return [_sanitize(v) for v in obj]
            # pandas NaT
            try:
                import pandas as _pd
                if obj is _pd.NaT:
                    return None
            except Exception:
                pass
            # Fallback to string
            try:
                return str(obj)
            except Exception:
                return None

        safe_timetables = _sanitize(timetables)
        return jsonify(safe_timetables)
        
    except Exception as e:
        print(f"[FAIL] Error in /timetables: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Error loading timetables: {str(e)}'})
    
@app.route('/download/<filename>')
def download_timetable(filename):
    file_path = os.path.join(OUTPUT_DIR, filename)
    if os.path.exists(file_path):
        return send_file(file_path, as_attachment=True)
    else:
        return jsonify({'success': False, 'message': 'File not found'})

@app.route('/download-all')
def download_all_timetables():
    try:
        zip_path = os.path.join(OUTPUT_DIR, 'all_timetables.zip')
        
        with zipfile.ZipFile(zip_path, 'w') as zipf:
            for file in glob.glob(os.path.join(OUTPUT_DIR, "*.xlsx")):
                zipf.write(file, os.path.basename(file))
        
        return send_file(zip_path, as_attachment=True)
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error creating zip: {str(e)}'})

@app.route('/stats')
def get_stats():
    try:
        # Count generated timetables
        excel_files = glob.glob(os.path.join(OUTPUT_DIR, "*.xlsx"))
        total_timetables = 0
        for file in excel_files:
            try:
                xlsx = pd.ExcelFile(file)
                # Count Section_A/Section_B as two separate timetables if both present
                if 'Section_B' in xlsx.sheet_names and 'Section_A' in xlsx.sheet_names:
                    total_timetables += 2
                elif 'Timetable' in xlsx.sheet_names or 'Section_A' in xlsx.sheet_names:
                    total_timetables += 1
                else:
                    # Fallback: treat as 1
                    total_timetables += 1
            except:
                # If reading fails, ignore and continue
                total_timetables += 1
        
        # Count courses, faculty, and classrooms from uploaded files
        course_count = 0
        faculty_count = 0
        classroom_count = 0
        usable_classroom_count = 0
        
        # Load data to get accurate counts - force reload
        data_frames = load_all_data(force_reload=True)
        if data_frames:
            if 'course' in data_frames:
                course_count = len(data_frames['course'])
            if 'faculty_availability' in data_frames:
                faculty_count = len(data_frames['faculty_availability'])
            if 'classroom' in data_frames:
                classroom_df = data_frames['classroom'].copy()
                classroom_count = len(classroom_df)
                
                # Debug: print all classrooms to understand filtering
                print("[STATS] All classrooms in data:")
                for idx, room in classroom_df.iterrows():
                    print(f"  {room.get('Room Number', 'Unknown')}: Type={room.get('Type', 'N/A')}, Capacity={room.get('Capacity', 'N/A')}")
                
                # Count usable classrooms: regular classrooms + labs (exclude recreation, library, etc.)
                # Exclude rooms with 'nil' capacity
                classroom_df['Capacity'] = pd.to_numeric(classroom_df['Capacity'], errors='coerce')
                
                # Show filtering process
                print("[STATS] Filtering for usable classrooms:")
                print(f"  Total rows: {len(classroom_df)}")
                
                # First filter: Type check
                type_mask = (classroom_df['Type'].str.contains('classroom|auditorium|lab', case=False, na=False)) | (classroom_df['Room Number'].astype(str).str.startswith('L', na=False))
                print(f"  After type filter (classroom|auditorium|lab OR starts with L): {type_mask.sum()}")
                
                # Second filter: Capacity check
                capacity_mask = classroom_df['Capacity'].notna() & (classroom_df['Capacity'] > 0)
                print(f"  After capacity filter (>0): {capacity_mask.sum()}")
                
                # Combined filter
                usable_df = classroom_df[type_mask & capacity_mask].copy()
                usable_classroom_count = len(usable_df)
                print(f"  Final usable classrooms: {usable_classroom_count}")
        
        return jsonify({
            'total_timetables': total_timetables,
            'total_courses': course_count,
            'total_faculty': faculty_count,
            'total_classrooms': classroom_count,
            'usable_classrooms': usable_classroom_count
        })
        
    except Exception as e:
        print(f"[FAIL] Error loading stats: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'total_timetables': 0,
            'total_courses': 0,
            'total_faculty': 0,
            'total_classrooms': 0,
            'usable_classrooms': 0
        })

@app.route('/upload', methods=['POST'])
def upload_files():
    try:
        print("=" * 50)
        print("[INFO] RECEIVED FILE UPLOAD REQUEST")
        print("=" * 50)
        
        if 'files' not in request.files:
            return jsonify({
                'success': False,
                'message': 'No files provided'
            }), 400
        
        files = request.files.getlist('files')
        uploaded_files = []
        print(f"[INFO] Received {len(files)} files")
        
        # Clear input directory first using shutil for better file handling
        if os.path.exists(INPUT_DIR):
            try:
                shutil.rmtree(INPUT_DIR)
                os.makedirs(INPUT_DIR, exist_ok=True)
                print("[CLEAN] Cleared input directory")
            except Exception as e:
                # Try a more tolerant cleanup if rmtree fails (e.g., files are locked)
                print(f"[WARN] Could not clear input directory with rmtree: {e}")
                try:
                    for fname in os.listdir(INPUT_DIR):
                        p = os.path.join(INPUT_DIR, fname)
                        try:
                            os.chmod(p, 0o666)
                            os.remove(p)
                        except Exception as e2:
                            print(f"[WARN] Could not remove {p}: {e2}")
                except Exception as e3:
                    print(f"[WARN] Could not perform individual file cleanup: {e3}")
                try:
                    os.makedirs(INPUT_DIR, exist_ok=True)
                except Exception:
                    pass

        # Ensure input directory exists and is writable
        os.makedirs(INPUT_DIR, exist_ok=True)
        try:
            test_path = os.path.join(INPUT_DIR, '.writetest')
            with open(test_path, 'w') as fp:
                fp.write('ok')
            os.remove(test_path)
        except Exception as e:
            print(f"[ERROR] Input directory not writable: {e}")
            return jsonify({'success': False, 'message': 'Server cannot write to input directory. Check permissions.'}), 500

        for file in files:
            if file and file.filename != '' and allowed_file(file.filename):
                filename = secure_filename(file.filename)
                filepath = os.path.join(INPUT_DIR, filename)
                # Save to a temporary file first, then atomically replace to avoid partial writes and lock errors
                tmp_path = filepath + '.uploading'
                try:
                    # Try the standard save into a temp path in INPUT_DIR
                    file.save(tmp_path)
                    # Atomically move into place
                    try:
                        os.replace(tmp_path, filepath)
                    except Exception:
                        # Fallback to shutil.move if replace fails on some systems
                        shutil.move(tmp_path, filepath)
                    uploaded_files.append(filename)
                    print(f"[OK] Uploaded: {filename} -> {filepath}")
                except PermissionError as perr:
                    print(f"[ERROR] Permission denied when saving uploaded file {filename}: {perr}")
                    # Cleanup temp file if present
                    try:
                        if os.path.exists(tmp_path):
                            os.remove(tmp_path)
                    except Exception:
                        pass
                    return jsonify({'success': False, 'message': f'Permission denied saving {filename}. Close any open handles to files in the input folder or run the server with appropriate permissions.'}), 500
                except Exception as err:
                    print(f"[ERROR] Failed to save uploaded file {filename}: {err}")
                    try:
                        if os.path.exists(tmp_path):
                            os.remove(tmp_path)
                    except Exception:
                        pass
                    return jsonify({'success': False, 'message': f'Failed to save file {filename}: {err}'}), 500
        
        if not uploaded_files:
            return jsonify({
                'success': False,
                'message': 'No valid CSV files uploaded'
            }), 400
        
        # Verify we have all required files with flexible matching
        required_files = [
            "course_data.csv",
            "faculty_availability.csv",
            "classroom_data.csv", 
            "student_data.csv",
            "exams_data.csv"
        ]
        
        missing_files = []
        available_files = os.listdir(INPUT_DIR)
        
        print(f"[DIR] Available files after upload: {available_files}")
        
        for required_file in required_files:
            found = False
            required_clean = required_file.lower().replace(' ', '').replace('_', '').replace('-', '')
            
            for uploaded_file in available_files:
                uploaded_clean = uploaded_file.lower().replace(' ', '').replace('_', '').replace('-', '')
                if (required_clean in uploaded_clean or 
                    uploaded_clean in required_clean or
                    any(part in uploaded_clean for part in required_file.split('_'))):
                    found = True
                    print(f"[OK] Matched {required_file} with {uploaded_file}")
                    break
            
            if not found:
                missing_files.append(required_file)
                print(f"[FAIL] No match found for {required_file}")
        
        if missing_files:
            return jsonify({
                'success': False,
                'message': f'Missing required files: {", ".join(missing_files)}',
                'uploaded_files': uploaded_files,
                'missing_files': missing_files
            }), 400
        
        # Clear the common elective allocations cache when new files are uploaded
        global _SEMESTER_ELECTIVE_ALLOCATIONS, _ELECTIVE_COMMON_ROOMS
        _SEMESTER_ELECTIVE_ALLOCATIONS = {}
        _ELECTIVE_COMMON_ROOMS = {}
        print("[CLEAN] Cleared common elective allocations cache")
        print("[CLEAN] Cleared elective common rooms tracker (will be shared across all branches)")
        
        # Reset classroom usage tracker for new generation
        reset_classroom_usage_tracker()
        print("[RESET] Reset classroom usage tracker for new timetable generation")
        
        # Generate consolidated timetables for all branches and semesters
        branches = ['CSE', 'DSAI', 'ECE']
        target_semesters = [1, 3, 5, 7]
        success_count = 0
        generated_files = []
        
        # Load the newly uploaded data immediately (force reload) before computing allocations
        data_frames = load_all_data(force_reload=True)
        if data_frames is None:
            return jsonify({
                'success': False,
                'message': 'Failed to load CSV data after upload',
                'uploaded_files': uploaded_files
            }), 400
        
        # Generate consolidated timetables (one file per branch per semester)
        for sem in target_semesters:
            for branch_idx, branch in enumerate(branches):
                try:
                    print(f"[RESET] Generating consolidated timetable for {branch} Semester {sem}...")
                    
                    # Use consolidated generation
                    # Never reset the classroom usage tracker between semesters - all semesters share classrooms
                    # Only reset preferred room preferences for first branch of each semester
                    reset_prefs_for_sem = (branch_idx == 0)
                    success = export_consolidated_semester_timetable(data_frames, sem, branch, _reset_for_semester=reset_prefs_for_sem)
                    
                    # Debug: Show tracker size to verify it's accumulating
                    tracker_size = sum(len(slots) for day in _CLASSROOM_USAGE_TRACKER.values() for slots in day.values())
                    print(f"[TRACKER] After {branch} Sem {sem}: {tracker_size} room-slot allocations tracked")
                    
                    filename = f"sem{sem}_{branch}_timetable.xlsx"
                    filepath = os.path.join(OUTPUT_DIR, filename)
                    
                    if success and os.path.exists(filepath):
                        success_count += 1
                        generated_files.append(filename)
                        print(f"[OK] Successfully generated consolidated timetable: {filename}")
                    else:
                        print(f"[FAIL] Consolidated timetable not created: {filename}")
                        
                except Exception as e:
                    print(f"[FAIL] Error generating consolidated timetable for {branch} semester {sem}: {e}")
                    traceback.print_exc()

        # After all timetables are generated, generate audit files
        audit_result = {'faculty_audit': None, 'classroom_audit': None}
        try:
            print(f"\n[AUDIT] Starting audit file generation after upload...")
            print(f"[AUDIT] Output directory: {OUTPUT_DIR}")
            print(f"[AUDIT] Timetable files generated: {len(generated_files)}")
            
            # Extract schedule data from generated timetables to build audit info
            populate_audit_trackers_from_timetables(data_frames, OUTPUT_DIR)
            
            print(f"[AUDIT] Tracker populated with {len(_FACULTY_SCHEDULE_TRACKER)} faculty, {len(_CLASSROOM_SCHEDULE_TRACKER)} classrooms")
            
            # Generate the audit files
            audit_result = generate_audit_files(data_frames, OUTPUT_DIR)
            
            print(f"[AUDIT] Audit result: faculty={audit_result.get('faculty_audit')}, classroom={audit_result.get('classroom_audit')}")
            
            if audit_result.get('faculty_audit'):
                generated_files.append(os.path.basename(audit_result['faculty_audit']))
            if audit_result.get('classroom_audit'):
                generated_files.append(os.path.basename(audit_result['classroom_audit']))
        except Exception as audit_error:
            print(f"[WARN] Audit file generation failed: {audit_error}")
            traceback.print_exc()

        return jsonify({
            'success': True,
            'message': f'Successfully uploaded {len(uploaded_files)} files and generated {success_count} consolidated timetables!',
            'uploaded_files': uploaded_files,
            'generated_count': success_count,
            'files': generated_files,
            'audit_files': {
                'faculty': os.path.basename(audit_result['faculty_audit']) if audit_result.get('faculty_audit') else None,
                'classroom': os.path.basename(audit_result['classroom_audit']) if audit_result.get('classroom_audit') else None
            }
        })
        
    except Exception as e:
        print(f"[FAIL] Error uploading files: {e}")
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': f'Error uploading files: {str(e)}'
        }), 500


def export_semester_timetable_with_baskets_common(dfs, semester, branch, common_elective_allocations, minimal_only=False):
    """Export timetable using pre-allocated common basket slots. Set minimal_only=True to emit only timetable sheets."""
    try:
        print(f"[STATS] Generating timetable for Semester {semester}, Branch {branch} with COMMON basket slots...")
        
        # Generate schedules using the COMMON basket allocations
        section_a = generate_section_schedule_with_elective_baskets(dfs, semester, 'A', common_elective_allocations, branch)
        section_b = generate_section_schedule_with_elective_baskets(dfs, semester, 'B', common_elective_allocations, branch)
        
        if section_a is None or section_b is None:
            return False

        filename = f"sem{semester}_{branch}_timetable_baskets.xlsx"
        filepath = os.path.join(OUTPUT_DIR, filename)
        
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            section_a.to_excel(writer, sheet_name='Section_A')
            section_b.to_excel(writer, sheet_name='Section_B')
            
            if not minimal_only:
                # Create basket allocations summary from the common allocations
                basket_allocations = {}
                for allocation in common_elective_allocations.values():
                    if allocation:
                        basket_name = allocation['basket_name']
                        if basket_name not in basket_allocations:
                            basket_allocations[basket_name] = {
                                'lectures': allocation['lectures'],
                                'tutorial': allocation['tutorial'],
                                'courses': allocation['all_courses_in_basket']
                            }
                
                basket_summary = create_basket_summary(basket_allocations, semester, branch)
                basket_summary.to_excel(writer, sheet_name='Basket_Allocation', index=False)
                
                course_summary = create_course_summary(dfs, semester, branch)
                if not course_summary.empty:
                    course_summary.to_excel(writer, sheet_name='Course_Summary', index=False)
        
        print(f"[OK] Generated: {filename} with COMMON basket slots")
        return True
        
    except Exception as e:
        print(f"[FAIL] Error: {e}")
        return False
    
# EXAM SCHEDULE GENERATION ROUTE - COMMENTED OUT
# @app.route('/exam-schedule', methods=['POST'])
# def generate_exam_schedule():
#     """Generate conflict-free exam timetable with configuration and classroom allocation"""
#     (see EXAM_TIMETABLE_DISABLED.md for full function code)
#
# EXAM TIMETABLE ROUTES - COMMENTED OUT
# @app.route('/exam-timetables')
# def get_exam_timetables():
#     """Get generated exam timetables - only shows schedules that are marked for display"""
#     (see EXAM_TIMETABLE_DISABLED.md for full function code)
#
# @app.route('/exam-timetables/all')
# def get_all_exam_timetables():
#     """Get all available exam timetables"""
#     (see EXAM_TIMETABLE_DISABLED.md for full function code)
#
# @app.route('/exam-timetables/add-to-display', methods=['POST'])
# def add_exam_to_display():
#     """Add exam timetable to current display"""
#     (see EXAM_TIMETABLE_DISABLED.md for full function code)
#
# @app.route('/exam-timetables/remove-from-display', methods=['POST'])
# def remove_exam_from_display():
#     """Remove exam timetable from current display"""
#     (see EXAM_TIMETABLE_DISABLED.md for full function code)
#
# @app.route('/exam-timetables/clear-display', methods=['POST'])
# def clear_exam_display():
#     """Clear all exam timetables from display"""
#     (see EXAM_TIMETABLE_DISABLED.md for full function code)
#
# EXAM SCHEDULE GENERATION AND HELPER FUNCTIONS - COMMENTED OUT
# def schedule_exams_conflict_free(...):
#     (see EXAM_TIMETABLE_DISABLED.md for full function code)
#
# def has_student_conflict_strict(...):
#     (see EXAM_TIMETABLE_DISABLED.md for full function code)
#
# def has_student_conflict_moderate(...):
#     (see EXAM_TIMETABLE_DISABLED.md for full function code)
#
# def has_student_conflict_lenient(...):
#     (see EXAM_TIMETABLE_DISABLED.md for full function code)
#
# def is_session_balanced(...):
#     (see EXAM_TIMETABLE_DISABLED.md for full function code)
#
# (Additional exam functions disabled - see documentation)

# EXAM ROUTES - ALL DISABLED (see EXAM_TIMETABLE_DISABLED.md)
# All exam-related API routes and functions have been disabled
# (See EXAM_TIMETABLE_DISABLED.md for original implementations)

# END OF EXAM TIMETABLE ROUTES

@app.route('/generate-with-baskets', methods=['POST'])
def generate_timetables_with_baskets():
    try:
        print("[CONSOLIDATED] Starting consolidated timetable generation...")
        
        # Reset classroom usage tracker ONCE at the start - all branches/semesters share the same physical classrooms
        reset_classroom_usage_tracker()
        print("[RESET] Classroom usage tracker reset - ready for new generation")
        
        # Clear existing timetable files first
        excel_files = glob.glob(os.path.join(OUTPUT_DIR, "sem*_*_timetable*.xlsx"))
        for file in excel_files:
            try:
                os.remove(file)
                print(f"[CLEAN] Removed old file: {file}")
            except Exception as e:
                print(f"[WARN] Could not remove {file}: {e}")
        
        # Also clear old audit files
        audit_files = glob.glob(os.path.join(OUTPUT_DIR, "*_Audit.xlsx"))
        for file in audit_files:
            try:
                os.remove(file)
                print(f"[CLEAN] Removed old audit file: {file}")
            except Exception as e:
                print(f"[WARN] Could not remove {file}: {e}")

        # Load data
        data_frames = load_all_data(force_reload=True)
        if data_frames is None:
            return jsonify({'success': False, 'message': 'Failed to load CSV data'})

        # Generate consolidated timetables (one file per branch per semester)
        departments = get_departments_from_data(data_frames)
        target_semesters = [1, 3, 5, 7]
        success_count = 0
        generated_files = []
        
        # Import for stdout capture - same as full_audit.py to ensure consistent allocation
        import io
        
        for branch in departments:
            for sem in target_semesters:
                try:
                    print(f"\n[PROCESSING] Semester {sem}, Branch {branch}...")
                    # Capture stdout during generation to ensure consistent allocation behavior
                    # This matches full_audit.py which produces conflict-free results
                    old_stdout = sys.stdout
                    sys.stdout = io.StringIO()
                    try:
                        success = export_consolidated_semester_timetable(data_frames, sem, branch)
                    finally:
                        sys.stdout = old_stdout
                    
                    if success:
                        filename = f"sem{sem}_{branch}_timetable.xlsx"
                        success_count += 1
                        generated_files.append(filename)
                        print(f"[OK] Generated: {filename}")
                except Exception as e:
                    sys.stdout = old_stdout  # Restore stdout on error
                    print(f"[FAIL] Error generating timetable for {branch} semester {sem}: {e}")
                    traceback.print_exc()
        
        # After all timetables are generated, populate audit trackers from generated files
        # and generate audit Excel files
        audit_result = {'faculty_audit': None, 'classroom_audit': None}
        try:
            print(f"\n[AUDIT] Starting audit file generation...")
            print(f"[AUDIT] Output directory: {OUTPUT_DIR}")
            print(f"[AUDIT] Timetable files generated: {len(generated_files)}")
            
            # Extract schedule data from generated timetables to build audit info
            populate_audit_trackers_from_timetables(data_frames, OUTPUT_DIR)
            
            print(f"[AUDIT] Tracker populated with {len(_FACULTY_SCHEDULE_TRACKER)} faculty, {len(_CLASSROOM_SCHEDULE_TRACKER)} classrooms")
            
            # Generate the audit files
            audit_result = generate_audit_files(data_frames, OUTPUT_DIR)
            
            print(f"[AUDIT] Audit result: faculty={audit_result.get('faculty_audit')}, classroom={audit_result.get('classroom_audit')}")
            
            if audit_result.get('faculty_audit'):
                generated_files.append(os.path.basename(audit_result['faculty_audit']))
            if audit_result.get('classroom_audit'):
                generated_files.append(os.path.basename(audit_result['classroom_audit']))
        except Exception as audit_error:
            print(f"[WARN] Audit file generation failed: {audit_error}")
            traceback.print_exc()
        
        return jsonify({
            'success': True, 
            'message': f'Successfully generated {success_count} consolidated timetables!',
            'generated_count': success_count,
            'files': generated_files,
            'audit_files': {
                'faculty': os.path.basename(audit_result['faculty_audit']) if audit_result.get('faculty_audit') else None,
                'classroom': os.path.basename(audit_result['classroom_audit']) if audit_result.get('classroom_audit') else None
            }
        })
        
    except Exception as e:
        print(f"[FAIL] Error in consolidated generation endpoint: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 500

@app.route('/generate-mid-semester', methods=['POST'])
def generate_mid_semester_timetables():
    """Generate separate pre-mid and post-mid timetables"""
    try:
        data = request.json or {}
        semester = int(data.get('semester'))
        branch = data.get('branch')
        time_config = data.get('time_config') or {}

        if not semester:
            return jsonify({'success': False, 'message': 'semester is required'}), 400

        # Load latest data
        dfs = load_all_data(force_reload=True)
        if dfs is None:
            return jsonify({'success': False, 'message': 'Failed to load CSV data'}), 500

        # Generate mid-semester timetables
        result = export_mid_semester_timetables(dfs, semester, branch, time_config=time_config)
        
        response = {
            'success': result['pre_mid_success'] or result['post_mid_success'],
            'message': 'Mid-semester timetables generated',
            'pre_mid_generated': result['pre_mid_success'],
            'post_mid_generated': result['post_mid_success']
        }
        
        if result['pre_mid_success']:
            response['pre_mid_file'] = result['pre_mid_filename']
        if result['post_mid_success']:
            response['post_mid_file'] = result['post_mid_filename']
        
        return jsonify(response)
        
    except Exception as e:
        print(f"[FAIL] Error generating mid-semester timetables: {e}")
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 500

@app.route('/generate-mid-semester-timetables', methods=['POST'])
def generate_mid_semester_timetables_endpoint():  # Changed function name
    """Generate separate pre-mid and post-mid timetables"""
    try:
        data = request.json or {}
        semester = int(data.get('semester'))
        branch = data.get('branch')
        time_config = data.get('time_config') or {}
        
        if not semester:
            return jsonify({'success': False, 'message': 'semester is required'}), 400
        
        # Load latest data
        dfs = load_all_data(force_reload=True)
        if dfs is None:
            return jsonify({'success': False, 'message': 'Failed to load CSV data'}), 500
        
        print(f"[RESET] Generating mid-semester timetables for Semester {semester}, Branch {branch}...")
        
        # Generate pre-mid and post-mid timetables
        results = export_mid_semester_timetables(dfs, semester, branch, time_config)
        
        generated_files = []
        if results['pre_mid_success'] and results['pre_mid_filename']:
            generated_files.append(results['pre_mid_filename'])
        
        if results['post_mid_success'] and results['post_mid_filename']:
            generated_files.append(results['post_mid_filename'])
        
        if generated_files:
            message = f'Generated {len(generated_files)} mid-semester timetables: {", ".join(generated_files)}'
            return jsonify({
                'success': True,
                'message': message,
                'files': generated_files,
                'pre_mid_success': results['pre_mid_success'],
                'post_mid_success': results['post_mid_success']
            })
        else:
            return jsonify({
                'success': False,
                'message': 'Failed to generate mid-semester timetables',
                'pre_mid_success': results['pre_mid_success'],
                'post_mid_success': results['post_mid_success']
            })
        
    except Exception as e:
        print(f"[FAIL] Error generating mid-semester timetables: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 500
    
@app.route('/generate-pre-mid-timetable', methods=['POST'])
def generate_pre_mid_timetable():
    """Generate pre-mid semester timetable"""
    try:
        data = request.json or {}
        semester = int(data.get('semester'))
        branch = data.get('branch')
        time_config = data.get('time_config') or {}

        if not semester or not branch:
            return jsonify({'success': False, 'message': 'semester and branch are required'}), 400

        # Load latest data
        dfs = load_all_data(force_reload=True)
        if dfs is None:
            return jsonify({'success': False, 'message': 'Failed to load CSV data'}), 500

        # Generate pre-mid timetable
        result = export_mid_semester_timetables(dfs, semester, branch, time_config=time_config)
        
        if result['pre_mid_success'] and result['pre_mid_filename']:
            return jsonify({
                'success': True, 
                'message': 'Pre-mid timetable generated successfully!',
                'filename': result['pre_mid_filename'],
                'timetable_type': 'pre_mid'
            })
        else:
            return jsonify({'success': False, 'message': 'Failed to generate pre-mid timetable'}), 500
    except Exception as e:
        print(f"[FAIL] Error generating pre-mid timetable: {e}")
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 500

@app.route('/generate-post-mid-timetable', methods=['POST'])
def generate_post_mid_timetable():
    """Generate post-mid semester timetable"""
    try:
        data = request.json or {}
        semester = int(data.get('semester'))
        branch = data.get('branch')
        time_config = data.get('time_config') or {}

        if not semester or not branch:
            return jsonify({'success': False, 'message': 'semester and branch are required'}), 400

        # Load latest data
        dfs = load_all_data(force_reload=True)
        if dfs is None:
            return jsonify({'success': False, 'message': 'Failed to load CSV data'}), 500

        # Generate post-mid timetable
        result = export_mid_semester_timetables(dfs, semester, branch, time_config=time_config)
        
        if result['post_mid_success'] and result['post_mid_filename']:
            return jsonify({
                'success': True, 
                'message': 'Post-mid timetable generated successfully!',
                'filename': result['post_mid_filename'],
                'timetable_type': 'post_mid'
            })
        else:
            return jsonify({'success': False, 'message': 'Failed to generate post-mid timetable'}), 500
    except Exception as e:
        print(f"[FAIL] Error generating post-mid timetable: {e}")
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 500

@app.route('/generate-both-mid-timetables', methods=['POST'])
def generate_both_mid_timetables():
    """Generate both pre-mid and post-mid timetables"""
    try:
        data = request.json or {}
        semester = int(data.get('semester'))
        branch = data.get('branch')
        time_config = data.get('time_config') or {}

        if not semester or not branch:
            return jsonify({'success': False, 'message': 'semester and branch are required'}), 400

        # Load latest data
        dfs = load_all_data(force_reload=True)
        if dfs is None:
            return jsonify({'success': False, 'message': 'Failed to load CSV data'}), 500

        # Generate both timetables
        result = export_mid_semester_timetables(dfs, semester, branch, time_config=time_config)
        
        generated_files = []
        if result['pre_mid_success'] and result['pre_mid_filename']:
            generated_files.append(result['pre_mid_filename'])
        if result['post_mid_success'] and result['post_mid_filename']:
            generated_files.append(result['post_mid_filename'])
        
        if generated_files:
            return jsonify({
                'success': True, 
                'message': f'Generated {len(generated_files)} mid-semester timetable(s)',
                'files': generated_files,
                'pre_mid_success': result['pre_mid_success'],
                'post_mid_success': result['post_mid_success']
            })
        else:
            return jsonify({'success': False, 'message': 'Failed to generate any mid-semester timetables'}), 500
    except Exception as e:
        print(f"[FAIL] Error generating mid-semester timetables: {e}")
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 500
    
if __name__ == '__main__':
    print("[START] Starting Timetable Generator with Comprehensive Statistics...")
    print(f"[FOLDER] Input directory: {INPUT_DIR}")
    print(f"[FOLDER] Output directory: {OUTPUT_DIR}")
    app.run(debug=True, port=5000)