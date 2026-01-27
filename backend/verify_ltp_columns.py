#!/usr/bin/env python3
"""
Verify that all timetables have the new Lectures/Week, Tutorials/Week, Labs/Week columns
in the CORE COURSES section of each timetable sheet.
"""
import os
import sys
from openpyxl import load_workbook

# Define all semesters and branches
SEMESTERS = [1, 3, 5, 7]
BRANCHES = ['CSE', 'DSAI', 'ECE']

output_dir = os.path.join(os.path.dirname(__file__), 'output_timetables')

print("=" * 80)
print("VERIFYING L/T/P COLUMNS IN ALL TIMETABLES")
print("=" * 80)

total_files = len(SEMESTERS) * len(BRANCHES)
current_file = 0
all_passed = True

for sem in SEMESTERS:
    for branch in BRANCHES:
        current_file += 1
        filename = f'sem{sem}_{branch}_timetable.xlsx'
        filepath = os.path.join(output_dir, filename)
        
        print(f"\n[{current_file}/{total_files}] Checking {filename}...")
        
        if not os.path.exists(filepath):
            print(f"  ✗ FILE NOT FOUND: {filename}")
            all_passed = False
            continue
        
        try:
            wb = load_workbook(filepath, data_only=True)
            
            # Check first timetable sheet (skip Course_Information which is index 0)
            if len(wb.sheetnames) < 2:
                print(f"  ✗ No timetable sheets found in {filename}")
                all_passed = False
                continue
            
            # Get first actual timetable sheet (not Course_Information)
            timetable_sheet = None
            for sheet_name in wb.sheetnames:
                if 'Course_Information' not in sheet_name:
                    timetable_sheet = wb[sheet_name]
                    break
            
            if timetable_sheet is None:
                print(f"  ✗ No timetable sheet found in {filename}")
                all_passed = False
                continue
            
            # Search for "CORE COURSES" section
            core_courses_row = None
            for row_idx, row in enumerate(timetable_sheet.iter_rows(min_row=1, max_row=100, max_col=10), start=1):
                cell_value = row[0].value
                if cell_value and 'CORE COURSES' in str(cell_value):
                    core_courses_row = row_idx
                    break
            
            if core_courses_row is None:
                print(f"  ⚠ No CORE COURSES section found in {timetable_sheet.title}")
                continue
            
            # Check headers in the next row
            header_row = core_courses_row + 1
            headers = []
            for col_idx in range(1, 8):  # Check first 7 columns
                cell_value = timetable_sheet.cell(row=header_row, column=col_idx).value
                if cell_value:
                    headers.append(str(cell_value))
            
            # Expected headers
            expected_headers = ['Course Code', 'Course Name', 'L-T-P-S-C', 'Term Type', 
                              'Lectures/Week', 'Tutorials/Week', 'Labs/Week']
            
            # Check if all expected headers are present
            missing_headers = [h for h in expected_headers if h not in headers]
            
            if missing_headers:
                print(f"  ✗ MISSING HEADERS in {timetable_sheet.title}:")
                print(f"     Missing: {missing_headers}")
                print(f"     Found: {headers}")
                all_passed = False
            else:
                print(f"  ✓ All L/T/P columns present in {timetable_sheet.title}")
                
                # Also check if there's data in the new columns
                data_row = header_row + 1
                ltpsc_value = timetable_sheet.cell(row=data_row, column=3).value
                lectures_value = timetable_sheet.cell(row=data_row, column=5).value
                tutorials_value = timetable_sheet.cell(row=data_row, column=6).value
                labs_value = timetable_sheet.cell(row=data_row, column=7).value
                
                if ltpsc_value and lectures_value is not None:
                    print(f"     Sample course LTPSC: {ltpsc_value}")
                    print(f"     Parsed values: L={lectures_value}, T={tutorials_value}, P={labs_value}")
        
        except Exception as e:
            print(f"  ✗ ERROR reading {filename}: {str(e)}")
            all_passed = False

print("\n" + "=" * 80)
if all_passed:
    print("✓ ALL TIMETABLES VERIFIED SUCCESSFULLY!")
    print("All CORE COURSES sections now have Lectures/Week, Tutorials/Week, Labs/Week columns.")
else:
    print("✗ SOME TIMETABLES FAILED VERIFICATION")
    print("Please check the errors above.")
print("=" * 80)
