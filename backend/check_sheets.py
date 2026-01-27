import openpyxl

wb = openpyxl.load_workbook(r'C:\timetable-generator-Logic-Beyond\timetable-generator-Logic-Beyond\backend\output_timetables\sem3_CSE_timetable.xlsx')

print(f"Number of sheets: {len(wb.worksheets)}")
for i, ws in enumerate(wb.worksheets):
    print(f"  Sheet {i}: {ws.title}")

# Check the FIRST sheet (Section A timetable)
ws = wb.worksheets[0]
print(f"\nChecking sheet '{ws.title}'...")

# Find CORE COURSES
for row_idx in range(30, 80):
    cell = ws.cell(row=row_idx, column=1)
    if cell.value and 'CORE' in str(cell.value).upper():
        print(f"\nFound at row {row_idx}: {cell.value}")
        # Get headers
        header_row = row_idx + 1
        headers = [ws.cell(row=header_row, column=i).value for i in range(1, 10)]
        print(f"Headers: {headers}")
        
        # Get first course
        data_row = header_row + 1
        course_data = [ws.cell(row=data_row, column=i).value for i in range(1, 10)]
        print(f"First course row: {course_data}")
        break
