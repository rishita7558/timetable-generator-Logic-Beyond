import openpyxl

wb = openpyxl.load_workbook(r'C:\timetable-generator-Logic-Beyond\timetable-generator-Logic-Beyond\backend\output_timetables\sem3_CSE_timetable.xlsx')

# Check Regular_Section_A
ws = wb['Regular_Section_A']
print(f"Checking sheet '{ws.title}'...")

# Find CORE COURSES at bottom of sheet
for row_idx in range(1, 100):
    cell = ws.cell(row=row_idx, column=1)
    if cell.value and 'CORE' in str(cell.value).upper() and 'COURSES' in str(cell.value).upper():
        print(f"\nFound CORE COURSES at row {row_idx}")
        # Get headers
        header_row = row_idx + 1
        headers = [ws.cell(row=header_row, column=i).value for i in range(1, 10)]
        print(f"Headers: {headers}")
        
        # Get first few courses
        for data_row in range(header_row + 1, header_row + 7):
            course_data = [ws.cell(row=data_row, column=i).value for i in range(1, 8)]
            if course_data[0]:
                print(f"\nRow {data_row}:")
                for i, (h, v) in enumerate(zip(headers[:7], course_data)):
                    print(f"  {h}: {v}")
        break
