#!/usr/bin/env python3
"""
Examine the actual timetable structure in the Excel file
"""
from openpyxl import load_workbook

filepath = 'output_timetables/sem1_CSE_timetable.xlsx'
wb = load_workbook(filepath)

# Get the first timetable sheet
sheet_name = None
for name in wb.sheetnames:
    if 'Course_Information' not in name:
        sheet_name = name
        break

if sheet_name:
    ws = wb[sheet_name]
    
    print(f"Sheet: {sheet_name}")
    print(f"Dimensions: {ws.dimensions}")
    print(f"\nFirst 15 rows (time slot column and first 3 course columns):")
    
    for row_idx in range(1, 16):
        row_data = []
        for col_idx in range(1, 5):
            cell = ws.cell(row=row_idx, column=col_idx)
            row_data.append(str(cell.value)[:30] if cell.value else "")
        print(f"Row {row_idx}: {row_data}")
    
    print(f"\n\nLooking for 'MA161' course entries...")
    found = False
    for row_idx in range(1, min(50, ws.max_row + 1)):
        for col_idx in range(1, min(10, ws.max_column + 1)):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value and 'MA161' in str(cell.value):
                print(f"  Row {row_idx}, Col {col_idx}: {cell.value}")
                found = True
    
    if not found:
        print("  MA161 not found in first 50 rows")
