import openpyxl

wb = openpyxl.load_workbook(r'C:\timetable-generator-Logic-Beyond\timetable-generator-Logic-Beyond\backend\output_timetables\sem3_CSE_timetable.xlsx')
ws = wb.worksheets[0]

# Find CORE COURSES header
for row_idx in range(1, 100):
    cell = ws.cell(row=row_idx, column=1)
    if cell.value == 'CORE COURSES':
        print(f"CORE COURSES header at row {row_idx}")
        # Next row should be column headers
        header_row = row_idx + 1
        headers = [ws.cell(row=header_row, column=i).value for i in range(1, 8)]
        print(f"Headers (row {header_row}): {headers}")
        
        # First data row
        data_row = header_row + 1
        for r in range(data_row, data_row + 6):
            values = [ws.cell(row=r, column=i).value for i in range(1, 8)]
            if values[0]:  # If course code exists
                print(f"\nRow {r}:")
                for i, (h, v) in enumerate(zip(headers, values)):
                    print(f"  {h}: {v}")
        break
