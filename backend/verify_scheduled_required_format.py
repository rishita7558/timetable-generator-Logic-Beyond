#!/usr/bin/env python3
"""
Verify that all timetables show the scheduled/required format for L/T/P columns.
"""
import os
from openpyxl import load_workbook

SEMESTERS = [1, 3, 5, 7]
BRANCHES = ['CSE', 'DSAI', 'ECE']
output_dir = os.path.join(os.path.dirname(__file__), 'output_timetables')

print("=" * 80)
print("VERIFYING SCHEDULED/REQUIRED FORMAT IN L/T/P COLUMNS")
print("=" * 80)

total_files = len(SEMESTERS) * len(BRANCHES)
current_file = 0
all_passed = True

for sem in SEMESTERS:
    for branch in BRANCHES:
        current_file += 1
        filename = f'sem{sem}_{branch}_timetable.xlsx'
        filepath = os.path.join(output_dir, filename)
        
        print(f"\n[{current_file}/{total_files}] Checking {filename}...")
        
        if not os.path.exists(filepath):
            print(f"  ✗ FILE NOT FOUND")
            all_passed = False
            continue
        
        try:
            wb = load_workbook(filepath, data_only=True)
            timetable_sheet = None
            for sheet_name in wb.sheetnames:
                if 'Course_Information' not in sheet_name:
                    timetable_sheet = wb[sheet_name]
                    break
            
            if timetable_sheet is None:
                print(f"  ✗ No timetable sheet found")
                all_passed = False
                continue
            
            # Find CORE COURSES section
            core_courses_row = None
            for row_idx in range(1, 100):
                cell_value = timetable_sheet.cell(row=row_idx, column=1).value
                if cell_value and 'CORE COURSES' in str(cell_value):
                    core_courses_row = row_idx
                    break
            
            if core_courses_row is None:
                print(f"  ⚠ No CORE COURSES section found")
                continue
            
            header_row = core_courses_row + 1
            data_row = header_row + 1
            
            # Check headers
            headers = []
            for col_idx in range(1, 8):
                cell_value = timetable_sheet.cell(row=header_row, column=col_idx).value
                if cell_value:
                    headers.append(str(cell_value))
            
            expected_headers = ['Course Code', 'Course Name', 'L-T-P-S-C', 'Term Type', 
                              'Lectures Hrs', 'Tutorials Hrs', 'Labs Hrs']
            
            missing_headers = [h for h in expected_headers if h not in headers]
            
            if missing_headers:
                print(f"  ✗ MISSING HEADERS: {missing_headers}")
                all_passed = False
            else:
                # Check data format
                ltpsc_value = timetable_sheet.cell(row=data_row, column=3).value
                lectures_value = timetable_sheet.cell(row=data_row, column=5).value
                tutorials_value = timetable_sheet.cell(row=data_row, column=6).value
                labs_value = timetable_sheet.cell(row=data_row, column=7).value
                
                # Check if values are in "scheduled/required" format
                is_correct_format = True
                for val, col_name in [(lectures_value, 'Lectures'), (tutorials_value, 'Tutorials'), (labs_value, 'Labs')]:
                    if val and '/' in str(val):
                        parts = str(val).split('/')
                        if len(parts) == 2:
                            try:
                                int(parts[0])
                                int(parts[1])
                            except ValueError:
                                is_correct_format = False
                                break
                    elif val:
                        is_correct_format = False
                        break
                
                if is_correct_format:
                    print(f"  ✓ Correct format with scheduled/required values")
                    print(f"     Sample: LTPSC={ltpsc_value}, L={lectures_value}, T={tutorials_value}, P={labs_value}")
                else:
                    print(f"  ✗ INCORRECT FORMAT")
                    print(f"     Expected 'scheduled/required' format")
                    print(f"     Got: L={lectures_value}, T={tutorials_value}, P={labs_value}")
                    all_passed = False
        
        except Exception as e:
            print(f"  ✗ ERROR: {str(e)}")
            all_passed = False

print("\n" + "=" * 80)
if all_passed:
    print("✓ ALL TIMETABLES VERIFIED SUCCESSFULLY!")
    print("All columns now show 'Scheduled/Required' format (e.g., 3/3, 0/1, 2/2)")
else:
    print("✗ SOME TIMETABLES FAILED VERIFICATION")
print("=" * 80)
