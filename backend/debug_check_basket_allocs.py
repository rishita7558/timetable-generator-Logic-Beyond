import os
import sys
import uuid
import openpyxl
sys.path.insert(0, os.path.join(os.getcwd(), 'backend'))
from app import app

unique = uuid.uuid4().hex[:8]
filename = os.path.join(os.getcwd(), 'output_timetables', f'sem5_CSE_{unique}_classrooms.xlsx')
if os.path.exists(filename):
    os.remove(filename)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Timetable'
ws['A1'] = 'Time'
ws['B1'] = 'Mon'
ws['A2'] = '09:00-10:00'
ws['B2'] = 'ELECTIVE_B4'

basket_ws = wb.create_sheet('Basket_Allocation')
basket_ws.append(['Basket Name', 'Courses in Basket', 'Day', 'Time Slot'])
basket_ws.append(['ELECTIVE_B4', 'CS301, CS302', 'Mon', '09:00-10:00'])

wb.save(filename)

c = app.test_client()
r = c.get('/timetables')
if r.status_code != 200:
    print('ERROR: /timetables returned', r.status_code)
    raise SystemExit(1)

d = r.get_json()
entries = [t for t in d if t.get('filename') == os.path.basename(filename)]
print('Found entries:', len(entries))
if entries:
    t = entries[0]
    print('has_classroom_allocation:', t.get('has_classroom_allocation'))
    print('basket_course_allocations:', t.get('basket_course_allocations'))
    print('classroom_details (summary count):', len(t.get('classroom_details', [])))
    print('\nHTML sample:\n', t.get('html'))
else:
    print('No timetables matched the filename:', os.path.basename(filename))
