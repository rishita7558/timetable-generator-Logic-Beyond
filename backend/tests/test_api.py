import os
import json
import glob
import openpyxl
import pytest
import uuid

from backend.app import app, OUTPUT_DIR, reset_classroom_usage_tracker, load_all_data, create_basket_courses_sheet, export_semester_timetable_with_baskets


def make_workbook_with_sheets(path, sheet_names):
    wb = openpyxl.Workbook()
    # Remove the default sheet if we will provide different names
    default_sheet = wb.active
    if default_sheet.title not in sheet_names:
        wb.remove(default_sheet)

    for idx, name in enumerate(sheet_names):
        if name in wb.sheetnames:
            continue
        ws = wb.create_sheet(name)
        # Add minimal content so DataFrame isn't empty when read
        ws['A1'] = 'Time'
        ws['B1'] = 'Mon 09:00-10:00'
        ws['A2'] = '09:00-10:00'
        # Alternate course code between sheets to avoid duplicate-free issues
        ws['B2'] = 'CS101 [R-101]' if idx % 2 == 0 else 'CS102 [R-102]'

    # Ensure at least one sheet persists (openpyxl requires at least one)
    if not wb.sheetnames:
        wb.create_sheet('Timetable')

    wb.save(path)


def test_create_basket_courses_sheet_handles_none_tutorial():
    # Basket allocation with None tutorial and one lecture
    basket_allocations = {
        'Basket1': {
            'lectures': [('Mon', '09:00-10:30')],
            'tutorial': None,
            'courses': ['CSE101', 'CSE102']
        }
    }

    df = create_basket_courses_sheet(basket_allocations)
    assert not df.empty
    assert all(df['Tutorial Slot'] == '-')
    assert all(df['Lecture Slots'].str.contains('Mon 09:00-10:30'))
    assert df['Total Courses in Basket'].iloc[0] == 2


def test_timetables_endpoint_returns_whole_branch_and_sections(tmp_path):
    # Create a unique filename to avoid collisions with existing outputs
    # Use unique filenames per test to avoid collisions and in-use errors
    unique_id = uuid.uuid4().hex[:8]
    # Ensure filenames match the glob pattern used in get_timetables: *_baskets.xlsx
    ece_filename = os.path.join(OUTPUT_DIR, f'sem3_ECE_{unique_id}_baskets.xlsx')
    cse_filename = os.path.join(OUTPUT_DIR, f'sem3_CSE_{unique_id}_baskets.xlsx')

    # Cleanup if exist
    for f in [ece_filename, cse_filename]:
        if os.path.exists(f):
            os.remove(f)

    # Create ECE workbook with 'Timetable' sheet (Whole)
    make_workbook_with_sheets(ece_filename, ['Timetable'])

    # Create CSE workbook with both Section_A and Section_B
    make_workbook_with_sheets(cse_filename, ['Section_A', 'Section_B'])

    # Use Flask test client
    client = app.test_client()
    resp = client.get('/timetables')
    assert resp.status_code == 200
    timetables = resp.get_json()
    # Find ECE entry
    ece_entries = [t for t in timetables if t.get('filename') == os.path.basename(ece_filename)]
    assert len(ece_entries) == 1
    assert ece_entries[0]['section'] == 'Whole'

    # Find CSE entries (A and B)
    cse_entries = [t for t in timetables if t.get('filename') == os.path.basename(cse_filename)]
    # There should be 2 entries for CSE
    assert len([e for e in cse_entries if e['section'] in ['A', 'B']]) >= 2


def test_stats_counts_timetables_correctly(tmp_path):
    # Create files to match sem3 as above
    unique_id = uuid.uuid4().hex[:8]
    ece_filename = os.path.join(OUTPUT_DIR, f'sem3_ECE_timetable_baskets_{unique_id}.xlsx')
    cse_filename = os.path.join(OUTPUT_DIR, f'sem3_CSE_timetable_baskets_{unique_id}.xlsx')

    # Cleanup
    for f in [ece_filename, cse_filename]:
        if os.path.exists(f):
            os.remove(f)

    make_workbook_with_sheets(ece_filename, ['Timetable'])
    make_workbook_with_sheets(cse_filename, ['Section_A', 'Section_B'])

    client = app.test_client()
    resp = client.get('/stats')
    assert resp.status_code == 200
    stats = resp.get_json()
    # For our two files, total timetables should be 3 (ECE=1, CSE=2)
    assert stats['total_timetables'] >= 3
    # Course/faculty/classroom counts should be non-negative integers
    assert isinstance(stats['total_courses'], int) and stats['total_courses'] >= 0
    assert isinstance(stats['total_faculty'], int) and stats['total_faculty'] >= 0
    assert isinstance(stats['total_classrooms'], int) and stats['total_classrooms'] >= 0


def test_whole_branch_timetable_populates_courses_and_baskets(tmp_path):
    # Create a simple workbook for ECE with a filled Timetable sheet and a Basket_Allocation sheet
    filename = os.path.join(OUTPUT_DIR, f'sem3_ECE_{uuid.uuid4().hex[:8]}_baskets.xlsx')
    # Cleanup
    if os.path.exists(filename):
        os.remove(filename)

    wb = openpyxl.Workbook()
    # Default sheet will be used - rename to Timetable
    ws = wb.active
    ws.title = 'Timetable'
    # Add a few rows/cols with course entries
    ws['A1'] = 'Time'
    ws['B1'] = 'Mon 09:00-10:00'
    ws['A2'] = '09:00-10:00'
    ws['B2'] = 'EC101 [R-101]'
    # Create a Basket_Allocation sheet for elective mapping
    basket_ws = wb.create_sheet('Basket_Allocation')
    basket_ws.append(['Basket Name', 'Courses in Basket', 'Day', 'Time Slot'])
    basket_ws.append(['ELECTIVE_B3', 'EC201, EC202', 'Mon', '09:00-10:00'])
    wb.save(filename)

    client = app.test_client()
    resp = client.get('/timetables')
    assert resp.status_code == 200
    timetables = resp.get_json()
    # Find the ECE entry
    ece_entries = [t for t in timetables if t.get('filename') == os.path.basename(filename)]
    assert len(ece_entries) == 1
    ece = ece_entries[0]
    # Ensure courses array is not empty and basket_courses_map contains ELECTIVE_B3
    assert isinstance(ece.get('courses', []), list) and len(ece.get('courses', [])) > 0
    assert isinstance(ece.get('basket_courses_map', {}), dict)
    assert ('ELECTIVE_B3' in ece.get('basket_courses_map', {})) or (any(b for b in ece.get('baskets', []) if 'B3' in b))
