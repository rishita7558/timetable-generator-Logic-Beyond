import os
import sys
import openpyxl
import uuid
import pandas as pd

# Add parent directory to path for module imports
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from app import app, OUTPUT_DIR


def test_all_scheduled_courses_get_classrooms(tmp_path):
    unique = uuid.uuid4().hex[:8]
    filename = os.path.join(OUTPUT_DIR, f'sem3_ECE_{unique}_classrooms.xlsx')
    if os.path.exists(filename):
        os.remove(filename)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Section_A'
    ws['A1'] = 'Time'
    ws['B1'] = 'Mon'
    ws['C1'] = 'Tue'
    ws['A2'] = '09:00-10:00'
    ws['B2'] = 'CS101'
    ws['C2'] = 'CS201'
    ws['A3'] = '10:00-11:00'
    ws['B3'] = 'CS101'
    ws['C3'] = 'Free'

    wb.save(filename)

    client = app.test_client()
    resp = client.get('/timetables')
    assert resp.status_code == 200
    timetables = resp.get_json()
    entries = [t for t in timetables if t.get('filename') == os.path.basename(filename)]
    assert len(entries) >= 1

    # Ensure classroom_details is present for each entry
    for t in entries:
        assert isinstance(t.get('classroom_details', []), list)
    assert len(entries) >= 1

    # Check allocations exist and rooms are assigned
    for t in entries:
        details = t.get('classroom_details', [])
        assert isinstance(details, list)
        # Every allocation must have a non-empty room
        assert all(d.get('room') for d in details)

    # CS101 should have same room across its allocations
    all_allocs = []
    for t in entries:
        all_allocs.extend([d for d in t.get('classroom_details', []) if d.get('course') == 'CS101'])
    assert len(all_allocs) >= 2
    rooms = set(d['room'] for d in all_allocs)
    assert len(rooms) == 1


def test_classrooms_show_in_html_and_details(tmp_path):
    unique = uuid.uuid4().hex[:8]
    filename = os.path.join(OUTPUT_DIR, f'sem5_CSE_{unique}_classrooms.xlsx')
    if os.path.exists(filename):
        os.remove(filename)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Timetable'
    ws['A1'] = 'Time'
    ws['B1'] = 'Mon'
    ws['A2'] = '09:00-10:00'
    ws['B2'] = 'ELECTIVE_B4'

    basket_ws = wb.create_sheet('Basket_Allocation')
    basket_ws.append(['Basket Name', 'Courses in Basket', 'Day', 'Time Slot'])
    basket_ws.append(['ELECTIVE_B4', 'CS301, CS302', 'Mon', '09:00-10:00'])

    wb.save(filename)

    client = app.test_client()
    resp = client.get('/timetables')
    assert resp.status_code == 200
    timetables = resp.get_json()

    entries = [t for t in timetables if t.get('filename') == os.path.basename(filename)]
    assert len(entries) == 1
    t = entries[0]

    # classroom_details should exist and be non-empty
    details = t.get('classroom_details', [])
    assert isinstance(details, list)
    assert len(details) > 0

    # HTML should NOT contain a bracketed room for the basket slot (rooms shown per-course in legend)
    html = t.get('html', '')
    # Ensure the basket cell itself isn't showing a room
    assert 'ELECTIVE_B4 [' not in html

    # Check basket_course_allocations exists and contains entries for the individual courses
    basket_allocs = t.get('basket_course_allocations', {})
    assert isinstance(basket_allocs, dict)
    assert 'ELECTIVE_B4' in basket_allocs
    # Courses listed in the basket should be present in the allocation map (may be None if unallocated)
    assert 'CS301' in basket_allocs['ELECTIVE_B4'] and 'CS302' in basket_allocs['ELECTIVE_B4']


def test_basket_course_allocations_written_to_excel(tmp_path):
    """Ensure per-course basket allocations are persisted into the output Excel file."""
    import pandas as pd
    unique = uuid.uuid4().hex[:8]
    filename = os.path.join(OUTPUT_DIR, f'sem5_CSE_{unique}_classrooms.xlsx')
    if os.path.exists(filename):
        os.remove(filename)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Timetable'
    ws['A1'] = 'Time'
    ws['B1'] = 'Mon'
    ws['A2'] = '09:00-10:00'
    ws['B2'] = 'ELECTIVE_B4'

    basket_ws = wb.create_sheet('Basket_Allocation')
    basket_ws.append(['Basket Name', 'Courses in Basket', 'Day', 'Time Slot'])
    basket_ws.append(['ELECTIVE_B4', 'CS301, CS302', 'Mon', '09:00-10:00'])

    wb.save(filename)

    client = app.test_client()
    resp = client.get('/timetables')
    assert resp.status_code == 200

    # Re-open the persisted file and validate contents of Basket_Course_Allocations sheet
    df = pd.read_excel(filename, sheet_name='Basket_Course_Allocations')
    assert 'Basket Name' in df.columns and 'Course' in df.columns and 'Allocated Rooms' in df.columns
    # Should contain rows for CS301 and CS302
    courses = set(df['Course'].astype(str).tolist())
    assert 'CS301' in courses and 'CS302' in courses


def test_persist_allocations_to_excel(tmp_path):
    unique = uuid.uuid4().hex[:8]
    filename = os.path.join(OUTPUT_DIR, f'sem4_ECE_{unique}_persist.xlsx')
    if os.path.exists(filename):
        os.remove(filename)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Section_A'
    ws['A1'] = 'Time'
    ws['B1'] = 'Mon'
    ws['A2'] = '09:00-10:00'
    ws['B2'] = 'CS101'
    wb.save(filename)

    client = app.test_client()
    resp = client.get('/timetables')
    assert resp.status_code == 200

    # Re-open the file and check that Classroom_Allocation sheet now exists
    wb2 = openpyxl.load_workbook(filename)
    assert 'Classroom_Allocation' in wb2.sheetnames
    assert 'Classroom_Utilization' in wb2.sheetnames

    # Section_A should have bracketed room values OR corresponding entry in Classroom_Allocation
    ws_a = wb2['Section_A']
    found_room = False
    for row in ws_a.iter_rows(min_row=2, values_only=True):
        for cell in row:
            if isinstance(cell, str) and '[' in cell and ']' in cell:
                found_room = True
                break
        if found_room:
            break
    if not found_room:
        # Fall back: check that Classroom_Allocation sheet contains an entry for CS101
        allocation_df = pd.read_excel(filename, sheet_name='Classroom_Allocation')
        assert any(allocation_df['Course'] == 'CS101'), "Expected allocation for CS101 either in-sheet or in Classroom_Allocation"
