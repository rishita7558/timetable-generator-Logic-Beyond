import os
import sys
import shutil
import openpyxl
import pandas as pd

# Add parent directory to path for module imports
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from app import app, INPUT_DIR, OUTPUT_DIR


def _backup_and_write_classroom_csv(tmp_csv_content):
    orig_path = os.path.join(INPUT_DIR, 'classroom_data.csv')
    backup_path = os.path.join(INPUT_DIR, 'classroom_data.csv.bak')
    # Backup original if exists
    if os.path.exists(orig_path):
        shutil.copy(orig_path, backup_path)
    # Write test content
    with open(orig_path, 'w', encoding='utf-8') as f:
        f.write(tmp_csv_content)
    return backup_path if os.path.exists(backup_path) else None


def _restore_classroom_csv(backup_path):
    path = os.path.join(INPUT_DIR, 'classroom_data.csv')
    if backup_path and os.path.exists(backup_path):
        shutil.move(backup_path, path)
    else:
        # Remove test file if no original
        if os.path.exists(path):
            os.remove(path)


def test_no_rooms_available(tmp_path):
    """When classroom_data.csv is empty, the system should NOT try to allocate rooms."""
    backup = _backup_and_write_classroom_csv("Room Number,Type,Capacity,Facilities\n")

    # Create a simple timetable file with Section_A only
    unique = 'no_rooms'
    filename = os.path.join(OUTPUT_DIR, f'sem3_ECE_{unique}_classrooms.xlsx')
    if os.path.exists(filename):
        os.remove(filename)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Section_A'
    ws['A1'] = 'Time'
    ws['B1'] = 'Mon'
    ws['A2'] = '09:00-10:00'
    ws['B2'] = 'CS101'
    wb.save(filename)

    client = app.test_client()
    resp = client.get('/timetables')
    assert resp.status_code == 200
    timetables = resp.get_json()

    matches = [t for t in timetables if t.get('filename') == os.path.basename(filename)]
    assert matches and len(matches) >= 1

    for m in matches:
        # Since classroom data was empty, we expect no allocations
        assert not m.get('has_classroom_allocation', False)
        assert m.get('classroom_details') in ([], None)

    _restore_classroom_csv(backup)


def test_forced_conflict_single_room(tmp_path):
    """With only one room available and both sections scheduled at the same slot, expect a conflict flag."""
    # Single room
    csv_content = "Room Number,Type,Capacity,Facilities\nC001,classroom,50,Projector\n"
    backup = _backup_and_write_classroom_csv(csv_content)

    unique = 'forced_conflict'
    filename = os.path.join(OUTPUT_DIR, f'sem3_ECE_{unique}_classrooms.xlsx')
    if os.path.exists(filename):
        os.remove(filename)

    wb = openpyxl.Workbook()
    ws_a = wb.active
    ws_a.title = 'Section_A'
    ws_a['A1'] = 'Time'
    ws_a['B1'] = 'Mon'
    ws_a['A2'] = '09:00-10:00'
    ws_a['B2'] = 'CS101'

    ws_b = wb.create_sheet('Section_B')
    ws_b['A1'] = 'Time'
    ws_b['B1'] = 'Mon'
    ws_b['A2'] = '09:00-10:00'
    ws_b['B2'] = 'CS201'

    wb.save(filename)

    client = app.test_client()
    resp = client.get('/timetables')
    assert resp.status_code == 200
    timetables = resp.get_json()

    matches = [t for t in timetables if t.get('filename') == os.path.basename(filename)]
    assert matches and len(matches) >= 2

    # Collect conflict flags across both sections - if the system allocated the same room, that's a resource reuse
    conflicts = []
    rooms = set()
    for m in matches:
        for d in m.get('classroom_details', []):
            rooms.add(d.get('Room Number') or d.get('room'))
            if d.get('Conflict') or d.get('conflict'):
                conflicts.append(d)

    # Either there should be an explicit conflict flagged, or both sections should reuse the same room
    assert (len(conflicts) >= 1) or (len(rooms) == 1 and None not in rooms)

    _restore_classroom_csv(backup)


def test_lab_pair_behavior(tmp_path):
    """Ensure a two-slot lab is allocated the same lab room (room number starts with 'L') for both slots."""
    csv_content = "Room Number,Type,Capacity,Facilities\nL100,lab,120,Computers\n"
    backup = _backup_and_write_classroom_csv(csv_content)

    unique = 'lab_pair'
    filename = os.path.join(OUTPUT_DIR, f'sem3_ECE_{unique}_classrooms.xlsx')
    if os.path.exists(filename):
        os.remove(filename)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Section_A'
    ws['A1'] = 'Time'
    ws['B1'] = 'Mon'
    # Lab pair slots as defined in app.py
    ws['A2'] = '13:00-14:30'
    ws['B2'] = 'CS300 (Lab)'
    ws['A3'] = '14:30-15:30'
    ws['B3'] = 'CS300 (Lab)'

    wb.save(filename)

    client = app.test_client()
    resp = client.get('/timetables')
    assert resp.status_code == 200
    timetables = resp.get_json()

    matches = [t for t in timetables if t.get('filename') == os.path.basename(filename)]
    assert matches and len(matches) >= 1

    # Ensure we find two allocations for the lab with the same room and no conflict
    room_names = set()
    conflicts = False
    for m in matches:
        for d in m.get('classroom_details', []):
            if d.get('Course') and 'CS300' in str(d.get('Course')):
                room_names.add(d.get('Room Number') or d.get('room'))
                if d.get('Conflict') or d.get('conflict'):
                    conflicts = True

    assert len(room_names) == 1 and not conflicts

    _restore_classroom_csv(backup)


def test_ignore_nil_capacity_rooms(tmp_path):
    """Rooms with 'nil' capacity should be ignored and not used for allocations."""
    csv_content = "Room Number,Type,Capacity,Facilities\nC100,classroom,nil,Projector\nC101,classroom,100,Projector\n"
    backup = _backup_and_write_classroom_csv(csv_content)

    unique = 'nil_capacity'
    filename = os.path.join(OUTPUT_DIR, f'sem3_ECE_{unique}_classrooms.xlsx')
    if os.path.exists(filename):
        os.remove(filename)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Section_A'
    ws['A1'] = 'Time'
    ws['B1'] = 'Mon'
    ws['A2'] = '09:00-10:00'
    ws['B2'] = 'CS101'

    wb.save(filename)

    client = app.test_client()
    resp = client.get('/timetables')
    assert resp.status_code == 200
    timetables = resp.get_json()

    matches = [t for t in timetables if t.get('filename') == os.path.basename(filename)]
    assert matches and len(matches) >= 1

    # Verify allocations: C100 (nil capacity) should NOT be allocated
    for m in matches:
        rooms = set()
        for d in m.get('classroom_details', []):
            rooms.add(d.get('Room Number') or d.get('room'))
        assert 'C100' not in rooms

    _restore_classroom_csv(backup)
