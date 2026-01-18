import os
import openpyxl
import uuid

from backend.app import app, OUTPUT_DIR


def make_workbook_with_sheets(path, sheet_names):
    wb = openpyxl.Workbook()
    default = wb.active
    if default.title not in sheet_names:
        wb.remove(default)
    for idx, name in enumerate(sheet_names):
        if name in wb.sheetnames:
            continue
        ws = wb.create_sheet(name)
        ws['A1'] = 'Time'
        ws['B1'] = 'Mon'
        ws['A2'] = '09:00-10:00'
        ws['B2'] = f'CS{100+idx} [R-1]'

    if not wb.sheetnames:
        wb.create_sheet('Timetable')
    wb.save(path)


def test_sem3_excludes_b4_b5(tmp_path):
    unique = uuid.uuid4().hex[:8]
    filename = os.path.join(OUTPUT_DIR, f'sem3_ECE_{unique}_baskets.xlsx')
    if os.path.exists(filename):
        os.remove(filename)

    # Create timetable and basket allocation containing B3, B4, B5
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Timetable'
    ws['A1'] = 'Time'
    ws['B1'] = 'Mon'
    ws['A2'] = '09:00-10:00'
    ws['B2'] = 'ELECTIVE_B3'

    basket_ws = wb.create_sheet('Basket_Allocation')
    basket_ws.append(['Basket Name', 'Courses in Basket', 'Day', 'Time Slot'])
    basket_ws.append(['ELECTIVE_B3', 'EC201, EC202', 'Mon', '09:00-10:00'])
    basket_ws.append(['ELECTIVE_B4', 'EC301, EC302', 'Tue', '10:00-11:00'])
    basket_ws.append(['ELECTIVE_B5', 'EC401, EC402', 'Wed', '11:00-12:00'])

    wb.save(filename)

    client = app.test_client()
    resp = client.get('/timetables')
    assert resp.status_code == 200
    timetables = resp.get_json()

    entries = [t for t in timetables if t.get('filename') == os.path.basename(filename)]
    assert len(entries) == 1
    t = entries[0]
    # B4 and B5 must not be present for semester 3
    assert all('ELECTIVE_B4' not in b for b in t.get('baskets', []))
    assert all('ELECTIVE_B5' not in b for b in t.get('baskets', []))
    # B3 should be present
    assert any('ELECTIVE_B3' == b for b in t.get('baskets', [])) or any('ELECTIVE_B3' in lst for lst in t.get('basket_courses_map', {}).keys())


def test_sem5_excludes_b3(tmp_path):
    unique = uuid.uuid4().hex[:8]
    filename = os.path.join(OUTPUT_DIR, f'sem5_CSE_{unique}_baskets.xlsx')
    if os.path.exists(filename):
        os.remove(filename)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Section_A'
    ws['A1'] = 'Time'
    ws['B1'] = 'Mon'
    ws['A2'] = '09:00-10:00'
    ws['B2'] = 'ELECTIVE_B3'

    basket_ws = wb.create_sheet('Basket_Allocation')
    basket_ws.append(['Basket Name', 'Courses in Basket', 'Day', 'Time Slot'])
    basket_ws.append(['ELECTIVE_B3', 'CS201, CS202', 'Mon', '09:00-10:00'])
    basket_ws.append(['ELECTIVE_B4', 'CS301, CS302', 'Tue', '10:00-11:00'])
    basket_ws.append(['ELECTIVE_B5', 'CS401, CS402', 'Wed', '11:00-12:00'])

    wb.save(filename)

    client = app.test_client()
    resp = client.get('/timetables')
    assert resp.status_code == 200
    timetables = resp.get_json()

    entries = [t for t in timetables if t.get('filename') == os.path.basename(filename)]
    assert len(entries) >= 1
    # For all entries belonging to this file, none should include ELECTIVE_B3 in baskets
    for t in entries:
        assert all('ELECTIVE_B3' not in b for b in t.get('baskets', []))
        assert ('ELECTIVE_B3' not in t.get('basket_courses_map', {}))


def test_sem1_allows_only_b1_and_sem7_allows_b6_b7(tmp_path):
    # Sem 1: only ELECTIVE_B1 allowed
    unique1 = uuid.uuid4().hex[:8]
    filename1 = os.path.join(OUTPUT_DIR, f'sem1_ECE_{unique1}_baskets.xlsx')
    if os.path.exists(filename1):
        os.remove(filename1)

    wb1 = openpyxl.Workbook()
    ws1 = wb1.active
    ws1.title = 'Timetable'
    ws1['A1'] = 'Time'
    ws1['B1'] = 'Mon'
    ws1['A2'] = '09:00-10:00'
    ws1['B2'] = 'ELECTIVE_B1'

    basket_ws1 = wb1.create_sheet('Basket_Allocation')
    basket_ws1.append(['Basket Name', 'Courses in Basket', 'Day', 'Time Slot'])
    basket_ws1.append(['ELECTIVE_B1', 'EC101, EC102', 'Mon', '09:00-10:00'])
    basket_ws1.append(['ELECTIVE_B3', 'EC201, EC202', 'Tue', '10:00-11:00'])

    wb1.save(filename1)

    client = app.test_client()
    resp = client.get('/timetables')
    assert resp.status_code == 200
    timetables1 = resp.get_json()

    entries1 = [t for t in timetables1 if t.get('filename') == os.path.basename(filename1)]
    assert len(entries1) == 1
    t1 = entries1[0]
    # Should only include B1 and not B3
    assert any('ELECTIVE_B1' == b for b in t1.get('baskets', [])) or ('ELECTIVE_B1' in t1.get('basket_courses_map', {}))
    assert all('ELECTIVE_B3' not in b for b in t1.get('baskets', []))
    assert ('ELECTIVE_B3' not in t1.get('basket_courses_map', {}))

    # Sem 7: only ELECTIVE_B6 and ELECTIVE_B7 allowed
    unique7 = uuid.uuid4().hex[:8]
    filename7 = os.path.join(OUTPUT_DIR, f'sem7_CSE_{unique7}_baskets.xlsx')
    if os.path.exists(filename7):
        os.remove(filename7)

    wb7 = openpyxl.Workbook()
    ws7 = wb7.active
    ws7.title = 'Section_A'
    ws7['A1'] = 'Time'
    ws7['B1'] = 'Mon'
    ws7['A2'] = '09:00-10:00'
    ws7['B2'] = 'ELECTIVE_B6'

    basket_ws7 = wb7.create_sheet('Basket_Allocation')
    basket_ws7.append(['Basket Name', 'Courses in Basket', 'Day', 'Time Slot'])
    basket_ws7.append(['ELECTIVE_B6', 'CS601, CS602', 'Mon', '09:00-10:00'])
    basket_ws7.append(['ELECTIVE_B3', 'CS201, CS202', 'Tue', '10:00-11:00'])

    wb7.save(filename7)

    resp = client.get('/timetables')
    assert resp.status_code == 200
    timetables7 = resp.get_json()

    entries7 = [t for t in timetables7 if t.get('filename') == os.path.basename(filename7)]
    assert len(entries7) >= 1
    for t7 in entries7:
        assert any('ELECTIVE_B6' == b for b in t7.get('baskets', [])) or ('ELECTIVE_B6' in t7.get('basket_courses_map', {}))
        assert all('ELECTIVE_B3' not in b for b in t7.get('baskets', []))
        assert ('ELECTIVE_B3' not in t7.get('basket_courses_map', {}))
