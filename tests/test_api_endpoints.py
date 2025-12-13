import os
import openpyxl
import pytest

from backend.app import app, OUTPUT_DIR, create_basket_courses_sheet


def make_workbook_with_sheets(path, sheet_names):
    wb = openpyxl.Workbook()
    # Remove the default sheet if we will provide different names
    default_sheet = wb.active
    if default_sheet.title not in sheet_names:
        wb.remove(default_sheet)

    for name in sheet_names:
        if name in wb.sheetnames:
            continue
        wb.create_sheet(name)

    # Ensure at least one sheet persists
    if not wb.sheetnames:
        wb.create_sheet('Timetable')

    wb.save(path)


def test_create_basket_courses_sheet_none_tutorial():
    basket_allocations = {
        'Basket1': {
            'lectures': [('Mon', '09:00-10:30')],
            'tutorial': None,
            'courses': ['CSE101']
        }
    }
    df = create_basket_courses_sheet(basket_allocations)
    assert not df.empty
    assert df.iloc[0]['Tutorial Slot'] == '-'


def test_timetables_and_stats_endpoints(tmp_path):
    ece_filename = os.path.join(OUTPUT_DIR, 'sem3_ECE_timetable_baskets.xlsx')
    cse_filename = os.path.join(OUTPUT_DIR, 'sem3_CSE_timetable_baskets.xlsx')

    # Cleanup if exist
    for f in [ece_filename, cse_filename]:
        if os.path.exists(f):
            os.remove(f)

    # Create ECE workbook with 'Timetable' sheet (Whole)
    make_workbook_with_sheets(ece_filename, ['Timetable'])
    # Create CSE workbook with both Section_A and Section_B
    make_workbook_with_sheets(cse_filename, ['Section_A', 'Section_B'])

    # Populate temp_inputs with test CSVs so load_all_data() works
    test_data_dir = os.path.join(os.getcwd(), 'backend', 'test_data')
    temp_inputs_dir = os.path.join(os.getcwd(), 'temp_inputs')
    os.makedirs(temp_inputs_dir, exist_ok=True)
    for csv_file in os.listdir(test_data_dir):
        if csv_file.lower().endswith('.csv'):
            src = os.path.join(test_data_dir, csv_file)
            dst = os.path.join(temp_inputs_dir, csv_file)
            with open(src, 'rb') as sf, open(dst, 'wb') as df:
                df.write(sf.read())

    client = app.test_client()
    resp = client.get('/timetables')
    assert resp.status_code == 200
    timetables = resp.get_json()
    ece_entries = [t for t in timetables if t.get('filename') == os.path.basename(ece_filename)]
    assert len(ece_entries) == 1
    assert ece_entries[0]['section'] == 'Whole'

    cse_entries = [t for t in timetables if t.get('filename') == os.path.basename(cse_filename)]
    assert len(cse_entries) >= 1
    assert any(e['section'] in ['A', 'B'] for e in cse_entries)

    resp = client.get('/stats')
    assert resp.status_code == 200
    stats = resp.get_json()
    assert stats['total_timetables'] >= 3


def test_pre_and_post_mid_generate_statistics_sheets(tmp_path):
    # Use dataframes loaded from CSVs
    from backend.app import load_all_data, export_mid_semester_timetables
    dfs = load_all_data(force_reload=True)

    # Run mid semester export for ECE (Whole) and CSE (Sec A/B)
    res_ece = export_mid_semester_timetables(dfs, 3, 'ECE')
    res_cse = export_mid_semester_timetables(dfs, 3, 'CSE')

    # Validate file creation and new sheets
    import openpyxl

    if res_ece.get('pre_mid_success') and res_ece.get('pre_mid_filename'):
        ece_pre = os.path.join(OUTPUT_DIR, res_ece['pre_mid_filename'])
        assert os.path.exists(ece_pre)
        wb = openpyxl.load_workbook(ece_pre, read_only=True)
        assert 'Classroom_Utilization' in wb.sheetnames
        assert 'Classroom_Allocation' in wb.sheetnames
        # Basket sheets may be present if electives exist
        # Assert presence if basket allocations were applied (best-effort)
        assert 'Course_Summary' in wb.sheetnames
        wb.close()

    if res_ece.get('post_mid_success') and res_ece.get('post_mid_filename'):
        ece_post = os.path.join(OUTPUT_DIR, res_ece['post_mid_filename'])
        assert os.path.exists(ece_post)
        wb = openpyxl.load_workbook(ece_post, read_only=True)
        assert 'Classroom_Utilization' in wb.sheetnames
        assert 'Classroom_Allocation' in wb.sheetnames
        assert 'Course_Summary' in wb.sheetnames
        wb.close()

    if res_cse.get('pre_mid_success') and res_cse.get('pre_mid_filename'):
        cse_pre = os.path.join(OUTPUT_DIR, res_cse['pre_mid_filename'])
        assert os.path.exists(cse_pre)
        wb = openpyxl.load_workbook(cse_pre, read_only=True)
        assert 'Classroom_Utilization' in wb.sheetnames
        assert 'Classroom_Allocation' in wb.sheetnames
        assert 'Course_Summary' in wb.sheetnames
        wb.close()

    if res_cse.get('post_mid_success') and res_cse.get('post_mid_filename'):
        cse_post = os.path.join(OUTPUT_DIR, res_cse['post_mid_filename'])
        assert os.path.exists(cse_post)
        wb = openpyxl.load_workbook(cse_post, read_only=True)
        assert 'Classroom_Utilization' in wb.sheetnames
        assert 'Classroom_Allocation' in wb.sheetnames
        assert 'Course_Summary' in wb.sheetnames
        wb.close()
